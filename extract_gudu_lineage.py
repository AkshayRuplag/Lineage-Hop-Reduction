"""
Extract column-level lineage from Gudu SQLFlow JSON output.

Produces a CSV with columns:
  Target Table, Target Column, Source Table, Source Column, Transformation
"""

import json
import csv
import sys
import os
import re
from collections import defaultdict


def build_id_map(dbobjs):
    """Build a map from object ID -> {type, name, parent_name, parent_id}."""
    id_map = {}

    for srv in dbobjs.get("servers", []):
        for db in srv.get("databases", []):
            for schema in db.get("schemas", []):
                # Tables
                for tbl in schema.get("tables", []):
                    tname = tbl["name"]
                    tid = tbl["id"]
                    id_map[tid] = {
                        "type": "table",
                        "name": tname,
                        "parent_name": None,
                    }
                    for col in tbl.get("columns", []):
                        id_map[col["id"]] = {
                            "type": "column",
                            "name": col["name"],
                            "parent_name": tname,
                            "parent_id": tid,
                        }

                # Views
                for vw in schema.get("views", []):
                    vname = vw["name"]
                    vid = vw["id"]
                    id_map[vid] = {
                        "type": "view",
                        "name": vname,
                        "parent_name": None,
                    }
                    for col in vw.get("columns", []):
                        id_map[col["id"]] = {
                            "type": "column",
                            "name": col["name"],
                            "parent_name": vname,
                            "parent_id": vid,
                        }

                # Others (functions, expressions, result-sets)
                for other in schema.get("others", []):
                    oname = other["name"]
                    oid = other["id"]
                    otype = other.get("type", "")
                    id_map[oid] = {
                        "type": "other",
                        "subtype": otype,
                        "name": oname,
                        "parent_name": None,
                    }
                    for col in other.get("columns", []):
                        id_map[col["id"]] = {
                            "type": "other_col",
                            "name": col["name"],
                            "parent_name": oname,
                            "parent_id": oid,
                        }

                # Procedures
                for proc in schema.get("procedures", []):
                    pname = proc["name"]
                    pid = proc["id"]
                    id_map[pid] = {
                        "type": "procedure",
                        "name": pname,
                        "parent_name": None,
                    }
                    for arg in proc.get("arguments", []):
                        id_map[arg["id"]] = {
                            "type": "argument",
                            "name": arg["name"],
                            "parent_name": pname,
                            "parent_id": pid,
                        }

    return id_map


def is_real_table_column(id_map, obj_id):
    """Check if an ID corresponds to a real table/view column."""
    entry = id_map.get(obj_id)
    if not entry:
        return False
    if entry["type"] == "column":
        return True
    return False


def resolve_name(id_map, obj_id, fallback_col=None, fallback_parent=None):
    """Resolve an ID to (parent_name, column_name)."""
    entry = id_map.get(obj_id)
    if entry:
        return entry.get("parent_name", fallback_parent), entry.get("name", fallback_col)
    return fallback_parent, fallback_col


def extract_lineage(json_path, sql_object_name):
    """Extract lineage rows from a single Gudu SQLFlow JSON file.
    Supports both the full API format (data.sqlflow wrapper) and the
    Gudu Lite / dlineage.py format (flat dbobjs/relationships at root).
    Returns a list of dicts with SQL Object Name included."""
    data = None
    # Try loading the entire file first (handles multi-line pretty-printed JSON
    # from Gudu Enterprise, as well as single-line Gudu Lite format).
    with open(json_path, "r", encoding="utf-8", errors="replace") as f:
        raw = f.read()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        # Fall back to line-by-line scan (some Lite outputs embed log text
        # before/after the JSON blob).
        for line in raw.splitlines():
            line = line.strip()
            if line.startswith('{'):
                try:
                    data = json.loads(line)
                    break
                except json.JSONDecodeError:
                    continue
    if data is None:
        raise ValueError(f"No valid JSON found in {json_path}")

    # Support both API wrapper format and flat Lite format
    if "data" in data and "sqlflow" in data.get("data", {}):
        sqlflow = data["data"]["sqlflow"]
        dbobjs = sqlflow["dbobjs"]
        relationships = sqlflow["relationships"]
    else:
        dbobjs = data["dbobjs"]
        relationships = data.get("relationships", [])

    id_map = build_id_map(dbobjs)

    # ── Step 1: Build adjacency from relationships ──
    # target_id -> list of (source_id, rel_type, effect_type)
    # source_id -> list of (target_id, rel_type, effect_type)
    fwd = defaultdict(list)  # source -> targets
    bwd = defaultdict(list)  # target -> sources

    for rel in relationships:
        rel_type = rel.get("type", "")
        # Skip 'call' relationships - they use caller/callees, not target/sources
        # Skip 'fdr' (functional dependency reference) — these are GROUP BY / WHERE
        # clause dependencies, NOT actual data-flow lineage. Following fdr edges
        # causes GROUP BY columns to appear as false sources for aggregate targets.
        if rel_type in ("call", "fdr") or "target" not in rel:
            continue
        effect_type = rel.get("effectType", "")
        target_id = rel["target"]["id"]
        target_col = rel["target"].get("column", "")
        target_parent = rel["target"].get("parentName", "")

        for src in rel.get("sources", []):
            src_id = src["id"]
            src_col = src.get("column", "")
            src_parent = src.get("parentName", "")

            fwd[src_id].append({
                "target_id": target_id,
                "rel_type": rel_type,
                "effect_type": effect_type,
                "target_col": target_col,
                "target_parent": target_parent,
            })
            bwd[target_id].append({
                "source_id": src_id,
                "rel_type": rel_type,
                "effect_type": effect_type,
                "source_col": src_col,
                "source_parent": src_parent,
            })

    # Patterns for internal SQLFlow construct names (not real transformations)
    _INTERNAL_PATTERN = re.compile(
        r'^(INSERT-SELECT-?\d*|SELECT-INTO-?\d*|MERGE-INSERT-?\d*'
        r'|MERGE-UPDATE-?\d*|UPDATE-SET-?\d*'
        r'|RS-\d+|RESULT_OF_\S*|RelationRows'
        r'|CTE-\S*|SQL_CONSTANTS-?\d*'
        r'|pseudo_table_include_orphan_column'
        r'|RAISE_APPLICATION_ERROR|GATHER_TABLE_STATS'
        r'|SEQ_\S+|I'
        r'|prc_force_indexes_unusable|prc_rebuild_indexes'
        r'|prc_set_global_idx_to_no_parallel)$',
        re.IGNORECASE
    )

    def _is_internal_name(name):
        return bool(_INTERNAL_PATTERN.match(name))

    # Match simple single-function expressions: sum(col), sum( col ), "sum(col)"
    _FUNC_EXPR_RE = re.compile(r'^"?\s*\w+\s*\(\s*(?:\w+\.)?\s*(\w+)\s*\)\s*"?$', re.IGNORECASE)

    def _normalize_col_name(name):
        """Strip simple aggregate/function wrappers from expression-based column names.

        Gudu uses the expression itself as a column name when the MV SELECT has no alias:
          sum(n_paid_amount_r)    -> n_paid_amount_r
          "sum( n_paid_amount_r)" -> n_paid_amount_r
        Only strips single-function, single-argument expressions to avoid false matches.
        """
        if not name:
            return name
        m = _FUNC_EXPR_RE.match(name.strip())
        return m.group(1) if m else name

    # ── Step 2: For each real-table target column, trace back to find
    #            real-table source columns, collecting transformations ──
    lineage_rows = []
    seen = set()

    # Reverse lookup: (parentName_upper, columnName_upper) -> id
    # Used to resolve wildcard (*) source edges by column name.
    _name_to_id = {}
    for _nid, _info in id_map.items():
        _pn = (_info.get("parent_name") or "")
        _cn = (_info.get("name") or "")
        if _cn and _cn != "*":
            _name_to_id[(_pn.upper(), _cn.upper())] = _nid

    def trace_sources(node_id, visited=None, path=None, col_hint=None):
        """Trace backwards from a node to find real table/view column sources.
        Returns list of (source_table, source_col, transformation_parts).

        col_hint: the column name we are currently looking for.  Used to
        resolve wildcard (*) source edges by name instead of expanding all
        columns of the wildcard node (which would cause a cartesian explosion).
        """
        if visited is None:
            visited = set()
        if path is None:
            path = []
        if node_id in visited:
            return []
        visited.add(node_id)

        results = []
        for edge in bwd.get(node_id, []):
            src_id = edge["source_id"]
            src_col = edge.get("source_col", "")
            src_parent = edge.get("source_parent", "")

            # ── Wildcard (*) source: Gudu emits this when a subquery uses
            # SELECT * and it cannot resolve individual columns.  Following
            # all 84+ parents of the wildcard node causes a cartesian
            # explosion.  Instead, look for the specific column (col_hint)
            # in the same parent namespace. ──
            if src_col == "*":
                hint = col_hint  # column we are currently tracing
                if hint:
                    specific_id = _name_to_id.get(((src_parent or "").upper(), hint.upper()))
                    if specific_id and specific_id not in visited:
                        deeper = trace_sources(specific_id, visited.copy(), path, hint)
                        results.extend(deeper)
                # Skip expanding the full wildcard regardless of hint
                continue

            if is_real_table_column(id_map, src_id):
                src_table, src_col_r = resolve_name(id_map, src_id,
                                                    src_col,
                                                    src_parent)
                results.append((src_table, src_col_r, list(path)))
            else:
                # Intermediate node (function, expression, etc.)
                node_entry = id_map.get(src_id)
                func_name = ""
                if node_entry:
                    if node_entry["type"] in ("other", "other_col"):
                        func_name = node_entry.get("parent_name", node_entry["name"])
                        if node_entry["type"] == "other":
                            func_name = node_entry["name"]
                else:
                    func_name = src_col

                new_path = path + [func_name] if func_name else path
                # Propagate col_hint: use the source column name when available
                next_hint = src_col if src_col and src_col != "*" else col_hint
                deeper = trace_sources(src_id, visited.copy(), new_path, next_hint)
                results.extend(deeper)

        return results

    # Collect all real table column IDs that appear as targets
    target_col_ids = set()
    for rel in relationships:
        if "target" not in rel:
            continue
        tid = rel["target"]["id"]
        if is_real_table_column(id_map, tid):
            target_col_ids.add(tid)

    # Pseudo-columns to skip when they appear as source columns
    _PSEUDO_SOURCE_COLS = {"RelationRows", "relationrows", "*"}

    for tgt_id in sorted(target_col_ids):
        tgt_table, tgt_col = resolve_name(id_map, tgt_id)

        # Normalize expression-based column names (e.g. sum(col) -> col)
        tgt_col = _normalize_col_name(tgt_col)

        # Skip internal pseudo-columns
        if _is_internal_name(tgt_col or ""):
            continue

        # Skip backup snapshot tables — _BKP tables are intermediate copies,
        # not lineage targets for the RPT layer.
        if tgt_table and tgt_table.upper().endswith('_BKP'):
            continue

        sources = trace_sources(tgt_id, col_hint=tgt_col)

        if sources:
            for src_table, src_col, transforms in sources:
                # Skip pseudo source columns (RelationRows, *)
                if src_col in _PSEUDO_SOURCE_COLS:
                    continue

                # Deduplicate
                key = (tgt_table, tgt_col, src_table, src_col)
                if key in seen:
                    continue
                seen.add(key)

                # Build transformation string - filter out internal names
                transform_parts = [t for t in transforms
                                   if t and not _is_internal_name(t)]
                # Remove duplicates preserving order
                seen_t = set()
                unique_transforms = []
                for t in transform_parts:
                    if t.upper() not in seen_t:
                        seen_t.add(t.upper())
                        unique_transforms.append(t)

                if unique_transforms:
                    transformation = " -> ".join(unique_transforms)
                elif src_col and tgt_col and src_col.upper() == tgt_col.upper():
                    transformation = "Direct"
                else:
                    transformation = "Direct"

                lineage_rows.append({
                    "SQL Object Name": sql_object_name,
                    "Target Table": tgt_table,
                    "Target Column": tgt_col,
                    "Source Table": src_table,
                    "Source Column": src_col,
                    "Transformation": transformation,
                })
        else:
            # No traceable source - might be a constant or expression
            lineage_rows.append({
                "SQL Object Name": sql_object_name,
                "Target Table": tgt_table,
                "Target Column": tgt_col,
                "Source Table": "",
                "Source Column": "",
                "Transformation": "Constant / Expression",
            })

    return lineage_rows


def extract_object_name(filename):
    """Extract SQL object name from various filename patterns.

    Supported patterns:
      mssql_<OBJECT_NAME>_gudu.json     -> <OBJECT_NAME>
      PKG_..._PRC_..._cleaned_lineage.json -> PKG_..._PRC_...
      PKG_..._PRC_..._partN_cleaned_lineage.json -> PKG_..._PRC_..._partN
    """
    # Original mssql pattern
    match = re.match(r'^mssql_(.+)_gudu\.json$', filename, re.IGNORECASE)
    if match:
        return match.group(1)
    # New Gudu Lite pattern: strip _cleaned_lineage.json or _lineage.json suffixes
    name = filename
    for suffix in ('_cleaned_lineage.json', '_lineage.json'):
        if name.lower().endswith(suffix):
            name = name[:-len(suffix)]
            break
    else:
        name = os.path.splitext(name)[0]
    return name


def extract_pkg_proc(sql_object_name):
    """Split a SQL object name into (package, procedure).

    e.g. PKG_GRP_LOAD_RPT_AGENT_POLICY_R_PRC_UPD_DEL_DATA_part1
         -> ('PKG_GRP_LOAD_RPT_AGENT_POLICY_R', 'PRC_UPD_DEL_DATA_part1')

    Non-PKG names (PRC_, PROC_, MV names, etc.) have no package wrapper:
         -> ('', 'PRC_GRP_LOAD_...')
    """
    # Only treat as PKG+proc if the name starts with PKG_
    # Standalone procs (PRC_/PROC_) and MVs/Views have no package wrapper.
    if not sql_object_name.upper().startswith('PKG_'):
        return '', sql_object_name

    idx = sql_object_name.rfind('_PRC_')
    if idx != -1:
        return sql_object_name[:idx], sql_object_name[idx + 1:]
    idx = sql_object_name.rfind('_MAIN')
    if idx != -1:
        return sql_object_name[:idx], sql_object_name[idx + 1:]
    # PKG_ prefix but no explicit proc suffix -> this is the package's MAIN entry
    return sql_object_name, 'MAIN'


def detect_object_type(sql_object_name):
    """Detect object type from filename prefix/suffix.
    Used as a fallback when the actual cleaned SQL file is not available.
    Returns one of: 'PKG Procedure', 'Standalone Procedure', 'Materialized View', 'View'
    """
    upper = sql_object_name.upper()
    # Strip _partN and _union_N suffixes before classifying
    base = re.sub(r'(_PART\d+|_UNION_\d+)$', '', upper)
    if base.startswith('PKG_'):
        return 'PKG Procedure'
    if base.startswith('PRC_') or base.startswith('PROC_'):
        return 'Standalone Procedure'
    # _union_N suffix is only produced for plain VIEWs (never MViews) by the cleaning pipeline
    if re.search(r'_UNION_\d+$', upper):
        return 'View'
    # Objects ending in _VW or _VW_N are plain Views (e.g. DIM_X_VW, DIM_X_VW_2)
    if re.search(r'_VW(_\d+)?$', base):
        return 'View'
    if any(base.startswith(p) for p in ('DIM_', 'FCT_', 'MVW_', 'RPT_')):
        return 'Materialized View'
    if base.startswith('VW_'):
        return 'View'
    return 'SQL Object'


def _detect_view_type_from_sql_file(sql_path: str) -> str:
    """Read the actual cleaned SQL file and return its DDL-based object type.

    Returns 'Materialized View', 'View', or '' (empty) when the file does not
    exist, is unreadable, or starts with a non-view statement (procedure, etc.).

    This is the authoritative source for view vs materialized-view classification
    because filenames like VW_* can contain CREATE MATERIALIZED VIEW and vice versa.
    """
    if not os.path.exists(sql_path):
        return ''
    try:
        with open(sql_path, 'r', encoding='utf-8', errors='replace') as f:
            text = f.read(1200)  # only need the first CREATE line
        for line in text.splitlines():
            stripped = line.strip()
            if not stripped or stripped.startswith('--'):
                continue
            upper = stripped.upper()
            if re.match(r'CREATE\s+MATERIALIZED\s+VIEW\b', upper):
                return 'Materialized View'
            if re.match(
                r'CREATE\s+(?:OR\s+REPLACE\s+)?'
                r'(?:(?:FORCE|NOFORCE)\s+)?(?:(?:EDITIONABLE|NONEDITIONABLE)\s+)?VIEW\b',
                upper
            ):
                return 'View'
            # First non-comment line is not a view DDL — let caller fall back to filename
            return ''
    except Exception:
        pass
    return ''


def _is_lineage_json(filename):
    """Return True if the file is a parseable Gudu lineage JSON.

    Accepted:
      *_cleaned_lineage.json  — Gudu SQLFlow Lite per-part output (valid JSON)
      *_lineage.json          — Stitched output from stitch_gudu_part_lineages()
                                (clean JSON written by our own code, no Error log)
      mssql_*_gudu.json       — Gudu Enterprise
    """
    fl = filename.lower()
    if not fl.endswith('.json'):
        return False
    # Gudu Lite per-part output
    if fl.endswith('_cleaned_lineage.json'):
        return True
    # Stitched output (produced by stitch_gudu_part_lineages — always clean JSON)
    if fl.endswith('_lineage.json'):
        return True
    # Gudu Enterprise: mssql_*_gudu.json
    if re.match(r'^mssql_.+_gudu\.json$', fl):
        return True
    return False


def collect_rows_from_dir(input_dir, pipeline='pkg'):
    """Collect lineage rows from all lineage JSON files in one directory.

    Accepts both Gudu SQLFlow Lite (*_lineage.json) and
    Gudu Enterprise (mssql_*_gudu.json) JSON files.

    pipeline='pkg'    -> split name into Package + Procedure via extract_pkg_proc()
    pipeline='nonpkg' -> Package='', Procedure=full object name, type auto-detected
    """
    all_json = sorted(f for f in os.listdir(input_dir) if _is_lineage_json(f))
    if not all_json:
        print(f"  [SKIP] No lineage JSON files found in {input_dir}")
        return []

    # Build the set of stitched base names.
    # Stitched files end with _lineage.json but NOT _cleaned_lineage.json.
    # When a stitched file exists for a base (e.g. PKG_X_PROC_Y_lineage.json),
    # skip the individual part files (PKG_X_PROC_Y_partN_cleaned_lineage.json)
    # to avoid duplicate/partial rows in the output.
    _stitched_re = re.compile(r'^(.+?)_lineage\.json$', re.IGNORECASE)
    _part_file_re = re.compile(r'^(.+?)_part\d+_cleaned_lineage\.json$', re.IGNORECASE)
    stitched_bases = set()
    for f in all_json:
        if not f.lower().endswith('_cleaned_lineage.json'):
            m = _stitched_re.match(f)
            if m:
                stitched_bases.add(m.group(1).upper())

    json_files = []
    for f in all_json:
        m = _part_file_re.match(f)
        if m and m.group(1).upper() in stitched_bases:
            continue  # stitched version covers all parts — skip individual part file
        json_files.append(f)

    rows = []
    for filename in json_files:
        json_path = os.path.join(input_dir, filename)
        sql_object_name = extract_object_name(filename)
        obj_type = detect_object_type(sql_object_name)
        # For union branch files (FCT_X_union_1), the SQL Object Name in the CSV
        # should be the parent view name (FCT_X), not the branch-suffixed name.
        display_name = re.sub(r'_union_\d+$', '', sql_object_name, flags=re.IGNORECASE)

        # For view-like objects, override with content-based detection when the
        # corresponding cleaned SQL file is present.  This catches cases where the
        # filename convention is misleading, e.g. VW_* containing CREATE MATERIALIZED VIEW
        # or FCT_* containing CREATE VIEW (plain view, not an MView).
        # Procedure types (PKG Procedure / Standalone Procedure) are never overridden
        # because their cleaned SQL starts with INSERT/procedure DDL, not a view DDL.
        if obj_type not in ('PKG Procedure', 'Standalone Procedure') and filename.lower().endswith('_cleaned_lineage.json'):
            sql_path = os.path.join(input_dir, filename[:-len('_lineage.json')] + '.sql')
            content_type = _detect_view_type_from_sql_file(sql_path)
            if content_type:
                obj_type = content_type

        if sql_object_name.upper().startswith('PKG_'):
            # PKG objects (regardless of pipeline): split into Package + Procedure.
            pkg, proc = extract_pkg_proc(sql_object_name)
        elif obj_type in ('Standalone Procedure',):
            # Standalone procedures: Package is empty, Procedure holds the object name.
            pkg = ''
            proc = re.sub(r'_part\d+$', '', sql_object_name, flags=re.IGNORECASE)
        else:
            # Materialized Views, Views, and other non-procedure objects:
            # neither Package nor Procedure applies — SQL Object Name is sufficient.
            pkg = ''
            proc = ''

        print(f"  [{pipeline.upper()}] {filename}")
        try:
            file_rows = extract_lineage(json_path, display_name)
            for r in file_rows:
                r['Package']     = pkg
                r['Procedure']   = proc
                r['Object Type'] = obj_type
            rows.extend(file_rows)
            print(f"    {len(file_rows)} lineage rows")
        except Exception as e:
            print(f"    ERROR: {e}")
    return rows


def process_directory(input_dir, output_path, fmt='excel',
                      nonpkg_input_dir=None, gudu_input_dir=None):
    """Process Gudu SQLFlow JSON files from one, two, or three directories and write output.

    input_dir       : PKG split-procs cleaned dir (Lite format, always used)
    nonpkg_input_dir: Non-PKG cleaned dir (Lite format, PRC + MViews); optional
    gudu_input_dir  : Gudu Enterprise JSON dir (mssql_*_gudu.json); optional
    """
    print(f"\n[PKG Pipeline] {input_dir}")
    all_rows = collect_rows_from_dir(input_dir, pipeline='pkg')

    if (nonpkg_input_dir
            and os.path.abspath(nonpkg_input_dir) != os.path.abspath(input_dir)):
        if os.path.isdir(nonpkg_input_dir):
            print(f"\n[Non-PKG Pipeline] {nonpkg_input_dir}")
            all_rows.extend(collect_rows_from_dir(nonpkg_input_dir, pipeline='nonpkg'))
        else:
            print(f"  [WARN] Non-PKG dir not found: {nonpkg_input_dir}")

    if gudu_input_dir and os.path.isdir(gudu_input_dir):
        print(f"\n[Gudu Enterprise Pipeline] {gudu_input_dir}")
        all_rows.extend(collect_rows_from_dir(gudu_input_dir, pipeline='nonpkg'))
    elif gudu_input_dir:
        print(f"  [WARN] Gudu Enterprise dir not found: {gudu_input_dir}")

    if not all_rows:
        print("No lineage rows extracted.")
        return

    # Sort
    all_rows.sort(key=lambda r: (r.get('Object Type', ''), r.get('Package', ''),
                                  r.get('Procedure', ''),
                                  r['Target Table'], r['Target Column'],
                                  r['Source Table'], r['Source Column']))

    fieldnames = ['Object Type', 'Package', 'Procedure', 'SQL Object Name',
                  'Target Table', 'Target Column',
                  'Source Table', 'Source Column', 'Transformation']

    if fmt == 'csv':
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(all_rows)
        print(f"\nCSV: {output_path} ({len(all_rows)} rows)")
    else:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("openpyxl not installed — falling back to CSV")
            output_path = re.sub(r'\.xlsx$', '.csv', output_path)
            process_directory(input_dir, output_path, fmt='csv')
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Column Lineage'

        hdr_fill = PatternFill('solid', fgColor='1F4E79')
        hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        for ci, col in enumerate(fieldnames, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal='center')

        for ri, row in enumerate(all_rows, 2):
            for ci, col in enumerate(fieldnames, 1):
                ws.cell(row=ri, column=ci, value=row.get(col, ''))

        # Auto-width
        for ci, col in enumerate(fieldnames, 1):
            width = max(len(col) + 2,
                        max((len(str(r.get(col, ''))) for r in all_rows[:300]), default=10))
            ws.column_dimensions[get_column_letter(ci)].width = min(width + 2, 60)
        ws.freeze_panes = 'A2'

        # Summary sheet: unique target table → source table mappings
        ws2 = wb.create_sheet('Summary by Target Table')
        s_fill = PatternFill('solid', fgColor='2E75B6')
        s_hdrs = ['Object Type', 'Package', 'Procedure', 'Target Table', 'Target Column',
                  'Source Table', 'Source Column', 'Transformation']
        for ci, h in enumerate(s_hdrs, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.fill = s_fill
            c.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)

        seen_sum = set()
        ri = 2
        for row in all_rows:
            key = (row.get('Object Type',''), row.get('Package',''), row.get('Procedure',''),
                   row['Target Table'], row['Target Column'],
                   row['Source Table'], row['Source Column'],
                   row['Transformation'])
            if key not in seen_sum:
                seen_sum.add(key)
                for ci, col in enumerate(s_hdrs, 1):
                    ws2.cell(row=ri, column=ci, value=row.get(col, ''))
                ri += 1
        ws2.freeze_panes = 'A2'
        for ci in range(1, len(s_hdrs) + 1):
            ws2.column_dimensions[get_column_letter(ci)].width = 40

        wb.save(output_path)
        print(f"\nExcel: {output_path}")
        print(f"  'Column Lineage' sheet: {len(all_rows)} rows")
        print(f"  'Summary by Target Table' sheet: {ri - 2} unique mappings")

    # Stats
    pkgs  = set(r.get('Package', '') for r in all_rows)
    tgts  = set(r['Target Table'] for r in all_rows)
    srcs  = set(r['Source Table'] for r in all_rows if r['Source Table'])
    print(f"\nPackages : {len(pkgs)}")
    print(f"Target tables : {len(tgts)}")
    print(f"Source tables : {len(srcs)}")
    print(f"Direct mappings : {sum(1 for r in all_rows if r['Transformation'] == 'Direct')}")
    print(f"With transforms : {sum(1 for r in all_rows if r['Transformation'] not in ('Direct', 'Constant / Expression'))}")

    # Print summary
    objects = set(r["SQL Object Name"] for r in all_rows)
    target_tables = set(r["Target Table"] for r in all_rows)
    source_tables = set(r["Source Table"] for r in all_rows if r["Source Table"])
    print(f"SQL Objects: {len(objects)}")
    print(f"Target tables: {len(target_tables)}")
    print(f"Source tables: {len(source_tables)}")
    print(f"Direct mappings: {sum(1 for r in all_rows if r['Transformation'] == 'Direct')}")
    print(f"With transformations: {sum(1 for r in all_rows if r['Transformation'] not in ('Direct', 'Constant / Expression'))}")

    # Print first 20 rows as preview
    print("\n-- Preview (first 20 rows) --")
    print(f"{'SQL Object Name':<50} {'Target Table':<40} {'Target Column':<30} {'Source Table':<40} {'Source Column':<30} {'Transformation'}")
    print("-" * 220)
    for r in all_rows[:20]:
        print(f"{r['SQL Object Name']:<50} {r['Target Table']:<40} {r['Target Column']:<30} {r['Source Table']:<40} {r['Source Column']:<30} {r['Transformation']}")


if __name__ == "__main__":
    import argparse
    _base = r"C:\Users\HM295EJ\OneDrive - EY\Desktop\RSLI-DataLineage\SQLObjectParser"
    parser = argparse.ArgumentParser(description='Extract column lineage from Gudu SQLFlow JSON files')
    _all_meta = os.path.join(_base, 'output', 'All_Metadata_cleaned')
    parser.add_argument('--input', '-i',
                        default=_all_meta,
                        help='PKG split-procs cleaned dir (contains *_lineage.json)')
    parser.add_argument('--nonpkg-input', '-n',
                        default=None,
                        help='Non-PKG cleaned dir (standalone PRC + MViews); omit if same as --input')
    parser.add_argument('--gudu-input', '-g',
                        default=None,
                        help='Gudu Enterprise JSON dir (mssql_*_gudu.json files); disabled by default as All_Metadata_cleaned supersedes these')
    parser.add_argument('--output', '-o',
                        default=os.path.join(_base, 'output', 'gudu_lineage_output.csv'),
                        help='Output file path (.xlsx or .csv)')
    parser.add_argument('--format', '-f', default='csv', choices=['excel', 'csv'],
                        help='Output format (default: csv)')
    args = parser.parse_args()
    process_directory(args.input, args.output, args.format,
                      nonpkg_input_dir=args.nonpkg_input,
                      gudu_input_dir=args.gudu_input)
