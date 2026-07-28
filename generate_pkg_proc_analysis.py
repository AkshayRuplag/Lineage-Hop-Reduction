"""
Scans all PKG_*.sql/.txt and PRC_*.sql/.txt files in All_Metadata, extracts PROCEDURE declarations,
cross-references with standalone PRC_* scripts, and writes an Excel report.
"""
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

METADATA_DIR = os.path.join(os.path.dirname(__file__), "All_Metadata")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "output", "PKG_PROC_Analysis_All_Metadata.xlsx")

# Extensions treated as SQL source files
SQL_EXTENSIONS = (".SQL", ".TXT")


def is_sql_file(filename: str) -> bool:
    """Return True for .sql/.txt files, skipping copy/backup duplicates."""
    name_upper = filename.upper()
    if not any(name_upper.endswith(ext) for ext in SQL_EXTENSIONS):
        return False
    # Skip files that are clearly backup copies (contain space or 'copy' in stem)
    stem = os.path.splitext(filename)[0]
    if ' ' in stem or 'copy' in stem.lower():
        return False
    return True

def extract_procedures(filepath):
    """Extract procedure names and line numbers, noting if they are inside comment blocks."""
    procs = []
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()

    in_block_comment = False
    for i, line in enumerate(lines, 1):
        # Track block comments
        if "/*" in line and "*/" not in line:
            in_block_comment = True
        if "*/" in line:
            in_block_comment = False
            continue

        # Check for single-line comment
        stripped = line.strip()
        is_commented = in_block_comment or stripped.startswith("--")

        match = re.match(r"^\s*(PROCEDURE)\s+(\w+)", line, re.IGNORECASE)
        if match:
            proc_name = match.group(2).upper()
            procs.append({
                "name": proc_name,
                "line": i,
                "commented": is_commented
            })
    return procs


def find_main_call_sequence(filepath, proc_names):
    """Find the order in which 'main' calls other sibling procedures."""
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        content = f.read()
        f.seek(0)
        lines = f.readlines()

    # Find main procedure boundaries
    main_start = None
    main_end = None
    proc_starts = []
    for i, line in enumerate(lines, 1):
        m = re.match(r"^\s*PROCEDURE\s+(\w+)", line, re.IGNORECASE)
        if m:
            proc_starts.append((i, m.group(1).upper()))
            if m.group(1).upper() == "MAIN":
                main_start = i

    if main_start is None:
        return []

    # main ends at the next PROCEDURE declaration or END of package
    for start_line, name in proc_starts:
        if start_line > main_start:
            main_end = start_line
            break
    if main_end is None:
        main_end = len(lines)

    main_body = lines[main_start - 1: main_end - 1]

    # Find calls to sibling procs within main body, in order of appearance
    sibling_names = [p for p in proc_names if p != "MAIN"]
    calls = []
    for line_offset, line in enumerate(main_body):
        stripped = line.strip()
        # skip comments and procedure declarations
        if stripped.startswith("--") or re.match(r"^\s*PROCEDURE\s+", stripped, re.IGNORECASE):
            continue
        for sib in sibling_names:
            if re.search(r'\b' + re.escape(sib) + r'\b', line, re.IGNORECASE):
                if sib not in [c["name"] for c in calls]:
                    calls.append({
                        "name": sib,
                        "call_line": main_start + line_offset
                    })
    return calls


# Regex for DML statements that indicate main contains its own lineage-relevant logic.
# Handles Oracle optimizer hints between INSERT and INTO:
#   INSERT /*+APPEND_VALUES*/ INTO ...  or  INSERT /* + APPEND */ INTO ...
_DML_PAT = re.compile(
    r'\b(INSERT\s+(?:/\*[^*]*\*/\s*)?INTO|UPDATE\s+\w|DELETE\s+FROM|MERGE\s+INTO)\b',
    re.IGNORECASE | re.DOTALL,
)


def main_has_dml(filepath: str, proc_starts: list) -> bool:
    """
    Return True if ANY MAIN procedure body contains DML statements
    (INSERT INTO, UPDATE, DELETE FROM, MERGE INTO) outside of comments.

    Handles forward declarations (PROCEDURE main; — package spec header with no body)
    by skipping them and checking every MAIN occurrence.  Multi-line Oracle hints
    like INSERT /*+APPEND_VALUES*/ \\n    INTO are detected because the full MAIN
    body is searched as a single text block (not line-by-line).
    """
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()

    main_lnos = [lno for lno, n in proc_starts if n == "MAIN"]
    if not main_lnos:
        return False

    all_lnos = [lno for lno, _ in proc_starts]

    for main_lno in main_lnos:
        main_end = next((lno for lno in all_lnos if lno > main_lno), len(lines) + 1)
        main_body = ''.join(lines[main_lno - 1: main_end - 1])

        # Skip forward declarations — a forward decl has no BEGIN block
        if not re.search(r'\bBEGIN\b', main_body, re.IGNORECASE):
            continue

        # Strip block comments (/* ... */) — may span multiple lines
        clean = re.sub(r'/\*.*?\*/', ' ', main_body, flags=re.DOTALL)
        # Strip single-line comments (-- ...)
        clean = re.sub(r'--[^\n]*', '', clean)

        if _DML_PAT.search(clean):
            return True
    return False


def classify_proc(name):
    """Classify procedure for lineage relevance."""
    utility_procs = [
        "PRC_REBUILD_INDEXES", "PRC_TRUNC_PARTITION", "PRC_DEBUG_TRACE",
        "PRC_DEBUG_EXEC_TIME"
    ]
    name_upper = name.upper()
    if name_upper in utility_procs:
        return "Utility (not lineage-relevant)"
    return "Lineage-relevant"


# ---------------------------------------------------------------------------
# Utility packages / built-ins that are NOT lineage-relevant
# ---------------------------------------------------------------------------
UTILITY_PKG_PREFIXES = [
    "PKG_GRP_LOG_UTIL", "PKG_LOG_ERROR_UTILITY", "PKG_GRP_COMMON_UTIL",
    "DBMS_STATS", "DBMS_MVIEW", "DBMS_OUTPUT", "DBMS_UTILITY",
    "UTL_FILE", "SYS.", "STANDARD."
]


def classify_external_call(called_name: str) -> str:
    """Classify an external procedure/function call by its relevance."""
    upper = called_name.upper()
    for prefix in UTILITY_PKG_PREFIXES:
        if upper.startswith(prefix):
            return "Utility / Logging"
    if upper.startswith("PKG_"):
        return "Cross-Package (Functional)"
    if upper.startswith("PRC_") or upper.startswith("PROC_"):
        return "Cross-PROC (Standalone)"
    return "Other"


def find_external_calls(filepath: str, self_name: str) -> list:
    """
    Scan a SQL file for calls to external procedures / packages.
    Returns a list of dicts with keys: called_name, call_type, line_no, is_comment, context.
    """
    self_upper = self_name.upper()
    results = []
    seen = {}   # called_name_upper -> first occurrence dict (dedup per name)

    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()

    in_block_comment = False
    for lno, line in enumerate(lines, 1):
        if "/*" in line and "*/" not in line:
            in_block_comment = True
        if "*/" in line:
            in_block_comment = False
            continue

        stripped = line.strip()
        is_comment = in_block_comment or stripped.startswith("--")

        # Match dotted calls: pkg.proc or schema.pkg.proc followed by ( or ;
        matches = re.findall(r'\b(\w+(?:\.\w+)+)\s*[;(]', line, re.IGNORECASE)
        for m in matches:
            m_upper = m.upper()
            # Skip self-references and string literals  
            if m_upper == self_upper or m_upper.startswith(self_upper + "."):
                continue
            # Skip column references like TABLE.COLUMN (all caps, no PKG_ prefix)
            # Only keep things that look like a proc/function call (pkg_ prefix or known util)
            root = m_upper.split(".")[0]
            if not any(root.startswith(pfx.split(".")[0].upper()) for pfx in
                       ["PKG_", "PRC_", "PROC_", "DBMS_", "UTL_", "SYS.", "STANDARD."]):
                continue
            # Skip CREATE OR REPLACE PROCEDURE/FUNCTION definitions
            if re.search(r'CREATE\s+(OR\s+REPLACE\s+)?(EDITIONABLE\s+)?(PROCEDURE|FUNCTION)\s+\w*' +
                         re.escape(m.split(".")[-1]), line, re.IGNORECASE):
                continue

            key = m_upper
            if key not in seen:
                call_type = classify_external_call(m)
                entry = {
                    "called_name": m.upper(),
                    "call_type": call_type,
                    "first_line_no": lno,
                    "is_comment": is_comment,
                    "context": stripped[:120],
                    "occurrence_count": 1
                }
                seen[key] = entry
                results.append(entry)
            else:
                seen[key]["occurrence_count"] += 1
                # Promote to active if we later find a non-comment occurrence
                if not is_comment:
                    seen[key]["is_comment"] = False

    return results


def main():
    # Collect all standalone PRC_* files
    standalone_prcs = set()
    all_files = os.listdir(METADATA_DIR)
    for f in all_files:
        if (f.upper().startswith("PRC_") or f.upper().startswith("PROC_")) and is_sql_file(f):
            standalone_prcs.add(os.path.splitext(f)[0].upper())

    # Collect all PKG_* files
    pkg_files = sorted([f for f in all_files if f.upper().startswith("PKG_") and is_sql_file(f)])

    # Extract procedures from each package
    rows = []
    pkg_summary = []
    serial = 0
    for pkg_file in pkg_files:
        pkg_name = os.path.splitext(pkg_file)[0].upper()
        filepath = os.path.join(METADATA_DIR, pkg_file)
        procs = extract_procedures(filepath)

        lineage_count = 0
        for proc in procs:
            serial += 1
            classification = classify_proc(proc["name"])
            has_script = "Yes" if proc["name"].upper() in standalone_prcs else "No"

            if proc["commented"]:
                status = "Commented Out"
                classification = "N/A (commented)"
            else:
                status = "Active"

            if classification == "Lineage-relevant" and status == "Active":
                lineage_count += 1

            rows.append({
                "serial": serial,
                "package": pkg_name,
                "procedure": proc["name"],
                "line_no": proc["line"],
                "status": status,
                "classification": classification,
                "separate_script": has_script
            })

        pkg_summary.append({
            "package": pkg_name,
            "total_procs": len(procs),
            "lineage_procs": lineage_count,
            "file": pkg_file
        })

    # Also list standalone PRC_* files not found in any package
    all_pkg_procs = set(r["procedure"].upper() for r in rows)
    standalone_only = []
    for prc in sorted(standalone_prcs):
        in_pkg = "Yes" if prc in all_pkg_procs else "No"
        standalone_only.append({"name": prc, "in_package": in_pkg})

    # --- Orchestrator analysis ---
    orchestrator_data = []
    for pkg_file in pkg_files:
        pkg_name = os.path.splitext(pkg_file)[0].upper()
        filepath = os.path.join(METADATA_DIR, pkg_file)
        procs = extract_procedures(filepath)
        active_procs = [p for p in procs if not p["commented"]]
        proc_names = [p["name"] for p in active_procs]
        has_main = "MAIN" in proc_names
        non_main = [p for p in proc_names if p != "MAIN"]

        if has_main:
            # Build (line_no, name) list for main_has_dml helper
            proc_starts_list = [(p["line"], p["name"]) for p in active_procs]
            dml_in_main = main_has_dml(filepath, proc_starts_list)

            call_seq = find_main_call_sequence(filepath, proc_names)
            call_order = [c["name"] for c in call_seq]
            not_called = [p for p in non_main if p not in call_order]

            if len(non_main) == 0:
                pattern = "Self-contained (main only)"
                recommendation = "Parse MAIN directly"
            elif dml_in_main:
                pattern = "Hybrid (main has DML + calls children)"
                recommendation = (
                    "Parse MAIN directly (contains DML) AND parse child procs individually"
                )
            else:
                pattern = "Orchestrator (main calls children)"
                recommendation = "Parse child procs individually; skip main (orchestrator/logging only)"

            orchestrator_data.append({
                "package": pkg_name,
                "pattern": pattern,
                "has_main": "Yes",
                "main_has_dml": "Yes" if dml_in_main else "No",
                "total_active_procs": len(active_procs),
                "entry_point": "MAIN",
                "call_sequence": " → ".join(call_order) if call_order else "N/A",
                "procs_not_called_from_main": ", ".join(not_called) if not_called else "None",
                "lineage_recommendation": recommendation,
            })
        else:
            # No main — independent procs
            lineage_procs = [p for p in non_main if classify_proc(p) == "Lineage-relevant"]
            orchestrator_data.append({
                "package": pkg_name,
                "pattern": "Independent procs (no main)",
                "has_main": "No",
                "main_has_dml": "N/A",
                "total_active_procs": len(active_procs),
                "entry_point": ", ".join(lineage_procs) if lineage_procs else ", ".join(non_main),
                "call_sequence": "N/A (each called externally)",
                "procs_not_called_from_main": "N/A",
                "lineage_recommendation": "Parse each procedure independently"
            })

    # --- External PROC / Package Call Analysis ---
    # Covers both standalone PRC_* files AND PKG_* files
    ext_call_rows = []
    prc_files = sorted([f for f in all_files if (f.upper().startswith("PRC_") or f.upper().startswith("PROC_")) and is_sql_file(f)])

    for fname in prc_files + pkg_files:
        obj_name = os.path.splitext(fname)[0].upper()
        fpath = os.path.join(METADATA_DIR, fname)
        calls = find_external_calls(fpath, obj_name)
        source_type = "Standalone PRC" if fname in prc_files else "Package (PKG)"
        for c in calls:
            ext_call_rows.append({
                "source_object": obj_name,
                "source_type": source_type,
                "called_name": c["called_name"],
                "call_type": c["call_type"],
                "status": "Comment only" if c["is_comment"] else "Active call",
                "occurrences": c["occurrence_count"],
                "first_line_no": c["first_line_no"],
                "context": c["context"],
            })

    # --- Write Excel ---
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    active_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    commented_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    utility_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header(ws, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    # Sheet 1: Detailed Procedure List
    ws1 = wb.active
    ws1.title = "Package Procedures"
    headers1 = ["#", "Package Name", "Procedure Name", "Line #", "Status",
                 "Classification", "Separate Script Exists?"]
    ws1.append(headers1)
    style_header(ws1, len(headers1))

    for r in rows:
        row_num = ws1.max_row + 1
        ws1.append([r["serial"], r["package"], r["procedure"], r["line_no"],
                     r["status"], r["classification"], r["separate_script"]])
        for col in range(1, len(headers1) + 1):
            ws1.cell(row=row_num, column=col).border = thin_border

        # Color status
        status_cell = ws1.cell(row=row_num, column=5)
        if r["status"] == "Active":
            status_cell.fill = active_fill
        else:
            status_cell.fill = commented_fill

        # Color classification
        class_cell = ws1.cell(row=row_num, column=6)
        if "Utility" in r["classification"]:
            class_cell.fill = utility_fill
        elif "commented" in r["classification"]:
            class_cell.fill = commented_fill

        # Color separate script
        script_cell = ws1.cell(row=row_num, column=7)
        if r["separate_script"] == "Yes":
            script_cell.fill = yes_fill
        else:
            script_cell.fill = no_fill

    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 50
    ws1.column_dimensions["C"].width = 55
    ws1.column_dimensions["D"].width = 8
    ws1.column_dimensions["E"].width = 15
    ws1.column_dimensions["F"].width = 30
    ws1.column_dimensions["G"].width = 22

    # Sheet 2: Package Summary
    ws2 = wb.create_sheet("Package Summary")
    headers2 = ["#", "Package Name", "File Name", "Total Procedures", "Lineage-Relevant Procedures"]
    ws2.append(headers2)
    style_header(ws2, len(headers2))

    for i, s in enumerate(pkg_summary, 1):
        row_num = ws2.max_row + 1
        ws2.append([i, s["package"], s["file"], s["total_procs"], s["lineage_procs"]])
        for col in range(1, len(headers2) + 1):
            ws2.cell(row=row_num, column=col).border = thin_border

    # Totals row
    total_row = ws2.max_row + 1
    ws2.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=4, value=sum(s["total_procs"] for s in pkg_summary)).font = Font(bold=True)
    ws2.cell(row=total_row, column=5, value=sum(s["lineage_procs"] for s in pkg_summary)).font = Font(bold=True)
    for col in range(1, len(headers2) + 1):
        ws2.cell(row=total_row, column=col).border = thin_border

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 55
    ws2.column_dimensions["D"].width = 20
    ws2.column_dimensions["E"].width = 30

    # Sheet 3: Standalone PRC Scripts
    ws3 = wb.create_sheet("Standalone PRC Scripts")
    headers3 = ["#", "Standalone PRC Script", "Also Inside a Package?"]
    ws3.append(headers3)
    style_header(ws3, len(headers3))

    for i, s in enumerate(standalone_only, 1):
        row_num = ws3.max_row + 1
        ws3.append([i, s["name"], s["in_package"]])
        for col in range(1, len(headers3) + 1):
            ws3.cell(row=row_num, column=col).border = thin_border
        pkg_cell = ws3.cell(row=row_num, column=3)
        if s["in_package"] == "Yes":
            pkg_cell.fill = utility_fill
        else:
            pkg_cell.fill = active_fill

    ws3.column_dimensions["A"].width = 5
    ws3.column_dimensions["B"].width = 60
    ws3.column_dimensions["C"].width = 25

    # Sheet 4: Orchestrator Analysis
    ws4 = wb.create_sheet("Orchestrator Analysis")
    headers4 = ["#", "Package Name", "Pattern", "Has main?", "Main has DML?",
                 "Total Active Procs", "Entry Point", "Call Sequence from main",
                 "Procs NOT Called from main", "Lineage Parsing Recommendation"]
    ws4.append(headers4)
    style_header(ws4, len(headers4))

    orch_fill     = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")  # blue
    selfcontained_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # green
    independent_fill   = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # orange
    hybrid_fill   = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # amber
    dml_yes_fill  = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # amber
    dml_no_fill   = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")  # grey

    for i, od in enumerate(orchestrator_data, 1):
        row_num = ws4.max_row + 1
        ws4.append([i, od["package"], od["pattern"], od["has_main"],
                     od["main_has_dml"], od["total_active_procs"], od["entry_point"],
                     od["call_sequence"], od["procs_not_called_from_main"],
                     od["lineage_recommendation"]])
        for col in range(1, len(headers4) + 1):
            ws4.cell(row=row_num, column=col).border = thin_border
            ws4.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        # Color pattern column (col 3)
        pattern_cell = ws4.cell(row=row_num, column=3)
        if "Hybrid" in od["pattern"]:
            pattern_cell.fill = hybrid_fill
        elif "Orchestrator" in od["pattern"]:
            pattern_cell.fill = orch_fill
        elif "Self-contained" in od["pattern"]:
            pattern_cell.fill = selfcontained_fill
        else:
            pattern_cell.fill = independent_fill

        # Color "Main has DML?" column (col 5)
        dml_cell = ws4.cell(row=row_num, column=5)
        if od["main_has_dml"] == "Yes":
            dml_cell.fill = dml_yes_fill
        elif od["main_has_dml"] == "No":
            dml_cell.fill = dml_no_fill

    ws4.column_dimensions["A"].width = 5
    ws4.column_dimensions["B"].width = 50
    ws4.column_dimensions["C"].width = 38
    ws4.column_dimensions["D"].width = 12
    ws4.column_dimensions["E"].width = 15
    ws4.column_dimensions["F"].width = 18
    ws4.column_dimensions["G"].width = 55
    ws4.column_dimensions["H"].width = 70
    ws4.column_dimensions["I"].width = 35
    ws4.column_dimensions["J"].width = 60

    # Sheet 5: External PROC / Package Call Analysis
    ws5 = wb.create_sheet("External PROC Calls")
    headers5 = [
        "#", "Source Object", "Source Type", "Called Object (pkg.proc)",
        "Call Type", "Status", "Occurrences", "First Line #", "Sample Context"
    ]
    ws5.append(headers5)
    style_header(ws5, len(headers5))

    # Colour fills for call type
    cross_pkg_fill  = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")  # blue-ish
    cross_prc_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow
    util_fill_ext   = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")  # grey
    comment_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # red-ish

    for i, ec in enumerate(ext_call_rows, 1):
        row_num = ws5.max_row + 1
        ws5.append([
            i,
            ec["source_object"],
            ec["source_type"],
            ec["called_name"],
            ec["call_type"],
            ec["status"],
            ec["occurrences"],
            ec["first_line_no"],
            ec["context"],
        ])
        for col in range(1, len(headers5) + 1):
            ws5.cell(row=row_num, column=col).border = thin_border
            ws5.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        # Colour by call type
        type_cell = ws5.cell(row=row_num, column=5)
        status_cell = ws5.cell(row=row_num, column=6)

        if ec["status"] == "Comment only":
            for col in range(1, len(headers5) + 1):
                ws5.cell(row=row_num, column=col).fill = comment_fill
        elif "Cross-Package" in ec["call_type"]:
            type_cell.fill = cross_pkg_fill
        elif "Cross-PROC" in ec["call_type"]:
            type_cell.fill = cross_prc_fill
        else:
            type_cell.fill = util_fill_ext

        # Green = active call, red = comment only
        if ec["status"] == "Active call":
            status_cell.fill = active_fill
        else:
            status_cell.fill = commented_fill

    ws5.column_dimensions["A"].width = 5
    ws5.column_dimensions["B"].width = 60
    ws5.column_dimensions["C"].width = 18
    ws5.column_dimensions["D"].width = 60
    ws5.column_dimensions["E"].width = 28
    ws5.column_dimensions["F"].width = 15
    ws5.column_dimensions["G"].width = 12
    ws5.column_dimensions["H"].width = 12
    ws5.column_dimensions["I"].width = 80

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"Excel report saved to: {OUTPUT_FILE}")
    print(f"  Sheet 1: {len(rows)} procedures from {len(pkg_files)} packages")
    print(f"  Sheet 2: Package summary")
    print(f"  Sheet 3: {len(standalone_only)} standalone PRC scripts")
    print(f"  Sheet 4: Orchestrator analysis for {len(orchestrator_data)} packages")
    print(f"  Sheet 5: {len(ext_call_rows)} external PROC/package call entries"
          f" ({len(prc_files)} standalone PRC files + {len(pkg_files)} PKG files scanned)")


if __name__ == "__main__":
    main()
