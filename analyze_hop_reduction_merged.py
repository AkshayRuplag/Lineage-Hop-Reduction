"""
Hop Reduction Analyzer — Mature / Merged
========================================

Detectors (14):
  A. Redundant multi-source loads (skips legit MULTI_SRC splits)
  B. Pass-through hops (SRC∩TGT + graph-walked chain detection)
  C. Near-duplicate procedures (Jaccard column-set clustering)
  D. Serial chains with NO column data dependency
  E. MV elimination (with consumer column-usage %)
  F. LLM-identified issues (target-table fallback + per-type counts)
  G. Duplicate column-lineage paths
  H. Stale intermediate tables
  I. Cyclic refresh dependencies
  J. Over-fanout MVs
  K. Repeated transformations (same src->tgt column mapping across jobs)
  L. Overlapping preprocessing on same target tables
  M. Orchestration edge without data handoff (MV->MV dependency mismatch)
  N. Post-consumption update on same intermediate table

Pipeline:
    load -> detectors -> standardise_metrics -> dedup -> attach_supporting_evidence
             -> critical-path-aware scoring -> what-if simulation -> Excel

Inputs:
  output/tidal_dependency_graph.json
  output/combined_lineage_latest.xlsx
  output/script_summaries_all.json
  output/hop_reduction_from_summaries_all.json

Output:
  output/hop_reduction_recommendations_RPT_TABLE_NAME.xlsx
"""

import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR / "output"
HOP_REC_DIR = OUTPUT_DIR / "hop_reduction_recommendations"
GRAPH_JSON = OUTPUT_DIR / "tidal_dependency_graph.json"
LINEAGE_XLSX = OUTPUT_DIR / "combined_lineage_latest.xlsx"
SUMMARIES_JSON = OUTPUT_DIR / "script_summaries_all.json"
HOP_SUMMARIES_JSON = OUTPUT_DIR / "hop_reduction_from_summaries_all.json"

# Configurable RPT table (default: RPT_CLAIM_PAYMENT_R)
# Can be overridden via command-line arg: python script.py RPT_TABLE_NAME
RPT_TABLE = os.environ.get("RPT_TABLE", "RPT_CLAIM_PAYMENT_R")
if len(sys.argv) > 1:
    RPT_TABLE = sys.argv[1]


# ── Loaders ──────────────────────────────────────────────────────────────────


def load_graph():
    with open(GRAPH_JSON, "r", encoding="utf-8") as f:
        return json.load(f)[RPT_TABLE]


def load_all_graphs():
    with open(GRAPH_JSON, "r", encoding="utf-8") as f:
        return json.load(f)


def load_lineage():
    wb = openpyxl.load_workbook(str(LINEAGE_XLSX), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [str(h) for h in rows[0]]
    records = []
    for r in rows[1:]:
        rec = {headers[i]: r[i] for i in range(len(headers))}
        if rec.get("RPT_TABLE") == RPT_TABLE:
            records.append(rec)
    
    if not records:
        print(f"WARNING: No lineage data found for RPT_TABLE='{RPT_TABLE}'")
        print(f"Available columns: {headers}")
    return records


def load_all_lineage():
    wb = openpyxl.load_workbook(str(LINEAGE_XLSX), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [str(h) for h in rows[0]]
    records = []
    for r in rows[1:]:
        rec = {headers[i]: r[i] for i in range(len(headers))}
        # Filter out rows where TGT_TABLE contains commas — these are
        # lineage data quality artefacts where multiple table names were concatenated
        # into a single cell.  They produce phantom table names that pollute detector
        # output with false "potential dead write" signals.
        # NOTE: SRC_TABLE may validly contain comma-separated source tables (a job
        # reading from multiple sources); we must NOT filter those rows.
        tgt = str(rec.get("TGT_TABLE") or "")
        if "," in tgt:
            continue
        records.append(rec)
    return records


def load_summaries():
    if SUMMARIES_JSON.exists():
        with open(SUMMARIES_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def load_hop_findings():
    if HOP_SUMMARIES_JSON.exists():
        with open(HOP_SUMMARIES_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


# ── Helpers ──────────────────────────────────────────────────────────────────


def parse_col_list(raw):
    """Parse '(COL1,COL2,...)' or comma-separated -> set of upper-case columns."""
    if not raw:
        return set()
    s = str(raw).strip().strip("()")
    if not s:
        return set()
    return {c.strip().upper() for c in s.split(",") if c.strip()}


def parse_table_list(raw):
    """Parse comma-separated table strings into a normalized upper-case set."""
    if not raw:
        return set()
    out = set()
    for p in str(raw).split(","):
        p = p.strip()
        if p and p not in ("None", "N/A", "-"):
            out.add(p.upper())
    return out


def _fmt_job_sql(job: str, sql_index: dict) -> str:
    """Format a Tidal job name with its SQL procedure for readable step output.
    Returns: 'JobName  [SQL: proc1, proc2]' if SQL info is available, else just 'JobName'.
    """
    objs = sorted((sql_index or {}).get(job, {}).get("full_objects", set()))
    if objs:
        sql_str = ", ".join(objs[:2]) + ("..." if len(objs) > 2 else "")
        return f"{job}  [SQL: {sql_str}]"
    return job


def jaccard(a: set, b: set) -> float:
    if not a and not b:
        return 1.0
    union = a | b
    if not union:
        return 0.0
    return len(a & b) / len(union)


def build_adjacency(graph):
    children = defaultdict(set)
    parents = defaultdict(set)
    for link in graph["links"]:
        children[link["source"]].add(link["target"])
        parents[link["target"]].add(link["source"])
    return children, parents


def build_table_producers_consumers(graph):
    table_map = defaultdict(lambda: {"producers": set(), "consumers": set()})
    for n in graph["nodes"]:
        for t in n.get("tgt_tables", []) or []:
            table_map[t]["producers"].add(n["id"])
        for t in n.get("src_tables", []) or []:
            table_map[t]["consumers"].add(n["id"])
    return table_map


def build_global_children(all_graphs):
    global_children = defaultdict(set)
    for g in all_graphs.values():
        for link in g.get("links", []):
            global_children[link["source"]].add(link["target"])
    return global_children


def build_global_table_usage(all_graphs):
    global_map = defaultdict(lambda: {"producers": set(), "consumers": set()})
    for g in all_graphs.values():
        for n in g.get("nodes", []):
            for t in n.get("tgt_tables", []) or []:
                global_map[t]["producers"].add(n["id"])
            for t in n.get("src_tables", []) or []:
                global_map[t]["consumers"].add(n["id"])
    return global_map


def build_global_job_rpt_map(all_graphs):
    """Build job -> set(RPTs) index across all graphs."""
    job_rpts = defaultdict(set)
    for rpt, g in all_graphs.items():
        for n in g.get("nodes", []):
            jid = n.get("id")
            if jid:
                job_rpts[jid].add(rpt)
    return job_rpts


def build_global_runtime_index(all_graphs):
    """Build global job -> avg_runtime_seconds (max observed across RPT graphs)."""
    runtime = {}
    for g in all_graphs.values():
        for n in g.get("nodes", []):
            jid = n.get("id")
            r = n.get("avg_runtime")
            if jid and isinstance(r, (int, float)) and r > 0:
                runtime[jid] = max(runtime.get(jid, 0.0), float(r))
    return runtime


def build_global_mv_context(all_graphs, global_runtime_index):
    """Build global MV context used for cross-RPT E-category decisions."""
    context = defaultdict(
        lambda: {
            "rpts": set(),
            "consumers": set(),
            "consumers_by_rpt": defaultdict(set),
            "tgt_tables": set(),
            "mv_runtime_seconds": 0.0,
        }
    )

    for rpt, g in all_graphs.items():
        children = defaultdict(set)
        for link in g.get("links", []):
            src = link.get("source")
            tgt = link.get("target")
            if src and tgt:
                children[src].add(tgt)

        node_map_local = {x["id"]: x for x in g.get("nodes", [])}

        for n in g.get("nodes", []):
            if n.get("category") != "MV":
                continue
            mv = n.get("id")
            if not mv:
                continue

            info = context[mv]
            info["rpts"].add(rpt)
            info["tgt_tables"] |= set(n.get("tgt_tables", []) or [])
            mv_rt = global_runtime_index.get(mv, 0.0)
            if mv_rt > info["mv_runtime_seconds"]:
                info["mv_runtime_seconds"] = mv_rt

            # Only count consumers that actually READ from this MV's target table(s)
            # in their SQL (src_tables).  Jobs connected only via orchestration/Tidal
            # scheduling edges or through an intermediate SQL VIEW are NOT data consumers
            # and cannot be inlined — they represent a lineage gap.
            mv_tgts_local = {t.upper() for t in (n.get("tgt_tables") or [])}
            consumers = {
                c for c in children.get(mv, set())
                if not c.upper().startswith("VIEW_")
                and mv_tgts_local & {
                    s.upper()
                    for s in (node_map_local.get(c, {}).get("src_tables") or [])
                }
            }
            info["consumers"] |= consumers
            info["consumers_by_rpt"][rpt] |= consumers

    return context


def build_runtime_index(graph):
    """Build job -> avg_runtime_seconds mapping from graph nodes."""
    runtime = {}
    for n in graph.get("nodes", []):
        r = n.get("avg_runtime")
        if isinstance(r, (int, float)) and r > 0:
            runtime[n["id"]] = float(r)
    return runtime


def attach_runtime_estimates(recommendations, runtime_index):
    """
    Attach runtime visibility to each recommendation.
    Heuristics:
      - D. Serial chain: savings approximated by sum(chain) - max(chain) (parallelization effect)
      - Other hop-reduction recs: savings approximated by sum of lowest N runtimes where
        N = potential_hop_savings (conservative estimate of removable jobs)
      - LLM-only recs: no deterministic runtime savings estimate
    """
    for rec in recommendations:
        jobs = rec.get("affected_jobs", []) or []
        runtimes = [runtime_index[j] for j in jobs if j in runtime_index]
        rec["runtime_affected_seconds"] = round(sum(runtimes), 2) if runtimes else 0.0
        rec["runtime_jobs_with_data"] = len(runtimes)

        est_saved = 0.0
        if rec.get("recommendation_type") == "HOP_REDUCTION" and runtimes:
            if str(rec.get("category", "")).startswith("D. Serial Orchestration Chain"):
                # Critical path after parallelizing a serial chain ~ longest branch runtime.
                est_saved = max(sum(runtimes) - max(runtimes), 0.0)
            elif str(rec.get("category", "")).startswith("M. Orchestration Edge Without Data Handoff"):
                # Edge-only dependency cleanup can at most save overlap equal to shorter branch.
                est_saved = min(runtimes)
            else:
                hops = int(rec.get("verified_hop_savings", 0) or 0)
                if hops > 0:
                    k = min(hops, len(runtimes))
                    # Conservative: assume removable work are the smallest k jobs.
                    est_saved = sum(sorted(runtimes)[:k])

        rec["est_runtime_saved_seconds"] = round(est_saved, 2)
        rec["est_runtime_saved_minutes"] = round(est_saved / 60.0, 2)

    return recommendations


# ── Phase-2 category-aware execution policy ────────────────────────────────
#
# Rationale: A job appearing in N RPT graphs does NOT mean that changing the
# orchestration or SQL within *this* RPT automatically affects those other RPTs.
# Only changes to shared assets (MVs consumed cross-RPT, shared SQL procs) truly
# require a global program.  All other patterns are local to the RPT under analysis.
#
# LOCAL:   The change is scoped to this RPT's SQL or scheduler only.
#          Affected_jobs may be shared infra, but the edit itself is local.
# GLOBAL:  The asset being removed/modified is consumed by multiple RPTs.
#          The E-detector pre-populates these via global_mv_context.
# DEFER:   Informational findings with no immediate action.

_LOCAL_EXECUTION_CATEGORIES = frozenset([
    "B.",   # Pass-through fold — local INSERT/MERGE rewrite
    "D.",   # Serial chain — local Tidal schedule parallelization
    "F.",   # LLM PLSQL — local script optimization
    "G.",   # Duplicate column path — local consolidation
    "H.",   # Stale intermediate — local dead-write removal
    "I.",   # Cyclic dependency — local cycle break
    "J.",   # Over-fanout MV — local split decision
    "M.",   # Orchestration edge removal — Tidal edge is per-RPT-schedule
    "N.",   # Post-consumption sequencing — local to this RPT's execution order
    "K.",   # Repeated transformations — local SQL refactor
    "L.",   # Overlapping preprocessing — local pipeline rationalization
])

_DEFER_CATEGORY_SUBSTRINGS = frozenset([
    "Already Parameterized via Tidal",   # No SQL action needed
    "Source System Architectural Split", # Intentional per-source separation — do not combine
    "View Naming Convention Violation",  # O-b: rename only, no hop savings
    # NOTE: "No Blind Elimination" (E. MV Shared) is intentionally NOT here.
    # Its implementation_decision is pre-set to EXECUTE_GLOBAL_PROGRAM by the detector
    # and must flow through the `existing_decision == "EXECUTE_GLOBAL_PROGRAM"` branch
    # in enrich_global_execution_context(). Adding it here would override that to DEFER_OR_KEEP.
])

_WAVE_BY_RISK = {"LOW": "WAVE_1", "MEDIUM": "WAVE_2", "HIGH": "WAVE_3"}


def _category_execution_policy(category: str) -> str | None:
    """Return the forced policy for a given category, or None to auto-decide."""
    if any(category.startswith(p) for p in _LOCAL_EXECUTION_CATEGORIES):
        return "EXECUTE_LOCALLY"
    if any(s in category for s in _DEFER_CATEGORY_SUBSTRINGS):
        return "DEFER_OR_KEEP"
    return None  # let caller decide (E-global is already pre-set by detector)


def enrich_global_execution_context(recommendations, global_job_rpt_map, global_runtime_index):
    """
    Phase-2 category-aware global execution context enrichment.

    Key rules
    ---------
    1. Categories that are always local (B/D/F/G/H/I/J/K/L/M/N):
       - impacted_rpts = {RPT_TABLE}, global_rpt_count_impacted = 1
       - implementation_decision = EXECUTE_LOCALLY
       - blast_radius_score based on affected_jobs count only (no RPT spread)
    2. Already-parameterised / informational categories:
       - implementation_decision = DEFER_OR_KEEP, execution_wave = DEFER
    3. E-category: decision and impacted_rpts already set by detect_mv_elimination.
       Enrich does NOT override those pre-computed fields.
    4. A/C: default to EXECUTE_LOCALLY (SQL consolidation is an intra-RPT change;
       if later we want cross-RPT proc alignment we can upgrade via a separate flag).
    5. execution_wave is risk-aware for local decisions:
       LOW → WAVE_1, MEDIUM → WAVE_2, HIGH → WAVE_3.
    6. EXECUTE_GLOBAL_PROGRAM always maps to GLOBAL_WAVE_1_PLUS.
    """
    for rec in recommendations:
        category = rec.get("category", "")
        risk = rec.get("risk", "MEDIUM")
        jobs = [j for j in (rec.get("affected_jobs", []) or []) if j and j != "UNMAPPED"]

        # ── Step 1: determine policy ─────────────────────────────────────────
        forced_policy = _category_execution_policy(category)
        # Respect detector pre-set decision (E-shared sets EXECUTE_GLOBAL_PROGRAM)
        existing_decision = rec.get("implementation_decision")

        if forced_policy == "EXECUTE_LOCALLY":
            # Local change: scope everything to this RPT only
            rec["impacted_rpts"] = rec.get("impacted_rpts") or [RPT_TABLE]
            if not any(r == RPT_TABLE for r in rec["impacted_rpts"]):
                rec["impacted_rpts"] = [RPT_TABLE]
            rec["impacted_rpts"] = [RPT_TABLE]   # always reset for local categories
            rec["global_rpt_count_impacted"] = 1
            rec["global_consumer_count"] = rec.get("global_consumer_count") or max(len(jobs) - 1, 0)
            local_rt = sum(global_runtime_index.get(j, 0.0) for j in jobs)
            rec["global_runtime_affected_minutes"] = round(local_rt / 60.0, 2)
            rec["global_est_runtime_saved_minutes"] = rec.get(
                "est_runtime_saved_minutes", 0.0
            )
            rec["blast_radius_score"] = min(len(jobs), 5)   # local blast = job count, capped
            rec["implementation_decision"] = "EXECUTE_LOCALLY"
            rec["execution_wave"] = _WAVE_BY_RISK.get(risk, "WAVE_2")
            rec.setdefault(
                "prerequisites",
                "Row-count parity and cycle-time comparison for 3+ runs within this RPT",
            )

        elif forced_policy == "DEFER_OR_KEEP":
            # No action: informational/already-done finding
            rec.setdefault("impacted_rpts", [RPT_TABLE])
            rec.setdefault("global_rpt_count_impacted", 1)
            rec.setdefault("global_consumer_count", 0)
            rec.setdefault("global_runtime_affected_minutes", 0.0)
            rec.setdefault("global_est_runtime_saved_minutes", 0.0)
            rec.setdefault("blast_radius_score", 0)
            rec["implementation_decision"] = "DEFER_OR_KEEP"
            rec["execution_wave"] = "DEFER"
            rec.setdefault("prerequisites", "None — no change required")

        elif existing_decision == "EXECUTE_GLOBAL_PROGRAM":
            # Pre-set by detector (E-shared): respect all pre-computed global fields,
            # just fill in any missing standard fields.
            if not rec.get("impacted_rpts"):
                impacted_rpts = set()
                for j in jobs:
                    impacted_rpts |= set(global_job_rpt_map.get(j, set()))
                rec["impacted_rpts"] = sorted(impacted_rpts or {RPT_TABLE})
            rec.setdefault("global_rpt_count_impacted", max(len(rec["impacted_rpts"]), 1))
            rec.setdefault("global_consumer_count", max(len(jobs) - 1, 0))
            if not rec.get("global_runtime_affected_minutes"):
                g_rt = sum(global_runtime_index.get(j, 0.0) for j in jobs)
                rec["global_runtime_affected_minutes"] = round(g_rt / 60.0, 2)
            rec.setdefault("global_est_runtime_saved_minutes", rec.get("est_runtime_saved_minutes", 0.0))
            rec.setdefault(
                "blast_radius_score",
                min(
                    rec.get("global_consumer_count", 0)
                    + max(rec.get("global_rpt_count_impacted", 1) - 1, 0),
                    10,
                ),
            )
            rec["execution_wave"] = "GLOBAL_WAVE_1_PLUS"
            rec.setdefault(
                "prerequisites",
                "Complete parity checks per wave (row counts, key metrics, SLA cycle-time)",
            )

        else:
            # A, C, and any unclassified category → conservative local execution
            rec["impacted_rpts"] = [RPT_TABLE]
            rec["global_rpt_count_impacted"] = 1
            rec["global_consumer_count"] = rec.get("global_consumer_count") or max(len(jobs) - 1, 0)
            local_rt = sum(global_runtime_index.get(j, 0.0) for j in jobs)
            rec["global_runtime_affected_minutes"] = round(local_rt / 60.0, 2)
            rec["global_est_runtime_saved_minutes"] = rec.get("est_runtime_saved_minutes", 0.0)
            rec["blast_radius_score"] = min(len(jobs), 5)
            rec["implementation_decision"] = "EXECUTE_LOCALLY"
            rec["execution_wave"] = _WAVE_BY_RISK.get(risk, "WAVE_2")
            rec.setdefault(
                "prerequisites",
                "Row-count parity, key-metric parity, and cycle-time comparison for 3+ runs",
            )

    return recommendations


def build_job_sql_index(lineage_rows):
    """
    Build a mapping: job_name -> {full_object, package_name, proc_name, sql_type, is_parameterized}
    Uses FULL_OBJECT, PACKAGE_NAME, PROC_NAME, SQL_OBJECT_TYPE, IS_PARAMETERIZED columns.
    A job may appear multiple times (once per src/tgt row), so we collect all SQL objects per job.
    """
    index = defaultdict(lambda: {
        "full_objects": set(),
        "package_names": set(),
        "proc_names": set(),
        "sql_types": set(),
        "is_parameterized": False,
    })
    for r in lineage_rows:
        job = r.get("DEPENDENT_JOB")
        if not job:
            continue
        fo = r.get("FULL_OBJECT") or ""
        pn = r.get("PACKAGE_NAME") or ""
        pr = r.get("PROC_NAME") or ""
        st = r.get("SQL_OBJECT_TYPE") or ""
        ip = r.get("IS_PARAMETERIZED")
        if fo:
            index[job]["full_objects"].add(fo)
        if pn:
            index[job]["package_names"].add(pn)
        if pr:
            index[job]["proc_names"].add(pr)
        if st:
            index[job]["sql_types"].add(st)
        if ip and str(ip).lower() == "true":
            index[job]["is_parameterized"] = True
    return dict(index)


def get_jobs_sql_summary(jobs, sql_index):
    """Return sorted list of unique FULL_OBJECT values across a set of jobs."""
    objs = set()
    for j in jobs:
        objs |= sql_index.get(j, {}).get("full_objects", set())
    return sorted(objs)


def classify_confidence_scope(jobs, sql_index, global_consumer_count=1, rpt_consumer_count=1):
    """
    Classify the confidence/actionability scope of a recommendation.
    Returns one of:
      ALREADY_PARAMETERIZED  - all jobs call same SQL obj, IS_PARAMETERIZED=True (Tidal IS the param layer)
      SAME_PKG_VARIANT       - all jobs call same base package but different sub-variants
      DIFF_SQL_OBJECTS       - jobs call genuinely different SQL objects -> real consolidation opportunity
      CROSS_RPT_SHARED       - MV/table shared across multiple RPT lineages
      LOCAL_ONLY             - stale/single-consumer only in this RPT slice
    """
    if global_consumer_count > 1:
        return "CROSS_RPT_SHARED"
    if rpt_consumer_count == 1:
        return "LOCAL_ONLY"

    full_objs = set()
    all_parameterized = True
    for j in jobs:
        info = sql_index.get(j, {})
        full_objs |= info.get("full_objects", set())
        if not info.get("is_parameterized", False):
            all_parameterized = False

    if not full_objs:
        return "DIFF_SQL_OBJECTS"  # no info -> assume actionable

    if len(full_objs) == 1 and all_parameterized:
        return "ALREADY_PARAMETERIZED"

    base_pkgs = set()
    for fo in full_objs:
        base = fo.split(".")[0]  # PKG_GRP_LOAD_DIFW_PD -> PKG_GRP_LOAD_DIFW_PD
        base_pkgs.add(base)
    if len(base_pkgs) == 1 and len(full_objs) > 1:
        return "SAME_PKG_VARIANT"

    return "DIFF_SQL_OBJECTS"


def enrich_sql_objects_called(recommendations, graph, sql_index):
    """
    Ensure every recommendation has a populated `sql_objects_called` list when possible.
    Priority:
      1) Existing `sql_objects_called`
      2) Existing `sql_objects` (legacy detector field)
      3) SQL objects inferred from affected jobs via lineage `sql_index`
      4) Graph node `primary_sql`
    """
    node_map = {n.get("id"): n for n in graph.get("nodes", [])}

    for rec in recommendations:
        existing = rec.get("sql_objects_called") or rec.get("sql_objects") or []
        objs = set(existing)

        for j in rec.get("affected_jobs", []) or []:
            if j == "UNMAPPED":
                continue

            objs |= set(sql_index.get(j, {}).get("full_objects", set()))

            node = node_map.get(j, {})
            for so in node.get("sql_objects", []) or []:
                name = so.get("name")
                if name:
                    objs.add(name)

            primary = node.get("primary_sql")
            if primary:
                objs.add(primary)

        rec["sql_objects_called"] = sorted(o for o in objs if o)

    return recommendations


def generate_concrete_actions_for_f_findings(recommendations, summaries):
    """
    Convert generic F recommendations to concrete action items by mining script summaries.
    For each F recommendation, extract and format specific optimization steps.
    """
    for rec in recommendations:
        if not rec.get("category", "").startswith("F."):
            continue

        script = rec.get("target_table", "")
        summary = summaries.get(script, {})

        # Extract concrete optimization actions
        actions = []

        # From HOP_REDUCTION_OPPORTUNITIES
        hop_opps = summary.get("HOP_REDUCTION_OPPORTUNITIES", [])
        if isinstance(hop_opps, list):
            for opp in hop_opps[:3]:  # top 3 opportunities
                if isinstance(opp, dict):
                    issue = opp.get("issue_type", "")
                    action = opp.get("recommended_action", "")
                    risk = opp.get("risk_level", "medium")
                    if action:
                        actions.append(f"[{risk.upper()}] {issue}: {action}")

        # From REFACTORING_RECOMMENDATIONS
        refactor = summary.get("REFACTORING_RECOMMENDATIONS", {})
        if isinstance(refactor, dict):
            sql_ideas = refactor.get("sql_optimization_ideas", [])
            if isinstance(sql_ideas, list):
                for idea in sql_ideas[:2]:
                    if idea:
                        actions.append(f"[SQL] {idea}")

        # From OPTIMIZATION_SIGNALS bottlenecks
        perf_char = summary.get("PERFORMANCE_CHARACTERISTICS", {})
        if isinstance(perf_char, dict):
            bottlenecks = perf_char.get("potential_bottlenecks", [])
            if isinstance(bottlenecks, list):
                for bn in bottlenecks[:2]:
                    if bn:
                        actions.append(f"[BOTTLENECK] {bn}")

        # Enrich recommendation text with numbered actions
        if actions:
            action_text = "\n".join([f"{i+1}. {a}" for i, a in enumerate(actions)])
            rec["recommendation"] = (
                f"Concrete optimization steps:\n{action_text}\n\n"
                f"→ Affected findings: {rec.get('total_findings', 0)} issues "
                f"({rec.get('redundancy_findings', 0)} redundancy, "
                f"{rec.get('perf_findings', 0)} performance)"
            )

    return recommendations


def enrich_recommendation_specificity(recommendations, graph, sql_index):
    """
    Make A/B/C/D/E recommendations more specific and actionable by adding concrete details.
    """
    node_map = {n.get("id"): n for n in graph.get("nodes", [])}

    for rec in recommendations:
        category = rec.get("category", "")
        jobs = rec.get("affected_jobs", []) or []
        existing_rec = rec.get("recommendation", "")

        # A/C: Add specific consolidation guidance
        if "Redundant Multi-Source" in category or "Near-Duplicate" in category:
            if len(jobs) >= 2:
                sql_objs = rec.get("sql_objects_called", [])
                if sql_objs:
                    if len(sql_objs) == 1:
                        rec["recommendation"] = (
                            existing_rec + f"\n\n[CONCRETE] All {len(jobs)} jobs call {sql_objs[0]}. "
                            f"Action: Review Tidal parameter consolidation—can source-system or region params be merged? "
                            f"Affected jobs: {', '.join(jobs[:3])}"
                            + (f", ..." if len(jobs) > 3 else "")
                        )
                    else:
                        rec["recommendation"] = (
                            existing_rec + f"\n\n[CONCRETE] Jobs call different SQL objects: {', '.join(sql_objs[:3])}. "
                            f"Action: Create single parameterized procedure accepting variant/offset/source as parameter. "
                            f"Affected jobs: {', '.join(jobs[:3])}"
                            + (f", ..." if len(jobs) > 3 else "")
                        )

        # B: Step-by-step pass-through fold with job→SQL mapping
        elif "Pass-Through" in category:
            tgt = rec.get("target_table", "")
            affected_jobs = rec.get("affected_jobs", [])
            if len(affected_jobs) >= 1:
                main_job = affected_jobs[0]
                main_sql_line = _fmt_job_sql(main_job, sql_index)
                rec["recommendation"] = (
                    f"SCOPE: LOCAL \u2014 execute within this RPT only.\n\n"
                    f"Step 1  [UNDERSTAND THE PATTERN]\n"
                    f"  Job: {main_sql_line}\n"
                    f"  This job both reads AND writes '{tgt}' in the same execution \u2014 "
                    f"a post-load update hop that adds an unnecessary pipeline step.\n\n"
                    f"Step 2  [IDENTIFY THE UPSTREAM PRODUCER]\n"
                    f"  Find the upstream job that first INSERTs / MERGEs rows into '{tgt}'.\n"
                    f"  (Check the Tidal dependency graph \u2014 it will be a direct parent of '{main_job}'.)\n\n"
                    f"Step 3  [MERGE THE LOGIC]\n"
                    f"  Open the upstream job's SQL procedure.\n"
                    f"  Move the UPDATE columns and WHERE conditions from '{main_job}'\n"
                    f"  into the upstream INSERT/MERGE statement\n"
                    f"  (e.g. as CASE expressions or additional SET clauses in a MERGE).\n\n"
                    f"Step 4  [VALIDATE]\n"
                    f"  Run 3+ full pipeline cycles.\n"
                    f"  Verify row counts and all affected columns match the pre-change baseline.\n\n"
                    f"Step 5  [DECOMMISSION]\n"
                    f"  \u2022 Remove '{main_job}' from the Tidal schedule.\n"
                    f"  \u2022 Remove its Tidal dependency edges (both inbound and outbound)."
                )

        # D: List the specific chain
        elif "Serial Orchestration Chain" in category:
            chain = rec.get("affected_jobs", [])
            if chain:
                depth_range = rec.get("depth_range", "")
                rec["recommendation"] = (
                    existing_rec + f"\n\n[CONCRETE] Parallelize this chain: "
                    f"{' → '.join(chain[:5])}{'...' if len(chain) > 5 else ''}. "
                    f"Action: Rewrite Tidal/orchestration to run jobs concurrently (no data dependencies). "
                    f"Expected critical-path reduction: {depth_range}."
                )

        # E: Append job→SQL detail lines to MV elimination step text
        elif "MV Elimination Candidate" in category:
            affected_jobs = rec.get("affected_jobs", [])
            if len(affected_jobs) >= 2:
                mv_job = affected_jobs[0]
                consumer_job = affected_jobs[1]
                mv_sql = _fmt_job_sql(mv_job, sql_index)
                consumer_sql = _fmt_job_sql(consumer_job, sql_index)
                if "[SQL:" in mv_sql or "[SQL:" in consumer_sql:
                    rec["recommendation"] = (
                        existing_rec
                        + f"\n\n[JOB \u2192 SQL MAPPING]\n"
                        f"  MV refresh job  : {mv_sql}\n"
                        f"  Consumer job    : {consumer_sql}"
                    )

        # E: Append job→SQL detail to MV shared wave steps
        elif "MV Shared Across RPTs" in category:
            affected_jobs = rec.get("affected_jobs", []) or []
            if len(affected_jobs) > 1:
                consumer_lines = []
                for j in affected_jobs[1:]:  # first entry is the MV job itself
                    line = _fmt_job_sql(j, sql_index)
                    consumer_lines.append(f"  {line}")
                if any("[SQL:" in ln for ln in consumer_lines):
                    rec["recommendation"] = (
                        existing_rec
                        + f"\n\n[JOB \u2192 SQL MAPPING \u2014 Consumer job SQL procedures]\n"
                        + "\n".join(consumer_lines)
                    )

        # N: Add sequencing-focused actions
        elif "Post-Consumption Update Pattern" in category:
            y = rec.get("target_table", "")
            ordering = rec.get("ordering_signal", "UNORDERED")
            consumer = rec.get("consumer_job", "")
            updater = rec.get("updater_job", "")
            targets = rec.get("consumer_targets", [])
            rec["recommendation"] = (
                existing_rec
                + f"\n\n[CONCRETE] Table sequencing check for {y}: consumer={consumer}, updater={updater}, ordering={ordering}. "
                + f"Consumer outputs: {', '.join(targets[:2]) if targets else 'unknown'}. "
                + "Action: compare column-level dependencies; if updater columns are needed by consumer output, execute updater before consumer or fold into consumer SQL; otherwise isolate updater as post-processing and remove redundant orchestration hops."
            )

    return recommendations


def sanitize_console_text(text):
    if text is None:
        return ""
    return (
        str(text)
        .replace("\u2265", ">=")
        .replace("\u2264", "<=")
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u2019", "'")
        .replace("\u2192", "->")
    )


# ── Log / debug / audit table filter ────────────────────────────────────────
# Execution-log, debug-trace, and audit side-effect tables are written as a
# by-product of running ETL jobs (progress tracking, error logging, debug
# tracing).  They are NOT part of the data-movement pipeline and should be
# suppressed from hop-reduction recommendations to avoid noise.
# Examples: PRCS_GRP_TBL_LOAD_DEBUG_TRC, FCT_PROC_EXEC_STATUS_LOG_R

_LOG_DEBUG_SUBSTRINGS: tuple[str, ...] = (
    "_DEBUG_TRC", "_TRC",
    "EXEC_STATUS_LOG", "PROC_EXEC_STATUS",
    "_JOB_LOG", "PRCS_JOB_LOG", "JOB_LOG_MESSAGE",
    "_AUDIT_LOG", "AUDIT_LOG", "LOG_TABLE",
    "PRCS_GRP_TBL_LOAD_DEBUG",
    # ── Watermark / incremental-load control tables ──────────────────────────
    # SSL_PACKAGE_MILESTONE_TABLE stores job start/end timestamps used as
    # high-water marks for incremental delta filtering.  The pattern
    # "job reads END_DATE → loads data → updates END_DATE" is intentional
    # control-flow, not a data-pipeline hop.  Suppress from all detectors.
    "MILESTONE_TABLE", "_MILESTONE_",
)

_LOG_DEBUG_PREFIXES: tuple[str, ...] = (
    "PRCS_JOB_LOG_",
    "PRCS_GRP_TBL_",
    "SSL_PACKAGE_",    # SSL_PACKAGE_MILESTONE_TABLE and any similar SSL control tables
)


def _is_log_or_debug_table(table_name: str) -> bool:
    """Return True for execution-log, debug-trace, audit side-effect tables,
    AND for generic SQL alias artifacts that are not real Oracle table names.
    These should not be flagged as hop-reduction targets."""
    t = table_name.upper().strip()
    # ── Log / debug / audit tables ───────────────────────────────────────────
    if (
        any(sub in t for sub in _LOG_DEBUG_SUBSTRINGS)
        or any(t.startswith(pfx.upper()) for pfx in _LOG_DEBUG_PREFIXES)
    ):
        return True
    # ── Generic SQL alias artifacts ──────────────────────────────────────────
    # Names like 'TAB1', 'TAB', 'T1', 'T4418483', 'A', 'B' are SQL aliases
    # captured incorrectly by the lineage parser — not real Oracle table names.
    # Real Oracle table names have business-meaningful names (4+ chars with
    # underscores and domain terms). Filter out short generic patterns.
    import re as _re_alias
    if _re_alias.fullmatch(r'TAB\d*', t):          # TAB, TAB1, TAB2 ...
        return True
    if _re_alias.fullmatch(r'T\d+', t):             # T1, T2, T4418483 ...
        return True
    if _re_alias.fullmatch(r'[A-Z]\d*', t):         # A, B, C, A1 (single-letter aliases)
        return True
    if len(t) <= 3 and t.isalpha():                 # Very short non-underscore names
        return True
    return False


# ── Source-system split guard ────────────────────────────────────────────────
# Jobs split along these source-system boundaries are INTENTIONALLY separated
# to allow independent hold management per source.  Consolidating them would
# couple their scheduling dependencies — if one source is placed on hold it
# would block all others.  (Validated by data team: Jul-2026)

_SOURCE_SYSTEM_MARKERS: dict[str, list[str]] = {
    "CV":     ["_CV_", "_CV-", "-CV-", "-CV_"],
    "PACS":   ["_PACS_", "_PACS-", "-PACS-", "-PACS_"],
    "VUE":    ["_VUE_", "_VUE-", "-VUE-"],
    "APS":    ["_APS_", "_APS-", "-APS-"],
    "SHINKA": ["_SHINKA_", "_SHINKA-"],
    "EIS":    ["_EIS_", "_EIS-"],
}


def _detect_source_system(job_name: str) -> str | None:
    """Return the source-system label embedded in a Tidal job name, or None."""
    j = job_name.upper()
    for sys_name, markers in _SOURCE_SYSTEM_MARKERS.items():
        if any(m in j for m in markers):
            return sys_name
    return None


def _is_source_system_split(jobs) -> bool:
    """Return True when the job set spans ≥2 distinct source systems."""
    systems = {_detect_source_system(j) for j in jobs}
    systems.discard(None)
    return len(systems) >= 2


def _source_system_names(jobs) -> str:
    """Comma-separated sorted list of distinct source systems found in the job names."""
    systems = {_detect_source_system(j) for j in jobs}
    systems.discard(None)
    return ", ".join(sorted(systems)) if systems else "unknown"


# ── Full-load / truncate-and-reload job guard (used by N detector) ───────────
# Jobs whose names match these patterns perform a TRUNCATE + full INSERT rather
# than an incremental UPDATE.  They are producers, not updaters, so they must
# not be classified as "updaters" in the post-consumption-update (N) pattern.

_FULL_LOAD_JOB_TOKENS = (
    "_LOAD_RPT_", "_FULLLOAD_", "_FULL_LOAD_", "_TRUNC_LOAD_", "_RELOAD_",
)

# Tokens that identify post-load in-place updater jobs (used by both
# detect_mv_elimination and detect_post_load_mv_indicator_chain).
_UPD_JOB_TOKENS = ("_UPD_", "_UPDATE_", "UPD_IND", "UPD_COL", "UPDATE_IND",)


def _is_full_load_job(job_name: str) -> bool:
    """Return True for jobs that truncate-and-reload a table (not update-in-place)."""
    j = job_name.upper()
    return any(tok in j for tok in _FULL_LOAD_JOB_TOKENS)


def _is_updater_job(job_name: str) -> bool:
    """Return True if the job name signals an in-place update (not a full reload)."""
    j = job_name.upper()
    return any(tok in j for tok in _UPD_JOB_TOKENS)


# ── Detector A ──────────────────────────────────────────────────────────────


def detect_redundant_multi_source(lineage_rows, sql_index=None):
    """A. Multiple jobs loading same target table. Skip legit MULTI_SRC splits.
    Uses SQL object index to suppress recommendations where consolidation is already done.
    """
    if sql_index is None:
        sql_index = {}
    tgt_jobs = defaultdict(list)
    for r in lineage_rows:
        tgt = r.get("TGT_TABLE")
        job = r.get("DEPENDENT_JOB")
        if tgt and tgt not in ("None", "N/A", "-") and job:
            # Skip synthetic VIEW rows — a VIEW definition is not a "load" job
            if r.get("LINEAGE_SOURCE") == "VIEW_EXPANSION" or r.get("JOB_CATEGORY") == "VIEW":
                continue
            tgt_jobs[tgt].append(
                {
                    "job": job,
                    "pattern": r.get("PATTERN_TYPE", ""),
                    "src": r.get("SRC_TABLE", ""),
                    "params": (r.get("TIDAL_PARAMS") or "").upper(),
                }
            )

    recommendations = []
    for tgt, jobs in sorted(tgt_jobs.items()):
        unique_jobs = {j["job"] for j in jobs}
        if len(unique_jobs) < 2:
            continue
        params_set = {j["params"] for j in jobs}
        if all("MULTI_SRC" in p for p in params_set):
            continue
        # Skip log/debug/audit side-effect tables — not part of the data pipeline
        if _is_log_or_debug_table(tgt):
            continue

        sql_objs = get_jobs_sql_summary(unique_jobs, sql_index)
        scope = classify_confidence_scope(unique_jobs, sql_index, rpt_consumer_count=len(unique_jobs))

        # If all jobs call same SQL object and already parameterized -> Tidal IS the param layer
        if scope == "ALREADY_PARAMETERIZED":
            sql_obj_str = sql_objs[0] if sql_objs else "unknown"
            if _is_source_system_split(unique_jobs):
                # Intentional per-source-system separation — do NOT suggest job reduction
                sys_names = _source_system_names(unique_jobs)
                recommendations.append(
                    {
                        "id": f"A-{len(recommendations)+1}",
                        "category": "A. Source System Architectural Split (By Design)",
                        "target_table": tgt,
                        "description": (
                            f"Table '{tgt}' loaded by {len(unique_jobs)} jobs from distinct "
                            f"source systems ({sys_names}), all calling '{sql_obj_str}'. "
                            f"Jobs are intentionally separated for independent source-system "
                            f"hold management — combining them would couple SLA dependencies."
                        ),
                        "affected_jobs": sorted(unique_jobs),
                        "sql_objects_called": sql_objs,
                        "confidence_scope": "SOURCE_SYSTEM_SPLIT",
                        "current_hops": len(unique_jobs),
                        "potential_hop_savings": 0,
                        "risk": "LOW",
                        "recommendation": (
                            f"Do not consolidate. Jobs are intentionally separated by source "
                            f"system ({sys_names}) so each source can be held independently "
                            f"without impacting others. Combining them would introduce SLA risk."
                        ),
                        "_dedup_key": ("A", tgt),
                    }
                )
            else:
                recommendations.append(
                    {
                        "id": f"A-{len(recommendations)+1}",
                        "category": "A. Already Parameterized via Tidal (No Action Needed)",
                        "target_table": tgt,
                        "description": (
                            f"Table '{tgt}' loaded by {len(unique_jobs)} jobs all calling "
                            f"'{sql_obj_str}' with different Tidal params. "
                            f"Consolidation is ALREADY done at SQL level - Tidal handles parameterization."
                        ),
                        "affected_jobs": sorted(unique_jobs),
                        "sql_objects_called": sql_objs,
                        "confidence_scope": scope,
                        "current_hops": len(unique_jobs),
                        "potential_hop_savings": 0,
                        "risk": "LOW",
                        "recommendation": (
                            "No SQL-level action needed. Review if the number of Tidal jobs "
                            "can be reduced (e.g. by combining source-system params into one call)."
                        ),
                        "_dedup_key": ("A", tgt),
                    }
                )
            continue

        # Source-system architectural split guard for non-parameterized cases.
        # ONLY fire when jobs all share the same SQL object family (SAME_PKG_VARIANT
        # or a single unique SQL object) — that means the source split is the ONLY
        # difference and combining them truly adds no value.
        # When jobs call DIFFERENT SQL objects (DIFF_SQL_OBJECTS scope), the procedures
        # are genuinely separate and can be consolidated: treat as Redundant Multi-Source Load.
        same_sql_family = (
            scope in ("SAME_PKG_VARIANT",)
            or (len({o.split(".")[0] for o in sql_objs if o}) <= 1)  # single package prefix
        )
        if _is_source_system_split(unique_jobs) and same_sql_family:
            sys_names = _source_system_names(unique_jobs)
            recommendations.append(
                {
                    "id": f"A-{len(recommendations)+1}",
                    "category": "A. Source System Architectural Split (By Design)",
                    "target_table": tgt,
                    "description": (
                        f"Table '{tgt}' loaded by {len(unique_jobs)} jobs from distinct "
                        f"source systems ({sys_names}). Separation is intentional for "
                        f"independent source-system hold management."
                    ),
                    "affected_jobs": sorted(unique_jobs),
                    "sql_objects_called": sql_objs,
                    "confidence_scope": "SOURCE_SYSTEM_SPLIT",
                    "current_hops": len(unique_jobs),
                    "potential_hop_savings": 0,
                    "risk": "LOW",
                    "recommendation": (
                        f"Do not consolidate. Jobs are intentionally separated by source "
                        f"system ({sys_names}) so each source can be held independently "
                        f"without impacting others."
                    ),
                    "_dedup_key": ("A", tgt),
                }
            )
            continue

        cv_jobs = [j for j in jobs if "SHINKA" in j["job"] or "CV" in j["job"]]
        pacs_jobs = [j for j in jobs if "PACS" in j["job"]]
        if cv_jobs and pacs_jobs:
            desc = (
                f"Table '{tgt}' loaded by {len(unique_jobs)} jobs "
                f"(CV_SHINKA + PACS). SQL objects called: {sql_objs or ['unknown']}."
            )
            if scope == "SAME_PKG_VARIANT":
                rec_text = (
                    "Jobs call different variants of the same base package. "
                    "Evaluate if PKG and PKG_PD can be unified with a source-system parameter."
                )
            else:
                rec_text = (
                    "Jobs call different SQL objects. "
                    "Investigate if loads can be consolidated into a single parameterized procedure."
                )
        else:
            desc = f"Table '{tgt}' loaded by {len(unique_jobs)} jobs: {sorted(unique_jobs)}. SQL: {sql_objs or ['unknown']}"
            rec_text = "Investigate if loads can be consolidated."

        recommendations.append(
            {
                "id": f"A-{len(recommendations)+1}",
                "category": "A. Redundant Multi-Source Load",
                "target_table": tgt,
                "description": desc,
                "affected_jobs": sorted(unique_jobs),
                "sql_objects_called": sql_objs,
                "confidence_scope": scope,
                "current_hops": len(unique_jobs),
                "potential_hop_savings": len(unique_jobs) - 1,
                "risk": "MEDIUM",
                "recommendation": rec_text,
                "_dedup_key": ("A", tgt),
            }
        )
    return recommendations


# ── Detector B ──────────────────────────────────────────────────────────────


def detect_pass_through(graph, lineage_rows):
    """B. SRC∩TGT update-in-place + graph-walked single-producer chain pass-through."""
    recommendations = []
    table_map = build_table_producers_consumers(graph)

    col_lineage = defaultdict(set)
    for r in lineage_rows:
        tgt_t, src_t = r.get("TGT_TABLE"), r.get("SRC_TABLE")
        if not (tgt_t and src_t):
            continue
        tgt_cols = parse_col_list(r.get("TARGET_COL"))
        src_cols = parse_col_list(r.get("SOURCE_COL"))
        if len(tgt_cols) == len(src_cols) and tgt_cols:
            for s, t in zip(sorted(src_cols), sorted(tgt_cols)):
                col_lineage[(tgt_t, t)].add((src_t, s))

    for node in graph["nodes"]:
        src = set(node.get("src_tables", []) or [])
        tgt = set(node.get("tgt_tables", []) or [])
        overlap = src & tgt
        if not overlap:
            continue
        # Filter out log/debug/audit side-effect tables and final RPT output tables.
        # Also skip GTT (Oracle Global Temporary Tables) — they are session-scoped
        # and reading/writing the same GTT within one job is by design.
        overlap = {t for t in overlap
                   if not _is_log_or_debug_table(t)
                   and not t.upper().startswith("RPT_")
                   and not t.upper().endswith("_GTT")}
        if not overlap:
            continue
        sql_objs = [o["name"] for o in node.get("sql_objects", [])]
        recommendations.append(
            {
                "id": f"B-{len(recommendations)+1}",
                "category": "B. Pass-Through / Update-in-Place",
                "target_table": ", ".join(sorted(overlap)),
                "description": (
                    f"Job '{node['id']}' reads and writes the same table(s): "
                    f"{sorted(overlap)}. This is a post-load update hop."
                ),
                "affected_jobs": [node["id"]],
                "sql_objects": sql_objs,
                "depth": node.get("depth"),
                "current_hops": 1,
                "potential_hop_savings": 1,
                "risk": "MEDIUM",
                "recommendation": (
                    "Fold UPDATE logic into the upstream INSERT/MERGE job "
                    "that originally created the table rows."
                ),
                "_dedup_key": ("B", tuple(sorted(overlap)), node["id"]),
            }
        )

    for node in graph["nodes"]:
        for src_t in node.get("src_tables", []) or []:
            producers = table_map[src_t]["producers"]
            consumers = table_map[src_t]["consumers"]
            if len(producers) != 1 or len(consumers) != 1:
                continue
            for tgt_t in node.get("tgt_tables", []) or []:
                pass_cols = sum(
                    1
                    for (tt, _tc), srcs in col_lineage.items()
                    if tt == tgt_t and any(s == src_t for s, _ in srcs)
                )
                if pass_cols < 5:
                    continue
                key = ("B-chain", src_t, tgt_t)
                if any(r.get("_dedup_key") == key for r in recommendations):
                    continue
                recommendations.append(
                    {
                        "id": f"B-{len(recommendations)+1}",
                        "category": "B. Pass-Through / Update-in-Place",
                        "target_table": tgt_t,
                        "description": (
                            f"Single-producer table '{src_t}' is consumed only by "
                            f"job '{node['id']}' (-> {tgt_t}). {pass_cols} columns "
                            f"flow through unchanged - candidate for inlining."
                        ),
                        "affected_jobs": sorted(producers | {node["id"]}),
                        "depth": node.get("depth"),
                        "current_hops": 2,
                        "potential_hop_savings": 1,
                        "risk": "MEDIUM",
                        "recommendation": (
                            f"Inline producer logic of '{src_t}' into the "
                            f"'{node['id']}' INSERT/MERGE statement."
                        ),
                        "_dedup_key": key,
                    }
                )
    return recommendations


# ── Detector C ──────────────────────────────────────────────────────────────


def detect_offset_consolidation(graph, lineage_rows, sql_index=None):
    """C. Near-duplicate procedures loading same target (Jaccard col-set clustering).
    Suppresses recs where all jobs call the same SQL object (already parameterized via Tidal).
    """
    if sql_index is None:
        sql_index = {}
    job_col_sets = defaultdict(set)
    for r in lineage_rows:
        job = r.get("DEPENDENT_JOB")
        tgt = r.get("TGT_TABLE")
        if not (job and tgt):
            continue
        job_col_sets[(job, tgt)] |= parse_col_list(r.get("TARGET_COL"))
        
        # Track if this job is parameterized (new column in enhanced data)
        is_param_val = r.get("IS_PARAMETERIZED")
        if is_param_val and str(is_param_val).lower() == "true":
            job_col_sets[(job, tgt)].add("__IS_PARAMETERIZED__")  # marker

    tgt_to_jobs = defaultdict(list)
    for (job, tgt), cols in job_col_sets.items():
        if cols:
            tgt_to_jobs[tgt].append((job, cols))

    recommendations = []
    seen = set()
    for tgt, jobs in tgt_to_jobs.items():
        if len(jobs) < 2:
            continue
        # Skip log/debug/audit side-effect tables
        if _is_log_or_debug_table(tgt):
            continue
        clusters = []
        for j_a, cols_a in jobs:
            placed = False
            for cluster in clusters:
                if all(jaccard(cols_a, c) >= 0.85 for _, c in cluster):
                    cluster.append((j_a, cols_a))
                    placed = True
                    break
            if not placed:
                clusters.append([(j_a, cols_a)])

        for cluster in clusters:
            if len(cluster) < 2:
                continue
            jobs_in_cluster = sorted(j for j, _ in cluster)
            key = (tgt, tuple(jobs_in_cluster))
            if key in seen:
                continue
            seen.add(key)

            is_offset = any("OFFSET" in j.upper() for j in jobs_in_cluster)
            is_param = any(
                "__IS_PARAMETERIZED__" in parse_col_list(r.get("TARGET_COL"))
                for r in lineage_rows
                if r.get("DEPENDENT_JOB") in jobs_in_cluster
            )

            # Check SQL object scope
            sql_objs = get_jobs_sql_summary(jobs_in_cluster, sql_index)
            scope = classify_confidence_scope(
                jobs_in_cluster, sql_index, rpt_consumer_count=len(jobs_in_cluster)
            )

            if scope == "ALREADY_PARAMETERIZED":
                # Same SQL object, already parameterized - Tidal IS the consolidation layer
                sql_obj_str = sql_objs[0] if sql_objs else "unknown"
                if _is_source_system_split(jobs_in_cluster):
                    sys_names = _source_system_names(jobs_in_cluster)
                    recommendations.append(
                        {
                            "id": f"C-{len(recommendations)+1}",
                            "category": "C. Source System Architectural Split (By Design)",
                            "target_table": tgt,
                            "description": (
                                f"{len(cluster)} jobs load '{tgt}' from distinct source "
                                f"systems ({sys_names}), all calling '{sql_obj_str}'. "
                                f"Jobs are intentionally separated for independent "
                                f"source-system hold management."
                            ),
                            "affected_jobs": jobs_in_cluster,
                            "sql_objects_called": sql_objs,
                            "confidence_scope": "SOURCE_SYSTEM_SPLIT",
                            "current_hops": len(cluster),
                            "potential_hop_savings": 0,
                            "risk": "LOW",
                            "recommendation": (
                                f"Do not consolidate. Jobs are intentionally separated by "
                                f"source system ({sys_names}) so each source can be placed "
                                f"on hold independently without impacting others. "
                                f"Combining them would couple SLA dependencies."
                            ),
                            "_dedup_key": ("C", tgt, tuple(jobs_in_cluster)),
                        }
                    )
                else:
                    recommendations.append(
                        {
                            "id": f"C-{len(recommendations)+1}",
                            "category": "C. Already Parameterized via Tidal (No Action Needed)",
                            "target_table": tgt,
                            "description": (
                                f"{len(cluster)} jobs load '{tgt}' via the SAME SQL object "
                                f"'{sql_obj_str}' with different Tidal parameters. "
                                f"SQL-level parameterization is ALREADY in place."
                            ),
                            "affected_jobs": jobs_in_cluster,
                            "sql_objects_called": sql_objs,
                            "confidence_scope": scope,
                            "current_hops": len(cluster),
                            "potential_hop_savings": 0,
                            "risk": "LOW",
                            "recommendation": (
                                "No SQL consolidation needed. If reducing Tidal job count is the goal, "
                                "evaluate whether multiple source systems can be combined in a single Tidal job call."
                            ),
                            "_dedup_key": ("C", tgt, tuple(jobs_in_cluster)),
                        }
                    )
                continue

            label = (
                "OFFSET / Near-Duplicate Procedures (Parameterized)"
                if is_offset and is_param
                else ("OFFSET / Near-Duplicate Procedures" if is_offset else "Near-Duplicate Procedures")
            )
            if scope == "SAME_PKG_VARIANT":
                rec_text = (
                    "Jobs call different variants of the same base package. "
                    "Evaluate if variants can be unified with a single parameterized package entry point."
                )
            else:
                rec_text = (
                    "Consolidate into single parameterized procedure that accepts "
                    "partition / offset key as a parameter."
                )
            recommendations.append(
                {
                    "id": f"C-{len(recommendations)+1}",
                    "category": f"C. {label}",
                    "target_table": tgt,
                    "description": (
                        f"{len(cluster)} jobs load '{tgt}' with >=85% identical "
                        f"target column sets. SQL objects called: {sql_objs or ['unknown']}."
                    ),
                    "affected_jobs": jobs_in_cluster,
                    "sql_objects_called": sql_objs,
                    "confidence_scope": scope,
                    "current_hops": len(cluster),
                    "potential_hop_savings": len(cluster) - 1,
                    "risk": "LOW",
                    "recommendation": rec_text,
                    "_dedup_key": ("C", tgt, tuple(jobs_in_cluster)),
                }
            )
    return recommendations


# ── Detector D ──────────────────────────────────────────────────────────────


def detect_serial_orchestration_chains(graph, lineage_rows):
    """D. Serial chains of OTHER-category jobs with NO column data dependency."""
    children, parents = build_adjacency(graph)
    node_map = {n["id"]: n for n in graph["nodes"]}

    job_writes = defaultdict(set)
    for r in lineage_rows:
        job = r.get("DEPENDENT_JOB")
        tgt = r.get("TGT_TABLE")
        if job and tgt:
            for c in parse_col_list(r.get("TARGET_COL")):
                job_writes[job].add((tgt, c))

    other_nodes = {
        n["id"]
        for n in graph["nodes"]
        if n["category"] == "OTHER"
        and not n.get("sql_objects")
        and n.get("depth", -1) >= 3
    }
    if len(other_nodes) < 2:
        return []

    visited = set()
    chains = []
    for nid in sorted(other_nodes):
        if nid in visited:
            continue
        chain = [nid]
        visited.add(nid)
        current = nid
        while True:
            ds = [
                c
                for c in children.get(current, set())
                if c in other_nodes and c not in visited
            ]
            if len(ds) == 1:
                current = ds[0]
                chain.append(current)
                visited.add(current)
            else:
                break
        current = nid
        while True:
            us = [
                p
                for p in parents.get(current, set())
                if p in other_nodes and p not in visited
            ]
            if len(us) == 1:
                current = us[0]
                chain.insert(0, current)
                visited.add(current)
            else:
                break
        if len(chain) >= 2:
            chains.append(chain)

    recommendations = []
    for chain in chains:
        has_real_data_dep = any(
            (job_writes.get(chain[i]) or set())
            & (job_writes.get(chain[i + 1]) or set())
            for i in range(len(chain) - 1)
        )
        if has_real_data_dep:
            continue
        depths = [node_map[j].get("depth", -1) for j in chain if j in node_map]
        if not depths:
            continue
        recommendations.append(
            {
                "id": f"D-{len(recommendations)+1}",
                "category": "D. Serial Orchestration Chain (No Data Dep)",
                "target_table": "N/A (orchestration only)",
                "description": (
                    f"Chain of {len(chain)} jobs with no overlapping column "
                    f"writes. Safe to parallelize."
                ),
                "affected_jobs": chain,
                "depth_range": f"{min(depths)}-{max(depths)}",
                "current_hops": len(chain),
                "potential_hop_savings": max(depths) - min(depths),
                "risk": "LOW",
                "recommendation": (
                    f"Run in parallel. Critical-path depth: "
                    f"{max(depths)} -> {min(depths)}."
                ),
                "_dedup_key": ("D", tuple(chain)),
            }
        )
    return recommendations


# ── Detector E ──────────────────────────────────────────────────────────────


def detect_mv_elimination(
    graph,
    lineage_rows,
    global_children=None,
    global_mv_context=None,
    global_job_rpt_map=None,
    global_runtime_index=None,
    summaries=None,
):
    """E. MVs with single consumer; downgrade risk if consumer uses <50% of cols."""
    children, _ = build_adjacency(graph)
    node_map = {n["id"]: n for n in graph.get("nodes", [])}
    global_mv_context = global_mv_context or {}
    global_job_rpt_map = global_job_rpt_map or {}
    global_runtime_index = global_runtime_index or {}
    summaries = summaries or {}

    # Build a set of all table names mentioned as INPUT sources in ANY script
    # summary.  Column-level lineage has parser gaps (JOIN-only references are
    # not captured), so a table appearing in summaries' INPUT_DEPENDENCIES is
    # evidence it IS being actively used even if lineage shows 0 readers.
    # This set is used in the dead-MV check below to suppress false positives.
    def _extract_tbl(entry: str) -> str:
        name = entry.split("(")[0].split(" ")[0].strip().upper()
        return name.rsplit(".", 1)[-1] if "." in name else name

    _summary_input_sources: set[str] = set()
    for sname, sdata in summaries.items():
        if not isinstance(sdata, dict):
            continue
        deps = sdata.get("INPUT_DEPENDENCIES", {})
        if isinstance(deps, dict):
            src_tbls = deps.get("source_tables", {})
            if isinstance(src_tbls, dict):
                for cat, lst in src_tbls.items():
                    if isinstance(lst, list):
                        for entry in lst:
                            if isinstance(entry, str):
                                nm = _extract_tbl(entry)
                                if nm:
                                    _summary_input_sources.add(nm)
            proc_deps = deps.get("procedure_level_dependencies") or {}
            if isinstance(proc_deps, dict):
                for items in proc_deps.values():
                    if isinstance(items, list):
                        for entry in items:
                            if isinstance(entry, str):
                                nm = _extract_tbl(entry)
                                if nm and not nm.startswith("PKG_"):
                                    _summary_input_sources.add(nm)

    job_writes_cols = defaultdict(lambda: defaultdict(set))
    for r in lineage_rows:
        job = r.get("DEPENDENT_JOB")
        tgt = r.get("TGT_TABLE")
        if job and tgt:
            job_writes_cols[job][tgt] |= parse_col_list(r.get("TARGET_COL"))

    job_reads_cols = defaultdict(lambda: defaultdict(set))
    for r in lineage_rows:
        job = r.get("DEPENDENT_JOB")
        src = r.get("SRC_TABLE")
        if job and src:
            job_reads_cols[job][src] |= parse_col_list(r.get("SOURCE_COL"))

    mv_nodes = [n for n in graph["nodes"] if n["category"] == "MV"]
    recommendations = []
    for mv in mv_nodes:
        # Step 1 — filter out VIEW_ synthetic nodes (SQL VIEW definitions, not Tidal jobs)
        # Step 2 — filter to DIRECT data consumers only: the consumer's src_tables must
        #           contain at least one of the MV's tgt_tables.  Jobs connected only via
        #           Tidal scheduling/orchestration edges or through an intermediate SQL VIEW
        #           are NOT real data consumers and would produce a false positive E. finding.
        mv_tgts = {t.upper() for t in (mv.get("tgt_tables") or [])}
        consumers = {
            c for c in children.get(mv["id"], set())
            if not c.upper().startswith("VIEW_")
            and mv_tgts & {
                s.upper()
                for s in (node_map.get(c, {}).get("src_tables") or [])
            }
        }
        g_info = global_mv_context.get(mv["id"], {})
        global_consumers = {c for c in g_info.get("consumers", set())
                             if not c.upper().startswith("VIEW_")}
        if not global_consumers:
            base = global_children.get(mv["id"], set()) if global_children else consumers
            # Apply the same direct-consumer filter to the fallback base set
            global_consumers = {
                c for c in base
                if not c.upper().startswith("VIEW_")
                and mv_tgts & {
                    s.upper()
                    for s in (node_map.get(c, {}).get("src_tables") or [])
                }
            }

        # If no direct data consumers exist anywhere, skip: this MV is consumed only
        # through a SQL VIEW (lineage gap) — not a Tidal hop reduction opportunity.
        if not consumers and not global_consumers:
            continue

        impacted_rpts = set(g_info.get("rpts", set()))
        if not impacted_rpts:
            for c in global_consumers:
                impacted_rpts |= set(global_job_rpt_map.get(c, set()))
        if not impacted_rpts:
            impacted_rpts = {RPT_TABLE}

        mv_runtime = global_runtime_index.get(mv["id"], 0.0)
        consumer_runtime_pairs = sorted(
            [(c, global_runtime_index.get(c, 0.0)) for c in global_consumers],
            key=lambda x: x[1],
        )
        global_runtime_affected_min = round(
            (mv_runtime + sum(rt for _c, rt in consumer_runtime_pairs)) / 60.0,
            2,
        )
        global_est_saved_min = round(mv_runtime / 60.0, 2) if mv_runtime else 0.0
        blast_radius = min(len(global_consumers) + max(len(impacted_rpts) - 1, 0), 10)

        wave1_n = max(1, (len(consumer_runtime_pairs) + 2) // 3) if consumer_runtime_pairs else 0
        wave1 = [c for c, _rt in consumer_runtime_pairs[:wave1_n]]
        wave2 = [c for c, _rt in consumer_runtime_pairs[wave1_n:]]

        # Pre-build reusable display variables for both E-category branches
        _tgt_display = ", ".join(sorted(mv.get("tgt_tables", []) or [])) or mv["id"]
        _wave1_bullets = "\n".join(f"     \u2022 {c}" for c in wave1) if wave1 else "     (none)"
        _wave2_bullets = "\n".join(f"     \u2022 {c}" for c in wave2) if wave2 else ""
        _decom_steps = (
            f"  \u2022 Remove Tidal job '{mv['id']}' from the Tidal schedule.\n"
            f"  \u2022 Drop table '{_tgt_display}' from the database.\n"
            f"  \u2022 Remove all Tidal scheduler dependency edges pointing to this job."
        )

        # If globally shared, do not recommend elimination based on single-RPT view.
        if len(global_consumers) > 1:
            # MV complexity for shared MVs
            mv_src_count_shared = len(mv.get("src_tables") or [])
            if mv_src_count_shared >= 4:
                shared_complexity_note = (
                    f"\nCOMPLEXITY: HIGH — MV reads from {mv_src_count_shared} source tables. "
                    f"Full SQL analysis required before inlining across all {len(global_consumers)} consumers."
                )
            elif mv_src_count_shared >= 2:
                shared_complexity_note = (
                    f"\nCOMPLEXITY: MEDIUM — MV reads from {mv_src_count_shared} source tables. "
                    f"Review JOIN logic carefully before inlining."
                )
            else:
                shared_complexity_note = ""
            _shared_wave_steps = (
                f"Step 2  [WAVE 1 \u2014 migrate lowest-runtime consumers first]\n{_wave1_bullets}\n"
                f"  \u2192 In each consumer's SQL procedure, replace all references to\n"
                f"    '{_tgt_display}' with the MV's source query inlined as a CTE or subquery.\n"
                f"  \u2192 Keep the MV refresh job running until this wave is validated.\n\n"
            )
            if _wave2_bullets:
                _shared_wave_steps += (
                    f"Step 3  [WAVE 2 \u2014 remaining consumers]\n{_wave2_bullets}\n"
                    f"  \u2192 Same as Step 2: inline source query, remove MV reference.\n\n"
                    f"Step 4  [PARITY] After each wave completes, run 3+ full pipeline cycles.\n"
                    f"  Compare row counts and key metrics against the Step 1 baseline.\n\n"
                    f"Step 5  [DECOMMISSION] After all consumers pass parity:\n{_decom_steps}"
                )
            else:
                _shared_wave_steps += (
                    f"Step 3  [PARITY] Run 3+ full pipeline cycles.\n"
                    f"  Compare row counts and key metrics against the Step 1 baseline.\n\n"
                    f"Step 4  [DECOMMISSION] After all consumers pass parity:\n{_decom_steps}"
                )

            recommendations.append(
                {
                    "id": f"E-{len(recommendations)+1}",
                    "category": "E. MV Shared Across RPTs (No Blind Elimination)",
                    "target_table": _tgt_display,
                    "description": (
                        f"WHAT: '{_tgt_display}' is a pre-computed snapshot table "
                        f"(Materialized View) refreshed by Tidal job '{mv['id']}'.\n"
                        f"SHARED: It feeds {len(global_consumers)} consumer(s) across "
                        f"{len(impacted_rpts)} RPT pipeline(s) "
                        f"({', '.join(sorted(impacted_rpts))}).\n"
                        f"IN THIS RPT: {len(consumers)} consumer(s) \u2014 "
                        f"{', '.join(sorted(consumers)[:4])}"
                        f"{'...' if len(consumers) > 4 else ''}.\n"
                        f"Cannot be removed for one RPT in isolation \u2014 a coordinated "
                        f"global migration across all {len(global_consumers)} consumer(s) is required."
                        + shared_complexity_note
                    ),
                    "affected_jobs": [mv["id"]] + sorted(global_consumers),
                    "impacted_rpts": sorted(impacted_rpts),
                    "consumers_by_rpt": {
                        rpt: sorted(jobs)
                        for rpt, jobs in (g_info.get("consumers_by_rpt") or {}).items()
                        if jobs
                    },
                    "confidence_scope": "CROSS_RPT_SHARED",
                    "depth": mv.get("depth"),
                    "current_hops": 1,
                    "potential_hop_savings": 0,
                    "risk": "HIGH" if len(global_consumers) >= 4 else "MEDIUM",
                    "implementation_decision": "EXECUTE_GLOBAL_PROGRAM",
                    "execution_wave": "GLOBAL_WAVE_1_PLUS",
                    "global_rpt_count_impacted": len(impacted_rpts),
                    "global_consumer_count": len(global_consumers),
                    "global_runtime_affected_minutes": global_runtime_affected_min,
                    "global_est_runtime_saved_minutes": global_est_saved_min,
                    "blast_radius_score": blast_radius,
                    "prerequisites": (
                        "Capture row-count + key-metric baseline before Wave 1; "
                        "run parity checks after each wave before proceeding to next."
                    ),
                    "recommendation": (
                        f"SCOPE: GLOBAL MIGRATION \u2014 '{_tgt_display}' feeds "
                        f"{len(global_consumers)} consumer(s) across {len(impacted_rpts)} "
                        f"RPT(s). Do NOT remove it locally.\n\n"
                        f"Step 1  [BASELINE] Before any code changes, record row-count and "
                        f"key-metric baseline for all {len(global_consumers)} consumer(s) "
                        f"listed in the Affected Jobs field.\n\n"
                        + _shared_wave_steps
                    ),
                    "_dedup_key": ("E-shared", mv["id"]),
                }
            )
            continue

        if len(consumers) > 1:
            continue
        sql_objs = [o["name"] for o in mv.get("sql_objects", [])]
        mv_cols = set()
        for tgts in job_writes_cols.get(mv["id"], {}).values():
            mv_cols |= tgts

        # ── MV complexity score ────────────────────────────────────────────────
        # Number of source tables the MV reads from is a proxy for SQL complexity.
        # Complex MVs (≥4 src tables) require detailed analysis before inlining;
        # they should be flagged with HIGH risk and a complexity warning.
        mv_src_count = len(mv.get("src_tables") or [])
        if mv_src_count >= 4:
            complexity_label = "HIGH"
            complexity_note = (
                f"COMPLEXITY: HIGH — MV reads from {mv_src_count} source tables. "
                f"Full SQL analysis required before inlining; logic may be too "
                f"complex to safely embed as a single CTE. Engage the SQL author."
            )
        elif mv_src_count >= 2:
            complexity_label = "MEDIUM"
            complexity_note = (
                f"COMPLEXITY: MEDIUM — MV reads from {mv_src_count} source tables. "
                f"Review JOIN logic carefully before inlining."
            )
        else:
            complexity_label = "LOW"
            complexity_note = ""  # simple projection/filter — no extra warning needed

        usage_note = ""
        risk = "HIGH"
        if consumers and mv_cols:
            consumer = next(iter(consumers))
            consumed = set()
            for mv_tgt in mv.get("tgt_tables") or []:
                consumed |= job_reads_cols.get(consumer, {}).get(mv_tgt, set())
            if consumed:
                ratio = len(consumed & mv_cols) / max(len(mv_cols), 1)
                usage_note = f" Consumer uses {ratio:.0%} of MV columns."
                if ratio < 0.5:
                    risk = "MEDIUM"

        # Complexity always escalates risk to HIGH regardless of column-usage ratio
        if complexity_label == "HIGH":
            risk = "HIGH"

        recommendations.append(
            {
                "id": f"E-{len(recommendations)+1}",
                "category": "E. MV Elimination Candidate",
                "target_table": _tgt_display,
                "description": (
                    f"WHAT: '{_tgt_display}' is a pre-computed snapshot table "
                    f"(Materialized View) refreshed by Tidal job '{mv['id']}'.\n"
                    f"CONSUMERS: {len(consumers)} downstream job(s) in this RPT \u2014 "
                    f"{', '.join(sorted(consumers))}.\n"
                    f"OPPORTUNITY: Single-consumer MVs can be eliminated by inlining "
                    f"the MV's source query directly into the consumer, removing one "
                    f"pre-computation hop from the pipeline."
                    + (f"\n{usage_note.strip()}" if usage_note else "")
                    + (f"\n{complexity_note}" if complexity_note else "")
                ),
                "affected_jobs": [mv["id"]] + sorted(consumers),
                "impacted_rpts": sorted(impacted_rpts),
                "consumers_by_rpt": {
                    rpt: sorted(jobs)
                    for rpt, jobs in (g_info.get("consumers_by_rpt") or {}).items()
                    if jobs
                },
                "confidence_scope": "LOCAL_ONLY" if len(impacted_rpts) == 1 else "CROSS_RPT_SHARED",
                "sql_objects": sql_objs,
                "depth": mv.get("depth"),
                "current_hops": 2,
                "potential_hop_savings": 1,
                "risk": risk,
                "implementation_decision": "EXECUTE_LOCALLY" if len(impacted_rpts) == 1 else "EXECUTE_GLOBAL_PROGRAM",
                "execution_wave": "WAVE_1" if len(impacted_rpts) == 1 else "GLOBAL_WAVE_1_PLUS",
                "global_rpt_count_impacted": len(impacted_rpts),
                "global_consumer_count": len(global_consumers),
                "global_runtime_affected_minutes": global_runtime_affected_min,
                "global_est_runtime_saved_minutes": global_est_saved_min,
                "blast_radius_score": blast_radius,
                "prerequisites": (
                    "Row-count parity, key-metric parity, and 3-cycle runtime comparison"
                    + (
                        f"\nNote: EXECUTE_GLOBAL_PROGRAM is set because the MV refresh "
                        f"job '{mv['id']}' appears as a Tidal scheduling dependency in "
                        f"{len(impacted_rpts)} RPT pipeline(s) "
                        f"({', '.join(sorted(impacted_rpts))}). "
                        f"Even though only {len(global_consumers)} job(s) consume the MV "
                        f"data, removing the refresh job from Tidal requires updating the "
                        f"dependency edges in ALL those RPT pipelines."
                        if len(impacted_rpts) > 1 else ""
                    )
                ),
                "recommendation": (
                    f"SCOPE: LOCAL \u2014 execute within this RPT only.\n\n"
                    f"Step 1  [BASELINE] Record row-count and key-metric baseline for "
                    f"consumer job: {', '.join(sorted(consumers))}.\n\n"
                    f"Step 2  [EXTRACT MV DEFINITION] Open the SQL procedure called by "
                    f"Tidal job '{mv['id']}'. Capture the SELECT query that populates "
                    f"'{_tgt_display}' (the INSERT \u2026 SELECT or MERGE source).\n\n"
                    f"Step 3  [INLINE INTO CONSUMER] Open the SQL procedure called by "
                    f"consumer '{next(iter(sorted(consumers)), 'CONSUMER')}'. "
                    f"Replace the direct table reference to '{_tgt_display}' with the "
                    f"extracted query as a WITH (CTE) clause or subquery.\n\n"
                    f"Step 4  [VALIDATE] Run 3+ full pipeline cycles. "
                    f"Compare row counts and key metrics against the Step 1 baseline.\n\n"
                    f"Step 5  [DECOMMISSION] After parity is confirmed:\n{_decom_steps}"
                ),
                "_dedup_key": ("E", mv["id"]),
            }
        )

    # ── E-sub: Dead-write MV (0 consumers, but has a refresh job) ─────────────
    # MVs that are refreshed on schedule but never READ by any downstream job.
    seen_dead: set = set()
    # These are scheduler-topology bugs: the refresh creates the MV but nobody
    # actually queries it (the pipeline bypasses it, reading from base tables
    # directly).  Flag as "E. Dead MV — Zero Consumers" so they can be
    # decommissioned without affecting data flows.
    tbl_writers_e: dict = defaultdict(set)
    tbl_readers_e: dict = defaultdict(set)
    for n in graph.get("nodes", []):
        nid = n.get("id", "")
        for t in n.get("tgt_tables") or []:
            tbl_writers_e[t.upper()].add(nid)
        for t in n.get("src_tables") or []:
            tbl_readers_e[t.upper()].add(nid)
    for r in lineage_rows:
        job = r.get("DEPENDENT_JOB") or ""
        for col in ["TGT_TABLE", "SRC_TABLE"]:
            val = (r.get(col) or "").upper().strip()
            if job and val and "," not in val:
                if col == "TGT_TABLE":
                    tbl_writers_e[val].add(job)
                else:
                    tbl_readers_e[val].add(job)

    for n in graph.get("nodes", []):
        if n.get("category") != "MV":
            continue
        nid = n.get("id", "")
        tgts = n.get("tgt_tables") or []
        for tbl in tgts:
            tbl_up = tbl.upper()
            readers = tbl_readers_e.get(tbl_up, set())
            if readers:
                continue  # has consumers — covered by the main E loop above
            if _is_log_or_debug_table(tbl_up):
                continue
            # Confirm there IS a refresh job writing this MV
            writers = tbl_writers_e.get(tbl_up, set())
            if not writers:
                continue

            # ── Indicator-update consumer check ──────────────────────────────
            # Column-level lineage does NOT capture the MV → UPD_IND_COLS link
            # because the UPD job's src_tables point to the base table it updates,
            # not to the MV it reads.  Detect this pattern via Tidal edges:
            # if the MV refresh job's Tidal children include a UPD_IND/UPD_COL
            # job, that job IS consuming the MV — just not via lineage.
            tidal_upd_consumers = {
                c for c in children.get(nid, set())
                if _is_updater_job(c)
                and c in node_map  # must be in the current graph
            }
            if tidal_upd_consumers:
                key_indic = ("E-INDIC", tbl_up)
                if key_indic in seen_dead:
                    continue
                seen_dead.add(key_indic)
                upd_consumer = sorted(tidal_upd_consumers)[0]
                upd_node    = node_map.get(upd_consumer, {})
                upd_tgts    = [t for t in (upd_node.get("tgt_tables") or [])
                               if not _is_log_or_debug_table(t.upper())]
                upd_tgt_str = upd_tgts[0] if upd_tgts else upd_consumer
                # Build recommendation identical in style to regular elimination candidate
                _indic_decom = (
                    f"  \u2022 Remove Tidal job '{nid}' from the Tidal schedule.\n"
                    f"  \u2022 Drop the MV table '{tbl}' from the database.\n"
                    f"  \u2022 Remove all Tidal scheduler dependency edges pointing to '{nid}'."
                )
                recommendations.append({
                    "id": f"E-{len(recommendations)+1}",
                    "category": "E. MV Elimination Candidate",
                    "target_table": tbl,
                    "description": (
                        f"WHAT: '{tbl}' is a Materialized View refreshed by Tidal job '{nid}'.\n"
                        f"CONSUMER: 1 downstream indicator-update job — '{upd_consumer}' "
                        f"(updates '{upd_tgt_str}').\n"
                        f"NOTE: This consumer pattern is NOT captured in column-level lineage "
                        f"because '{upd_consumer}' lists '{upd_tgt_str}' (the table it writes "
                        f"back to) as its source, not the MV it reads. The link is confirmed "
                        f"via the Tidal scheduling dependency edge.\n"
                        f"OPPORTUNITY: Single-consumer MV with indicator-update pattern. "
                        f"The MV can be eliminated by folding its aggregation/filter logic "
                        f"directly into the main load INSERT SELECT, removing both the MV "
                        f"refresh hop and the separate indicator-update hop. "
                        f"(See also: E. MV Indicator-Chain Elimination Candidate findings "
                        f"for the full chain context.)"
                    ),
                    "affected_jobs": [nid, upd_consumer],
                    "impacted_rpts": [RPT_TABLE],
                    "consumers_by_rpt": {},
                    "confidence_scope": "LOCAL_ONLY",
                    "sql_objects": [],
                    "depth": n.get("depth"),
                    "current_hops": 2,
                    "potential_hop_savings": 2,   # removes MV refresh + UPD job
                    "risk": "HIGH",
                    "implementation_decision": "EXECUTE_LOCALLY",
                    "execution_wave": "WAVE_3",
                    "global_rpt_count_impacted": 1,
                    "global_consumer_count": 1,
                    "global_runtime_affected_minutes": 0.0,
                    "global_est_runtime_saved_minutes": 0.0,
                    "blast_radius_score": 2,
                    "prerequisites": (
                        "Row-count parity and indicator-column distribution comparison "
                        "for 3+ cycles; co-ordinate with the team that owns the "
                        f"'{upd_consumer}' procedure."
                    ),
                    "recommendation": (
                        f"INDICATOR-CONSUMER MV ELIMINATION — Action required:\n"
                        f"Step 1  [TRACE] Open the SQL procedure called by '{upd_consumer}'. "
                        f"Find every SELECT/JOIN that references '{tbl}' and identify which "
                        f"indicator columns it computes (e.g. V_HAS_LTD_IND_R, V_HAS_WAIVER_IND_R).\n\n"
                        f"Step 2  [INLINE] Move the derivation logic into the main load "
                        f"INSERT SELECT that populates '{upd_tgt_str}'. The MV's aggregation "
                        f"query becomes a WITH (CTE) clause or a subquery inside the main load.\n\n"
                        f"Step 3  [VALIDATE] Run 3+ full pipeline cycles. Confirm indicator "
                        f"column values match the baseline from Step 1.\n\n"
                        f"Step 4  [DECOMMISSION] After parity confirmed:\n{_indic_decom}"
                    ),
                    "_dedup_key": ("E", tbl_up, "INDIC_CONSUMER"),
                })
                continue  # Do NOT fall through to dead/parser-gap classification
            # ── Summaries cross-check ─────────────────────────────────────────
            # Column-level lineage may miss JOIN-only references (parser gap).
            # If the MV name appears in any script summary's INPUT_DEPENDENCIES,
            # the table IS actively used somewhere — do NOT call it dead.
            # Instead, flag it as a "parser gap" finding so the team can verify.
            in_summaries = tbl_up in _summary_input_sources
            key = ("E-DEAD", tbl_up)
            if key in seen_dead:
                continue
            seen_dead.add(key)
            global_ctx = global_mv_context.get(tbl_up, {})
            n_rpts = len(global_ctx.get("consumer_rpts", set()))
            if in_summaries:
                # MV appears in at least one script summary as an input — we have
                # evidence of lineage from the LLM summaries. This is a parser gap
                # (column-level extractor missed a JOIN-only reference), NOT a dead MV.
                # Do NOT surface this to the data team; track separately for parser
                # improvement only.
                continue  # Skip: evidence of lineage exists; no data-team action needed
            else:
                # Not found in summaries either — likely a genuine dead MV.
                recommendations.append({
                    "id": f"E-{len(recommendations)+1}",
                    "category": "E. Dead MV — Zero Consumers (Decommission Candidate)",
                    "target_table": tbl,
                    "description": (
                        f"MV '{tbl}' is refreshed by job '{nid}' but has NO consumers "
                        f"in the pipeline: zero readers found in column-level lineage "
                        f"AND zero references found in script summaries. "
                        f"The pipeline bypasses this MV and reads from the underlying "
                        f"base table directly. "
                        f"This is a scheduler-topology issue: the refresh job runs on "
                        f"schedule, consuming compute resources, but the output is never used. "
                        f"{'Confirmed across all RPT graphs: 0 consumers globally.' if n_rpts == 0 else ''}"
                    ),
                    "affected_jobs": sorted(writers)[:4],
                    "current_hops": 1,
                    "potential_hop_savings": 1,
                    "risk": "MEDIUM",
                    "recommendation": (
                        f"DECOMMISSION CANDIDATE — Action required:\n"
                        f"(1) Confirm '{tbl}' has no consumers by searching all SQL "
                        f"packages and procedures for any SELECT/JOIN referencing '{tbl}'.\n"
                        f"(2) Cross-verify with the script LLM summaries: the table does "
                        f"NOT appear in any summary's INPUT_DEPENDENCIES — "
                        f"supporting the dead-write assessment.\n"
                        f"(3) If confirmed zero usage, remove the refresh job '{nid}' "
                        f"from the Tidal schedule and drop the MV definition.\n"
                        f"(4) Verify that removing '{nid}' does not break any downstream "
                        f"Tidal dependencies (check children jobs: "
                        f"{sorted(children.get(nid, set()))[:3]}).\n"
                        f"(5) If children jobs depend on '{nid}' purely for ordering, "
                        f"replace with a direct dependency on the upstream base-table "
                        f"load job instead."
                    ),
                    "_dedup_key": ("E", tbl_up, "DEAD"),
                })

    return recommendations


def detect_post_load_mv_indicator_chain(graph, all_graphs):
    """
    E-sub. Post-load indicator-update MV chain.

    Detects the pattern:
        T (full-load job) → MV_chain (refresh MVs with no lineage tgt_tables)
                          → UPD job (reads MV, updates T in-place)

    This multi-hop pattern is a strong MV elimination candidate: the indicator
    or derived-column logic in the UPD job could be folded directly into the
    original full-load INSERT SELECT, eliminating the intermediate MV refresh
    steps and the separate update job.

    Lineage extractors often miss the src/tgt tables for these MVs (they appear
    as src=[], tgt=[] in graph nodes).  We infer their table name from the job
    name pattern:
        EDP_*_MV_REFRESH_<TABLE_NAME>[-N] → <TABLE_NAME>

    Only flags chains that affect tables present in the CURRENT graph so the
    finding is relevant to the RPT being analysed.
    """
    recommendations = []

    # Build Tidal edge sets across ALL graphs (for chain following)
    all_edges: set = set()
    all_nodes: dict = {}
    for g in all_graphs.values():
        for lnk in g.get("links", []):
            all_edges.add((lnk.get("source"), lnk.get("target")))
        for n in g.get("nodes", []):
            all_nodes[n["id"]] = n

    # Tables that appear in the CURRENT graph (as tgt/src tables)
    def _is_plsql_collection_type(name: str) -> bool:
        """Return True for PL/SQL bulk-collect variable type names (not real tables)."""
        n = name.upper()
        return (n.startswith("LT_") or n.endswith("_TYP") or n.endswith("_TYPE")
                or "TBL_TYP" in n)

    current_tables: set = set()
    for n in graph.get("nodes", []):
        for t in (n.get("tgt_tables") or []):
            if not _is_plsql_collection_type(t):
                current_tables.add(t.upper())
        for t in (n.get("src_tables") or []):
            if not _is_plsql_collection_type(t):
                current_tables.add(t.upper())

    def _children_of(job_id):
        return {tgt for src, tgt in all_edges if src == job_id}

    def _parents_of(job_id):
        return {src for src, tgt in all_edges if tgt == job_id}

    def _infer_mv_name(job_id: str) -> str | None:
        """Extract MV table name from MV_REFRESH job name. Returns None if not applicable."""
        j = job_id.upper()
        for prefix in ("EDP_GRP_EDW_MV_REFRESH_", "EDP_EDW_GRP_MV_REFRESH_",
                       "EDP_EDW_MV_REFRESH_", "MV_REFRESH_"):
            if j.startswith(prefix):
                remainder = job_id[len(prefix):]
                # Strip trailing -N version suffix
                import re as _re
                remainder = _re.sub(r"-\d+(\.\d+)?$", "", remainder)
                return remainder.upper()
        return None

    def _is_updater(job_id: str) -> bool:
        """Return True if the job name signals an in-place update (not a full reload)."""
        return _is_updater_job(job_id)  # delegate to module-level helper

    # Find all MV_REFRESH jobs that have empty tgt_tables (lineage gap)
    mv_jobs_no_lineage = {
        nid: n for nid, n in all_nodes.items()
        if ("MV_REFRESH" in nid.upper())
        and not (n.get("tgt_tables") or [])
        and _infer_mv_name(nid) is not None
    }

    # For each such MV job, check if it leads to an UPD job that updates a
    # table present in the current graph.
    seen_chains: set = set()
    for mv_job_id, mv_node in mv_jobs_no_lineage.items():
        mv_table = _infer_mv_name(mv_job_id)
        if not mv_table:
            continue

        # Walk downstream from mv_job_id (BFS, max 3 hops) looking for updaters
        queue = [mv_job_id]
        visited_chain = {mv_job_id}
        chain_mvs = [mv_job_id]
        found_upd: list = []

        for _ in range(4):  # max chain depth
            next_q = []
            for node in queue:
                for child in _children_of(node):
                    if child in visited_chain:
                        continue
                    visited_chain.add(child)
                    child_node = all_nodes.get(child, {})
                    if _is_updater(child):
                        # Check if this updater modifies a current-graph table.
                        # Exclude PL/SQL collection types — see _is_plsql_collection_type.
                        upd_tgts = {t.upper() for t in (child_node.get("tgt_tables") or [])
                                    if not _is_plsql_collection_type(t)}
                        upd_srcs = {t.upper() for t in (child_node.get("src_tables") or [])
                                    if not _is_plsql_collection_type(t)}
                        target_tables = (upd_tgts | upd_srcs) & current_tables
                        if target_tables:
                            found_upd.append((child, target_tables))
                    elif "MV_REFRESH" in child.upper():
                        chain_mvs.append(child)
                        next_q.append(child)
                    else:
                        next_q.append(child)
            queue = next_q

        if not found_upd:
            continue

        # Also collect upstream producers of the MV (they tell us what triggers the chain)
        chain_producers = []
        for parent in _parents_of(mv_job_id):
            parent_node = all_nodes.get(parent, {})
            parent_tgts = {t.upper() for t in (parent_node.get("tgt_tables") or [])}
            if parent_tgts & current_tables:
                chain_producers.append(parent)

        for upd_job, target_tbls in found_upd:
            for target_tbl in target_tbls:
                chain_key = (target_tbl, mv_table)
                if chain_key in seen_chains:
                    continue
                seen_chains.add(chain_key)

                # Collect all inferred MV names in the chain
                inferred_mvs = []
                for cj in chain_mvs:
                    nm = _infer_mv_name(cj)
                    if nm:
                        inferred_mvs.append(nm)

                # Find OTHER target tables updated by the same starting MV,
                # so the description can cross-reference them.
                sibling_targets = []
                for other_upd_job, other_tgts in found_upd:
                    for t in other_tgts:
                        if t != target_tbl:
                            sibling_targets.append(t)

                single_mv_flag = (len(inferred_mvs) == 1)
                sibling_note = (
                    f"\nCROSS-TABLE: The same MV chain also updates indicator columns "
                    f"for {', '.join(sorted(set(sibling_targets)))}. "
                    f"Implementing this fix ONCE addresses all affected target tables."
                    if sibling_targets else ""
                )
                priority_note = (
                    "\nPRIORITY: Single-MV chain — this is the simplest elimination case. "
                    "Only 1 MV refresh hop stands between the producer and the indicator "
                    "update job. This MV is also flagged individually as an "
                    "'E. MV Elimination Candidate' (see corresponding row)."
                    if single_mv_flag else ""
                )

                recommendations.append({
                    "id": f"E-MV-{len(recommendations)+1}",
                    "category": "E. MV Indicator-Chain Elimination Candidate",
                    "target_table": target_tbl,
                    "description": (
                        f"Table '{target_tbl}' is fully loaded by a producer job, then "
                        f"{len(inferred_mvs)} intermediate MV(s) are refreshed: "
                        f"{inferred_mvs}, and finally '{upd_job}' updates indicator or derived "
                        f"columns back into '{target_tbl}'. "
                        f"This multi-hop chain exists because the indicator/derived-column logic "
                        f"was implemented as a separate post-load step that reads pre-aggregated "
                        f"MV data rather than computing the values inline. "
                        f"Note: these MVs have no column-level lineage captured — they are "
                        f"identified via Tidal dependency edges and job-name inference."
                        + sibling_note
                        + priority_note
                    ),
                    "affected_jobs": list(dict.fromkeys(chain_producers + chain_mvs + [upd_job]))[:8],
                    "mv_chain": inferred_mvs,
                    "updater_job": upd_job,
                    "current_hops": len(chain_mvs) + 1,
                    "potential_hop_savings": len(chain_mvs) + 1,
                    "risk": "HIGH",
                    "recommendation": (
                        f"MV CHAIN ELIMINATION — Action required: "
                        f"(1) Identify which columns in '{target_tbl}' are filled by '{upd_job}' "
                        f"(typically indicator columns like V_HAS_LTD_IND_R, V_HAS_WAIVER_IND_R). "
                        f"(2) Trace the logic through the MV chain: "
                        f"{' → '.join(inferred_mvs)} → {upd_job}. "
                        f"(3) Fold the indicator derivation directly into the main load INSERT SELECT "
                        f"by joining to the base DIM/FCT tables that the MV chain aggregates, "
                        f"eliminating {len(inferred_mvs)} MV refresh step(s) and the separate "
                        f"update job. "
                        f"(4) Validate row counts and indicator column distributions across "
                        f"3+ consecutive cycles before promoting to production. "
                        f"(5) If the indicator logic is complex, consider creating a persistent "
                        f"SQL view encapsulating the derivation so it can be reused without "
                        f"materializing intermediate tables."
                        + (f"\n(6) CROSS-TABLE: same fix applies to {', '.join(sorted(set(sibling_targets)))}."
                           if sibling_targets else "")
                    ),
                    "_dedup_key": ("E", target_tbl, mv_table, upd_job),
                })

                # For single-MV chains, also emit a companion "E. MV Elimination Candidate"
                # for the MV itself, since its only consumer is an indicator-update job.
                # This provides the per-MV view alongside the per-chain view.
                if single_mv_flag:
                    mv_ref_job = chain_mvs[0]  # the single MV refresh job
                    upd_node_eic = all_nodes.get(upd_job, {})
                    upd_tgts_eic = [t for t in (upd_node_eic.get("tgt_tables") or [])
                                    if t.upper() not in ("", "NONE")]
                    upd_tgt_eic = upd_tgts_eic[0] if upd_tgts_eic else target_tbl
                    _eic_decom = (
                        f"  \u2022 Remove Tidal job '{mv_ref_job}' from the Tidal schedule.\n"
                        f"  \u2022 Drop the MV table '{mv_table}' from the database.\n"
                        f"  \u2022 Remove all Tidal scheduler dependency edges pointing to '{mv_ref_job}'."
                    )
                    recommendations.append({
                        "id": f"E-MV-{len(recommendations)+1}",
                        "category": "E. MV Elimination Candidate",
                        "target_table": mv_table,
                        "description": (
                            f"WHAT: '{mv_table}' is a Materialized View refreshed by "
                            f"Tidal job '{mv_ref_job}'.\n"
                            f"CONSUMER: 1 downstream indicator-update job — '{upd_job}' "
                            f"(updates '{upd_tgt_eic}').\n"
                            f"NOTE: Column-level lineage does not capture the '{mv_table}' → "
                            f"'{upd_job}' data link — confirmed via Tidal scheduling edge. "
                            f"This is the single-MV variant of the indicator-chain pattern; "
                            f"only 1 hop separates the producer from the indicator updater.\n"
                            f"OPPORTUNITY: Eliminating '{mv_table}' removes 2 hops (MV refresh + "
                            f"indicator-update job) by folding the indicator logic into the main "
                            f"INSERT SELECT. See the companion 'E. MV Indicator-Chain "
                            f"Elimination Candidate' finding for full chain context."
                        ),
                        "affected_jobs": [mv_ref_job, upd_job],
                        "impacted_rpts": [RPT_TABLE],
                        "consumers_by_rpt": {},
                        "confidence_scope": "LOCAL_ONLY",
                        "sql_objects": [],
                        "depth": None,
                        "current_hops": 2,
                        "potential_hop_savings": 2,
                        "risk": "HIGH",
                        "implementation_decision": "EXECUTE_LOCALLY",
                        "execution_wave": "WAVE_3",
                        "global_rpt_count_impacted": 1,
                        "global_consumer_count": 1,
                        "global_runtime_affected_minutes": 0.0,
                        "global_est_runtime_saved_minutes": 0.0,
                        "blast_radius_score": 2,
                        "prerequisites": (
                            "Row-count parity and indicator-column distribution comparison "
                            f"for 3+ cycles; co-ordinate with the team owning '{upd_job}'."
                        ),
                        "recommendation": (
                            f"INDICATOR-CONSUMER MV ELIMINATION — Action required:\n"
                            f"Step 1  [TRACE] Open the SQL procedure called by '{upd_job}'. "
                            f"Find every SELECT/JOIN referencing '{mv_table}' and identify "
                            f"which indicator columns it computes.\n\n"
                            f"Step 2  [INLINE] Move the derivation into the main load "
                            f"INSERT SELECT for '{upd_tgt_eic}'. The MV query becomes a "
                            f"WITH (CTE) clause or subquery inside the main load.\n\n"
                            f"Step 3  [VALIDATE] Run 3+ full cycles. Confirm indicator "
                            f"column values match baseline.\n\n"
                            f"Step 4  [DECOMMISSION] After parity confirmed:\n{_eic_decom}"
                        ),
                        "_dedup_key": ("E", mv_table, "INDIC_SINGLE_MV"),
                    })

    return recommendations


def detect_repeated_transformations(lineage_rows):
    """K. Detect repeated src->tgt column mapping patterns across multiple jobs."""
    mapping_jobs = defaultdict(set)

    for r in lineage_rows:
        src_t = r.get("SRC_TABLE")
        tgt_t = r.get("TGT_TABLE")
        job = r.get("DEPENDENT_JOB")
        if not (src_t and tgt_t and job):
            continue

        src_cols = sorted(parse_col_list(r.get("SOURCE_COL")))
        tgt_cols = sorted(parse_col_list(r.get("TARGET_COL")))
        if not src_cols or not tgt_cols:
            continue

        pairs = tuple(zip(src_cols, tgt_cols)) if len(src_cols) == len(tgt_cols) else ()
        if len(pairs) < 3:
            continue

        key = (src_t, tgt_t, pairs)
        mapping_jobs[key].add(job)

    recommendations = []
    for (src_t, tgt_t, pairs), jobs in mapping_jobs.items():
        if len(jobs) < 2:
            continue

        recommendations.append(
            {
                "id": f"K-{len(recommendations)+1}",
                "category": "K. Repeated Transformations",
                "target_table": tgt_t,
                "description": (
                    f"{len(jobs)} jobs implement the same {len(pairs)}-column "
                    f"mapping from '{src_t}' to '{tgt_t}'."
                ),
                "affected_jobs": sorted(jobs),
                "current_hops": len(jobs),
                "potential_hop_savings": len(jobs) - 1,
                "risk": "MEDIUM",
                "recommendation": (
                    "Externalize shared transformation into reusable SQL component "
                    "(CTE/view/UDF), keep job-specific filters separate, and validate row parity."
                ),
                "_dedup_key": ("K", src_t, tgt_t, tuple(sorted(jobs))),
            }
        )

    return recommendations


# -- Detector L ---------------------------------------------------------------


def detect_preprocessing_overlap(lineage_rows):
    """L. Detect same target table prepared by multiple preprocessing-style jobs."""
    tgt_jobs = defaultdict(set)
    for r in lineage_rows:
        tgt = r.get("TGT_TABLE")
        job = r.get("DEPENDENT_JOB")
        if tgt and job:
            tgt_jobs[tgt].add(job)

    preprocess_tokens = re.compile(r"(MV_REFRESH|PREP|STG|STAGE|WRK|WORK|TEMP)", re.I)
    recommendations = []
    for tgt, jobs in tgt_jobs.items():
        if len(jobs) < 2:
            continue
        # Skip log/debug/audit side-effect tables
        if _is_log_or_debug_table(tgt):
            continue
        prep_jobs = [j for j in jobs if preprocess_tokens.search(j)]
        if len(prep_jobs) < 2:
            continue

        recommendations.append(
            {
                "id": f"L-{len(recommendations)+1}",
                "category": "L. Overlapping Preprocessing",
                "target_table": tgt,
                "description": (
                    f"Target '{tgt}' is prepared by {len(prep_jobs)} preprocessing-style "
                    f"jobs before downstream loads."
                ),
                "affected_jobs": sorted(prep_jobs),
                "current_hops": len(prep_jobs),
                "potential_hop_savings": max(len(prep_jobs) - 1, 0),
                "risk": "MEDIUM",
                "recommendation": (
                    "Rationalize preprocessing into a canonical prep layer; preserve business "
                    "logic boundaries and add reconciliation checks before hop removal."
                ),
                "_dedup_key": ("L", tgt, tuple(sorted(prep_jobs))),
            }
        )

    return recommendations


# ── Detector O ──────────────────────────────────────────────────────────────

# Suffixes that reliably signal an intermediate / working / temp table.
# EXCLUDED by design: _EXG (Oracle partition-exchange — legitimate architecture),
# _GTT (Oracle Global Temporary Table — correct naming), _STG_ prefix (desired).
_INTERMEDIATE_SUFFIXES = (
    "_TMP", "_TABLE_TMP", "_DEBUG_TRC", "_TRC",
    "_UPD_PAYEE_NAME", "_BKP_PREV_MONTH", "_BKP",
)
_INTERMEDIATE_SUBSTRINGS = ("_INTERMEDIATE_MV_TBL", "_INTERMEDIATE_MV_TBL2",
                             "_INTERMEDIATE_")

# Hybrid naming: tables that combine two standard naming conventions, e.g.
# FCT_RPT_* (starts with FCT_ fact-table prefix but embeds RPT_ report infix).
# These are neither proper fact tables nor proper report tables — likely
# intermediate aggregation steps that should be renamed for clarity.
# Validated: FCT_RPT_CLAIM_SUMMARY_R, FCT_RPT_CROSS_SELL_SUMMARY_R (Jul-2026)
_HYBRID_NAMING_PREFIXES: tuple[str, ...] = (
    "FCT_RPT_",   # FCT_ prefix + RPT_ infix — ambiguous ownership and purpose
)

# View naming violation: objects ending _VW should start with VW_.
_VIEW_WRONG_SUFFIX = "_VW"
_VIEW_CORRECT_PREFIX = "VW_"


def _compute_global_table_impact(table_name, global_table_usage, global_job_rpt_map):
    """
    Return (impacted_rpts, all_jobs, blast_radius, implementation_decision).
    Uses the same logic as the E-shared MV global impact computation.
    """
    info = global_table_usage.get(table_name.upper(), {})
    all_jobs = (info.get("producers") or set()) | (info.get("consumers") or set())
    impacted_rpts = set()
    for j in all_jobs:
        impacted_rpts |= set(global_job_rpt_map.get(j, set()))
    blast = min(len(all_jobs) + max(len(impacted_rpts) - 1, 0), 10)
    decision = "EXECUTE_GLOBAL_PROGRAM" if len(impacted_rpts) > 1 else "EXECUTE_LOCALLY"
    return sorted(impacted_rpts), sorted(all_jobs), blast, decision


def detect_naming_and_intermediate_tables(
    graph,
    lineage_rows,
    global_table_usage=None,
    global_job_rpt_map=None,
    summaries=None,
):
    """
    O. Intermediate/working table detection with global cross-RPT impact scoring.

    Sub-type O-a — Intermediate/Working Table
    -----------------------------------------
    Flags tables whose names carry clear 'temp / working / debug' signals
    (_TMP, _TRC, _DEBUG_TRC, _UPD_PAYEE_NAME, _INTERMEDIATE_*, _BKP_*).  For
    each candidate:
      - Checks script summaries for intra-procedure staging usage (CTAS patterns
        where the same procedure creates and consumes the TMP table internally).
        These are NOT dead writes — they are staging steps within one procedure.
      - Applies a global cross-RPT impact check for the rename recommendation.
      - Applies a global cross-RPT impact check via global_table_usage:
          * If the table is referenced in > 1 RPT: EXECUTE_GLOBAL_PROGRAM
          * Otherwise: EXECUTE_LOCALLY
      - Recommendation: (1) rename to STG_ prefix, (2) evaluate eliminating the
        intermediate hop by folding logic into the upstream or downstream job.

    Sub-type O-b — View Naming Convention Violation
    ------------------------------------------------
    Flags SQL view objects whose names end in _VW instead of starting with VW_.
    These mislead developers into treating them as base tables.
    Also applies global scope check.
    """
    global_table_usage = global_table_usage or {}
    global_job_rpt_map = global_job_rpt_map or {}
    summaries = summaries or {}

    # ── Build intra-procedure staging index from script summaries ─────────────
    # Some TMP/working tables are created and consumed WITHIN a single stored
    # procedure (CTAS staging pattern).  The column-level lineage extractor only
    # captures INSERT/MERGE column mappings — CTAS creations inside a procedure
    # don’t produce a SRC_TABLE lineage row, so these appear as 0-reader tables.
    # We cross-reference with the LLM script summaries which DO capture
    # intermediate_tables_created and staging_layers explicitly.
    _intra_proc_staging: set[str] = set()  # upper-cased table names
    for sdata in summaries.values():
        if not isinstance(sdata, dict):
            continue
        pf = sdata.get("PROCESS_FLOW", {})
        # intermediate_tables_created: list of strings or single string
        interm = pf.get("intermediate_tables_created", [])
        if isinstance(interm, str):
            interm = [interm]
        for entry in (interm or []):
            _intra_proc_staging.add(str(entry).upper())
        # staging_layers: string description
        staging = str(pf.get("staging_layers", ""))
        if staging:
            _intra_proc_staging.add(staging.upper())

    # Build local table index from current graph
    local_tbl_writers = defaultdict(set)
    local_tbl_readers = defaultdict(set)
    for n in graph.get("nodes", []):
        nid = n.get("id", "")
        for t in n.get("tgt_tables") or []:
            local_tbl_writers[t.upper()].add(nid)
        for t in n.get("src_tables") or []:
            local_tbl_readers[t.upper()].add(nid)

    # Also collect from lineage rows (picks up lineage-only references)
    for r in lineage_rows:
        job = r.get("DEPENDENT_JOB") or ""
        tgt = (r.get("TGT_TABLE") or "").upper()
        src = (r.get("SRC_TABLE") or "").upper()
        if job and tgt:
            local_tbl_writers[tgt].add(job)
        if job and src:
            local_tbl_readers[src].add(job)

    # Collect view violations from lineage (VIEW_EXPANSION source rows)
    view_violations = set()
    for r in lineage_rows:
        ls = (r.get("LINEAGE_SOURCE") or "").upper()
        if ls == "VIEW_EXPANSION":
            for t in [r.get("SRC_TABLE") or "", r.get("TGT_TABLE") or ""]:
                t_up = t.upper()
                if t_up.endswith(_VIEW_WRONG_SUFFIX) and not t_up.startswith(_VIEW_CORRECT_PREFIX.upper()):
                    view_violations.add(t)
    # Also from graph nodes with VIEW_ prefix id
    for n in graph.get("nodes", []):
        nid = n.get("id", "")
        if nid.startswith("VIEW_"):
            tbl = nid[5:]
            if tbl.upper().endswith(_VIEW_WRONG_SUFFIX) and not tbl.upper().startswith(_VIEW_CORRECT_PREFIX.upper()):
                view_violations.add(tbl)

    recommendations = []

    # ── O-a: Intermediate/Working tables ─────────────────────────────────────
    # Collect all tables visible in this RPT's pipeline
    all_local_tables = (
        set(local_tbl_writers.keys()) | set(local_tbl_readers.keys())
    )
    for tbl_up in sorted(all_local_tables):
        has_intermediate_suffix = any(tbl_up.endswith(s.upper()) for s in _INTERMEDIATE_SUFFIXES)
        has_intermediate_substr = any(sub.upper() in tbl_up for sub in _INTERMEDIATE_SUBSTRINGS)
        has_hybrid_prefix       = any(tbl_up.startswith(p.upper()) for p in _HYBRID_NAMING_PREFIXES)

        if not (has_intermediate_suffix or has_intermediate_substr or has_hybrid_prefix):
            continue

        # Skip legitimate partition-exchange staging (correct Oracle naming, no action needed)
        if tbl_up.endswith("_EXG"):
            continue
        # Skip log/debug/audit side-effect tables — not part of the data pipeline
        if _is_log_or_debug_table(tbl_up):
            continue
        # Skip compound table names (lineage data quality artefacts where TGT_TABLE
        # contains a comma-separated list instead of a single table name)
        if "," in tbl_up:
            continue
        # Skip if no local writers (can't confirm it's produced in this pipeline)
        if not local_tbl_writers.get(tbl_up):
            continue

        writers = sorted(local_tbl_writers.get(tbl_up, set()))
        readers = sorted(local_tbl_readers.get(tbl_up, set()))

        # ── Intra-procedure staging check ─────────────────────────────────────
        # If any script summary lists this table in intermediate_tables_created
        # or staging_layers, the table IS consumed within the same procedure
        # (CTAS pattern).  The lineage extractor misses this because CTAS rows
        # don't produce INSERT/MERGE lineage entries.  This is NOT a dead write
        # — it's an ephemeral staging step inside one stored procedure.
        is_intra_proc_staging = any(
            tbl_up in entry for entry in _intra_proc_staging
        )

        impacted_rpts, all_jobs, blast, decision = _compute_global_table_impact(
            tbl_up, global_table_usage, global_job_rpt_map
        )

        rpt_count = len(impacted_rpts)
        scope_text = (
            f"across {rpt_count} RPT pipeline(s) ({', '.join(impacted_rpts[:4])}"
            f"{'...' if rpt_count > 4 else ''})"
            if rpt_count > 1
            else "within this RPT only"
        )

        # ── Determine rename suggestion and description ────────────────────────
        already_stg = tbl_up.startswith("STG_")

        if has_hybrid_prefix and not has_intermediate_suffix and not has_intermediate_substr:
            # O-c: Hybrid naming convention (e.g. FCT_RPT_*) — not temp but wrongly named
            rename_to = tbl_up
            for old_pref in ("FCT_RPT_", "FCT_GRP_RPT_"):
                if tbl_up.startswith(old_pref):
                    rename_to = "STG_" + tbl_up[len(old_pref):]
                    break
            category_label = "O. Hybrid Naming Convention Violation"
            desc_signal = (
                f"Table '{tbl_up}' combines two naming conventions in its name "
                f"(FCT_ prefix + RPT_ infix), making it ambiguous — it is neither "
                f"a standard fact table (FCT_) nor a standard report table (RPT_). "
                f"This pattern typically indicates an intermediate aggregation or "
                f"staging step used between fact-layer processing and report generation. "
                f"Written by {len(writers)} job(s) and read by {len(readers)} job(s) {scope_text}."
            )
            rename_note = (
                f"Rename '{tbl_up}' → '{rename_to}' to use the STG_ prefix, "
                f"which clearly signals that this is an intermediate staging table "
                f"rather than a permanent fact or report table."
            )
        elif already_stg:
            category_label = "O. Intermediate/Working Table — Naming & Hop Reduction"
            desc_signal = (
                f"Table '{tbl_up}' is written by {len(writers)} job(s) and read by "
                f"{len(readers)} job(s) {scope_text}. "
                + (
                    "It has active downstream readers — evaluate whether the intermediate "
                    "hop can be eliminated by folding logic into the consumer."
                    if bool(readers)
                    else (
                        "Script analysis confirms this is an intra-procedure staging table "
                        "(CTAS pattern): it is created and consumed within the same stored "
                        "procedure, so it does NOT appear as a reader in column-level lineage. "
                        "This is NOT a dead write."
                        if is_intra_proc_staging
                        else "It has no downstream readers in the pipeline — potential dead write."
                    )
                )
            )
            rename_to = tbl_up
            rename_note = (
                f"Table already uses STG_ prefix which is correct. "
                f"The '{next((s for s in _INTERMEDIATE_SUFFIXES if tbl_up.endswith(s.upper())), '')}' "
                f"suffix confirms its working/backup nature. No rename action required."
            )
        else:
            category_label = "O. Intermediate/Working Table — Naming & Hop Reduction"
            rename_to = tbl_up
            for old_pref in ("FCT_RPT_", "FCT_GRP_", "FCT_LG_", "FCT_",
                             "DIM_", "RPT_", "PRCS_"):
                if tbl_up.startswith(old_pref):
                    rename_to = "STG_" + tbl_up[len(old_pref):]
                    break
            else:
                rename_to = "STG_" + tbl_up
            rename_note = (
                f"'{tbl_up}' does not follow the standard naming convention for "
                f"temporary or working tables. Apply the project's standard naming "
                f"convention (e.g. STG_ prefix or similar agreed pattern) so the "
                f"table's transient nature is immediately clear to all developers."
            )
            desc_signal = (
                f"Table '{tbl_up}' carries a naming signal indicating it is a "
                f"temporary / working / debug table (not a permanent fact/dim/report table). "
                f"It is written by {len(writers)} job(s) and read by {len(readers)} job(s) "
                f"{scope_text}. "
                + (
                    "It has active downstream readers — evaluate whether the intermediate "
                    "hop can be eliminated by folding logic into the consumer."
                    if bool(readers)
                    else (
                        "Script analysis confirms this is an intra-procedure staging table "
                        "(CTAS pattern): created and consumed within the same stored procedure, "
                        "so it does NOT appear as a reader in column-level lineage. "
                        "This is NOT a dead write — it is an ephemeral staging step. "
                        "Hop reduction opportunity: replace the CTAS staging step with "
                        "a CTE or inline subquery to eliminate the physical table write."
                        if is_intra_proc_staging
                        else "It has no downstream readers in the pipeline — potential dead write."
                    )
                )
            )

        # Intra-procedure staging tables ARE hop-reduction candidates even though
        # they show 0 lineage readers (the consumer is the same procedure).
        hop_reduction = bool(readers) or is_intra_proc_staging

        recommendations.append(
            {
                "id": f"O-{len(recommendations)+1}",
                "category": category_label,
                "target_table": tbl_up,
                "description": desc_signal,
                "affected_jobs": list(dict.fromkeys(writers + readers)),
                "impacted_rpts": impacted_rpts,
                "current_hops": 1 if hop_reduction else 1,
                "potential_hop_savings": 1 if hop_reduction else 1,
                "risk": "HIGH" if rpt_count > 3 else "MEDIUM",
                "implementation_decision": decision,
                "execution_wave": (
                    "GLOBAL_WAVE_1_PLUS" if decision == "EXECUTE_GLOBAL_PROGRAM" else "WAVE_2"
                ),
                "global_rpt_count_impacted": rpt_count,
                "global_consumer_count": len(readers),
                "blast_radius_score": blast,
                "recommendation": (
                    f"SCOPE: {'GLOBAL — ' + str(rpt_count) + ' RPTs reference this table.' if rpt_count > 1 else 'LOCAL'}\n\n"
                    f"(1) NAMING: {rename_note}\n"
                    f"    Impact: {len(all_jobs)} job(s) across {rpt_count} RPT(s) reference this table — "
                    f"coordinate the rename as a {'global program' if rpt_count > 1 else 'single-sprint change'}.\n\n"
                    + (
                        f"(2) HOP REDUCTION: Evaluate folding '{tbl_up}' production logic into "
                        f"consumer job(s): {', '.join(readers[:3])}{'...' if len(readers) > 3 else ''}. "
                        f"If the intermediate result can be expressed as a CTE or inline view, "
                        f"the separate write step can be eliminated.\n"
                        if hop_reduction
                        else
                        f"(2) DEAD WRITE CHECK: No downstream consumers found in the Tidal pipeline. "
                        f"Verify the table is not used by ad-hoc queries or external processes. "
                        f"If confirmed unused, remove the writing step.\n"
                    )
                    + f"\nPrerequisites: Confirm no external (non-Tidal) consumers before renaming or removing."
                ),
                "_dedup_key": ("O-a", tbl_up),
            }
        )

    # ── O-b: View naming convention violations ────────────────────────────────

    # Pre-build a lookup: view_name_upper → set of scripts that reference it.
    # Sources: (1) lineage PACKAGE_NAME / FULL_OBJECT for rows where the view
    # appears in SRC_TABLE; (2) summaries INPUT_DEPENDENCIES text.
    def _scripts_referencing_view(v_upper: str) -> list[str]:
        """Return sorted list of unique SQL script names that reference a view."""
        scripts: set = set()
        # From lineage rows
        for r in lineage_rows:
            src = (r.get("SRC_TABLE") or "").upper()
            if v_upper not in src:
                continue
            pkg = (r.get("PACKAGE_NAME") or "").strip()
            obj = (r.get("FULL_OBJECT") or "").strip()
            for candidate in [pkg, obj]:
                if candidate and not candidate.upper().startswith("VIEW_") and candidate != v_upper:
                    # Normalise to .sql filename
                    fname = candidate if "." in candidate else candidate + ".sql"
                    scripts.add(fname)
        # From summaries INPUT_DEPENDENCIES (view may appear in source_tables)
        for script_name, sdata in (summaries or {}).items():
            if not isinstance(sdata, dict) or script_name.upper() == v_upper + ".SQL":
                continue  # skip the view's own summary
            deps = sdata.get("INPUT_DEPENDENCIES", {})
            if not isinstance(deps, dict):
                continue
            src_tbls = deps.get("source_tables", {})
            if isinstance(src_tbls, dict):
                for cat_lst in src_tbls.values():
                    if isinstance(cat_lst, list):
                        for entry in cat_lst:
                            if isinstance(entry, str) and v_upper in entry.upper():
                                scripts.add(script_name)
                                break
            # Also check procedure_level_dependencies
            proc_deps = deps.get("procedure_level_dependencies") or {}
            if isinstance(proc_deps, dict):
                for items in proc_deps.values():
                    if isinstance(items, list):
                        for entry in items:
                            if isinstance(entry, str) and v_upper in entry.upper():
                                scripts.add(script_name)
                                break
        return sorted(scripts)

    for v in sorted(view_violations):
        v_up = v.upper()
        # Suggested correct name: VW_ prefix + strip trailing _VW
        correct_name = _VIEW_CORRECT_PREFIX + v_up.replace(_VIEW_WRONG_SUFFIX, "").lstrip("_")

        impacted_rpts, all_jobs, blast, decision = _compute_global_table_impact(
            v_up, global_table_usage, global_job_rpt_map
        )
        rpt_count = len(impacted_rpts)

        # Collect actual script names from lineage + summaries
        referencing_scripts = _scripts_referencing_view(v_up)

        # Separate real Tidal jobs from synthetic VIEW_* tracking names.
        # Lineage uses artificial "VIEW_<view_name>" identifiers for logical views
        # since views are not orchestrated by Tidal — they have no real scheduler job.
        real_jobs  = [j for j in all_jobs if not j.upper().startswith("VIEW_")]
        view_nodes = [j for j in all_jobs if j.upper().startswith("VIEW_")]

        recommendations.append(
            {
                "id": f"O-{len(recommendations)+1}",
                "category": "O. View Naming Convention Violation",
                "target_table": v,
                "description": (
                    f"Object '{v}' is a SQL VIEW (identified from lineage VIEW_EXPANSION source) "
                    f"but its name ends in '_VW' rather than starting with 'VW_'. "
                    f"This misleads developers into treating it as a base table. "
                    f"Note: the 'Affected Jobs' column may show synthetic identifiers like "
                    f"'VIEW_{v}' — these are NOT real Tidal jobs; they are lineage "
                    f"tracking names used internally because logical views have no Tidal "
                    f"scheduler entry. The actual consumers are the SQL packages and procedures "
                    f"that reference this view in their FROM/JOIN clauses. "
                    f"Referenced across {rpt_count} RPT(s): {', '.join(impacted_rpts[:4])}."
                ),
                "affected_jobs": all_jobs[:10],
                "impacted_rpts": impacted_rpts,
                "current_hops": 0,
                "potential_hop_savings": 0,
                "risk": "LOW",
                "implementation_decision": decision,
                "execution_wave": (
                    "GLOBAL_WAVE_1_PLUS" if decision == "EXECUTE_GLOBAL_PROGRAM" else "WAVE_2"
                ),
                "global_rpt_count_impacted": rpt_count,
                "global_consumer_count": len(all_jobs),
                "blast_radius_score": blast,
                "recommendation": (
                    f"RENAME '{v}' → '{correct_name}' to follow the VW_ prefix convention.\n\n"
                    f"Impact: This view is used across {rpt_count} RPT(s) "
                    f"({', '.join(impacted_rpts[:4])}{'...' if rpt_count > 4 else ''}). "
                    f"{'This is a GLOBAL change — coordinate the rename across all referencing RPTs.' if rpt_count > 1 else 'Single-RPT change.'}\n\n"
                    f"Steps:\n"
                    f"(1) Rename the view DDL: ALTER VIEW {v} RENAME TO {correct_name} "
                    f"(or DROP and recreate with the new name).\n"
                    f"(2) Update all SQL references in the following scripts where "
                    f"'{v}' appears in FROM/JOIN/WITH clauses:\n"
                    + (
                        "".join(f"     - {s}\n" for s in referencing_scripts)
                        if referencing_scripts
                        else f"     (No scripts identified via lineage/summaries — search manually)\n"
                    )
                    + f"(3) Note: the 'Affected Jobs' column contains synthetic lineage tracking "
                    f"identifiers (VIEW_*) — NOT real Tidal jobs. The actual code changes are "
                    f"in the SQL source files listed above.\n"
                    + (f"(4) Real Tidal jobs that load from these packages: {real_jobs[:5]}\n" if real_jobs else "")
                    + f"(5) Validate each RPT pipeline ({', '.join(impacted_rpts[:4])}) runs "
                    f"correctly after the rename before marking complete."
                ),
                "_dedup_key": ("O-b", v_up),
            }
        )

    # ── O-d: Oracle GTT hop-reduction candidates ──────────────────────────────
    # Oracle Global Temporary Tables (_GTT suffix) are CORRECTLY named — no rename
    # needed.  But they ARE intra-procedure staging hops: a procedure populates the
    # GTT and then reads from it.  Flag as a hop reduction candidate only: evaluate
    # whether the GTT can be replaced by a WITH clause (CTE) or inline subquery.
    gtt_tables = {
        tbl for tbl in (set(local_tbl_writers.keys()) | set(local_tbl_readers.keys()))
        if tbl.endswith("_GTT") and local_tbl_writers.get(tbl)
    }
    for tbl_up in sorted(gtt_tables):
        if _is_log_or_debug_table(tbl_up):
            continue

        writers = sorted(local_tbl_writers.get(tbl_up, set()))
        readers = sorted(local_tbl_readers.get(tbl_up, set()))
        impacted_rpts, all_jobs, blast, decision = _compute_global_table_impact(
            tbl_up, global_table_usage, global_job_rpt_map
        )
        rpt_count = len(impacted_rpts)
        scope_text = (
            f"across {rpt_count} RPT pipeline(s) ({', '.join(impacted_rpts[:4])}"
            f"{'...' if rpt_count > 4 else ''})"
            if rpt_count > 1 else "within this RPT only"
        )
        is_intra_proc = any(tbl_up in entry for entry in _intra_proc_staging)

        recommendations.append(
            {
                "id": f"O-{len(recommendations)+1}",
                "category": "O. GTT Intra-Procedure Staging — Evaluate CTE Replacement",
                "target_table": tbl_up,
                "description": (
                    f"'{tbl_up}' is an Oracle Global Temporary Table (GTT) — the '_GTT' "
                    f"suffix is the correct Oracle naming convention, so NO rename is needed. "
                    + (
                        "Script analysis confirms it is created and consumed within the same "
                        "stored procedure (intra-procedure staging). "
                        if is_intra_proc
                        else f"It is written by {len(writers)} job(s) and read by "
                             f"{len(readers)} job(s) {scope_text}. "
                    )
                    + "GTTs represent a physical staging hop that may be eliminatable."
                ),
                "affected_jobs": list(dict.fromkeys(writers + readers)),
                "impacted_rpts": impacted_rpts,
                "current_hops": 1,
                "potential_hop_savings": 1,
                "risk": "LOW",
                "implementation_decision": decision,
                "execution_wave": (
                    "GLOBAL_WAVE_1_PLUS" if decision == "EXECUTE_GLOBAL_PROGRAM" else "WAVE_2"
                ),
                "global_rpt_count_impacted": rpt_count,
                "global_consumer_count": len(readers),
                "blast_radius_score": blast,
                "prerequisites": (
                    "Verify the GTT result set is referenced only once; "
                    "if referenced multiple times, keep the GTT (CTE re-evaluation risk in Oracle)."
                ),
                "recommendation": (
                    f"NAMING: No rename needed — '_GTT' is the correct Oracle convention.\n\n"
                    f"HOP REDUCTION: '{tbl_up}' stages data between steps within the same "
                    f"procedure. Evaluate replacing with a WITH clause (CTE):\n"
                    f"  (1) If the result is read MORE than once inside the procedure → "
                    f"keep the GTT (Oracle may re-execute a CTE each time it is referenced).\n"
                    f"  (2) If the result is read EXACTLY ONCE → rewrite as a CTE to "
                    f"eliminate the physical write/read cycle.\n"
                    f"  (3) Validate row counts and key metrics before and after the change."
                ),
                "_dedup_key": ("O-d", tbl_up),
            }
        )

    return recommendations


# ── Detector F ──────────────────────────────────────────────────────────────


def extract_llm_findings(summaries, hop_findings, graph, lineage_rows):
    """
    F. LLM-identified issues mapped to jobs.
    Uses NEW's mapping (direct + partial + target-table fallback)
    and stores SP's per-type count fields in the rec dict.
    """
    sql_to_jobs = defaultdict(list)
    for n in graph["nodes"]:
        for so in n.get("sql_objects", []) or []:
            name = (so.get("name") or "").upper()
            if name:
                sql_to_jobs[name].append(n["id"])

    tgt_to_jobs = defaultdict(set)
    for r in lineage_rows:
        if r.get("DEPENDENT_JOB") and r.get("TGT_TABLE"):
            tgt_to_jobs[r["TGT_TABLE"].upper()].add(r["DEPENDENT_JOB"])

    script_to_jobs = {}
    for sname in summaries:
        base = sname.replace(".sql", "").replace(".txt", "").upper()
        if base in sql_to_jobs:
            script_to_jobs[sname] = sql_to_jobs[base]
            continue
        matched = []
        for sql_name, jobs in sql_to_jobs.items():
            if base in sql_name or sql_name in base:
                matched.extend(jobs)
        if matched:
            script_to_jobs[sname] = list(dict.fromkeys(matched))
            continue
        for tgt_name, jobs in tgt_to_jobs.items():
            if tgt_name and tgt_name in base:
                script_to_jobs[sname] = sorted(jobs)
                break

    if not hop_findings or "findings" not in hop_findings:
        return []

    perf_kw = ("performance", "slow", "full scan", "index", "parallel", "bulk")
    logic_kw = ("bug", "error", "incorrect", "wrong", "missing", "mismatch")
    redund_kw = ("redundant", "duplicate", "repeated", "unnecessary", "remove")

    script_findings = defaultdict(list)
    for f in hop_findings["findings"]:
        script_findings[f.get("_source_script", "")].append(f)

    recommendations = []
    for script, findings in script_findings.items():
        jobs = script_to_jobs.get(script, ["UNMAPPED"])
        perf = logic = redund = 0
        for f in findings:
            txt = json.dumps(f, default=str).lower()
            if any(k in txt for k in redund_kw):
                redund += 1
            elif any(k in txt for k in logic_kw):
                logic += 1
            elif any(k in txt for k in perf_kw):
                perf += 1
        other = len(findings) - redund - logic - perf

        issues = []
        if redund:
            issues.append(f"{redund} redundancy signals")
        if logic:
            issues.append(f"{logic} logic risks")
        if perf:
            issues.append(f"{perf} perf concerns")
        if other > 0:
            issues.append(f"{other} other findings")

        risk = "HIGH" if logic > 2 else "MEDIUM" if perf > 2 else "LOW"
        top_concern = (
            "redundancy removal"
            if redund
            else (
                "performance tuning"
                if perf
                else "bug fixes" if logic else "general cleanup"
            )
        )
        recommendations.append(
            {
                "id": f"F-{len(recommendations)+1}",
                "category": "F. LLM-Identified Issues",
                "target_table": script,
                "description": (
                    f"Script '{script}' has {len(findings)} findings: "
                    + ", ".join(issues)
                    + "."
                ),
                "affected_jobs": jobs,
                "total_findings": len(findings),
                "perf_findings": perf,
                "logic_findings": logic,
                "redundancy_findings": redund,
                "current_hops": 0,
                "potential_hop_savings": redund,
                "risk": risk,
                "recommendation": (
                    f"Review script for optimization opportunities. "
                    f"Top concern: {top_concern}."
                ),
                "_dedup_key": ("F", script),
            }
        )
    return recommendations


# ── Detector E-IC: Indicator-Chain MV Elimination ───────────────────────────


def detect_indicator_chain_mv_elimination(graph, summaries=None, global_job_rpt_map=None):
    """
    E-IC. Detects MVs used solely as intermediate aggregations that feed an
    indicator-column UPDATE job on a report table.

    Pattern (identified via data team review):
      LoadJob  →  MV_REFRESH_1, MV_REFRESH_2, ...  →  UPD_IND_COLS
                                                          ↓
                                                   (updates RPT_TABLE in-place)

    These MV refresh jobs typically appear with empty src_tables/tgt_tables in
    column-level lineage (the lineage extractor missed them because the MVs are
    defined externally and not in All_Metadata).  We infer the MV table names
    from the Tidal job names (stripping the EDP_GRP_EDW_MV_REFRESH_ prefix and
    trailing -N suffix) and cross-validate against script summaries.

    Recommendation: fold the MV aggregation logic directly into the UPD_IND_COLS
    SQL (or into the initial INSERT SELECT of the load job) to eliminate the
    intermediate MV materialization hop entirely.
    """
    summaries = summaries or {}
    global_job_rpt_map = global_job_rpt_map or {}

    node_map = {n["id"]: n for n in graph.get("nodes", [])}
    edges = [(l["source"], l["target"]) for l in graph.get("links", [])]
    # parent → set(children)
    children_of = defaultdict(set)
    # child → set(parents)
    parents_of  = defaultdict(set)
    for src, tgt in edges:
        children_of[src].add(tgt)
        parents_of[tgt].add(src)

    # Build a set of all table names mentioned in ANY script summary's
    # INPUT_DEPENDENCIES to provide confirmation evidence.
    summary_source_tables: set[str] = set()
    for sdata in summaries.values():
        if not isinstance(sdata, dict):
            continue
        deps = sdata.get("INPUT_DEPENDENCIES", {})
        if isinstance(deps, dict):
            src_tbls = deps.get("source_tables", {})
            if isinstance(src_tbls, dict):
                for cat, lst in src_tbls.items():
                    if isinstance(lst, list):
                        for e in lst:
                            if isinstance(e, str):
                                # strip schema prefix and annotations
                                name = e.split("(")[0].split(" ")[0].strip().upper()
                                if "." in name:
                                    name = name.rsplit(".", 1)[-1]
                                if name:
                                    summary_source_tables.add(name)
            # Also check procedure_level_dependencies
            proc_deps = deps.get("procedure_level_dependencies") or {}
            if isinstance(proc_deps, dict):
                for proc_items in proc_deps.values():
                    if isinstance(proc_items, list):
                        for e in proc_items:
                            if isinstance(e, str):
                                name = e.split("(")[0].split(" ")[0].strip().upper()
                                if "." in name:
                                    name = name.rsplit(".", 1)[-1]
                                if name:
                                    summary_source_tables.add(name)

    def infer_mv_table_name(job_id: str) -> str | None:
        """Extract the MV table name from a Tidal MV_REFRESH job name.
        Pattern: EDP_GRP_EDW_MV_REFRESH_<TABLE_NAME>-N  →  <TABLE_NAME>
        """
        j = job_id.upper()
        prefix = "MV_REFRESH_"
        idx = j.find(prefix)
        if idx < 0:
            return None
        remainder = job_id[idx + len(prefix):]  # preserve original case
        # Strip trailing -N suffix
        import re
        remainder = re.sub(r"-\d+$", "", remainder)
        return remainder.upper() if remainder else None

    # Identify UPD_IND_COLS / indicator-update jobs
    UPD_SIGNALS = ("_UPD_IND_COLS", "_UPD_IND_", "_UPD_INDICATOR", "_UPDATE_IND")
    recommendations = []
    seen: set = set()

    for node in graph.get("nodes", []):
        upd_job = node.get("id", "")
        if not any(sig in upd_job.upper() for sig in UPD_SIGNALS):
            continue

        # Find all MV_REFRESH jobs that are UPSTREAM of this update job
        mv_refresh_parents = [
            p for p in parents_of.get(upd_job, set())
            if "MV_REFRESH" in p.upper()
        ]
        if not mv_refresh_parents:
            continue

        # Find the primary LOAD job that is upstream of the MV_REFRESH jobs
        load_parents = set()
        for mv_job in mv_refresh_parents:
            for p in parents_of.get(mv_job, set()):
                if "MV_REFRESH" not in p.upper():
                    load_parents.add(p)

        # The update job's target table (inferred from src/tgt if available,
        # or from the job name pattern _RPT_<TABLE>_UPD_)
        upd_node = node_map.get(upd_job, {})
        upd_tgt_tables = list(upd_node.get("tgt_tables") or [])
        if not upd_tgt_tables:
            # Try to infer from job name: EDP_GRP_EDW_LOAD_RPT_CLAIM_DTL_R_UPD_IND_COLS
            # → RPT_CLAIM_DTL_R
            import re
            m = re.search(r"LOAD_(RPT_[^_]+(?:_[^_]+)*?)_UPD", upd_job.upper())
            if m:
                upd_tgt_tables = [m.group(1)]

        # Collect MV table names (inferred + summary-confirmed)
        mv_details = []
        for mv_job in sorted(mv_refresh_parents):
            mv_tbl = infer_mv_table_name(mv_job)
            if not mv_tbl:
                continue
            confirmed = mv_tbl in summary_source_tables
            mv_details.append({
                "job": mv_job,
                "inferred_table": mv_tbl,
                "confirmed_in_summaries": confirmed,
            })

        if not mv_details:
            continue

        key = (upd_job, tuple(sorted(m["inferred_table"] for m in mv_details)))
        if key in seen:
            continue
        seen.add(key)

        confirmed_mvs = [m["inferred_table"] for m in mv_details if m["confirmed_in_summaries"]]
        all_mvs       = [m["inferred_table"] for m in mv_details]
        n_rpts = len(global_job_rpt_map.get(upd_job, set()))
        scope_text = (
            f"across {n_rpts} RPT pipeline(s)" if n_rpts > 1 else "within this RPT only"
        )

        recommendations.append({
            "id": f"E-IC-{len(recommendations)+1}",
            "category": "E. Indicator-Chain MV Elimination Candidate",
            "target_table": upd_tgt_tables[0] if upd_tgt_tables else upd_job,
            "description": (
                f"INDICATOR-CHAIN MV PATTERN: Job '{upd_job}' updates indicator "
                f"columns on '{upd_tgt_tables[0] if upd_tgt_tables else 'target'}' "
                f"AFTER refreshing intermediate MVs: {all_mvs}. "
                f"These MVs serve as aggregation intermediaries — they pre-compute "
                f"grouped counts/flags (e.g. waiver indicator, LTD indicator) that are "
                f"then applied as UPDATE SET on the report table. "
                f"The column-level lineage extractor did not capture src/tgt tables for "
                f"the MV refresh jobs (lineage gap); MV names were inferred from Tidal "
                f"job names. "
                + (f"Summaries confirm {confirmed_mvs} are active source tables. " if confirmed_mvs else "")
                + f"Pipeline scope: {scope_text}. "
                f"Upstream load job(s): {sorted(load_parents)[:3]}."
            ),
            "affected_jobs": [upd_job] + sorted(m["job"] for m in mv_details),
            "inferred_mv_tables": all_mvs,
            "confirmed_mv_tables": confirmed_mvs,
            "current_hops": len(mv_details) + 1,
            "potential_hop_savings": len(mv_details),
            "risk": "MEDIUM",
            "recommendation": (
                f"MV ELIMINATION — Fold the aggregation logic from "
                f"{all_mvs} directly into the UPDATE statement of '{upd_job}' "
                f"(or into the initial INSERT SELECT of the load job). "
                f"Steps: "
                f"(1) Locate the CREATE MATERIALIZED VIEW DDL for each MV in {all_mvs} "
                f"— the SELECT query defines the aggregation that can be inlined. "
                f"(2) Replace each MV join in '{upd_job}' with an inline subquery or CTE "
                f"that reproduces the same GROUP BY aggregation directly from the base "
                f"fact/dimension tables. "
                f"(3) Remove the MV_REFRESH Tidal jobs ({[m['job'] for m in mv_details][:3]}) "
                f"from the pipeline once the inline SQL is validated. "
                f"(4) Alternatively, consider folding the entire indicator-column update "
                f"into the main load job's INSERT SELECT to eliminate the separate update "
                f"step entirely — this also resolves the post-load self-reference pattern. "
                f"Validate: compare indicator column distributions on {upd_tgt_tables[0] if upd_tgt_tables else 'target'} "
                f"for 3+ consecutive cycles before removing the MVs."
            ),
            "_dedup_key": ("E-IC", upd_job),
        })

    return recommendations


# ── Detector P: Post-Load Cursor UPDATE ─────────────────────────────────────


def detect_cursor_post_load_updates(graph, sql_index, summaries):
    """
    P. Script-level: Cursor-based BULK COLLECT + FORALL UPDATE after initial load.

    Detects stored-procedure scripts where the main load (INSERT SELECT) is
    followed by one or more separate cursor loops that update individual columns
    via BULK COLLECT → FORALL UPDATE.  This pattern adds unnecessary procedural
    hops — the same columns could be derived and populated inline in the original
    INSERT SELECT, eliminating the separate update passes.

    Signal: PROCESS_FLOW.cursor_usage_bulk_logic text contains BULK COLLECT and
    (FORALL or FORALL) and UPDATE keywords together with post-load update context.
    """
    summaries = summaries or {}

    # Build reverse mapping: package_name_upper → (script_name, summary)
    pkg_to_summary: dict[str, tuple[str, dict]] = {}
    for script_name, sdata in summaries.items():
        if not isinstance(sdata, dict) or "error" in sdata:
            continue
        base = script_name.upper().removesuffix(".SQL").removesuffix(".TXT")
        pkg_to_summary[base] = (script_name, sdata)

    node_map = {n["id"]: n for n in graph.get("nodes", [])}
    recommendations = []
    seen: set = set()

    for job_id, job_info in sql_index.items():
        if job_id not in node_map:
            continue
        pkg_names = {p.upper() for p in (job_info.get("package_names") or set())}
        if not pkg_names:
            continue

        for pkg in pkg_names:
            if pkg not in pkg_to_summary:
                continue
            script_name, sdata = pkg_to_summary[pkg]

            pf = sdata.get("PROCESS_FLOW", {}) or {}
            # Find the cursor/bulk field — it may have slightly different key names
            cursor_text = ""
            for k, v in pf.items():
                if "cursor" in k.lower() or "bulk" in k.lower():
                    cursor_text += " " + (str(v) if v else "")

            cursor_up = cursor_text.upper()
            has_bulk   = "BULK COLLECT" in cursor_up
            has_forall = "FORALL" in cursor_up
            has_update = "UPDATE" in cursor_up
            if not (has_bulk and has_forall and has_update):
                continue

            # Look for HOP_REDUCTION_OPPORTUNITIES that mention update/cursor
            hop_ops = sdata.get("HOP_REDUCTION_OPPORTUNITIES", []) or []
            cursor_hop_hints = []
            for h in hop_ops:
                if not isinstance(h, dict):
                    continue
                txt = str(h).upper()
                if "UPDATE" in txt or "CURSOR" in txt or "BULK COLLECT" in txt or "FORALL" in txt:
                    hint = (h.get("recommended_action") or h.get("issue_type")
                            or h.get("candidate") or h.get("current_pattern") or "")
                    if hint:
                        cursor_hop_hints.append(str(hint)[:200])

            # Count how many separate cursor-based update passes exist
            import re
            n_cursors = max(1, len(re.findall(
                r"OPEN\s+\w+|CURSOR\s+\w+\s+IS", cursor_text, re.IGNORECASE
            )))
            update_pattern_count = len(re.findall(
                r"FORALL.{1,50}UPDATE", cursor_text, re.IGNORECASE | re.DOTALL
            ))
            update_count = max(1, update_pattern_count)

            key = (job_id, script_name)
            if key in seen:
                continue
            seen.add(key)

            node = node_map[job_id]
            tgt_tables = [t for t in (node.get("tgt_tables") or [])
                          if not t.upper().startswith("LT_")]
            tgt_str = tgt_tables[0] if tgt_tables else job_id

            recommendations.append({
                "id": f"P-{len(recommendations)+1}",
                "category": "P. Post-Load Cursor UPDATE — Fold into INSERT",
                "target_table": tgt_str,
                "description": (
                    f"Job '{job_id}' (script: {script_name}) performs its main load "
                    f"via INSERT SELECT, then executes ~{update_count} separate "
                    f"cursor-based BULK COLLECT → FORALL UPDATE pass(es) to populate "
                    f"additional columns after the initial load. "
                    f"Each cursor opens a separate query, fetches rows in batches, and "
                    f"issues per-batch UPDATEs — adding procedural overhead and extra "
                    f"table scan passes that could be eliminated. "
                    f"Script-level signal: BULK COLLECT + FORALL UPDATE detected in "
                    f"procedure cursor logic."
                ),
                "affected_jobs": [job_id],
                "script_name": script_name,
                "current_hops": update_count + 1,
                "potential_hop_savings": update_count,
                "risk": "MEDIUM",
                "recommendation": (
                    f"CURSOR UPDATE ELIMINATION — Action required:\n"
                    f"(1) Identify each post-load cursor UPDATE in '{script_name}' "
                    f"(e.g. cur_upd_prs_strs_ind_r, cur_upd_wavier_ltd_ind_cols etc.).\n"
                    f"(2) For each cursor, trace its source query — typically a simple "
                    f"GROUP BY / MIN / MAX aggregation over a DIM or plan-design table.\n"
                    f"(3) Merge that aggregation directly into the main INSERT SELECT as "
                    f"a LEFT JOIN subquery or CTE, computing the derived column inline.\n"
                    f"(4) Remove the cursor OPEN / FETCH / FORALL UPDATE / CLOSE block "
                    f"and the associated PL/SQL collection type declarations.\n"
                    f"(5) Result: the table is fully populated in a single pass "
                    f"(INSERT + inline JOIN), eliminating {update_count} separate "
                    f"UPDATE step(s) and reducing procedural context switches.\n"
                    + (f"LLM-identified optimisation hints: {'; '.join(cursor_hop_hints[:2])}"
                       if cursor_hop_hints else "")
                ),
                "_dedup_key": ("P", job_id, script_name),
            })

    return recommendations


# ── Detector Q: Hardcoded Value Mapping ──────────────────────────────────────


def _parse_hardcoded_entry(entry: str) -> dict:
    """
    Parse a single hardcoded_logic string from a script summary into
    structured components: column name, type, and values.
    Returns a dict with keys: column, kind, is_substantial, values_str

    is_substantial=True  → the entry is a genuine multi-branch code-translation
                           CASE/DECODE that is worth externalising to a lookup table.
    is_substantial=False → expected operational constant (active_status='Y',
                           NVL default, source-system filter) — do NOT flag as Q.
    """
    text = entry.strip()
    col_match = re.search(r'\b([VN]_[A-Z0-9_]+(?:_R)?)\b', text, re.IGNORECASE)
    column = col_match.group(1).upper() if col_match else None

    text_lower = text.lower()

    # ── Signals that indicate a GENUINE lookup-table candidate ────────────────
    # The LLM typically uses these phrases for 5+ branch business-rule CASE blocks.
    _SUBSTANTIAL_SIGNALS = (
        "many when",        # LLM says "many WHEN branches"
        "large case",       # LLM says "large CASE"
        "numerous when",
        "explicit mapping", # LLM says "explicit mapping of ..."
        "→",                # Source → target arrow in description
        "->",               # Same with ASCII arrow
        "\u2192",           # Unicode right arrow
        "maps ",            # "maps Agent/Beneficiary -> 'UNK'"
        "map to",
        "code translation",
        "code values",      # "explicit mapping of coverage codes to hard-coded three-digit strings"
        "payee type",       # known high-signal domain term
        "event cause",      # known high-signal domain term (30+ branches)
        "benefit code",     # known high-signal domain term
        "benefit codes",
        "benefit description",
        "coverage type",    # known high-signal domain term (maps 1/2/3 → LTD/STD/Life)
        "coverage code",
        "coverage level",
        "lob type",         # known high-signal domain term
        "class of business", # class_of_business mapping → face amount calculation
        "class_of_business",
        "gross benefit code",
        "payment status",   # when combined with DECODE
        "amount type",
        "premium mode",     # monthly/quarterly/annual → factor
        "mode mapping",
        "carrier name",     # company/carrier name normalization
        "company name",     # address/company name mappings
        "address",          # company address mapping
        "policy prefix",    # LOB mapping by policy prefix
        "line-of-business",
        "line of business",
        "product line",     # product line to indicator flags
        "option",           # option text → rate values
    )

    # ── Signals that indicate an OPERATIONAL CONSTANT (not a lookup candidate) ─
    _OPERATIONAL_SIGNALS = (
        "= 'y'",            # active_status = 'Y'
        "= 'n'",
        "= 'paid'",         # payment status filters (simple)
        "= 'void'",
        "= 'pacs'",         # source system constant
        "= 'cv'",
        "= 'eis'",
        "= 'mgis'",
        "= 'tpa'",
        "= 'vue'",
        "= 'claims'",       # business unit filter
        "= 'main'",
        "= 'mailing'",
        "nvl(",             # NVL/COALESCE default value
        "coalesce(",
        "1900-01-01",       # sentinel dates
        "31-dec-99",
        "31-dec-2999",
        "to_date(",
        "= -1",             # surrogate key default
        " -1)",
        "n_version_number", # technical operational constants
        "n_batch_id",
        "v_created_by",
        "v_last_modified_by",
        "v_source_system_name_r = ",  # source system constant (single value)
        "parallel rebuild",  # technical constant
        "surrogate",
        "set to 0",
        "snapshot_id",
        "partition name",
        # ── Calculation logic (not lookup-table candidates) ──
        "days elapsed",      # Task status → days elapsed = derived calc, not lookup
        "trunc(d_",          # Date truncation calculations
        "trunc(d_completed", # Date calculation
        "trunc(d_created",
        "sysdate - trunc",   # Days between dates = calculation
        "d_due_date_r",      # Due date arithmetic
        # ── Binary flags (1-branch mappings not worth externalizing) ──
        "n_national_account",  # binary indicator 1/0
        "v_individual_or_org_ind",  # binary O/I
        "fein",              # technical identity check
        "v_ach_payment_ind", # ACH binary flag
        "ach payment",       # ACH → Y/N single branch
        # ── Completely stable static values ──
        "month name",        # month names never change
        "january",           # static month label
    )

    # Check if entry looks like an operational constant → disqualify
    is_operational = any(sig in text_lower for sig in _OPERATIONAL_SIGNALS)

    # Check for substantial mapping signal
    has_substantial = any(sig in text_lower for sig in _SUBSTANTIAL_SIGNALS)

    # Also count quoted values: 4+ distinct quoted values suggests a multi-entry mapping
    quoted = re.findall(r"'([^']{1,40})'", text)
    quoted_clean = [v for v in quoted if v.strip() and v not in ("Y", "N", "1", "0")]
    many_values = len(quoted_clean) >= 4

    is_substantial = (has_substantial or many_values) and not is_operational

    # Classify entry kind
    if is_operational:
        if any(kw in text_lower for kw in ("nvl(", "coalesce(", "default", "1900-01-01", "31-dec")):
            kind = "DEFAULT_VALUE"
        else:
            kind = "CONSTANT_FILTER"
    elif any(kw in text_lower for kw in ("case", "decode", "mapping", "map", "→", "->")):
        kind = "VALUE_MAPPING"
    elif any(kw in text_lower for kw in ("exclusion", "exclude", "excluded", "not in")):
        kind = "EXCLUSION_LIST"
    elif any(kw in text_lower for kw in ("filter", "where")):
        kind = "CONSTANT_FILTER"
    else:
        kind = "INLINE_LOGIC"

    values_str = ", ".join(f"'{v}'" for v in quoted_clean[:6]) if quoted_clean else ""

    return {
        "column": column,
        "kind": kind,
        "is_substantial": is_substantial,
        "values_str": values_str,
        "raw": text,
    }


def _suggest_lookup_table_name(column: str | None, script_name: str) -> str:
    """
    Derive a meaningful lookup/reference table name from a column name.
    e.g. V_BENEFIT_GROUP_R → REF_BENEFIT_GROUP_MAP
         V_RECORD_TYPE_R   → REF_RECORD_TYPE_MAP
    """
    if not column:
        base = script_name.upper().removesuffix(".SQL").removesuffix(".TXT")
        # Strip package prefix for brevity
        for pfx in ("PKG_GRP_LOAD_", "PKG_GRP_", "PRC_GRP_LOAD_", "PRC_LOAD_", "PRC_"):
            if base.startswith(pfx):
                base = base[len(pfx):]
                break
        return f"REF_{base[:30]}_MAP"
    # Strip V_ / N_ prefix and _R suffix, then add REF_ + _MAP suffix
    col = re.sub(r"^[VN]_", "", column, flags=re.IGNORECASE)
    col = re.sub(r"_R$", "", col, flags=re.IGNORECASE)
    return f"REF_{col}_MAP"


def detect_hardcoded_value_mappings(graph, sql_index, summaries):
    """
    Q. Script-level: Extensive hardcoded CASE/WHEN value mappings inline in SQL.

    Detects scripts that contain long CASE expressions mapping source values to
    target codes (e.g. event_cause → elimination period code, benefit codes →
    categories) instead of joining to a lookup/reference table.  These inline
    mappings create maintenance debt: every time the business adds a new value
    the SQL package must be redeployed.

    Signal: TRANSFORMATION_SUMMARY.hardcoded_logic list has 3+ entries AND at
    least one entry mentions CASE, mapping, literal, or code values.
    """
    summaries = summaries or {}

    # Build reverse mapping: package_name_upper → (script_name, summary)
    pkg_to_summary: dict[str, tuple[str, dict]] = {}
    for script_name, sdata in summaries.items():
        if not isinstance(sdata, dict) or "error" in sdata:
            continue
        base = script_name.upper().removesuffix(".SQL").removesuffix(".TXT")
        pkg_to_summary[base] = (script_name, sdata)

    node_map = {n["id"]: n for n in graph.get("nodes", [])}
    recommendations = []
    seen: set = set()

    for job_id, job_info in sql_index.items():
        if job_id not in node_map:
            continue
        pkg_names = {p.upper() for p in (job_info.get("package_names") or set())}
        if not pkg_names:
            continue

        for pkg in pkg_names:
            if pkg not in pkg_to_summary:
                continue
            script_name, sdata = pkg_to_summary[pkg]

            ts = sdata.get("TRANSFORMATION_SUMMARY", {}) or {}
            hardcoded = ts.get("hardcoded_logic") or []
            if isinstance(hardcoded, str):
                hardcoded = [hardcoded] if hardcoded.strip() else []

            if not isinstance(hardcoded, list) or len(hardcoded) < 1:
                continue

            # Parse every entry — handle both string format (older summaries) and
            # dict format: {"column": ..., "approx_branch_count": N, "examples": [...]}
            parsed = []
            for h in hardcoded:
                if isinstance(h, str):
                    parsed.append(_parse_hardcoded_entry(h))
                elif isinstance(h, dict):
                    branch_count = h.get("approx_branch_count", 0) or 0
                    col = (h.get("column") or "").strip()
                    examples = h.get("examples") or []
                    vals_str = "; ".join(str(e) for e in examples[:3]) if examples else ""
                    is_subst = isinstance(branch_count, (int, float)) and branch_count >= 5
                    parsed.append({
                        "kind": "VALUE_MAPPING",
                        "column": col.split("(")[0].strip().upper() or None,
                        "is_substantial": is_subst,
                        "values_str": vals_str,
                        "raw": f"{col} ({branch_count} branches)" if col else str(h)[:80],
                    })

            # Require at least one SUBSTANTIAL value-mapping entry.
            # "Substantial" means a genuine multi-branch CASE/DECODE code translation
            # (5+ WHEN branches, e.g. event_cause → code, payee_type → 'UNK'/'GRP').
            # Operational constants (active_status='Y', NVL(-1), source-system='PACS')
            # are expected by design and are NOT lookup-table candidates.
            substantial_mappings = [
                p for p in parsed
                if p["kind"] in ("VALUE_MAPPING", "EXCLUSION_LIST")
                and p["is_substantial"]
            ]
            if not substantial_mappings:
                continue

            key = (job_id, script_name)
            if key in seen:
                continue
            seen.add(key)

            node = node_map[job_id]
            tgt_tables = [t for t in (node.get("tgt_tables") or [])
                          if not t.upper().startswith("LT_")]
            tgt_str = tgt_tables[0] if tgt_tables else job_id

            # ── Build clear description ──────────────────────────────────────
            # Group into mapping, exclusion, filter, default
            by_kind: dict[str, list] = {}
            for p in parsed:
                by_kind.setdefault(p["kind"], []).append(p)

            desc_lines = [
                f"Script '{script_name}' (loaded by job '{job_id}') contains "
                f"{len(parsed)} hardcoded logic block(s) that should be reviewed:",
                "",
            ]
            kind_labels = {
                "VALUE_MAPPING":   "Value Mappings (CASE/DECODE — business codes hardcoded in SQL)",
                "EXCLUSION_LIST":  "Exclusion Lists (hardcoded NOT IN / exclusion filters)",
                "CONSTANT_FILTER": "Constant Filters (literal flag/status comparisons)",
                "DEFAULT_VALUE":   "Default Values (NVL/COALESCE literal defaults)",
                "INLINE_LOGIC":    "Other Inline Logic",
            }
            for kind, label in kind_labels.items():
                entries = by_kind.get(kind, [])
                if not entries:
                    continue
                desc_lines.append(f"  [{label}]")
                for i, p in enumerate(entries, 1):
                    col_tag = f"  Column: {p['column']}" if p["column"] else ""
                    vals_tag = f"  Values: {p['values_str']}" if p["values_str"] else ""
                    desc_lines.append(f"    {i}. {p['raw']}{col_tag}{vals_tag}")
                desc_lines.append("")

            desc_lines.append(
                "Every business change (new code, renamed status, added category) "
                "currently requires a package recompile and a full deployment cycle "
                "instead of a simple data-table INSERT."
            )
            description = "\n".join(desc_lines)

            # ── Build specific recommendation ────────────────────────────────
            rec_lines = ["HARDCODED LOGIC REMEDIATION — Specific actions per block:\n"]

            # Step 1 — SUBSTANTIAL value mappings → lookup tables (only genuine candidates)
            vm_entries = [p for p in by_kind.get("VALUE_MAPPING", []) if p.get("is_substantial")]
            if vm_entries:
                rec_lines.append("STEP 1: Externalize value mappings to lookup tables")
                for p in vm_entries:
                    lut = _suggest_lookup_table_name(p["column"], script_name)
                    col_desc = p["column"] or "(see entry below)"
                    vals_hint = f" (currently maps: {p['values_str']})" if p["values_str"] else ""
                    rec_lines.append(
                        f"  • {p['raw']}\n"
                        f"    → Column affected : {col_desc}\n"
                        f"    → Suggested table : {lut} (SOURCE_CODE, TARGET_CODE, DESCRIPTION, EFFECTIVE_DT)\n"
                        f"    → Action          : Create {lut}{vals_hint}, "
                        f"then replace the CASE/DECODE block with:\n"
                        f"      LEFT JOIN {lut} lut ON lut.SOURCE_CODE = <src_col>\n"
                        f"      and reference lut.TARGET_CODE in the SELECT list."
                    )
                rec_lines.append("")

            # Step 2 — exclusion lists → config table
            ex_entries = by_kind.get("EXCLUSION_LIST", [])
            if ex_entries:
                rec_lines.append("STEP 2: Move exclusion/inclusion lists to a config table")
                for p in ex_entries:
                    col_desc = p["column"] or "(see entry below)"
                    vals_hint = f" — current values: {p['values_str']}" if p["values_str"] else ""
                    rec_lines.append(
                        f"  • {p['raw']}\n"
                        f"    → Column affected : {col_desc}{vals_hint}\n"
                        f"    → Action          : Create REF_EXCLUSION_CONFIG "
                        f"(CONTEXT_NAME, EXCLUDED_VALUE, ACTIVE_FLG) and replace "
                        f"the NOT IN (...) literal list with a subquery:\n"
                        f"      WHERE <col> NOT IN (SELECT EXCLUDED_VALUE FROM REF_EXCLUSION_CONFIG "
                        f"WHERE CONTEXT_NAME = '{col_desc}_EXCL' AND ACTIVE_FLG = 'Y')"
                    )
                rec_lines.append("")

            # Step 3 — constant filters → note only
            cf_entries = by_kind.get("CONSTANT_FILTER", [])
            if cf_entries:
                rec_lines.append("STEP 3: Review constant filters (lower priority)")
                for p in cf_entries:
                    rec_lines.append(
                        f"  • {p['raw']}\n"
                        f"    → This is a fixed-value filter. Confirm it is intentional "
                        f"and document it; only externalize if the value changes across environments."
                    )
                rec_lines.append("")

            # Step 4 — validation
            rec_lines.append(
                "VALIDATION: After each replacement, run a row-count and column-distribution "
                "comparison between the original CASE output and the JOIN-based output across "
                "at least 3 full load cycles before removing the old inline logic."
            )

            recommendation = "\n".join(rec_lines)

            recommendations.append({
                "id": f"Q-{len(recommendations)+1}",
                "category": "Q. Hardcoded Value Mapping — Externalize to Lookup Table",
                "target_table": tgt_str,
                "description": description,
                "affected_jobs": [job_id],
                "script_name": script_name,
                "hardcoded_entries": hardcoded[:5],
                "current_hops": 1,
                "potential_hop_savings": 0,  # no hop saving; maintainability improvement
                "risk": "LOW",
                "recommendation": recommendation,
                "_dedup_key": ("Q", job_id, script_name),
            })

    return recommendations


# ── Detector G ──────────────────────────────────────────────────────────────


def detect_duplicate_column_paths(lineage_rows):
    """G. Same (SRC.col -> TGT.col) produced by multiple jobs."""
    path_jobs = defaultdict(set)
    for r in lineage_rows:
        src_t, tgt_t = r.get("SRC_TABLE"), r.get("TGT_TABLE")
        job = r.get("DEPENDENT_JOB")
        if not (src_t and tgt_t and job):
            continue
        src_cols = sorted(parse_col_list(r.get("SOURCE_COL")))
        tgt_cols = sorted(parse_col_list(r.get("TARGET_COL")))
        if len(src_cols) == len(tgt_cols) and src_cols:
            for s, t in zip(src_cols, tgt_cols):
                path_jobs[(src_t, s, tgt_t, t)].add(job)

    pair_dup_count = defaultdict(lambda: {"jobs": set(), "cols": set()})
    for (src_t, s_c, tgt_t, t_c), jobs in path_jobs.items():
        if len(jobs) > 1:
            pair_dup_count[(src_t, tgt_t)]["jobs"] |= jobs
            pair_dup_count[(src_t, tgt_t)]["cols"].add(f"{s_c}->{t_c}")

    recommendations = []
    for (src_t, tgt_t), info in pair_dup_count.items():
        # Skip log/debug/audit side-effect tables
        if _is_log_or_debug_table(tgt_t) or _is_log_or_debug_table(src_t):
            continue
        recommendations.append(
            {
                "id": f"G-{len(recommendations)+1}",
                "category": "G. Duplicate Column-Lineage Paths",
                "target_table": tgt_t,
                "description": (
                    f"{len(info['cols'])} columns flow {src_t} -> {tgt_t} via "
                    f"{len(info['jobs'])} different jobs (same path)."
                ),
                "affected_jobs": sorted(info["jobs"]),
                "current_hops": len(info["jobs"]),
                "potential_hop_savings": len(info["jobs"]) - 1,
                "risk": "MEDIUM",
                "recommendation": (
                    f"Consolidate column writes - only 1 job needs to load "
                    f"{len(info['cols'])} columns from {src_t} -> {tgt_t}."
                ),
                "_dedup_key": ("G", src_t, tgt_t),
            }
        )
    return recommendations


# ── Detector H ──────────────────────────────────────────────────────────────


def detect_stale_intermediates(graph, global_table_usage=None, summaries=None):
    """H. Tables written but never read in this RPT graph."""
    table_map = build_table_producers_consumers(graph)

    # Set of jobs that have at least one downstream edge in this graph.
    jobs_with_children = {
        link["source"]
        for link in graph.get("links", [])
        if link.get("source") and link.get("target")
    }

    # Build intra-procedure staging index from script summaries.
    # Tables listed as intermediate_tables_created or staging_layers inside a
    # stored procedure are created AND consumed within that same procedure
    # (CTAS pattern).  They won't appear as lineage consumers but are NOT
    # dead writes — they are ephemeral staging steps.
    _intra_proc = set()
    # Also build a confirmed-input-source index from INPUT_DEPENDENCIES.
    # DIM/FCT tables used only as JOIN lookup sources are not captured by the
    # column-level SQL parser (which only traces INSERT/SELECT column mappings).
    # The LLM summaries DO list them under INPUT_DEPENDENCIES.source_tables.
    # If a DIM/FCT table appears here, it is an active JOIN source — suppress
    # the H finding entirely rather than emitting a misleading parser-gap warning.
    _input_sources: set[str] = set()

    def _extract_tbl_name(entry: str) -> str:
        """Strip annotations like '(target read/metadata)' and schema prefixes
        like 'ATOMIC.' or 'dbo.' from table entry strings."""
        # Strip anything after a space or open-paren (annotations)
        name = entry.split("(")[0].split(" ")[0].strip().upper()
        # Strip leading schema prefix (SCHEMA.TABLENAME → TABLENAME)
        if "." in name:
            name = name.rsplit(".", 1)[-1]
        return name

    for sdata in (summaries or {}).values():
        if not isinstance(sdata, dict):
            continue
        pf = sdata.get("PROCESS_FLOW", {})
        interm = pf.get("intermediate_tables_created", [])
        if isinstance(interm, str):
            interm = [interm]
        for entry in (interm or []):
            _intra_proc.add(str(entry).upper())
        staging = str(pf.get("staging_layers", ""))
        if staging:
            _intra_proc.add(staging.upper())

        # INPUT_DEPENDENCIES.source_tables (dict with FACT/DIM/OTHER lists)
        deps = sdata.get("INPUT_DEPENDENCIES", {})
        if isinstance(deps, dict):
            src_tables = deps.get("source_tables", {})
            if isinstance(src_tables, dict):
                for _cat, tbl_list in src_tables.items():
                    if isinstance(tbl_list, list):
                        for entry in tbl_list:
                            if isinstance(entry, str):
                                tbl_name = _extract_tbl_name(entry)
                                if tbl_name and not tbl_name.startswith("PKG_"):
                                    _input_sources.add(tbl_name)
            elif isinstance(src_tables, list):
                for entry in src_tables:
                    if isinstance(entry, str):
                        tbl_name = _extract_tbl_name(entry)
                        if tbl_name and not tbl_name.startswith("PKG_"):
                            _input_sources.add(tbl_name)
            # procedure_level_dependencies may also list per-procedure inputs
            proc_deps = deps.get("procedure_level_dependencies", {})
            if isinstance(proc_deps, dict):
                for _proc, items in proc_deps.items():
                    if isinstance(items, list):
                        for item in items:
                            if isinstance(item, str):
                                tbl_name = _extract_tbl_name(item)
                                if tbl_name and not tbl_name.startswith("PKG_"):
                                    _input_sources.add(tbl_name)

    recommendations = []
    for tbl, info in table_map.items():
        if not info["producers"] or info["consumers"]:
            continue

        if global_table_usage and global_table_usage.get(tbl, {}).get("consumers"):
            # Not stale globally even if stale in this single-RPT slice.
            continue

        # ── Downstream-edges guard ─────────────────────────────────────────────
        # If any producer job of this table has downstream edges in the Tidal
        # dependency graph, those child jobs likely consume this table (e.g.
        # as a JOIN source) even if the column-level lineage doesn't show it.
        # Suppress to avoid false "stale" findings for JOIN-only dependencies.
        if any(p in jobs_with_children for p in info["producers"]):
            continue

        upper = tbl.upper()
        # Skip RPT output tables, archive, staging prefixes, and log/debug side-effects
        if upper.startswith("RPT_") or upper.startswith("ARCHIVE") or upper.startswith("STG_"):
            continue
        if _is_log_or_debug_table(tbl):
            continue

        globally_confirmed = global_table_usage is not None

        # DIM_ and FCT_ tables are frequently used as JOIN lookup sources in
        # stored procedures.  Column-level SQL parsers only capture column
        # extractions (INSERT/SELECT column mappings), NOT JOIN-only references.
        # So a DIM_ table can appear as zero-consumer in lineage even when it IS
        # actively used — this is a known parser gap, not a genuine dead write.
        is_dim_or_fct = upper.startswith("DIM_") or upper.startswith("FCT_")

        # ── Intra-procedure staging check (from script summaries) ─────────────
        # If the table is listed in any script's intermediate_tables_created or
        # staging_layers, it is consumed within the same stored procedure (CTAS).
        # The column-level lineage extractor misses CTAS consumption, making the
        # table appear as a dead write even though it IS being used.
        is_intra_proc = any(upper in entry for entry in _intra_proc)

        # ── Confirmed-input-source check (from script summaries) ──────────────
        # DIM/FCT tables used only as JOIN lookup sources will appear under
        # INPUT_DEPENDENCIES.source_tables in the LLM summaries but produce zero
        # column-level consumers (the parser misses JOIN-only references).
        # If ANY script summary lists this table as an input source, it IS
        # actively used — suppress the H finding entirely.
        is_confirmed_input = upper in _input_sources

        if is_intra_proc:
            # Not a dead write — it's an ephemeral intra-procedure staging step.
            # Still flag it as a hop reduction candidate (replace CTAS with CTE).
            recommendations.append(
                {
                    "id": f"H-{len(recommendations)+1}",
                    "category": "H. Stale Intermediate Table",
                    "target_table": tbl,
                    "description": (
                        f"Table '{tbl}' is written by {len(info['producers'])} job(s) and "
                        f"shows NO column-level consumers in lineage. However, script analysis "
                        f"confirms this is an intra-procedure staging table (CTAS pattern): "
                        f"it is created and consumed within the same stored procedure. "
                        f"This is NOT a dead write."
                    ),
                    "affected_jobs": sorted(info["producers"]),
                    "current_hops": 1,
                    "potential_hop_savings": 1,
                    "risk": "MEDIUM",
                    "recommendation": (
                        f"This table is an ephemeral staging step created via CTAS inside a "
                        f"stored procedure. Hop reduction opportunity: replace the CTAS "
                        f"staging step with a CTE or inline subquery in the consuming SQL "
                        f"to eliminate the physical table write entirely. "
                        f"Risk: verify the CTE result is equivalent before removing."
                    ),
                    "_dedup_key": ("H", tbl),
                }
            )
            continue

        if is_dim_or_fct and is_confirmed_input:
            # LLM summary confirms this DIM/FCT table is an active JOIN source.
            # The zero-consumer appearance in column-level lineage is a parser gap,
            # not a dead write.  Suppress entirely — no H finding needed.
            continue

        if is_dim_or_fct:
            confidence_label = "POSSIBLE dead write — low confidence"
            scope_text = (
                "No column-level consumers found across all RPT lineages or the Tidal "
                "dependency graph. However, DIM/FCT tables are commonly referenced as "
                "JOIN lookup sources in SQL procedures — the column-level parser does NOT "
                "capture JOIN-only references, so this may be a parser gap rather than a "
                "genuine dead write."
            )
            risk = "MEDIUM"
            rec_text = (
                f"PARSER-GAP RISK: '{tbl}' has no column-level consumers in the lineage, "
                f"but DIM tables are frequently used as JOIN-only lookup sources that the "
                f"SQL parser does not trace. Before taking any action:\n"
                f"  (1) Search the SQL packages that load downstream RPT tables for any "
                f"JOIN to '{tbl}'.\n"
                f"  (2) If found in a JOIN clause — this is a parser gap, not a dead write. "
                f"Flag the lineage extractor to capture this dependency.\n"
                f"  (3) If not found anywhere — confirm with the data owner and remove the "
                f"write step only after sign-off."
            )
        else:
            confidence_label = "confirmed dead write" if globally_confirmed else "potential dead write (local view)"
            scope_text = (
                "across all RPT lineages (globally confirmed dead write)"
                if globally_confirmed
                else "in this RPT graph (local view only — global check not available)"
            )
            risk = "HIGH"
            rec_text = (
                f"{'Global check confirmed: no other RPT pipeline reads this table. ' if globally_confirmed else ''}"
                f"Safe to remove the write step for '{tbl}' after verifying it is not "
                f"consumed by any ad-hoc query, report, or external process outside Tidal."
            )

        recommendations.append(
            {
                "id": f"H-{len(recommendations)+1}",
                "category": "H. Stale Intermediate Table",
                "target_table": tbl,
                "description": (
                    f"Table '{tbl}' is written by {len(info['producers'])} job(s) but has "
                    f"NO column-level consumers {scope_text}."
                ),
                "affected_jobs": sorted(info["producers"]),
                "current_hops": 1,
                "potential_hop_savings": 1,
                "risk": risk,
                "recommendation": rec_text,
                "_dedup_key": ("H", tbl),
            }
        )
    return recommendations


# ── Detector I ──────────────────────────────────────────────────────────────


def detect_cycles(graph):
    """I. Cycles in the dependency graph."""
    children, _ = build_adjacency(graph)
    WHITE, GREY, BLACK = 0, 1, 2
    color = defaultdict(lambda: WHITE)
    cycles = []

    def dfs(node, path):
        color[node] = GREY
        for nxt in children.get(node, set()):
            if color[nxt] == GREY:
                if nxt in path:
                    cyc = path[path.index(nxt) :] + [nxt]
                    cycles.append(cyc)
            elif color[nxt] == WHITE:
                dfs(nxt, path + [nxt])
        color[node] = BLACK

    for n in graph["nodes"]:
        if color[n["id"]] == WHITE:
            dfs(n["id"], [n["id"]])

    seen = set()
    recommendations = []
    for cyc in cycles:
        key = tuple(sorted(set(cyc)))
        if key in seen:
            continue
        seen.add(key)
        recommendations.append(
            {
                "id": f"I-{len(recommendations)+1}",
                "category": "I. Cyclic Refresh Dependency",
                "target_table": "—",
                "description": f"Cycle detected: {' -> '.join(cyc)}.",
                "affected_jobs": list(key),
                "current_hops": len(cyc),
                "potential_hop_savings": 0,
                "risk": "HIGH",
                "recommendation": (
                    "Break cycle - cyclic refreshes cause redundant work or "
                    "incorrect ordering."
                ),
                "_dedup_key": ("I", key),
            }
        )
    return recommendations


# ── Detector J ──────────────────────────────────────────────────────────────


def detect_overfanout_mvs(graph, fanout_threshold=5):
    """J. MVs consumed by many jobs - candidate for splitting."""
    children, _ = build_adjacency(graph)
    recommendations = []
    for n in graph["nodes"]:
        if n["category"] != "MV":
            continue
        consumers = children.get(n["id"], set())
        if len(consumers) < fanout_threshold:
            continue
        recommendations.append(
            {
                "id": f"J-{len(recommendations)+1}",
                "category": "J. Over-Fanout MV",
                "target_table": ", ".join(n.get("tgt_tables", []) or []) or n["id"],
                "description": (
                    f"MV '{n['id']}' is consumed by {len(consumers)} downstream "
                    f"jobs - high blast radius if changed, possible candidate "
                    f"for splitting into focused MVs."
                ),
                "affected_jobs": [n["id"]] + sorted(consumers),
                "depth": n.get("depth"),
                "current_hops": 1,
                "potential_hop_savings": 0,
                "risk": "LOW",
                "recommendation": (
                    "Review consumer column usage. If consumers use disjoint "
                    "column subsets, split MV into 2+ smaller MVs."
                ),
                "_dedup_key": ("J", n["id"]),
            }
        )
    return recommendations


# ── Detector M ──────────────────────────────────────────────────────────────


def detect_orchestration_edge_no_data_handoff(graph, lineage_rows):
    """
    M. Detect MV->MV dependency edges where downstream does not read upstream output table.
    This flags likely scheduler-edge cleanup opportunities (edge removal), not blind job removal.
    """
    job_reads = defaultdict(set)
    job_writes = defaultdict(set)
    for r in lineage_rows:
        job = r.get("DEPENDENT_JOB")
        if not job:
            continue
        job_reads[job] |= parse_table_list(r.get("SRC_TABLE"))
        job_writes[job] |= parse_table_list(r.get("TGT_TABLE"))

    # Augment job_reads/job_writes with graph node src/tgt_tables so that MV
    # refresh jobs (which produce no column-level lineage rows) are captured.
    for n in graph.get("nodes", []):
        nid = n.get("id", "")
        if not nid:
            continue
        for t in n.get("tgt_tables") or []:
            job_writes[nid].add(t.upper())
        for t in n.get("src_tables") or []:
            job_reads[nid].add(t.upper())

    # Build table-level 1-hop reachability across ALL jobs and VIEW expansions.
    # Used for the multi-hop transitive data dependency check below.
    # Pattern: TABLE_A is read by Job_X which writes TABLE_B  →  A produces B in 1 hop.
    table_produces = defaultdict(set)
    for job, reads in job_reads.items():
        for src_t in reads:
            for tgt_t in job_writes.get(job, set()):
                table_produces[src_t.upper()].add(tgt_t.upper())

    nodes = {n.get("id"): n for n in graph.get("nodes", [])}
    children, parents = build_adjacency(graph)

    recommendations = []
    for link in graph.get("links", []):
        src = link.get("source")
        tgt = link.get("target")
        src_node = nodes.get(src, {})
        tgt_node = nodes.get(tgt, {})

        # Keep this detector high-confidence: match MV->MV orchestration mismatch only.
        if src_node.get("category") != "MV" or tgt_node.get("category") != "MV":
            continue

        src_writes_all = set(job_writes.get(src, set()))
        tgt_reads = set(job_reads.get(tgt, set()))
        if not tgt_reads:
            continue

        # Strip log/audit/status side-effect tables from upstream writes — these
        # are internal procedure side-effects, not data handoffs.
        _SKIP_PATS = ("LOG", "AUDIT", "EXEC_STATUS")
        src_writes = {t for t in src_writes_all
                      if not any(p in t.upper() for p in _SKIP_PATS)
                      and not t.upper().startswith("RPT_")}
        if not src_writes:
            continue

        # Normalise case for comparison (mixed-case from gudu vs uppercase from shell)
        src_writes_upper = {t.upper() for t in src_writes}
        tgt_reads_upper = {t.upper() for t in tgt_reads}

        # ── Graph-level src_table JOIN guard ──────────────────────────────────
        # Column lineage only captures column extraction, not JOIN-only references.
        # If the upstream node's output tables appear in the downstream node's graph
        # src_tables list (which reflects ALL SQL table dependencies including JOINs
        # as parsed by the SQL dependency extractor), treat this as a valid dependency
        # and suppress the M finding rather than falsely recommending edge removal.
        tgt_graph_src = {t.upper() for t in (tgt_node.get("src_tables") or [])}
        if src_writes_upper & tgt_graph_src:
            continue  # upstream output is a JOIN source for downstream — edge is justified

        overlap = src_writes_upper & tgt_reads_upper
        if overlap:
            continue

        # Check for an indirect data handoff routed through a VIEW node.
        # Pattern: upstream writes TABLE_X  →  VIEW sources TABLE_X  →  downstream reads VIEW_NAME
        for vn in graph.get("nodes", []):
            if vn.get("category") != "VIEW" and vn.get("source") != "VIEW_EXPANSION":
                continue
            v_srcs = {t.upper() for t in vn.get("src_tables", []) or []}
            v_tgts = {t.upper() for t in vn.get("tgt_tables", []) or []}
            # Also recognise the plain table name embedded in VIEW_<TableName> id
            vid = vn["id"]
            if vid.upper().startswith("VIEW_"):
                v_tgts.add(vid[5:].upper())
            if (src_writes_upper & v_srcs) and (tgt_reads_upper & v_tgts):
                overlap = src_writes_upper & v_srcs  # indirect handoff confirmed
                break

        if overlap:
            continue  # data handoff exists (direct or via VIEW)

        # Multi-hop transitive data dependency check.
        # The 1-hop VIEW check above misses chains like:
        #   upstream_tbl → VIEW → intermediate_tbl → another_job → downstream_tbl
        # Use a BFS through `table_produces` (built from all lineage rows) up to 4 hops.
        # If upstream's written tables can transitively reach any of downstream's read tables,
        # the Tidal scheduling edge IS justified — there is a real (indirect) data dependency.
        if not overlap:
            _frontier = set(src_writes_upper)
            _visited = set(src_writes_upper)
            for _hop in range(4):
                _next = set()
                for _t in _frontier:
                    _next |= table_produces.get(_t, set())
                _next -= _visited
                if _next & tgt_reads_upper:
                    overlap = _next & tgt_reads_upper  # transitive handoff confirmed
                    break
                _frontier = _next
                _visited |= _next
                if not _frontier:
                    break

        if overlap:
            continue  # transitive data dependency found — Tidal edge is valid

        src_tgts = sorted(src_writes)
        tgt_srcs = sorted(tgt_reads)
        succ_count = len(children.get(src, set()))
        parent_count = len(parents.get(tgt, set()))

        recommendations.append(
            {
                "id": f"M-{len(recommendations)+1}",
                "category": "M. Orchestration Edge Without Data Handoff (MV->MV)",
                "target_table": ", ".join(src_tgts[:3]),
                "description": (
                    f"Dependency edge '{src}' -> '{tgt}' has no table handoff. "
                    f"Upstream writes {src_tgts[:3]} but downstream reads {tgt_srcs[:4]}. "
                    f"Intersection is empty. Successors of source: {succ_count}; parents of target: {parent_count}."
                ),
                "affected_jobs": [src, tgt],
                "src_written_tables": src_tgts,
                "tgt_read_tables": tgt_srcs,
                "current_hops": 1,
                "potential_hop_savings": 1,
                "risk": "MEDIUM",
                "recommendation": (
                    f"Validate edge-level dependency cleanup for '{src}' -> '{tgt}'. "
                    f"Checklist: (1) confirm '{tgt}' SQL does not reference {src_tgts[:2]}, "
                    f"(2) verify no hidden side-effects/temp-table dependencies, "
                    f"(3) remove or relax this scheduler edge in Tidal, "
                    f"(4) run parity + cycle-time comparison for 3+ runs."
                ),
                "_dedup_key": ("M", src, tgt),
            }
        )

    return recommendations


# ── Detector N ──────────────────────────────────────────────────────────────


def detect_post_consumption_updates(graph, lineage_rows):
    """
    N. Detect tables where one job consumes Y to produce T, while another job updates Y.
    If update appears after the consumer (B->C edge), this may be a sequencing and hop-reduction candidate.

    Sub-type N-CL (Circular Update Chain):
    Detects full-load producer jobs that also read tables that are themselves
    DOWNSTREAM of the table being loaded.  This creates a lagged-cycle circular
    dependency: the current-cycle load uses last-cycle aggregated/derived data
    to fill in indicator/decision columns.  The recommendation is to fold that
    downstream-derived logic directly into the INSERT SELECT, removing the
    cross-cycle dependency.

    Sub-type N-PL (Post-Load Updater):
    Detects a separate UPDATE-only job that modifies a table after a full-load
    job, where the update uses downstream-derived data (same circular pattern
    but split across two Tidal jobs rather than embedded in one).
    """
    job_reads = defaultdict(set)
    job_writes = defaultdict(set)
    for r in lineage_rows:
        job = r.get("DEPENDENT_JOB")
        if not job:
            continue
        # Skip synthetic VIEW rows — they are SQL view definitions, not real
        # Tidal jobs.  Treating them as consumers/updaters creates false-positive
        # N. findings because a VIEW has no scheduling order constraint.
        if r.get("LINEAGE_SOURCE") == "VIEW_EXPANSION" or r.get("JOB_CATEGORY") == "VIEW":
            continue
        job_reads[job] |= parse_table_list(r.get("SRC_TABLE"))
        job_writes[job] |= parse_table_list(r.get("TGT_TABLE"))

    table_consumers = defaultdict(set)  # reads Y, writes T!=Y
    table_updaters = defaultdict(set)   # reads Y and writes Y
    table_producers = defaultdict(set)  # writes Y

    for job in set(job_reads.keys()) | set(job_writes.keys()):
        reads = job_reads.get(job, set())
        writes = job_writes.get(job, set())
        for t in writes:
            table_producers[t].add(job)
        for y in reads:
            if y in writes:
                if _is_full_load_job(job):
                    # Truncate-and-load (producer) job — not an updater.
                    # These jobs fully reload a table; classifying them as
                    # "updaters" produces N false positives.
                    table_producers[y].add(job)
                else:
                    table_updaters[y].add(job)
            elif writes:
                table_consumers[y].add(job)

    edges = {(l.get("source"), l.get("target")) for l in graph.get("links", [])}
    recommendations = []
    seen = set()

    for y in sorted(set(table_consumers.keys()) & set(table_updaters.keys())):
        consumers = sorted(table_consumers.get(y, set()))
        updaters = sorted(table_updaters.get(y, set()))
        producers = sorted(table_producers.get(y, set()))

        for b in consumers:
            b_targets = sorted(job_writes.get(b, set()) - {y})
            if not b_targets:
                continue

            for c in updaters:
                if b == c:
                    continue
                key = (y, b, c)
                if key in seen:
                    continue
                seen.add(key)

                update_after_consume = (b, c) in edges
                update_before_consume = (c, b) in edges
                ordering = (
                    "AFTER_CONSUME"
                    if update_after_consume
                    else ("BEFORE_CONSUME" if update_before_consume else "UNORDERED")
                )

                risk = "MEDIUM" if update_after_consume else "LOW"
                rec_text = (
                    f"Validate sequencing for table '{y}': consumer '{b}' materializes {b_targets[:2]} before updater '{c}' modifies '{y}'. "
                    "Actions: (1) verify whether updater columns are required by consumer outputs, "
                    "(2) if yes, move updater logic before consumer or fold into consumer SQL, "
                    "(3) if no, isolate updater into post-processing and remove unnecessary dependency hops, "
                    "(4) run row-count and key metric parity checks for 3+ cycles."
                    if update_after_consume
                    else (
                        f"Table '{y}' has both consumer '{b}' (Y->T) and updater '{c}' (Y->Y). "
                        "Review whether updater can be folded into the producer/consumer path to remove extra hops and simplify orchestration."
                    )
                )

                recommendations.append(
                    {
                        "id": f"N-{len(recommendations)+1}",
                        "category": "N. Post-Consumption Update Pattern",
                        "target_table": y,
                        "description": (
                            f"Table '{y}' is consumed by '{b}' to produce {b_targets[:3]}, while '{c}' also updates '{y}'. "
                            f"Ordering signal: {ordering}. Producers of '{y}': {producers[:3] if producers else ['unknown']}."
                        ),
                        "affected_jobs": [b, c] + producers[:2],
                        "ordering_signal": ordering,
                        "consumer_job": b,
                        "updater_job": c,
                        "consumer_targets": b_targets,
                        "current_hops": 2,
                        "potential_hop_savings": 1,
                        "risk": risk,
                        "recommendation": rec_text,
                        "_dedup_key": ("N", y, b, c),
                    }
                )

    # ── N-CL: Circular Update Chain ──────────────────────────────────────────
    # Pattern: Full-load job P writes table T, but also reads table Z where
    # Z is produced DOWNSTREAM of T in the same pipeline (i.e., Z = f(T) from
    # a previous cycle).  This is a lagged-cycle circular dependency:
    #   Current run:  P reads Z_prev (last cycle) to fill indicator columns → writes T_curr
    #   Later:        child job builds Z_curr from T_curr
    #   Next run:     P reads Z_curr (which contains T_curr-derived data)
    # Recommendation: fold Z's aggregation logic directly into P's INSERT SELECT
    # so the full-load job no longer depends on a previous-cycle artifact.
    #
    # N-PL: Post-Load Updater
    # Pattern: A separate updater job U (non-full-load, reads T AND writes T)
    # runs after full-load producer P, where U also reads downstream tables Z.
    # This is the same circular pattern split across two Tidal jobs.

    # Build table→table dependency map across the whole pipeline (src → tgt)
    tbl_feeds = defaultdict(set)
    for j, writes in job_writes.items():
        for src in job_reads.get(j, set()):
            for tgt in writes:
                if src.upper() != tgt.upper():
                    tbl_feeds[src.upper()].add(tgt.upper())
    # Augment tbl_feeds with Tidal graph node src_tables → tgt_tables so that
    # dependencies visible in the Tidal graph but not in column-level lineage
    # (e.g. RPT_CLAIM_DTL_R → RPT_FCT_RPT_CLAIM_SUMMARY_R via job-24) are
    # captured for accurate downstream BFS traversal.
    for n in graph.get("nodes", []):
        for src in n.get("src_tables") or []:
            for tgt in n.get("tgt_tables") or []:
                if src.upper() != tgt.upper():
                    tbl_feeds[src.upper()].add(tgt.upper())

    # Also build a graph-augmented job-reads map so that source tables captured
    # in the Tidal graph node metadata (but absent from column-level lineage)
    # are visible to the N-CL circular-reads check.
    # Example: EDP_GRP_EDW_LOAD_RPT_CLAIM_DTL_R-19 reads RPT_FCT_RPT_CLAIM_SUMMARY_R
    # in its stored procedure code — this shows up in graph.nodes[].src_tables but
    # not always in column-level lineage SRC_TABLE rows.
    graph_job_reads: dict[str, set] = defaultdict(set)
    for n in graph.get("nodes", []):
        nid = n.get("id", "")
        if nid:
            for t in n.get("src_tables") or []:
                graph_job_reads[nid].add(t.upper())
            for t in n.get("tgt_tables") or []:
                job_writes[nid].add(t.upper())  # also enrich job_writes

    def _job_reads_augmented(job: str) -> set:
        """Merge lineage-derived and graph-node-derived reads for a job."""
        return job_reads.get(job, set()) | graph_job_reads.get(job, set())

    def _downstream_tables(start_tbl: str, max_depth: int = 8) -> set:
        """BFS: all tables reachable from start_tbl via tbl_feeds."""
        visited: set = set()
        queue = [start_tbl.upper()]
        depth = 0
        while queue and depth < max_depth:
            nxt = []
            for t in queue:
                for child in tbl_feeds.get(t, set()):
                    if child not in visited:
                        visited.add(child)
                        nxt.append(child)
            queue = nxt
            depth += 1
        return visited

    edges_set = {(l.get("source"), l.get("target")) for l in graph.get("links", [])}

    # Collect all producer jobs per table
    all_producers: dict[str, set] = defaultdict(set)
    all_updaters:  dict[str, set] = defaultdict(set)
    for j in set(list(job_reads.keys()) + list(job_writes.keys())):
        writes = job_writes.get(j, set())
        reads  = job_reads.get(j, set())
        for t in writes:
            if t in reads:
                if _is_full_load_job(j):
                    all_producers[t].add(j)
                else:
                    all_updaters[t].add(j)
            else:
                all_producers[t].add(j)  # pure writer = producer

    seen_cl: set = set()
    for tbl_up in sorted(all_producers.keys() | all_updaters.keys()):
        # Skip log/debug/audit tables — these are infrastructure tables, not
        # data-pipeline tables.  Circular-dependency detection is only meaningful
        # for actual business data tables.
        if _is_log_or_debug_table(tbl_up):
            continue
        downstream = _downstream_tables(tbl_up)
        if not downstream:
            continue

        # N-CL: full-load producer that reads from its own downstream
        for prod_job in all_producers.get(tbl_up, set()):
            circular_reads = (
                {t for t in _job_reads_augmented(prod_job) if not _is_log_or_debug_table(t)}
                & downstream
            ) - {tbl_up}
            if not circular_reads:
                continue
            key = ("N-CL", tbl_up, prod_job)
            if key in seen_cl:
                continue
            seen_cl.add(key)

            downstream_list = sorted(circular_reads)[:4]
            recommendations.append({
                "id": f"N-{len(recommendations)+1}",
                "category": "N. Circular Update Chain (Post-Load Self-Reference)",
                "target_table": tbl_up,
                "description": (
                    f"Job '{prod_job}' is a full-load job for '{tbl_up}', but it also reads "
                    f"{downstream_list} — table(s) that are DOWNSTREAM of '{tbl_up}' "
                    f"(i.e., they are derived FROM '{tbl_up}' in a previous pipeline cycle). "
                    f"This creates a lagged-cycle circular dependency: the current-run load "
                    f"uses last-cycle aggregated/derived data to fill in indicator or decision "
                    f"columns, meaning the output of '{tbl_up}' in this cycle will differ from "
                    f"what it would be if fresh data were used. "
                    f"Note: '{tbl_up}' gets updated using "
                    f"{downstream_list[0] if downstream_list else '?'}, which is itself derived "
                    f"from '{tbl_up}' — this is a circular feedback loop across pipeline cycles."
                ),
                "affected_jobs": [prod_job],
                "circular_read_tables": sorted(circular_reads)[:6],
                "current_hops": 2,
                "potential_hop_savings": 1,
                "risk": "HIGH",
                "recommendation": (
                    f"CIRCULAR DEPENDENCY — Action required: "
                    f"(1) Identify which columns in '{tbl_up}' are derived from {downstream_list} "
                    f"(typically indicator/decision columns like V_HAS_LTD_IND_R, decision dates, "
                    f"turnaround buckets). "
                    f"(2) Fold that aggregation logic DIRECTLY into the INSERT SELECT of '{prod_job}' "
                    f"by joining to the underlying DIM/FCT tables that feed {downstream_list[0] if downstream_list else '?'}, "
                    f"rather than reading the pre-aggregated summary. "
                    f"(3) This eliminates the cross-cycle lag — the indicator columns will reflect "
                    f"the CURRENT cycle's source data instead of the previous cycle's summary. "
                    f"(4) Validate: run 3 consecutive daily cycles comparing row counts and indicator "
                    f"column distributions before/after the change. "
                    f"Risk: HIGH — the folded logic must produce equivalent results to avoid silent "
                    f"data drift in downstream reports."
                ),
                "_dedup_key": ("N", tbl_up, prod_job, "CL"),
            })

        # N-PL: separate post-load updater that reads downstream tables
        for upd_job in all_updaters.get(tbl_up, set()):
            circular_reads = (
                {t for t in _job_reads_augmented(upd_job) if not _is_log_or_debug_table(t)}
                & downstream
            ) - {tbl_up}
            if not circular_reads:
                continue
            key = ("N-PL", tbl_up, upd_job)
            if key in seen_cl:
                continue
            seen_cl.add(key)

            # Only flag if an upstream full-load producer also exists for tbl_up
            producers_of_tbl = sorted(all_producers.get(tbl_up, set()))[:3]
            downstream_list = sorted(circular_reads)[:4]
            recommendations.append({
                "id": f"N-{len(recommendations)+1}",
                "category": "N. Post-Load Updater — Downstream Circular Read",
                "target_table": tbl_up,
                "description": (
                    f"Job '{upd_job}' updates '{tbl_up}' (reads and writes same table) AFTER "
                    f"the full-load job(s) {producers_of_tbl}, AND it reads {downstream_list} "
                    f"which are downstream of '{tbl_up}'. "
                    f"This is a two-job split of the circular update pattern: "
                    f"the post-load update step uses data derived from '{tbl_up}' itself, "
                    f"creating a lagged-cycle dependency across separate Tidal jobs."
                ),
                "affected_jobs": [upd_job] + producers_of_tbl,
                "circular_read_tables": sorted(circular_reads)[:6],
                "current_hops": 2,
                "potential_hop_savings": 1,
                "risk": "HIGH",
                "recommendation": (
                    f"POST-LOAD UPDATER ANTI-PATTERN — Action required: "
                    f"(1) Merge the update logic from '{upd_job}' into the INSERT SELECT of the "
                    f"full-load job {producers_of_tbl[0] if producers_of_tbl else '?'} if the "
                    f"source data ({downstream_list}) can be sourced directly from DIM/FCT tables "
                    f"at the time of the initial load. "
                    f"(2) If merging is not feasible, restructure the pipeline so that '{upd_job}' "
                    f"reads from DIM/FCT tables directly instead of the pre-aggregated "
                    f"{downstream_list[0] if downstream_list else '?'}, breaking the circular chain. "
                    f"(3) Consider whether the 'update after load' pattern could be replaced by a "
                    f"MERGE statement within the main load job to eliminate the separate update step. "
                    f"(4) Validate: row counts and updated-column distributions should match across "
                    f"3+ consecutive cycles before and after the change."
                ),
                "_dedup_key": ("N", tbl_up, upd_job, "PL"),
            })

    return recommendations


def standardise_recommendation_metrics(recommendations):
    """
    Split metrics into verified-graph-evidence vs LLM-signal columns so that
    F (LLM-only) never inflates hop-savings totals.
    """
    for rec in recommendations:
        category = rec.get("category", "")
        rec["verified_hop_savings"] = 0
        rec["llm_signal_count"] = 0
        rec["llm_redundancy_signals"] = 0
        rec["llm_performance_signals"] = 0
        rec["llm_logic_signals"] = 0
        rec["recommendation_type"] = ""
        rec["evidence_level"] = ""

        if category.startswith("F."):
            rec["recommendation_type"] = "PLSQL_OPTIMISATION_REVIEW"
            rec["evidence_level"] = "LLM_ONLY"
            rec["verified_hop_savings"] = 0
            rec["llm_signal_count"] = rec.get("total_findings", 0)
            rec["llm_redundancy_signals"] = rec.get("redundancy_findings", 0)
            rec["llm_performance_signals"] = rec.get("perf_findings", 0)
            rec["llm_logic_signals"] = rec.get("logic_findings", 0)
        else:
            rec["recommendation_type"] = "HOP_REDUCTION"
            rec["evidence_level"] = "GRAPH_OR_LINEAGE"
            rec["verified_hop_savings"] = rec.get("potential_hop_savings", 0)
    return recommendations


# ── Cross-detector deduplication ────────────────────────────────────────────


def deduplicate_recommendations(recommendations):
    """
    Deduplicate within the SAME category only.

    Previously this merged any recs sharing (affected_jobs, target_table) regardless
    of category, producing combined rows like "N. Circular Update + P. Cursor UPDATE
    + Q. Hardcoded Values" that are hard to read.

    New behaviour:
    - Include the category prefix (letter(s) before the first '.') in the composite
      key, so recs from DIFFERENT detector families are NEVER merged.
    - Within the same category, recs with identical (category, sorted_jobs, target)
      ARE still merged (e.g. two C. Near-Duplicate findings for the same job pair).
    """
    def _cat_prefix(cat: str) -> str:
        """Return the letter prefix before the first '.' e.g. 'C' from 'C. Near-Dup'."""
        return cat.split(".")[0].strip() if cat else ""

    by_key = defaultdict(list)
    for r in recommendations:
        cat_pfx = _cat_prefix(r.get("category", ""))
        composite = (
            cat_pfx,
            tuple(sorted(r.get("affected_jobs", []))),
            r.get("target_table", ""),
        )
        by_key[composite].append(r)

    merged = []
    for _composite, group in by_key.items():
        if len(group) == 1:
            merged.append(group[0])
            continue
        primary = dict(group[0])
        # Within-category merges: keep same category label, just note duplicates
        primary["potential_hop_savings"] = max(
            g.get("potential_hop_savings", 0) for g in group
        )
        primary["verified_hop_savings"] = max(
            g.get("verified_hop_savings", 0) for g in group
        )
        merged_sql_objs = set()
        for g in group:
            merged_sql_objs |= set(g.get("sql_objects_called", []) or [])
            merged_sql_objs |= set(g.get("sql_objects", []) or [])
        if merged_sql_objs:
            primary["sql_objects_called"] = sorted(merged_sql_objs)
        risk_rank = {"LOW": 0, "MEDIUM": 1, "HIGH": 2}
        primary["risk"] = max(
            (g.get("risk", "MEDIUM") for g in group),
            key=lambda r: risk_rank.get(r, 1),
        )
        primary["_merged_count"] = len(group)
        merged.append(primary)
    return merged


# ── Attach LLM findings as supporting evidence (non-destructive) ────────────


def attach_supporting_evidence(recommendations):
    """
    For each graph/lineage rec (A-E, G-J), attach overlapping LLM findings (F)
    as supporting evidence. F recs are kept intact in the list.
    """
    for rec in recommendations:
        rec.setdefault("supporting_recommendations", [])
        rec.setdefault("supporting_llm_findings", [])

    f_recs = [r for r in recommendations if r.get("category", "").startswith("F.")]

    for rec in recommendations:
        if rec.get("category", "").startswith("F."):
            continue
        rec_jobs = {j for j in rec.get("affected_jobs", []) if j != "UNMAPPED"}
        if not rec_jobs:
            continue
        for f in f_recs:
            f_jobs = {j for j in f.get("affected_jobs", []) if j != "UNMAPPED"}
            if rec_jobs & f_jobs:
                rec["supporting_llm_findings"].append(
                    {
                        "id": f.get("id"),
                        "script": f.get("target_table"),
                        "total_findings": f.get("total_findings", 0),
                        "redundancy_findings": f.get("redundancy_findings", 0),
                        "perf_findings": f.get("perf_findings", 0),
                        "logic_findings": f.get("logic_findings", 0),
                    }
                )
    return recommendations


# ── Hybrid scoring (category weights + critical-path + LLM cap) ─────────────


def score_recommendations(recommendations, graph):
    """
    Hybrid scoring:
      - Category weight (SP)
      - Verified hop savings only (SP)
      - Risk weighting (both)
      - Critical-path proximity (NEW)
      - Breadth of impact (both)
      - Small capped LLM signal bonus for F (SP)
      - UNMAPPED penalty for F (SP)
      - Merged-evidence bonus across detectors (NEW)
            - Runtime impact bonus from avg_runtime-based estimate (NEW)
    """
    node_map = {n["id"]: n for n in graph["nodes"]}
    max_depth = max((n.get("depth", 0) for n in graph["nodes"]), default=1)

    category_weight_lookup = [
        ("D. Serial Orchestration Chain", 25),
        ("C. OFFSET", 22),
        ("C. Near-Duplicate", 22),
        ("A. Redundant Multi-Source", 18),
        ("B. Pass-Through", 16),
        ("G. Duplicate Column-Lineage", 14),
        ("H. Stale Intermediate", 12),
        ("I. Cyclic Refresh", 12),
        ("E. MV Elimination", 10),
        ("J. Over-Fanout MV", 10),
        ("K. Repeated Transformations", 16),
        ("L. Overlapping Preprocessing", 15),
        ("M. Orchestration Edge Without Data Handoff", 14),
        ("N. Post-Consumption Update Pattern", 14),
        ("F. LLM-Identified Issues", 8),
    ]
    risk_weight = {"LOW": 25, "MEDIUM": 15, "HIGH": 5}

    def _category_weight(category):
        for prefix, w in category_weight_lookup:
            if prefix in category:
                return w
        return 10

    for rec in recommendations:
        score = 0
        category = rec.get("category", "")
        risk = rec.get("risk", "MEDIUM")
        verified = rec.get("verified_hop_savings", 0)
        llm_count = rec.get("llm_signal_count", 0)

        score += _category_weight(category)
        score += min(verified * 12, 40)
        score += risk_weight.get(risk, 15)

        affected = rec.get("affected_jobs", []) or []
        depths = [
            node_map[j].get("depth", 0)
            for j in affected
            if j in node_map and node_map[j].get("depth", -1) >= 0
        ]
        if depths:
            critical_weight = max(depths) / max(max_depth, 1)
            score += int(critical_weight * 15)

        score += min(len([j for j in affected if j != "UNMAPPED"]), 8)

        if category.startswith("F."):
            score += min(llm_count // 5, 8)
            if "UNMAPPED" in affected:
                score -= 5

        if rec.get("_merged_count", 1) > 1:
            score += 10

        # Reward practical impact: larger estimated runtime savings -> higher priority.
        est_min = rec.get("est_runtime_saved_minutes", 0.0) or 0.0
        score += min(int(est_min / 5), 20)

        # Global impact bonus: only the *extra* cross-RPT savings beyond local savings
        # count here. For EXECUTE_LOCALLY categories global_est == local est, so
        # extra_global_min == 0 and no double-counting occurs.
        global_est_min = rec.get("global_est_runtime_saved_minutes", 0.0) or 0.0
        extra_global_min = max(global_est_min - est_min, 0.0)
        score += min(int(extra_global_min / 5), 15)
        score += min(max(rec.get("global_rpt_count_impacted", 1) - 1, 0) * 2, 10)
        if rec.get("implementation_decision") == "EXECUTE_GLOBAL_PROGRAM":
            score += 6
        if rec.get("implementation_decision") == "DEFER_OR_KEEP":
            score -= 4

        rec["priority_score"] = max(score, 0)

    recommendations.sort(key=lambda r: r["priority_score"], reverse=True)
    return recommendations


# ── What-if simulation ──────────────────────────────────────────────────────


def simulate_after_top_n(recommendations, graph, top_n=10):
    """
    Apply top-N recs conceptually and recompute the new critical-path depth.

    Removal strategy
    ----------------
    - D. Serial Orchestration Chain: jobs are ordered deepest→shallowest.
      We keep jobs[-1] (shallowest) which still connects to the rest of the
      pipeline, and remove jobs[:-1] (the redundant upstream chain members).
      This mirrors "parallelize the chain — only one slot on the critical path".
    - All other categories: remove jobs[1:] (keep the canonical / replacement
      job and discard the eliminated hops).

    Depth recomputation
    -------------------
    Graph edges go from upstream (high depth) to downstream (low depth / root).
    "Depth" = longest path from a node to the root (critical path, not shortest).
    Algorithm: root nodes start at 0; propagate upward using
      depth[node] = max(depth of downstream neighbours) + 1
    Only the connected subgraph (reachable back to root) contributes to max depth.
    """
    node_map = {n["id"]: n for n in graph["nodes"]}
    # Root nodes (depth=0) must never be removed — they are the anchor for depth
    # recomputation. Removing them would disconnect the entire remaining graph.
    root_set = {nid for nid, n in node_map.items() if n.get("depth", -1) == 0}

    removed_jobs = set()
    for rec in recommendations[:top_n]:
        if rec.get("verified_hop_savings", 0) > 0:
            jobs = rec.get("affected_jobs", [])
            if len(jobs) > 1:
                if rec.get("category", "").startswith("D."):
                    # Serial chain: keep shallowest job (jobs[-1]) to preserve
                    # pipeline connectivity; remove the upstream chain members.
                    candidates = jobs[:-1]
                else:
                    candidates = jobs[1:]
                removed_jobs.update(j for j in candidates if j not in root_set)

    remaining_nodes = {
        n["id"]: n for n in graph["nodes"] if n["id"] not in removed_jobs
    }
    remaining_links = [
        lnk
        for lnk in graph["links"]
        if lnk["source"] in remaining_nodes and lnk["target"] in remaining_nodes
    ]

    # Reachability: which remaining nodes can still reach a root node?
    # Walk edges BACKWARD from root (reverse the edge direction: target → source).
    # A node is "connected" if it lies on some path leading to the root.
    # Use original depth values — avoids depth-inflation from any cross/reverse edges.
    from collections import deque as _deque
    reverse_map = defaultdict(set)
    for lnk in remaining_links:
        reverse_map[lnk["target"]].add(lnk["source"])

    reachable = set(root_set & remaining_nodes.keys())
    queue = _deque(reachable)
    while queue:
        nid = queue.popleft()
        for parent in reverse_map.get(nid, set()):
            if parent not in reachable and parent in remaining_nodes:
                reachable.add(parent)
                queue.append(parent)

    connected_depths = [
        remaining_nodes[nid].get("depth", 0)
        for nid in reachable
        if remaining_nodes[nid].get("depth", -1) >= 0
    ]

    return {
        "original_node_count": len(graph["nodes"]),
        "original_max_depth": max((n["depth"] for n in graph["nodes"]), default=0),
        "removed_job_count": len(removed_jobs),
        "new_node_count": len(remaining_nodes),
        "new_max_depth": max(connected_depths, default=0),
        "removed_jobs_sample": sorted(removed_jobs)[:15],
    }


# ── Global Decisions enrichment ───────────────────────────────────────────


def enrich_global_decisions_detail(recommendations, sql_index):
    """
    Pre-compute two display fields for every EXECUTE_GLOBAL_PROGRAM recommendation
    so the Global Decisions Excel tab can render them without needing sql_index at
    write time.

    Uses consumers_by_rpt (stored by detect_mv_elimination) to show the TRUE global
    scope: which job in which RPT needs to change, and how many RPT pipelines each
    SQL procedure change actually fixes.

      impacted_jobs_detail  — MV refresh job + all consumer jobs grouped by RPT
      implementation_steps  — per-job migration plan with explicit RPT context per job
    """
    for rec in recommendations:
        if rec.get("implementation_decision") != "EXECUTE_GLOBAL_PROGRAM":
            continue

        affected_jobs = rec.get("affected_jobs", []) or []
        tgt = rec.get("target_table", "")
        impacted_rpts = sorted(rec.get("impacted_rpts", []) or [])
        # consumers_by_rpt: RPT -> [sorted list of consumer jobs in that RPT]
        consumers_by_rpt = rec.get("consumers_by_rpt") or {}

        if not affected_jobs:
            continue

        mv_job = affected_jobs[0]

        # All unique consumer jobs globally (union across all RPTs)
        all_consumer_jobs = sorted({j for jobs in consumers_by_rpt.values() for j in jobs})
        if not all_consumer_jobs:
            all_consumer_jobs = affected_jobs[1:]   # fallback

        # Build reverse map: job -> which RPTs it appears in
        job_to_rpts: dict = defaultdict(list)
        for rpt, jobs in consumers_by_rpt.items():
            for j in jobs:
                job_to_rpts[j].append(rpt)
        job_to_rpts = {j: sorted(rpts) for j, rpts in job_to_rpts.items()}

        total_instances = sum(len(jobs) for jobs in consumers_by_rpt.values())

        # ── Impacted Jobs Detail column (grouped by RPT) ───────────────────
        mv_procs = sorted((sql_index or {}).get(mv_job, {}).get("full_objects", set()))
        mv_proc_str = ", ".join(mv_procs) if mv_procs else "(see Tidal job definition)"
        lines = [
            f"MV Refresh Job (writes '{tgt}'):",
            f"  {mv_job}",
            f"  SQL: {mv_proc_str}",
        ]
        if consumers_by_rpt:
            lines.append(
                f"\nConsumer Jobs \u2014 {len(all_consumer_jobs)} unique job(s), "
                f"{total_instances} total job-instances across {len(impacted_rpts)} RPT(s)."
            )
            lines.append(
                "(Updating the SQL procedures below removes the MV dependency across all listed RPTs)"
            )
            for rpt in impacted_rpts:
                rpt_jobs = sorted(consumers_by_rpt.get(rpt, []))
                if not rpt_jobs:
                    continue
                lines.append(f"\n  {rpt} ({len(rpt_jobs)} job(s)):")
                for j in rpt_jobs:
                    procs = sorted((sql_index or {}).get(j, {}).get("full_objects", set()))
                    proc_str = ", ".join(procs) if procs else "(see Tidal job definition)"
                    lines.append(f"    \u2022 {j}")
                    lines.append(f"       SQL: {proc_str}")
        else:
            lines.append(f"\nConsumer Jobs \u2014 {len(all_consumer_jobs)} job(s):")
            for i, j in enumerate(all_consumer_jobs, 1):
                procs = sorted((sql_index or {}).get(j, {}).get("full_objects", set()))
                proc_str = ", ".join(procs) if procs else "(see Tidal job definition)"
                lines.append(f"  {i}. {j}")
                lines.append(f"       SQL: {proc_str}")
        rec["impacted_jobs_detail"] = "\n".join(lines)

        # ── Implementation Steps column ────────────────────────────────────
        # Wave split by runtime: same proportioning as the recommendation text
        wave1_n = max(1, (len(all_consumer_jobs) + 2) // 3) if all_consumer_jobs else 0
        wave1 = all_consumer_jobs[:wave1_n]
        wave2 = all_consumer_jobs[wave1_n:]

        steps = []
        step_n = 1

        # Step 1 — Baseline (across all RPTs)
        steps.append(
            f"STEP {step_n} \u2014 ESTABLISH BASELINE (all {len(impacted_rpts)} RPTs)"
            f"\n  Before any code change, capture these metrics for EVERY RPT pipeline listed:"
            "\n  \u2022 Row counts in each consumer job's target table"
            "\n  \u2022 End-to-end cycle-time per RPT (wall clock)"
            "\n  \u2022 SLA breach count over the last 5 runs per RPT"
            f"\n  RPTs to baseline: {', '.join(impacted_rpts)}"
            "\n  Store results in a shared comparison spreadsheet for parity sign-off."
        )
        step_n += 1

        # Step 2 — Extract MV definition
        steps.append(
            f"STEP {step_n} \u2014 EXTRACT MV SOURCE DEFINITION (one-time, reused by all consumers)"
            f"\n  Open SQL procedure: {mv_proc_str}"
            f"\n  Find the INSERT \u2026 SELECT (or MERGE) that populates '{tgt}'."
            "\n  Copy the core SELECT query and save it as a shared CTE template:"
            f"\n    WITH {tgt} AS ("
            "\n        <paste extracted SELECT here>"
            "\n    )"
            "\n  This single CTE definition will be added to EVERY consumer job below."
        )
        step_n += 1

        # Wave steps (per unique consumer job, with RPT context)
        for wave_idx, wave_jobs in enumerate(
            [w for w in [wave1, wave2] if w], start=1
        ):
            job_blocks = []
            for j in wave_jobs:
                procs = sorted((sql_index or {}).get(j, {}).get("full_objects", set()))
                proc_str = ", ".join(procs) if procs else "(see Tidal job definition)"
                j_rpts = job_to_rpts.get(j, [])
                rpt_note = (
                    f"{len(j_rpts)} RPT(s): {', '.join(j_rpts)}"
                    if j_rpts else "(RPT context unavailable)"
                )
                val_step = step_n + 1
                job_blocks.append(
                    f"  Job : {j}\n"
                    f"  SQL : {proc_str}\n"
                    f"  Impacts: {rpt_note}\n"
                    "  Changes required:\n"
                    f"    a) Open {proc_str} and add the CTE from Step 2 at the top.\n"
                    f"    b) Replace every FROM / JOIN reference to '{tgt}' with the CTE name.\n"
                    f"    c) This ONE SQL change resolves all {len(j_rpts)} RPT pipeline(s) above.\n"
                    f"    d) After parity (Step {val_step}), remove the Tidal dependency edge on\n"
                    f"       '{mv_job}' in each of the {len(j_rpts)} affected RPT Tidal schedule(s)."
                )
            steps.append(
                f"STEP {step_n} \u2014 WAVE {wave_idx} CODE CHANGES "
                f"({len(wave_jobs)} unique SQL procedure(s) to change)\n"
                + "\n\n".join(job_blocks)
            )
            step_n += 1

            steps.append(
                f"STEP {step_n} \u2014 WAVE {wave_idx} VALIDATION (all {len(impacted_rpts)} RPTs)\n"
                "  Deploy Wave changes and run 3+ full pipeline cycles for ALL affected RPTs.\n"
                "  Confirm across every RPT:\n"
                "    \u2022 Row counts match the Step 1 baseline (\u00b10).\n"
                "    \u2022 Key metrics (claim totals, payment amounts) are identical.\n"
                "  Do NOT proceed to the next wave until ALL RPTs have signed off on parity."
            )
            step_n += 1

        # Tidal scheduler cleanup (one per RPT)
        steps.append(
            f"STEP {step_n} \u2014 TIDAL SCHEDULER CLEANUP (repeat for each of the {len(impacted_rpts)} RPT schedules)\n"
            "  Once ALL consumer jobs have passed parity in PRODUCTION:\n"
            f"  For EACH RPT in [{', '.join(impacted_rpts)}]:\n"
            f"    \u2022 Remove Tidal job '{mv_job}' from that RPT's Tidal schedule.\n"
            f"    \u2022 Delete ALL inbound Tidal dependency edges pointing TO '{mv_job}'.\n"
            f"    \u2022 Delete ALL outbound Tidal dependency edges pointing FROM '{mv_job}'."
        )
        step_n += 1

        steps.append(
            f"STEP {step_n} \u2014 DATABASE CLEANUP (global, one-time)\n"
            f"  \u2022 EXEC sp_rename '{tgt}', 'DEPRECATED_{tgt}'  \u2014 30-day safety hold.\n"
            f"  \u2022 After 30 days with no incidents across ALL {len(impacted_rpts)} RPTs: DROP TABLE DEPRECATED_{tgt}\n"
            "  \u2022 Remove any SQL VIEWs, synonyms, or grants referencing this table."
        )

        rec["implementation_steps"] = "\n\n".join(steps)

    return recommendations


# ── Output ──────────────────────────────────────────────────────────────────


def write_xlsx(recommendations, graph, simulation):
    wb = openpyxl.Workbook()

    # Split for stakeholder readability: Hop Reduction vs PLSQL optimisation.
    hop_recommendations = [
        r for r in recommendations if r.get("recommendation_type") == "HOP_REDUCTION"
    ]
    plsql_recommendations = [
        r
        for r in recommendations
        if r.get("recommendation_type") == "PLSQL_OPTIMISATION_REVIEW"
    ]

    headers = [
        "ID",  # 1
        "Recommendation Type",  # 2
        "Evidence Level",  # 3
        "Confidence Scope",  # 4  NEW
        "Category",  # 5
        "Target Table",  # 6
        "SQL Objects Called",  # 7  NEW
        "Description",  # 8
        "Affected Jobs",  # 9
        "Current Hops",  # 10
        "Verified Hop Savings",  # 11
        "Runtime Affected (sec)",  # 12 NEW
        "Estimated Runtime Saved (sec)",  # 13 NEW
        "Estimated Runtime Saved (min)",  # 14 NEW
        "LLM Signals",  # 15
        "LLM Redundancy Signals",  # 16
        "LLM Performance Signals",  # 17
        "LLM Logic Signals",  # 18
        "Risk",  # 19
        "Priority Score",  # 20
        "Recommendation",  # 21
        "Supporting Evidence",  # 22
        "Edge Source Writes",  # 23 NEW
        "Edge Target Reads",  # 24 NEW
        "Implementation Decision",  # 25 NEW
        "Execution Wave",  # 26 NEW
        "Global RPT Count Impacted",  # 27 NEW
        "Global Consumer Count",  # 28 NEW
        "Global Runtime Affected (min)",  # 29 NEW
        "Global Est Runtime Saved (min)",  # 30 NEW
        "Blast Radius Score",  # 31 NEW
        "Prerequisites",  # 32 NEW
    ]

    def write_recommendation_sheet(ws, recs):
        hdr_fill = PatternFill(start_color="1F4E79", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center", wrap_text=True)

        risk_fills = {
            "LOW": PatternFill(start_color="C6EFCE", fill_type="solid"),
            "MEDIUM": PatternFill(start_color="FFEB9C", fill_type="solid"),
            "HIGH": PatternFill(start_color="FFC7CE", fill_type="solid"),
        }

        for row_idx, rec in enumerate(recs, 2):
            supporting = []
            for s in rec.get("supporting_recommendations", []):
                supporting.append(
                    f"{s.get('id')} | {s.get('category')} | {s.get('description')}"
                )
            for s in rec.get("supporting_llm_findings", []):
                supporting.append(
                    f"{s.get('id')} | {s.get('script')} | "
                    f"findings={s.get('total_findings', 0)}, "
                    f"redundancy={s.get('redundancy_findings', 0)}, "
                    f"performance={s.get('perf_findings', 0)}, "
                    f"logic={s.get('logic_findings', 0)}"
                )

            scope_fills = {
                "ALREADY_PARAMETERIZED": PatternFill(start_color="D9EAD3", fill_type="solid"),
                "DIFF_SQL_OBJECTS": PatternFill(start_color="FCE5CD", fill_type="solid"),
                "SAME_PKG_VARIANT": PatternFill(start_color="FFF2CC", fill_type="solid"),
                "CROSS_RPT_SHARED": PatternFill(start_color="CFE2F3", fill_type="solid"),
                "LOCAL_ONLY": PatternFill(start_color="F3F3F3", fill_type="solid"),
            }
            sql_objs_str = "\n".join(rec.get("sql_objects_called", []))
            scope = rec.get("confidence_scope", "")

            ws.cell(row=row_idx, column=1, value=rec.get("id", ""))
            ws.cell(row=row_idx, column=2, value=rec.get("recommendation_type", ""))
            ws.cell(row=row_idx, column=3, value=rec.get("evidence_level", ""))
            sc = ws.cell(row=row_idx, column=4, value=scope)
            sc.fill = scope_fills.get(scope, PatternFill())
            ws.cell(row=row_idx, column=5, value=rec.get("category", ""))
            ws.cell(row=row_idx, column=6, value=rec.get("target_table", ""))
            ws.cell(row=row_idx, column=7, value=sql_objs_str)
            ws.cell(row=row_idx, column=8, value=rec.get("description", ""))
            ws.cell(row=row_idx, column=9, value="\n".join(rec.get("affected_jobs", [])))
            ws.cell(row=row_idx, column=10, value=rec.get("current_hops", 0))
            ws.cell(row=row_idx, column=11, value=rec.get("verified_hop_savings", 0))
            ws.cell(row=row_idx, column=12, value=rec.get("runtime_affected_seconds", 0.0))
            ws.cell(row=row_idx, column=13, value=rec.get("est_runtime_saved_seconds", 0.0))
            ws.cell(row=row_idx, column=14, value=rec.get("est_runtime_saved_minutes", 0.0))
            ws.cell(row=row_idx, column=15, value=rec.get("llm_signal_count", 0))
            ws.cell(row=row_idx, column=16, value=rec.get("llm_redundancy_signals", 0))
            ws.cell(row=row_idx, column=17, value=rec.get("llm_performance_signals", 0))
            ws.cell(row=row_idx, column=18, value=rec.get("llm_logic_signals", 0))
            rc = ws.cell(row=row_idx, column=19, value=rec.get("risk", ""))
            rc.fill = risk_fills.get(rec.get("risk", ""), PatternFill())
            ws.cell(row=row_idx, column=20, value=rec.get("priority_score", 0))
            ws.cell(row=row_idx, column=21, value=rec.get("recommendation", ""))
            ws.cell(row=row_idx, column=22, value="\n".join(supporting))
            ws.cell(
                row=row_idx,
                column=23,
                value="\n".join(rec.get("src_written_tables", []) or []),
            )
            ws.cell(
                row=row_idx,
                column=24,
                value="\n".join(rec.get("tgt_read_tables", []) or []),
            )
            ws.cell(row=row_idx, column=25, value=rec.get("implementation_decision", ""))
            ws.cell(row=row_idx, column=26, value=rec.get("execution_wave", ""))
            ws.cell(row=row_idx, column=27, value=rec.get("global_rpt_count_impacted", 1))
            ws.cell(row=row_idx, column=28, value=rec.get("global_consumer_count", 0))
            ws.cell(row=row_idx, column=29, value=rec.get("global_runtime_affected_minutes", 0.0))
            ws.cell(row=row_idx, column=30, value=rec.get("global_est_runtime_saved_minutes", 0.0))
            ws.cell(row=row_idx, column=31, value=rec.get("blast_radius_score", 0))
            ws.cell(row=row_idx, column=32, value=rec.get("prerequisites", ""))
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

        widths = [
            8, 24, 18, 22, 32, 35, 40, 60, 50, 12, 18,
            18, 22, 22, 12, 18, 18, 18, 10, 14, 100, 70, 38, 38,
            24, 18, 14, 14, 20, 22, 14, 42,
        ]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(recs) + 1}"

    # ── Sheet 1: Hop Recommendations ──
    ws = wb.active
    ws.title = "Hop Recommendations"
    write_recommendation_sheet(ws, hop_recommendations)

    # ── Sheet 2: PLSQL Optimisation ──
    ws_plsql = wb.create_sheet("PLSQL Optimization")
    write_recommendation_sheet(ws_plsql, plsql_recommendations)

    # ── Sheet 3: Summary ─────────────────────────────────────────────────────
    # Separated into two sections: Non-F (structural, verified) and F (LLM findings).
    # Each section has its own subtotal; a grand total appears at the bottom.
    ws2 = wb.create_sheet("Summary")

    # Colour palette
    _HDR_FILL   = PatternFill(start_color="1F4E79", fill_type="solid")
    _HDR_FONT   = Font(color="FFFFFF", bold=True, size=11)
    _SEC_A_FILL = PatternFill(start_color="D6E4F0", fill_type="solid")  # light blue – Non-F
    _SEC_B_FILL = PatternFill(start_color="FFF2CC", fill_type="solid")  # light yellow – F
    _SUB_FILL   = PatternFill(start_color="BDD7EE", fill_type="solid")  # subtotal – Non-F
    _SUB_F_FILL = PatternFill(start_color="FFE699", fill_type="solid")  # subtotal – F
    _TOT_FILL   = PatternFill(start_color="1F4E79", fill_type="solid")  # grand total
    _DEFER_FILL = PatternFill(start_color="F2F2F2", fill_type="solid")  # grey – deferred

    _RISK_FILLS = {
        "LOW":    PatternFill(start_color="C6EFCE", fill_type="solid"),
        "MEDIUM": PatternFill(start_color="FFEB9C", fill_type="solid"),
        "HIGH":   PatternFill(start_color="FFC7CE", fill_type="solid"),
    }

    sum_cols = [
        "Category",
        "# Recs",
        "Verified Hop Savings",
        "Est. Runtime Saved (min)",
        "Unique Jobs Affected",
        "Risk Level",
        "Implementation Decision",
    ]
    for ci, h in enumerate(sum_cols, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    def _sum_stats(recs_subset):
        stats = defaultdict(lambda: {"count": 0, "savings": 0, "jobs": set(),
                                     "risks": [], "decisions": set()})
        for rec in recs_subset:
            cat = rec.get("category", "Unknown")
            stats[cat]["count"] += 1
            stats[cat]["savings"] += rec.get("verified_hop_savings", 0)
            stats[cat]["risks"].append(rec.get("risk", "MEDIUM"))
            stats[cat]["decisions"].add(rec.get("implementation_decision", ""))
            for j in rec.get("affected_jobs", []):
                stats[cat]["jobs"].add(j)
        return stats

    def _dominant_risk(risks):
        """Return HIGH > MEDIUM > LOW based on majority."""
        if "HIGH" in risks:
            return "HIGH"
        if "MEDIUM" in risks:
            return "MEDIUM"
        return "LOW"

    def _write_section_header(ws, row, label, fill):
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(bold=True, size=11, color="1F4E79")
        c.fill = fill
        for ci in range(2, len(sum_cols) + 1):
            ws.cell(row=row, column=ci).fill = fill
        return row + 1

    def _write_summary_row(ws, row, cat, s, rec_list, fill=None):
        cat_rt = round(
            sum(r.get("est_runtime_saved_minutes", 0.0)
                for r in rec_list if r.get("category") == cat), 2)
        risk_label = _dominant_risk(s["risks"])
        decision_label = ", ".join(sorted(d for d in s["decisions"] if d)) or "—"
        ws.cell(row=row, column=1, value=cat).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=2, value=s["count"])
        ws.cell(row=row, column=3, value=s["savings"])
        ws.cell(row=row, column=4, value=cat_rt)
        ws.cell(row=row, column=5, value=len(s["jobs"]))
        rc = ws.cell(row=row, column=6, value=risk_label)
        rc.fill = _RISK_FILLS.get(risk_label, PatternFill())
        ws.cell(row=row, column=7, value=decision_label)
        if fill:
            for ci in range(1, len(sum_cols) + 1):
                if ci != 6:  # keep risk colour
                    ws.cell(row=row, column=ci).fill = fill
        return row + 1

    def _write_subtotal(ws, row, label, recs_subset, fill):
        total_savings = sum(r.get("verified_hop_savings", 0) for r in recs_subset)
        total_rt = round(sum(r.get("est_runtime_saved_minutes", 0.0) for r in recs_subset), 2)
        all_j = {j for r in recs_subset for j in (r.get("affected_jobs") or [])}
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=len(recs_subset)).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_savings).font = Font(bold=True)
        ws.cell(row=row, column=4, value=total_rt).font = Font(bold=True)
        ws.cell(row=row, column=5, value=len(all_j)).font = Font(bold=True)
        for ci in range(1, len(sum_cols) + 1):
            ws.cell(row=row, column=ci).fill = fill
        return row + 1

    non_f_recs = [r for r in recommendations if not r.get("category", "").startswith("F.")]
    f_recs     = [r for r in recommendations if r.get("category", "").startswith("F.")]

    non_f_stats = _sum_stats(non_f_recs)
    f_stats     = _sum_stats(f_recs)

    cur_row = 2

    # ── Non-F section ──
    cur_row = _write_section_header(
        ws2, cur_row,
        f"HOP REDUCTION RECOMMENDATIONS  (Structural / Verified — {len(non_f_recs)} total)",
        _SEC_A_FILL,
    )
    for cat in sorted(non_f_stats):
        # Grey out deferred / already-done categories
        is_deferred = any(s in cat for s in _DEFER_CATEGORY_SUBSTRINGS)
        fill = _DEFER_FILL if is_deferred else None
        cur_row = _write_summary_row(ws2, cur_row, cat, non_f_stats[cat], non_f_recs, fill=fill)
    cur_row = _write_subtotal(ws2, cur_row, "Hop Reduction Subtotal", non_f_recs, _SUB_FILL)

    cur_row += 1  # blank separator

    # ── F section ──
    cur_row = _write_section_header(
        ws2, cur_row,
        f"F. OPTIMIZATION FINDINGS  (Require SQL / Code Review — {len(f_recs)} total)",
        _SEC_B_FILL,
    )
    for cat in sorted(f_stats):
        cur_row = _write_summary_row(ws2, cur_row, cat, f_stats[cat], f_recs)
    cur_row = _write_subtotal(ws2, cur_row, "Optimization Subtotal", f_recs, _SUB_F_FILL)

    cur_row += 1  # blank separator

    # ── Grand total ──
    grand_savings = sum(r.get("verified_hop_savings", 0) for r in recommendations)
    grand_rt      = round(sum(r.get("est_runtime_saved_minutes", 0.0) for r in recommendations), 2)
    all_jobs_grand = {j for r in recommendations for j in (r.get("affected_jobs") or [])}
    for ci, val in enumerate(
        ["GRAND TOTAL", len(recommendations), grand_savings, grand_rt, len(all_jobs_grand), "", ""],
        1,
    ):
        c = ws2.cell(row=cur_row, column=ci, value=val)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = _TOT_FILL

    for col, width in enumerate([62, 10, 20, 24, 20, 14, 30], 1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    ws2.row_dimensions[1].height = 28

    # ── Sheet 4: Global Decisions (cross-RPT execution guidance) ──
    ws_global = wb.create_sheet("Global Decisions")
    global_rows = [
        r
        for r in hop_recommendations
        if r.get("implementation_decision") == "EXECUTE_GLOBAL_PROGRAM"
    ]
    global_headers = [
        "ID",                               # 1
        "Category",                         # 2
        "Target Table",                     # 3
        "Impacted RPTs",                    # 4
        "Impacted Jobs",                    # 5  NEW
        "Global Consumer Count",            # 6
        "Global Est Runtime Saved (min)",   # 7
        "Blast Radius",                     # 8
        "Execution Wave",                   # 9
        "Recommendation",                   # 10
        "Implementation Steps",             # 11  NEW
    ]
    for i, h in enumerate(global_headers, 1):
        c = ws_global.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="1F4E79", fill_type="solid")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, rec in enumerate(global_rows, 2):
        ws_global.cell(row=i, column=1, value=rec.get("id", ""))
        ws_global.cell(row=i, column=2, value=rec.get("category", ""))
        ws_global.cell(row=i, column=3, value=rec.get("target_table", ""))
        ws_global.cell(row=i, column=4, value=", ".join(rec.get("impacted_rpts", []) or []))
        ws_global.cell(row=i, column=5, value=rec.get("impacted_jobs_detail", "\n".join(rec.get("affected_jobs", []) or [])))
        ws_global.cell(row=i, column=6, value=rec.get("global_consumer_count", 0))
        ws_global.cell(row=i, column=7, value=rec.get("global_est_runtime_saved_minutes", 0.0))
        ws_global.cell(row=i, column=8, value=rec.get("blast_radius_score", 0))
        ws_global.cell(row=i, column=9, value=rec.get("execution_wave", ""))
        ws_global.cell(row=i, column=10, value=rec.get("recommendation", ""))
        ws_global.cell(row=i, column=11, value=rec.get("implementation_steps", ""))
        for col in range(1, len(global_headers) + 1):
            ws_global.cell(row=i, column=col).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    # col widths: ID, Cat, Target, RPTs, Jobs, Count, Runtime, Blast, Wave, Rec, Steps
    for col, width in enumerate([10, 38, 38, 32, 72, 18, 24, 14, 20, 90, 100], 1):
        ws_global.column_dimensions[get_column_letter(col)].width = width

    # ── Sheet 5: RPT Rollout Plan ─────────────────────────────────────────────
    # Shows ALL actionable recommendations (LOCAL + GLOBAL) grouped by execution wave.
    # DEFER_OR_KEEP items are excluded — they require no action.
    ws_rollout = wb.create_sheet("RPT Rollout Plan")

    rollout_hdr_fill = PatternFill(start_color="1F4E79", fill_type="solid")
    rollout_headers = [
        "Execution Wave",        # 1
        "ID",                    # 2
        "Category",              # 3
        "Target Table",          # 4
        "Scope",                 # 5  LOCAL / GLOBAL
        "Risk",                  # 6
        "Priority Score",        # 7
        "# Affected Jobs",       # 8
        "Affected Jobs (List)",  # 9
        "Prerequisites",         # 10
    ]
    for ci, h in enumerate(rollout_headers, 1):
        c = ws_rollout.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = rollout_hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Wave display order and labels
    _WAVE_ORDER = ["WAVE_1", "WAVE_2", "WAVE_3", "GLOBAL_WAVE_1_PLUS"]
    _WAVE_LABEL = {
        "WAVE_1":            "Wave 1 — Low Risk (Local)",
        "WAVE_2":            "Wave 2 — Medium Risk (Local)",
        "WAVE_3":            "Wave 3 — High Risk (Local)",
        "GLOBAL_WAVE_1_PLUS":"Global Wave — Cross-RPT Program",
    }
    _WAVE_FILLS = {
        "WAVE_1":            PatternFill(start_color="C6EFCE", fill_type="solid"),
        "WAVE_2":            PatternFill(start_color="FFEB9C", fill_type="solid"),
        "WAVE_3":            PatternFill(start_color="FFC7CE", fill_type="solid"),
        "GLOBAL_WAVE_1_PLUS":PatternFill(start_color="CFE2F3", fill_type="solid"),
    }
    _SEC_WAVE_FILLS = {
        "WAVE_1":            PatternFill(start_color="E2EFDA", fill_type="solid"),
        "WAVE_2":            PatternFill(start_color="FFF2CC", fill_type="solid"),
        "WAVE_3":            PatternFill(start_color="FCE4D6", fill_type="solid"),
        "GLOBAL_WAVE_1_PLUS":PatternFill(start_color="D6E4F0", fill_type="solid"),
    }

    # Exclude DEFER_OR_KEEP, ALREADY_PARAMETERIZED, and F-category (Optimization) items
    _NO_ACTION_SUBSTR = ("Already Parameterized via Tidal",)
    rollout_recs = [
        r for r in recommendations
        if r.get("implementation_decision") != "DEFER_OR_KEEP"
        and not any(s in r.get("category", "") for s in _NO_ACTION_SUBSTR)
        and not r.get("category", "").startswith("F.")
    ]

    # Group by wave
    from collections import defaultdict as _dd
    by_wave = _dd(list)
    for r in rollout_recs:
        by_wave[r.get("execution_wave", "WAVE_2")].append(r)

    rollout_row = 2
    for wave_key in _WAVE_ORDER:
        wave_recs = sorted(by_wave.get(wave_key, []),
                           key=lambda r: -r.get("priority_score", 0))
        if not wave_recs:
            continue

        # Section header row
        sec_label = f"{_WAVE_LABEL[wave_key]}  ({len(wave_recs)} items)"
        c = ws_rollout.cell(row=rollout_row, column=1, value=sec_label)
        c.font = Font(bold=True, size=11, color="1F4E79")
        sec_fill = _SEC_WAVE_FILLS[wave_key]
        for ci in range(1, len(rollout_headers) + 1):
            ws_rollout.cell(row=rollout_row, column=ci).fill = sec_fill
        rollout_row += 1

        for rec in wave_recs:
            jobs = rec.get("affected_jobs", []) or []
            scope = (
                "GLOBAL" if rec.get("implementation_decision") == "EXECUTE_GLOBAL_PROGRAM"
                else "LOCAL"
            )
            wave_fill = _WAVE_FILLS[wave_key]
            ws_rollout.cell(row=rollout_row, column=1, value=_WAVE_LABEL[wave_key])
            ws_rollout.cell(row=rollout_row, column=1).fill = wave_fill
            ws_rollout.cell(row=rollout_row, column=2, value=rec.get("id", ""))
            ws_rollout.cell(row=rollout_row, column=3, value=rec.get("category", ""))
            ws_rollout.cell(row=rollout_row, column=4, value=rec.get("target_table", ""))
            sc = ws_rollout.cell(row=rollout_row, column=5, value=scope)
            sc.fill = (PatternFill(start_color="CFE2F3", fill_type="solid") if scope == "GLOBAL"
                       else PatternFill(start_color="E2EFDA", fill_type="solid"))
            rc = ws_rollout.cell(row=rollout_row, column=6, value=rec.get("risk", ""))
            rc.fill = _RISK_FILLS.get(rec.get("risk", ""), PatternFill())
            ws_rollout.cell(row=rollout_row, column=7, value=rec.get("priority_score", 0))
            ws_rollout.cell(row=rollout_row, column=8, value=len(jobs))
            ws_rollout.cell(row=rollout_row, column=9, value="\n".join(jobs))
            ws_rollout.cell(row=rollout_row, column=10, value=rec.get("prerequisites", ""))
            for ci in range(1, len(rollout_headers) + 1):
                ws_rollout.cell(row=rollout_row, column=ci).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )
            rollout_row += 1

        rollout_row += 1  # blank row between wave groups

    ws_rollout.auto_filter.ref = f"A1:{get_column_letter(len(rollout_headers))}1"
    for ci, w in enumerate([28, 8, 42, 38, 10, 10, 12, 12, 55, 65], 1):
        ws_rollout.column_dimensions[get_column_letter(ci)].width = w
    ws_rollout.row_dimensions[1].height = 28

    # ── Sheet 6: Action Items (Quick Reference) ──────────────────────────────
    # Sorted by priority score, excludes DEFER / Already-Parameterized items.
    ws_actions = wb.create_sheet("Action Items")

    action_headers = [
        "Rank",                # 1
        "ID",                  # 2
        "Category",            # 3
        "Target Table",        # 4
        "Risk",                # 5
        "Scope",               # 6
        "Priority Score",      # 7
        "What To Do",          # 8  – concrete steps from recommendation
        "SQL Objects",         # 9
        "Est. Saved (min)",    # 10
    ]
    for ci, h in enumerate(action_headers, 1):
        c = ws_actions.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill(start_color="1F4E79", fill_type="solid")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Only actionable items — exclude DEFER and already-done
    _NO_ACTION_CATS = ("DEFER_OR_KEEP",)
    _SKIP_CAT_SUBSTR = ("Already Parameterized via Tidal",)
    action_recs = [
        r for r in sorted(recommendations, key=lambda x: -x.get("priority_score", 0))
        if r.get("implementation_decision") not in _NO_ACTION_CATS
        and not any(s in r.get("category", "") for s in _SKIP_CAT_SUBSTR)
        and not r.get("category", "").startswith("F.")
    ]

    def _extract_action_text(rec_text: str) -> str:
        """
        Extract the most useful lines from a recommendation for the action column.
        Preference order:
          1. Lines that start a numbered step ("Step N" or "N.")
          2. Lines prefixed with "[" (annotations like [SCOPE], [CONCRETE])
          3. First non-blank line as fallback
        Truncated to 5 lines max.
        """
        lines = [l.rstrip() for l in rec_text.split("\n") if l.strip()]
        # Grab lines that look like action steps
        step_lines = [
            l for l in lines
            if (l.strip()[:4].lower().startswith("step")
                or (l.strip() and l.strip()[0].isdigit() and "." in l.strip()[:4])
                or l.strip().startswith("["))
        ]
        selected = step_lines[:6] if step_lines else lines[:4]
        return "\n".join(selected) if selected else rec_text[:300]

    for rank, rec in enumerate(action_recs, 1):
        scope = (
            "GLOBAL" if rec.get("implementation_decision") == "EXECUTE_GLOBAL_PROGRAM"
            else "LOCAL"
        )
        cat_short = rec.get("category", "")
        sql_objs = "\n".join(rec.get("sql_objects_called", [])[:3])
        action_text = _extract_action_text(rec.get("recommendation", ""))

        ws_actions.cell(row=rank + 1, column=1, value=rank)
        ws_actions.cell(row=rank + 1, column=2, value=rec.get("id", ""))
        ws_actions.cell(row=rank + 1, column=3, value=cat_short)
        ws_actions.cell(row=rank + 1, column=4, value=rec.get("target_table", ""))
        rc = ws_actions.cell(row=rank + 1, column=5, value=rec.get("risk", ""))
        rc.fill = _RISK_FILLS.get(rec.get("risk", ""), PatternFill())
        sc = ws_actions.cell(row=rank + 1, column=6, value=scope)
        sc.fill = (PatternFill(start_color="CFE2F3", fill_type="solid") if scope == "GLOBAL"
                   else PatternFill(start_color="E2EFDA", fill_type="solid"))
        ws_actions.cell(row=rank + 1, column=7, value=rec.get("priority_score", 0))
        ws_actions.cell(row=rank + 1, column=8, value=action_text)
        ws_actions.cell(row=rank + 1, column=9, value=sql_objs)
        ws_actions.cell(row=rank + 1, column=10, value=rec.get("est_runtime_saved_minutes", 0.0))

        for ci in range(1, len(action_headers) + 1):
            ws_actions.cell(row=rank + 1, column=ci).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    for ci, w in enumerate([6, 8, 40, 38, 10, 10, 12, 90, 35, 14], 1):
        ws_actions.column_dimensions[get_column_letter(ci)].width = w
    ws_actions.row_dimensions[1].height = 28
    ws_actions.auto_filter.ref = f"A1:{get_column_letter(len(action_headers))}1"
    ws3 = wb.create_sheet("Graph Overview")

    def _ws3_kv(row, key, val, bold_val=False):
        ws3.cell(row=row, column=1, value=key).font = Font(bold=True)
        c = ws3.cell(row=row, column=2, value=val)
        if bold_val:
            c.font = Font(bold=True)

    _ws3_kv(1, "RPT Table", RPT_TABLE, bold_val=True)
    _ws3_kv(2, "Total Nodes", len(graph["nodes"]))
    _ws3_kv(3, "Total Edges", len(graph["links"]))
    _ws3_kv(4, "Max Depth", max((n["depth"] for n in graph["nodes"]), default=0))

    # Category breakdown
    cat_counts = defaultdict(int)
    for n in graph["nodes"]:
        cat_counts[n.get("category", "UNKNOWN")] += 1
    ws3.cell(row=6, column=1, value="Node Category Breakdown").font = Font(bold=True, size=11)
    ws3.cell(row=7, column=1, value="Category").font = Font(bold=True)
    ws3.cell(row=7, column=2, value="Count").font = Font(bold=True)
    for i, (cat, cnt) in enumerate(sorted(cat_counts.items(), key=lambda x: -x[1]), 8):
        ws3.cell(row=i, column=1, value=cat)
        ws3.cell(row=i, column=2, value=cnt)
    depth_start = 8 + len(cat_counts) + 1

    # Depth breakdown
    ws3.cell(row=depth_start, column=1, value="Depth Breakdown").font = Font(bold=True, size=11)
    ws3.cell(row=depth_start + 1, column=1, value="Depth").font = Font(bold=True)
    ws3.cell(row=depth_start + 1, column=2, value="Node Count").font = Font(bold=True)
    ws3.cell(row=depth_start + 1, column=3, value="Categories").font = Font(bold=True)
    depth_info = defaultdict(lambda: {"count": 0, "cats": defaultdict(int)})
    for n in graph["nodes"]:
        depth_info[n["depth"]]["count"] += 1
        depth_info[n["depth"]]["cats"][n["category"]] += 1
    for i, d in enumerate(sorted(depth_info.keys()), depth_start + 2):
        ws3.cell(row=i, column=1, value=d)
        ws3.cell(row=i, column=2, value=depth_info[d]["count"])
        ws3.cell(row=i, column=3,
                 value=", ".join(f"{c}:{cnt}"
                                 for c, cnt in sorted(depth_info[d]["cats"].items())))
    for col, w in enumerate([30, 18, 50], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 7: Savings Simulation ──────────────────────────────────────────
    ws4 = wb.create_sheet("Savings Simulation")

    def _ws4_kv(row, key, val, bold_val=False, fill=None, note=None):
        c1 = ws4.cell(row=row, column=1, value=key)
        c1.font = Font(bold=True)
        c2 = ws4.cell(row=row, column=2, value=val)
        if bold_val:
            c2.font = Font(bold=True)
        if fill:
            c1.fill = fill
            c2.fill = fill
        if note:
            ws4.cell(row=row, column=3, value=note).font = Font(italic=True, color="595959")

    _hl_fill = PatternFill(start_color="D6E4F0", fill_type="solid")

    ws4.cell(row=1, column=1, value="Parameter").font = Font(bold=True, size=11)
    ws4.cell(row=1, column=2, value="Value").font = Font(bold=True, size=11)
    ws4.cell(row=1, column=3, value="What this means").font = Font(bold=True, size=11)

    _ws4_kv(2, "RPT Table", RPT_TABLE,
            note="The report pipeline being analysed")
    _ws4_kv(3, "Original node count", simulation["original_node_count"],
            note="Total Tidal jobs in the dependency graph (root + all upstream dependencies)")
    _ws4_kv(4, "Original max depth", simulation["original_max_depth"],
            note="Longest chain of dependent jobs (critical path). Determines minimum pipeline latency.")
    _ws4_kv(5, "Jobs removed in simulation (top 10 recs)", simulation["removed_job_count"],
            note="Jobs conceptually eliminated based on the top 10 recommendations (parallelise chain, fold pass-through, consolidate loaders, etc.)")
    _ws4_kv(6, "New node count (after simulation)", simulation["new_node_count"],
            note="Remaining jobs if all top 10 recommendations are implemented")
    _ws4_kv(7, "New max depth (after simulation)", simulation["new_max_depth"],
            note="Estimated new critical-path length after removals. Only jobs still connected to root are counted.")
    _ws4_kv(8, "Depth reduction", simulation["original_max_depth"] - simulation["new_max_depth"],
            bold_val=True, fill=_hl_fill,
            note="How many sequential hops are removed from the critical path (pipeline latency improvement)")
    _ws4_kv(9, "Node reduction", simulation["original_node_count"] - simulation["new_node_count"],
            bold_val=True, fill=_hl_fill,
            note="How many jobs are eliminated entirely (maintenance and scheduling overhead reduction)")

    # Runtime savings from top-10 recs
    top10 = [r for r in recommendations if r.get("recommendation_type") == "HOP_REDUCTION"][:10]
    top10_rt_saved = round(sum(r.get("est_runtime_saved_minutes", 0.0) for r in top10), 2)
    top10_hop_saved = sum(r.get("verified_hop_savings", 0) for r in top10)
    _ws4_kv(10, "Est. runtime saved — top 10 recs (min)", top10_rt_saved,
            bold_val=True, fill=_hl_fill,
            note="Sum of estimated pipeline time saved based on avg_runtime data for affected jobs")
    _ws4_kv(11, "Verified hop savings — top 10 recs", top10_hop_saved,
            bold_val=True, fill=_hl_fill,
            note="Total hops with hard lineage/graph evidence that they can be removed (excludes LLM-only estimates)")

    ws4.cell(row=13, column=1, value="Jobs removed in simulation (first 15):").font = Font(bold=True)
    for i, j in enumerate(simulation["removed_jobs_sample"], 14):
        ws4.cell(row=i, column=2, value=j)

    ws4.column_dimensions["A"].width = 42
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 85

    # ── Sheet 5: Job Optimization Plan ───────────────────────────────────────
    # Groups all recommendations by primary affected job and presents them as
    # ordered steps so the reader can see the full action plan for each job in
    # one place without flipping between rows.
    ws5 = wb.create_sheet("Job_Optimization_Plan")
    ws5.title = "Job_Optimization_Plan"

    # Header
    hdr5 = ["Job", "Script / SQL Object", "Step #", "Category", "Target Table",
            "Risk", "Verified Hop Savings", "Action Summary"]
    for col, h in enumerate(hdr5, 1):
        cell = ws5.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    # Group recommendations by primary affected job
    from collections import defaultdict as _dd
    job_recs = _dd(list)
    for rec in recommendations:
        primary_job = (rec.get("affected_jobs") or ["UNKNOWN"])[0]
        job_recs[primary_job].append(rec)

    # Sort jobs alphabetically; within each job sort by priority_score desc
    row5 = 2
    alt_fill = openpyxl.styles.PatternFill("solid", fgColor="EBF3FB")
    white_fill = openpyxl.styles.PatternFill("solid", fgColor="FFFFFF")
    job_fill = openpyxl.styles.PatternFill("solid", fgColor="D9E2F3")

    for job_idx, job_name in enumerate(sorted(job_recs)):
        recs_for_job = sorted(job_recs[job_name],
                              key=lambda x: -x.get("priority_score", 0))
        if not recs_for_job:
            continue

        for step_idx, rec in enumerate(recs_for_job, 1):
            cat = rec.get("category", "")
            tgt = rec.get("target_table", "")
            risk = rec.get("risk", "")
            savings = rec.get("verified_hop_savings", 0)

            # Build concise action summary: first sentence of recommendation
            full_rec = rec.get("recommendation", "") or ""
            # Take up to first 400 chars or first 2 newlines
            lines = [l.strip() for l in full_rec.split("\n") if l.strip()]
            action_summary = " | ".join(lines[:3])[:400]

            # SQL objects for this step
            sql_objs = sorted(set(
                (rec.get("sql_objects_called") or []) +
                (rec.get("sql_objects") or [])
            ))[:3]
            sql_str = ", ".join(sql_objs) if sql_objs else rec.get("script_name", "")

            fill = job_fill if step_idx == 1 else (alt_fill if step_idx % 2 == 0 else white_fill)
            row_data = [
                job_name if step_idx == 1 else "",  # only show job name on first step
                sql_str,
                step_idx,
                cat,
                tgt,
                risk,
                savings,
                action_summary,
            ]
            for col, val in enumerate(row_data, 1):
                c = ws5.cell(row=row5, column=col, value=val)
                c.fill = fill
                c.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
                if col == 1 and step_idx == 1:
                    c.font = openpyxl.styles.Font(bold=True)
            row5 += 1

        # Blank separator between jobs
        row5 += 1

    ws5.column_dimensions["A"].width = 55   # Job
    ws5.column_dimensions["B"].width = 35   # Script
    ws5.column_dimensions["C"].width = 8    # Step
    ws5.column_dimensions["D"].width = 45   # Category
    ws5.column_dimensions["E"].width = 35   # Target Table
    ws5.column_dimensions["F"].width = 10   # Risk
    ws5.column_dimensions["G"].width = 10   # Savings
    ws5.column_dimensions["H"].width = 90   # Action Summary

    ws5.freeze_panes = "A2"

    # Use RPT-specific output filename
    HOP_REC_DIR.mkdir(parents=True, exist_ok=True)
    output_file = HOP_REC_DIR / f"hop_reduction_recommendations_{RPT_TABLE}.xlsx"
    wb.save(str(output_file))
    print(f"Output: {output_file}")


# ── Main ────────────────────────────────────────────────────────────────────


def main(rpt_table=None, preloaded_all_graphs=None, preloaded_all_lineage=None):
    global RPT_TABLE
    if rpt_table is not None:
        RPT_TABLE = rpt_table

    print("=== Hop Reduction Analyzer (Mature / Merged) ===")
    print(f"RPT Table: {RPT_TABLE}")
    print(f"Graph JSON: {GRAPH_JSON}")
    print(f"Lineage XLSX: {LINEAGE_XLSX}\n")

    # Validate that RPT_TABLE exists in graph
    try:
        print("Loading data...")
        if preloaded_all_graphs is not None:
            all_graphs = preloaded_all_graphs
        else:
            with open(GRAPH_JSON, "r", encoding="utf-8") as f:
                all_graphs = json.load(f)
        if RPT_TABLE not in all_graphs:
            print(f"\nERROR: RPT_TABLE '{RPT_TABLE}' not found in graph JSON.")
            print(f"Available tables: {sorted(all_graphs.keys())}")
            sys.exit(1)
        graph = all_graphs[RPT_TABLE]
        global_children = build_global_children(all_graphs)
        global_table_usage = build_global_table_usage(all_graphs)
        global_job_rpt_map = build_global_job_rpt_map(all_graphs)
        global_runtime_index = build_global_runtime_index(all_graphs)
        global_mv_context = build_global_mv_context(all_graphs, global_runtime_index)
    except FileNotFoundError:
        print(f"ERROR: Graph JSON not found at {GRAPH_JSON}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"ERROR: Invalid JSON in {GRAPH_JSON}: {e}")
        sys.exit(1)

    # Load lineage — read the XLSX once and filter (avoids double load)
    if preloaded_all_lineage is not None:
        all_lineage = preloaded_all_lineage
    else:
        all_lineage = load_all_lineage()
    lineage = [r for r in all_lineage if r.get("RPT_TABLE") == RPT_TABLE]
    if not lineage:
        print(f"WARNING: No lineage rows found for {RPT_TABLE}. Proceeding with detectors that only need graph.\n")

    summaries = load_summaries()
    hop_findings = load_hop_findings()

    # Build SQL object index (job -> SQL PKG/PROC/MV metadata)
    sql_index = build_job_sql_index(lineage)

    print(f"  Graph: {len(graph['nodes'])} nodes, {len(graph['links'])} edges")
    print(f"  Lineage rows (this RPT): {len(lineage)}")
    print(f"  Script summaries: {len(summaries)}")
    print(f"  Hop findings: {len(hop_findings.get('findings', []))}\n")

    all_recs = []
    detectors = [
        (
            "A. Redundant multi-source loads",
            lambda: detect_redundant_multi_source(lineage, sql_index),
        ),
        (
            "B. Pass-through (with chain detection)",
            lambda: detect_pass_through(graph, lineage),
        ),
        (
            "C. Near-duplicate procedures (Jaccard)",
            lambda: detect_offset_consolidation(graph, lineage, sql_index),
        ),
        (
            "D. Serial chains (no data dep)",
            lambda: detect_serial_orchestration_chains(graph, lineage),
        ),
        (
            "E. MV elimination (with column usage)",
            lambda: detect_mv_elimination(
                graph,
                lineage,
                global_children,
                global_mv_context,
                global_job_rpt_map,
                global_runtime_index,
                summaries,
            ),
        ),
        (
            "E. MV indicator-chain elimination candidate",
            lambda: detect_post_load_mv_indicator_chain(graph, all_graphs),
        ),
        (
            "F. LLM findings (tgt-table fallback + stats)",
            lambda: extract_llm_findings(summaries, hop_findings, graph, lineage),
        ),
        (
            "G. Duplicate column-lineage paths",
            lambda: detect_duplicate_column_paths(lineage),
        ),
        ("K. Repeated transformations", lambda: detect_repeated_transformations(lineage)),
        (
            "L. Overlapping preprocessing on same target",
            lambda: detect_preprocessing_overlap(lineage),
        ),
        (
            "M. Orchestration edge without data handoff (MV->MV)",
            lambda: detect_orchestration_edge_no_data_handoff(graph, lineage),
        ),
        (
            "N. Post-consumption update on same table",
            lambda: detect_post_consumption_updates(graph, lineage),
        ),
        (
            "E-IC. Indicator-chain MV elimination",
            lambda: detect_indicator_chain_mv_elimination(
                graph, summaries, global_job_rpt_map
            ),
        ),
        (
            "H. Stale intermediate tables",
            lambda: detect_stale_intermediates(graph, global_table_usage, summaries),
        ),
        ("I. Cyclic refresh dependencies", lambda: detect_cycles(graph)),
        ("J. Over-fanout MVs", lambda: detect_overfanout_mvs(graph)),
        (
            "O. Intermediate/working table naming & global impact",
            lambda: detect_naming_and_intermediate_tables(
                graph, lineage, global_table_usage, global_job_rpt_map, summaries
            ),
        ),
        (
            "P. Post-load cursor UPDATE — fold into INSERT",
            lambda: detect_cursor_post_load_updates(graph, sql_index, summaries),
        ),
        (
            "Q. Hardcoded value mapping — externalize to lookup",
            lambda: detect_hardcoded_value_mappings(graph, sql_index, summaries),
        ),
    ]
    for name, fn in detectors:
        print(f"Detecting {name}...")
        recs = fn()
        print(f"  Found: {len(recs)}")
        all_recs.extend(recs)

    print(f"\nTotal raw recommendations: {len(all_recs)}")

    print("Standardising metrics (verified vs LLM split)...")
    all_recs = standardise_recommendation_metrics(all_recs)

    print("Deduplicating across detectors...")
    all_recs = deduplicate_recommendations(all_recs)
    print(f"After dedup: {len(all_recs)}")

    print("Normalizing SQL object mapping for all recommendations...")
    all_recs = enrich_sql_objects_called(all_recs, graph, sql_index)

    print("Generating concrete actions for F (LLM) recommendations...")
    all_recs = generate_concrete_actions_for_f_findings(all_recs, summaries)

    print("Enriching specificity for all other recommendation types...")
    all_recs = enrich_recommendation_specificity(all_recs, graph, sql_index)

    print("Estimating runtime impact from avg_runtime...")
    runtime_index = build_runtime_index(graph)
    all_recs = attach_runtime_estimates(all_recs, runtime_index)

    print("Enriching globally-resolved execution context...")
    all_recs = enrich_global_execution_context(
        all_recs, global_job_rpt_map, global_runtime_index
    )

    print("Attaching LLM findings as supporting evidence...")
    all_recs = attach_supporting_evidence(all_recs)

    print("Scoring (hybrid: category + critical-path + LLM cap)...")
    all_recs = score_recommendations(all_recs, graph)

    print("Running what-if simulation (top 10)...")
    simulation = simulate_after_top_n(all_recs, graph, top_n=10)
    print(
        f"  Max depth: {simulation['original_max_depth']} -> "
        f"{simulation['new_max_depth']}"
    )
    print(
        f"  Nodes: {simulation['original_node_count']} -> "
        f"{simulation['new_node_count']}"
    )

    print("Enriching Global Decisions tab with per-job implementation detail...")
    all_recs = enrich_global_decisions_detail(all_recs, sql_index)

    print("Writing Excel output...")
    write_xlsx(all_recs, graph, simulation)

    # Update output filename to reflect the RPT table being analyzed
    output_file = HOP_REC_DIR / f"hop_reduction_recommendations_{RPT_TABLE}.xlsx"
    
    print(f"\n{'=' * 80}\nTOP 10 RECOMMENDATIONS (by priority score)\n{'=' * 80}")
    for i, rec in enumerate(all_recs[:10], 1):
        print(
            f"\n{i}. [{rec['id']}] Score={rec['priority_score']}  "
            f"Risk={rec['risk']}  Verified Savings="
            f"{rec.get('verified_hop_savings', 0)} hops  "
            f"Est Time Saved={rec.get('est_runtime_saved_minutes', 0.0):.2f} min"
        )
        print(
            f"   Type: {rec.get('recommendation_type', '')}  "
            f"Evidence: {rec.get('evidence_level', '')}"
        )
        print(f"   Category: {sanitize_console_text(rec['category'])}")
        desc = sanitize_console_text(rec["description"])[:130]
        print(f"   {desc}")
        rec_text = sanitize_console_text(rec["recommendation"])[:130]
        print(f"   -> {rec_text}")

    print(f"\n{'=' * 80}")
    print("Summary by category (verified hop savings only):")
    cats = defaultdict(lambda: {"n": 0, "savings": 0, "runtime_min": 0.0})
    for r in all_recs:
        cats[r["category"]]["n"] += 1
        cats[r["category"]]["savings"] += r.get("verified_hop_savings", 0)
        cats[r["category"]]["runtime_min"] += r.get("est_runtime_saved_minutes", 0.0)
    for cat in sorted(cats):
        print(
            f"  {cat:55s} {cats[cat]['n']:3d} recs, "
            f"{cats[cat]['savings']:3d} verified hop savings, "
            f"{cats[cat]['runtime_min']:8.2f} min est saved"
        )


if __name__ == "__main__":
    main()
