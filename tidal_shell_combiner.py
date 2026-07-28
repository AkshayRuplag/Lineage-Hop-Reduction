"""
TIDAL-to-Shell-Parser Combiner
================================
Joins five data sources to build a unified column-level lineage map across 10 RPT tables:

  1. DIFW_Query_Results.xlsx      — Job dependency chains per RPT table + DIFW lineage
                                        (ROOT_JOB, DEPENDENT_JOB, CMD, SRC_TABLE, TGT_TABLE, SOURCE_COL, TARGET_COL)
  2. TIDAL_15_April_updated.xlsx      — Full TIDAL dump with CMD, PARAMS per job
  3. output/shell_parsed_objects.csv  — Shell parser output (shell_script → SQL object + table_references)
  4. output/gudu_lineage_output.csv   — Gudu-parsed column lineage (priority source)
  5. output/llm_lineage_output.csv    — LLM-parsed column lineage (fallback where Gudu has no coverage)

  Supplemental: PKG_PROC_Analysis_All_Metadata.xlsx — PKG → child proc call structure for orchestrator expansion

Logic:
  - DIFW jobs (cmd = edp_grp_plsql_difw_job_srcsys*.sh): keep existing lineage from DIFW_Query_Results.xlsx
  - Non-DIFW jobs: join on CMD (shell script name) → enrich with SQL objects from shell parser
  - Parameterized jobs (MV_REFRESH etc.): resolve object name from TIDAL PARAMS column;
    for hardcoded-table scripts, surface table_references from shell parser as SHELL_TABLE_REFS
  - PKG.main orchestrator rows are expanded into individual child proc rows
  - All non-DIFW rows are enriched with SRC/TGT from Gudu (priority) then LLM (fallback)
  - Recursive BFS from each root job adds upstream jobs not present in RPT_ALL_Schema

Outputs (in SQLObjectParser/output/):
  - combined_lineage_latest.csv   — Full combined lineage (all 10 RPT tables)
  - combined_lineage_latest.xlsx  — Same in Excel with conditional formatting
  - lineage_gaps_latest.csv       — Jobs still missing full lineage
  - lineage_summary_latest.csv    — Per-RPT-table summary statistics

Usage:
    python tidal_shell_combiner.py [--rpt-schema PATH] [--tidal PATH] [--shell-parsed PATH]
                                   [--gudu-lineage PATH] [--llm-lineage PATH]
                                   [--pkg-proc-analysis PATH] [--output-dir PATH]
"""

import os
import re
import csv
import json
import shutil
import argparse
from pathlib import Path
from dataclasses import dataclass, field, asdict
from collections import defaultdict, deque

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"

# 8 target RPT tables and their root jobs
RPT_ROOT_JOBS = {
    "RPT_CLAIM_DTL_R":              "EDP_GRP_EDW_LOAD_RPT_CLAIM_DTL_R_UPD_IND_COLS",  #"EDP_GRP_EDW_LOAD_RPT_CLAIM_DTL_R-19",
    "RPT_CLAIM_NOTE_R":             "EDP_GRP_EDW_LOAD_RPT_CLAIM_NOTE_R-18",
    "RPT_CLAIM_PAYMENT_DTL_R":      "EDP_GRP_EDW_LOAD_RPT_CLAIM_PAYMENT_DTL_R-21",
    "RPT_CLAIM_PAYMENT_R":          "EDP_GRP_EDW_LOAD_RPT_CLAIM_PAYMENT_R-20",
    "RPT_CLAIM_SUM_R":              "EDP_GRP_EDW_LOAD_RPT_CLAIM_SUM_R_UPD_COLS",   #"EDP_GRP_EDW_LOAD_RPT_CLAIM_SUM_R-22",
    "RPT_CLAIM_TASK_R":             "EDP_GRP_EDW_LOAD_RPT_CLAIM_TASK_R-UW-7.1",
    "RPT_CLAIMANT_DTL_R":           "EDP_GRP_EDW_LOAD_RPT_CLAIMANT_DTL_R-23",
    "RPT_FCT_RPT_CLAIM_SUMMARY_R":  "EDP_GRP_EDW_LOAD_RPT_FCT_RPT_CLAIM_SUMMARY_R-24",
    "RPT_DICS_CLAIMS_DETAIL_OAC_R": "EDP_GRP_EDW_LOAD_RPT_DICS_CLAIMS_DETAIL_OAC_R-1",
    "RPT_FCT_RPT_CLAIM_SUMMARY_R_HIST": "EDP_GRP_EDW_LOAD_RPT_FCT_RPT_CLAIM_SUMMARY_R_HIST-28",
    "RPT_CLIENT_DTL_R":           "EDP_GRP_EDW_LOAD_RPT_CLIENT_DTL_R-17",
    "RPT_GRP_PRODUCT_R":          "EDP_GRP_EDW_LOAD_RPT_GRP_PRODUCT_R-14",
    "RPT_EMPLOYEE_R":             "EDP_GRP_EDW_LOAD_RPT_EMPLOYEE_R-13",
    "RPT_POLICY_DTL_R":           "EDP_GRP_EDW_LOAD_RPT_POLICY_DTL_R-16",

}


# Mapping from full IN_TYPE names (as written in TIDAL/shell scripts) to the
# abbreviated suffixes used in Gudu SQL object names.
# Shell scripts use descriptive names; the SQL source code (and therefore Gudu)
# uses shorter abbreviations that were baked into the procedure/branch names.
_IN_TYPE_GUDU_ALIASES: dict[str, str] = {
    'DISBURSEMENT': 'DISBURS',
}


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_rpt_schema(path: Path) -> list[dict]:
    """Load RPT_ALL_Schema.xlsx — the dependency + DIFW lineage data."""
    path = _copy_if_locked(path)
    wb = openpyxl.load_workbook(str(path), read_only=True)
    sheet_name = 'Export Worksheet' if 'Export Worksheet' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = [str(h).strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        record = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        data.append(record)
    return data


def load_tidal_deps(path: Path) -> dict:
    """Load a TIDAL dependency file (xlsx or csv).

    Supports:
    - .xlsx files (openpyxl, single snapshot — one row per JOB_NAME/DEP_JOB pair)
    - .csv files (multi-run exports — many rows per job across different run dates)

    For CSV files the data is deduplicated by (JOB_NAME, DEPENDENT_JOB) so that
    each unique dependency relationship appears exactly once, with CMD/PARAMS taken
    from the first non-null occurrence.

    Returns dict: JOB_NAME -> list[dict] (one entry per unique DEPENDENT_JOB).
    """
    suffix = path.suffix.lower()
    if suffix == '.csv':
        return _load_tidal_csv(path)
    else:
        return _load_tidal_xlsx(path)


def _load_tidal_xlsx(path: Path) -> dict:
    """Load TIDAL deps from an xlsx workbook.

    Handles two export styles:
    - Older snapshots (TidalDeps_April15.xlsx): mixed-case headers, None for nulls
    - Newer exports (Tidal Deps 20270713.xlsx): UPPER headers, 'NULL' string for nulls,
      extra FULL_PATH column
    """
    path = _copy_if_locked(path)
    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = [str(h).strip() for h in rows[0]]
    # Normalise to uppercase so lookups work regardless of export style
    headers_up = [h.upper() for h in headers]
    _NULL_STRS = {'NULL', 'NONE', 'N/A', 'NAN', ''}

    def _clean(v):
        if v is None:
            return None
        s = str(v).strip()
        return None if s.upper() in _NULL_STRS else s

    tidal: defaultdict[str, list] = defaultdict(list)
    for row in rows[1:]:
        record = {headers_up[i]: _clean(row[i] if i < len(row) else None)
                  for i in range(len(headers_up))}
        job_name = record.get('JOB_NAME', '') or ''
        if job_name:
            tidal[job_name].append(record)
    return dict(tidal)


def _load_tidal_csv(path: Path) -> dict:
    """Load TIDAL deps from a CSV multi-run export.

    CSV exports contain one row per job execution per run date, so the same
    (JOB_NAME, DEPENDENT_JOB) pair appears many times.  We deduplicate and keep
    CMD/PARAMS/JOB_ID from the first row that has a non-null CMD value.
    """
    seen: dict[tuple, dict] = {}   # (job_name_upper, dep_job_upper) -> record
    with open(path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            job_name = str(row.get('JOB_NAME', '') or '').strip()
            dep_job  = str(row.get('DEPENDENT_JOB', '') or '').strip()
            if not job_name:
                continue
            key = (job_name.upper(), dep_job.upper())
            if key not in seen:
                seen[key] = {k: (v if v not in ('', 'nan', 'NaN', 'None') else None)
                             for k, v in row.items()}
            else:
                # Upgrade CMD/PARAMS if the existing record has nulls
                existing = seen[key]
                if not existing.get('CMD') and row.get('CMD', '').strip():
                    existing['CMD'] = row['CMD'].strip()
                if not existing.get('PARAMS') and row.get('PARAMS', '').strip():
                    existing['PARAMS'] = row['PARAMS'].strip()

    tidal: defaultdict[str, list] = defaultdict(list)
    for record in seen.values():
        job_name = str(record.get('JOB_NAME', '') or '').strip()
        tidal[job_name].append(record)
    return dict(tidal)


def merge_tidal_dicts(primary: dict, supplement: dict) -> dict:
    """Merge two TIDAL dicts, preferring primary for any (JOB_NAME, DEP_JOB) pairs
    that already exist there.  Supplement entries are only added when a job or a
    specific (JOB_NAME, DEP_JOB) pair is absent from primary.

    This lets a 30-day CSV (primary, more jobs) be supplemented with an older
    single-day xlsx (supplement) to cover any jobs that were cut off by the
    30-day CSV's Excel row-limit truncation.

    Returns merged dict in the same JOB_NAME -> list[dict] format.
    """
    # Build set of existing (job_name, dep_job) pairs in primary
    existing_pairs: set[tuple[str, str]] = set()
    for job_name, entries in primary.items():
        for entry in entries:
            dj = str(entry.get('DEPENDENT_JOB', '') or '').strip().upper()
            existing_pairs.add((job_name.upper(), dj))

    merged = defaultdict(list, {k: list(v) for k, v in primary.items()})

    added_jobs, added_pairs = 0, 0
    for job_name, entries in supplement.items():
        for entry in entries:
            dj = str(entry.get('DEPENDENT_JOB', '') or '').strip().upper()
            key = (job_name.upper(), dj)
            if key not in existing_pairs:
                merged[job_name].append(entry)
                existing_pairs.add(key)
                added_pairs += 1
        if job_name not in primary:
            added_jobs += 1

    if added_pairs:
        print(f"  -> TIDAL merge: added {added_pairs} new (job, dep) pairs "
              f"({added_jobs} entirely new jobs) from supplement file")
    return dict(merged)


def load_shell_parsed(path: Path) -> dict:
    """
    Load shell_parsed_objects.csv.
    Returns dict keyed by shell script filename -> list[dict].
    """
    parsed = defaultdict(list)
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            parsed[row['shell_script']].append(row)
    return dict(parsed)


def _copy_if_locked(path: Path) -> Path:
    """If a file is locked (open in Excel), copy it and return the copy path."""
    try:
        with open(path, 'rb') as f:
            f.read(1)
        return path
    except PermissionError:
        copy_path = path.parent / (path.stem + "_copy" + path.suffix)
        shutil.copy2(str(path), str(copy_path))
        return copy_path


# ─────────────────────────────────────────────────────────────────────────────
# RECURSIVE BFS DEPTH COMPUTATION
# ─────────────────────────────────────────────────────────────────────────────

def build_recursive_depths(tidal: dict) -> dict:
    """Run BFS from each of the 8 RPT root jobs to compute depth for every
    dependent job, excluding _DL_ (DataLake) jobs.

    Returns:
        dict of (root_job, dep_job) -> {
            'depth': int,           # 0 = root, 1 = direct dependency, 2+ = recursive
            'category': str,        # inferred layer category
            'inferred_table': str,  # table name inferred from job name
        }
    """
    # Build dependency map: job -> set of upstream prerequisites
    dep_map = defaultdict(set)
    for job_name, entries in tidal.items():
        for entry in entries:
            dj = entry.get('DEPENDENT_JOB')
            if dj:
                dep_map[job_name].add(dj)

    result = {}  # (root_job, dep_job) -> info dict

    for rpt_table, root_job in RPT_ROOT_JOBS.items():
        visited = {}
        seen = set()   # tracks all visited (incl _DL_) to prevent revisits
        queue = deque([(root_job, 0)])
        while queue:
            jn, d = queue.popleft()
            if jn in seen:
                continue
            seen.add(jn)
            is_dl = '_DL_' in jn.upper()
            is_month_end = 'MONTH_END' in jn.upper()
            if not is_dl and not is_month_end:
                visited[jn] = d      # only record non-excluded jobs
            # MONTH_END nodes: exclude and STOP (don't follow children)
            # _DL_ nodes: exclude but traverse through (same depth)
            if is_month_end:
                continue
            for dep in dep_map.get(jn, []):
                if dep not in seen:
                    next_d = d if is_dl else d + 1
                    queue.append((dep, next_d))

        for jn, d in visited.items():
            inferred = _infer_table_from_job(jn)
            result[(root_job, jn)] = {
                'depth': d,
                'category': _classify_job_category(jn, inferred),
                'inferred_table': inferred or '',
            }

    return result


def _infer_table_from_job(job_name: str) -> str:
    """Infer table/object name from a TIDAL job name."""
    name = job_name.upper()
    prefixes = [
        "EDP_GRP_EDW_MV_REFRESH_", "EDP_GRP_EDW_LOAD_", "EDP_GRP_EDW_REFRESH_",
        "EDP_EDW_GRP_PACS_LOAD_", "EDP_EDW_GRP_STACS_LOAD_", "EDP_EDW_GRP_LOAD_",
        "EDP_EDW_CV_SHINKA_LOAD_", "EDP_EDW_SHINKA_CV_LOAD_",
        "EDP_EDW_SHINKA_UPD_CV_PARTYSK_", "EDP_EDW_GRP_UPD_PACS_PARTYSK_",
        "EDP_EDW_GRP_UPD_CV_PAYEENAME_",
        "EDP_EDW_GRP_PACS_LOAD_", "EDP_EDW_PACS_LOAD_",
        "EDP_EDW_GRP_REFRESH_", "EDP_EDW_PACS_GRP_LOAD_",
        "EDP_GRP_EDW_",
    ]
    for prefix in prefixes:
        if name.startswith(prefix):
            rest = name[len(prefix):]
            rest = re.sub(r'[-_](MONTH_END|UPD_IND_COLS|UPD_COLS|HIST|INCR)$', '', rest)
            rest = re.sub(r'-[\dUW.]+$', '', rest)
            return rest if rest else None
    return None


def _classify_job_category(job_name: str, table_name: str) -> str:
    """Classify a job into a data layer category."""
    upper = job_name.upper()
    if table_name:
        tn = table_name.upper()
        if tn.startswith("RPT_"): return "RPT"
        if "_MV_SSL" in tn or tn.startswith("VW_"): return "MV"
        if tn.startswith("FCT_") or tn.startswith("VUE_FCT_"): return "FCT"
        if tn.startswith("DIM_"): return "DIM"
        if tn.startswith("STG_"): return "STG"
        if tn.startswith("REF_"): return "REF"
    if "MV_REFRESH" in upper: return "MV"
    if "MONTH_END" in upper: return "MONTH_END"
    if "LOAD_STG_" in upper: return "STG"
    if "LOAD_DIM_" in upper: return "DIM"
    if "LOAD_FCT_" in upper: return "FCT"
    if "LOOKUP" in upper: return "REF"
    return "OTHER"


# ─────────────────────────────────────────────────────────────────────────────
# LLM / STM ENRICHMENT
# ─────────────────────────────────────────────────────────────────────────────

def _strip_schema(value: str) -> str:
    """Strip schema prefix from a table name (e.g., 'ATOMIC.TABLE' -> 'TABLE')."""
    if value and '.' in value:
        return value.split('.', 1)[1]
    return value


def _strip_alias(value: str) -> str:
    """Strip SQL alias prefix(es) from a column name or comma-separated list.

    Handles single values ('t374156.col' -> 'col') and multi-value cells
    ('t374156.col1, b.col2' -> 'col1, col2').
    """
    if not value:
        return value
    parts = [p.strip() for p in value.split(',')]
    cleaned = []
    for p in parts:
        if '.' in p:
            left, right = p.split('.', 1)
            # Strip if prefix looks like a table alias (short, no spaces, not a schema name)
            if len(left) <= 20 and ' ' not in left and left.upper() not in ('ATOMIC', 'DBO', 'PUBLIC'):
                cleaned.append(right)
                continue
        cleaned.append(p)
    return ', '.join(cleaned)


def load_llm_lineage(path: Path) -> dict:
    """Load LLM lineage from CSV (llm_lineage_output.csv) or XLSX (LLM_Based_Parsing_Results_All.xlsx).

    CSV format: same column layout as gudu_lineage_output.csv — produced by llm_lineage_extractor.py.
    XLSX format: reads the 'Source-to-Target Mappings' sheet (legacy baseline file).

    Returns dict keyed by UPPER(object_name) -> {
        'src_tables': sorted unique source tables,
        'tgt_tables': sorted unique target tables,
        'src_cols':   sorted unique source columns,
        'tgt_cols':   sorted unique target columns,
        'row_count':  number of column-level rows,
    }
    """
    # ── CSV branch ──────────────────────────────────────────────────────────
    if str(path).lower().endswith('.csv'):
        agg = defaultdict(lambda: {
            'src_tables': set(), 'tgt_tables': set(),
            'src_cols': set(), 'tgt_cols': set(), 'row_count': 0,
            'col_map_raw': [],   # ordered (tgt_table, tgt_col, src_table, src_col, transformation) rows
        })
        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for rec in reader:
                pkg = str(rec.get('SQL Object Name') or '').strip()
                if not pkg or pkg == 'None':
                    continue
                key = pkg.upper()
                a = agg[key]
                a['row_count'] += 1
                tgt_t = _strip_schema(str(rec.get('Target Table') or '').strip())
                tgt_c = str(rec.get('Target Column') or '').strip()
                src_t = _strip_schema(str(rec.get('Source Table') or '').strip())
                src_c = _strip_alias(str(rec.get('Source Column') or '').strip())
                transf = str(rec.get('Transformation') or '').strip()
                for _v in (t.strip() for t in tgt_t.split(',') if t.strip()):
                    if _v and _v != 'None': a['tgt_tables'].add(_v)
                for _v in (t.strip() for t in tgt_c.split(',') if t.strip()):
                    if _v and _v != 'None': a['tgt_cols'].add(_v)
                for _v in (t.strip() for t in src_t.split(',') if t.strip()):
                    if _v and _v not in ('None', 'N/A', '-'): a['src_tables'].add(_v)
                for _v in (t.strip() for t in src_c.split(',') if t.strip()):
                    if _v and _v not in ('None', 'N/A', '-'): a['src_cols'].add(_v)
                # Preserve per-column mapping for aligned output
                if tgt_c and tgt_c != 'None':
                    a['col_map_raw'].append((tgt_t, tgt_c, src_t, src_c, transf))
        result = {}
        for key, a in agg.items():
            result[key] = {
                'src_tables': ', '.join(sorted(a['src_tables'])),
                'tgt_tables': ', '.join(sorted(a['tgt_tables'])),
                'src_cols':   ', '.join(sorted(a['src_cols'])),
                'tgt_cols':   ', '.join(sorted(a['tgt_cols'])),
                'row_count':  a['row_count'],
                'col_mappings': _build_col_mappings(a['col_map_raw']),
            }
        print(f"  -> LLM CSV: {len(result)} unique SQL objects")
        return result

    # ── XLSX branch (legacy LLM_Based_Parsing_Results_All.xlsx) ───────────────────
    path = _copy_if_locked(path)
    # data_only=True ensures Excel formula cells return their cached values
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    # Prefer named sheet, fall back to active
    if 'Source-to-Target Mappings' in wb.sheetnames:
        ws = wb['Source-to-Target Mappings']
    else:
        ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = [str(h).strip() for h in rows[0]]
    # Detect the object name column (handle multiple naming conventions)
    obj_col = None
    for candidate in ('SQL Object Name', 'SQL_PACKAGE', 'Package_Name'):
        if candidate in headers:
            obj_col = candidate
            break
    if obj_col is None:
        print(f"  WARNING: No object name column found in {path.name}. Tried: SQL Object Name, SQL_PACKAGE, Package_Name")
        return {}

    agg = defaultdict(lambda: {
        'src_tables': set(), 'tgt_tables': set(),
        'src_cols': set(), 'tgt_cols': set(), 'row_count': 0,
    })
    for row in rows[1:]:
        rec = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        pkg = str(rec.get(obj_col, '') or '').strip()
        if not pkg or pkg == 'None':
            continue
        key = pkg.upper()
        a = agg[key]
        a['row_count'] += 1
        tgt_t = _strip_schema(str(rec.get('Target Table', '') or '').strip())
        tgt_c = str(rec.get('Target Column', '') or '').strip()
        src_t = _strip_schema(str(rec.get('Source Table', '') or '').strip())
        src_c = _strip_alias(str(rec.get('Source Column', '') or '').strip())
        # Some LLM cells store multiple values separated by commas — add each individually
        for _v in (t.strip() for t in tgt_t.split(',') if t.strip()):
            if _v and _v != 'None':
                a['tgt_tables'].add(_v)
        for _v in (t.strip() for t in tgt_c.split(',') if t.strip()):
            if _v and _v != 'None':
                a['tgt_cols'].add(_v)
        for _v in (t.strip() for t in src_t.split(',') if t.strip()):
            if _v and _v not in ('None', 'N/A', '-'):
                a['src_tables'].add(_v)
        for _v in (t.strip() for t in src_c.split(',') if t.strip()):
            if _v and _v not in ('None', 'N/A', '-'):
                a['src_cols'].add(_v)

    # Convert sets to sorted comma-separated strings
    result = {}
    for key, a in agg.items():
        result[key] = {
            'src_tables': ', '.join(sorted(a['src_tables'])),
            'tgt_tables': ', '.join(sorted(a['tgt_tables'])),
            'src_cols': ', '.join(sorted(a['src_cols'])),
            'tgt_cols': ', '.join(sorted(a['tgt_cols'])),
            'row_count': a['row_count'],
        }

    # Also index by SQL Parent Object Name (e.g., PKG_..._MAIN) so that
    # self-contained 'MAIN' procs can be found via dot->underscore match.
    # Without this, all MAIN procs would collapse under a single 'MAIN' key.
    if 'SQL Parent Object Name' in headers:
        parent_agg = defaultdict(lambda: {
            'src_tables': set(), 'tgt_tables': set(),
            'src_cols': set(), 'tgt_cols': set(), 'row_count': 0,
        })
        for row in rows[1:]:
            rec = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            parent = str(rec.get('SQL Parent Object Name', '') or '').strip()
            pkg = str(rec.get(obj_col, '') or '').strip().upper()
            if not parent or parent.upper() in ('NONE', '', pkg):
                continue
            pk = parent.upper()
            a = parent_agg[pk]
            a['row_count'] += 1
            tgt_t = _strip_schema(str(rec.get('Target Table', '') or '').strip())
            tgt_c = str(rec.get('Target Column', '') or '').strip()
            src_t = _strip_schema(str(rec.get('Source Table', '') or '').strip())
            src_c = _strip_alias(str(rec.get('Source Column', '') or '').strip())
            for _v in (t.strip() for t in tgt_t.split(',') if t.strip()):
                if _v and _v != 'None': a['tgt_tables'].add(_v)
            for _v in (t.strip() for t in tgt_c.split(',') if t.strip()):
                if _v and _v != 'None': a['tgt_cols'].add(_v)
            for _v in (t.strip() for t in src_t.split(',') if t.strip()):
                if _v and _v not in ('None', 'N/A', '-'): a['src_tables'].add(_v)
            for _v in (t.strip() for t in src_c.split(',') if t.strip()):
                if _v and _v not in ('None', 'N/A', '-'): a['src_cols'].add(_v)
        for pk, a in parent_agg.items():
            if pk not in result:  # don't overwrite existing entries
                result[pk] = {
                    'src_tables': ', '.join(sorted(a['src_tables'])),
                    'tgt_tables': ', '.join(sorted(a['tgt_tables'])),
                    'src_cols': ', '.join(sorted(a['src_cols'])),
                    'tgt_cols': ', '.join(sorted(a['tgt_cols'])),
                    'row_count': a['row_count'],
                }
        print(f"  -> LLM: {len(result)} unique keys ({len(parent_agg)} extra parent-name keys added)")

    return result


def load_gudu_lineage(path: Path) -> dict:
    """Load gudu_lineage_output.csv (CSV) or .xlsx (Excel) and aggregate per SQL Object Name.

    CSV format: columns produced by extract_gudu_lineage.py.
    XLSX format: reads the 'Column Lineage' sheet.
    Returns dict with same structure as load_llm_lineage().
    """
    # ── CSV branch ──────────────────────────────────────────────────────────
    if str(path).lower().endswith('.csv'):
        agg = defaultdict(lambda: {
            'src_tables': set(), 'tgt_tables': set(),
            'src_cols': set(), 'tgt_cols': set(), 'row_count': 0,
            'col_map_raw': [],   # ordered (tgt_table, tgt_col, src_table, src_col, transformation) rows
        })
        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for rec in reader:
                pkg = str(rec.get('SQL Object Name') or '').strip()
                if not pkg or pkg == 'None':
                    continue
                key = pkg.upper()
                a = agg[key]
                a['row_count'] += 1
                tgt_t = str(rec.get('Target Table') or '').strip()
                tgt_c = str(rec.get('Target Column') or '').strip()
                src_t = str(rec.get('Source Table') or '').strip()
                src_c = str(rec.get('Source Column') or '').strip()
                transf = str(rec.get('Transformation') or '').strip()
                if tgt_t and tgt_t != 'None':
                    a['tgt_tables'].add(tgt_t)
                if tgt_c and tgt_c != 'None':
                    a['tgt_cols'].add(tgt_c)
                if src_t and src_t not in ('None', 'N/A', '-'):
                    a['src_tables'].add(src_t)
                if src_c and src_c not in ('None', 'N/A', '-'):
                    a['src_cols'].add(src_c)
                # Preserve per-column mapping for aligned output
                if tgt_c and tgt_c != 'None':
                    a['col_map_raw'].append((tgt_t, tgt_c, src_t, src_c, transf))
        result = {}
        for key, a in agg.items():
            result[key] = {
                'src_tables': ', '.join(sorted(a['src_tables'])),
                'tgt_tables': ', '.join(sorted(a['tgt_tables'])),
                'src_cols':   ', '.join(sorted(a['src_cols'])),
                'tgt_cols':   ', '.join(sorted(a['tgt_cols'])),
                'row_count':  a['row_count'],
                'col_mappings': _build_col_mappings(a['col_map_raw']),
            }

        # RC2 fix: also register each _partN key under its base name so that the
        # combiner can match PRC_FOO against PRC_FOO_PART1 / PRC_FOO_PART2 etc.
        import re as _re
        _part_re = _re.compile(r'_PART\d+$', _re.IGNORECASE)
        base_agg: dict = {}   # base_key -> merged agg entry
        for key, entry in result.items():
            base = _part_re.sub('', key)
            if base == key:          # not a _partN key, skip
                continue
            if base not in base_agg:
                base_agg[base] = {
                    'src_tables': set(), 'tgt_tables': set(),
                    'src_cols': set(), 'tgt_cols': set(), 'row_count': 0,
                }
            ba = base_agg[base]
            ba['row_count'] += entry['row_count']
            for t in entry['src_tables'].split(', '):
                if t: ba['src_tables'].add(t)
            for t in entry['tgt_tables'].split(', '):
                if t: ba['tgt_tables'].add(t)
            for c in entry['src_cols'].split(', '):
                if c: ba['src_cols'].add(c)
            for c in entry['tgt_cols'].split(', '):
                if c: ba['tgt_cols'].add(c)
        for base, ba in base_agg.items():
            if base not in result:   # don't overwrite a direct entry
                result[base] = {
                    'src_tables': ', '.join(sorted(ba['src_tables'])),
                    'tgt_tables': ', '.join(sorted(ba['tgt_tables'])),
                    'src_cols':   ', '.join(sorted(ba['src_cols'])),
                    'tgt_cols':   ', '.join(sorted(ba['tgt_cols'])),
                    'row_count':  ba['row_count'],
                }

        print(f"  -> Gudu CSV: {len(result)} unique SQL objects")
        return result

    # ── XLSX branch ─────────────────────────────────────────────────────────
    path = _copy_if_locked(path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    # Prefer 'Column Lineage' sheet, fall back to active
    if 'Column Lineage' in wb.sheetnames:
        ws = wb['Column Lineage']
    else:
        ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print(f"  WARNING: No rows found in {path.name}")
        return {}

    headers = [str(h).strip() for h in rows[0]]
    obj_col = None
    for candidate in ('SQL Object Name', 'SQL_PACKAGE', 'Package_Name'):
        if candidate in headers:
            obj_col = candidate
            break
    if obj_col is None:
        print(f"  WARNING: No object name column found in {path.name}")
        return {}

    agg = defaultdict(lambda: {
        'src_tables': set(), 'tgt_tables': set(),
        'src_cols': set(), 'tgt_cols': set(), 'row_count': 0,
    })
    for row in rows[1:]:
        rec = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        pkg = str(rec.get(obj_col) or '').strip()
        if not pkg or pkg == 'None':
            continue
        key = pkg.upper()
        a = agg[key]
        a['row_count'] += 1
        tgt_t = str(rec.get('Target Table') or '').strip()
        tgt_c = str(rec.get('Target Column') or '').strip()
        src_t = str(rec.get('Source Table') or '').strip()
        src_c = str(rec.get('Source Column') or '').strip()
        if tgt_t and tgt_t != 'None':
            a['tgt_tables'].add(tgt_t)
        if tgt_c and tgt_c != 'None':
            a['tgt_cols'].add(tgt_c)
        if src_t and src_t not in ('None', 'N/A', '-'):
            a['src_tables'].add(src_t)
        if src_c and src_c not in ('None', 'N/A', '-'):
            a['src_cols'].add(src_c)

    result = {}
    for key, a in agg.items():
        result[key] = {
            'src_tables': ', '.join(sorted(a['src_tables'])),
            'tgt_tables': ', '.join(sorted(a['tgt_tables'])),
            'src_cols': ', '.join(sorted(a['src_cols'])),
            'tgt_cols': ', '.join(sorted(a['tgt_cols'])),
            'row_count': a['row_count'],
        }
    return result


def load_pkg_proc_analysis(path: Path) -> dict:
    """Load PKG_PROC_Analysis_All_Metadata.xlsx.

    Reads 'Orchestrator Analysis' and 'Package Procedures' sheets to build
    a map of each package's internal procedure call structure.

    Returns dict keyed by UPPER(pkg_name) -> {
        'pattern':     str,   # 'Orchestrator', 'Hybrid', 'Independent', 'Self-contained'
        'has_main':    bool,
        'child_procs': list,  # lineage-relevant child procs in call order
    }
    """
    path = _copy_if_locked(path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    # Build (PKG, PROC) -> classification from 'Package Procedures' sheet
    proc_classification = {}
    if 'Package Procedures' in wb.sheetnames:
        ws_p = wb['Package Procedures']
        rows_p = list(ws_p.iter_rows(values_only=True))
        hdrs_p = list(rows_p[0])
        for r in rows_p[1:]:
            rec = dict(zip(hdrs_p, r))
            pkg  = str(rec.get('Package Name') or '').strip().upper()
            proc = str(rec.get('Procedure Name') or '').strip().upper()
            stat = str(rec.get('Status') or '').strip()
            cls  = str(rec.get('Classification') or '').strip()
            if pkg and proc and stat == 'Active':
                proc_classification[(pkg, proc)] = cls

    result = {}
    if 'Orchestrator Analysis' not in wb.sheetnames:
        wb.close()
        return result

    ws_o = wb['Orchestrator Analysis']
    rows_o = list(ws_o.iter_rows(values_only=True))
    hdrs_o = list(rows_o[0])
    wb.close()

    for r in rows_o[1:]:
        rec = dict(zip(hdrs_o, r))
        pkg      = str(rec.get('Package Name') or '').strip().upper()
        pattern  = str(rec.get('Pattern') or '').strip()
        has_main = str(rec.get('Has main?') or '').strip().upper() == 'YES'

        call_seq_str   = str(rec.get('Call Sequence from main') or '')
        not_main_str   = str(rec.get('Procs NOT Called from main') or '')

        # Parse call sequence (delimiter is → or comma)
        child_procs = []
        if call_seq_str and call_seq_str.upper() not in ('N/A', 'N/A (EACH CALLED EXTERNALLY)', 'NONE', ''):
            for p in re.split(r'[→,]', call_seq_str):
                p = p.strip().upper()
                if p and p not in ('N/A', 'NONE', ''):
                    child_procs.append(p)

        # Add procs not called from main
        if not_main_str and not_main_str.upper() not in ('N/A', 'NONE', ''):
            for p in re.split(r'[,;]', not_main_str):
                p = p.strip().upper()
                if p and p not in ('N/A', 'NONE', '') and p not in child_procs:
                    child_procs.append(p)

        # Keep only lineage-relevant procs (exclude Utility ones)
        lineage_procs = []
        for proc in child_procs:
            cls = proc_classification.get((pkg, proc), 'Lineage-relevant')
            if 'Utility' not in cls:
                lineage_procs.append(proc)

        if pkg:
            result[pkg] = {
                'pattern':     pattern,
                'has_main':    has_main,
                'child_procs': lineage_procs,
            }

    return result


def expand_pkg_main_rows(combined: list[dict], pkg_analysis: dict) -> list[dict]:
    """Expand PKG.main orchestrator rows into individual child proc rows.

    For each row where SQL_OBJECT_TYPE=PACKAGE and PROC_NAME=main and
    the package pattern is 'Orchestrator (main calls children)':
      - The original row is marked ORCHESTRATOR_ONLY (main has no data lineage)
      - One new row is added per lineage-relevant child proc
      - FULL_OBJECT is set to PKG_PROC (underscore) to match Gudu/LLM keys

    Packages with 'Self-contained' pattern (main IS the lineage object) are kept
    as-is.  Packages with 'Independent procs' pattern are already captured by
    DIRECT_PKG_PROC_CALL and need no expansion.
    """
    expanded = []
    for row in combined:
        pkg_name = (row.get('PACKAGE_NAME') or '').strip()
        sub_obj  = (row.get('PROC_NAME') or '').strip().upper()
        obj_type = (row.get('SQL_OBJECT_TYPE') or '').strip()

        pkg_info = pkg_analysis.get(pkg_name.upper()) if pkg_name else None

        # Only expand PACKAGE rows with sub=main that are Orchestrators or Hybrids.
        # DIFW_QUERY rows already carry full column lineage from RPT_ALL_Schema
        # and must never be touched — expanding them would overwrite COMPLETE
        # with ORCHESTRATOR_ONLY and lose the SOURCE_COL / TARGET_COL data.
        pattern = pkg_info.get('pattern', '') if pkg_info else ''
        is_orchestrator = 'Orchestrator' in pattern
        is_hybrid       = 'Hybrid' in pattern   # "Hybrid (main has DML + calls children)"

        if (sub_obj != 'MAIN'
                or obj_type not in ('PACKAGE',)
                or pkg_info is None
                or (not is_orchestrator and not is_hybrid)
                or row.get('LINEAGE_SOURCE') == 'DIFW_QUERY'):
            # For DIFW rows that have pkg_info, annotate with the framework proc sequence.
            if (row.get('LINEAGE_SOURCE') == 'DIFW_QUERY'
                    and pkg_info is not None
                    and pkg_info.get('child_procs')):
                procs = pkg_info['child_procs']
                seq_note = 'DIFW infra procs: ' + ' → '.join(procs)
                existing = row.get('NOTES') or ''
                if seq_note not in existing:
                    row['NOTES'] = (existing + ' | ' + seq_note).strip(' |')
            expanded.append(row)
            continue

        child_procs = pkg_info.get('child_procs', [])
        if not child_procs:
            # No child proc info available — keep as-is
            expanded.append(row)
            continue

        if is_hybrid:
            # Hybrid: main has its own DML lineage — keep main row unchanged
            # so Gudu/LLM can still enrich it, then also add child proc rows.
            main_row = dict(row)
            main_row['NOTES'] = ((main_row.get('NOTES') or '')
                                 + f' | Hybrid: main has DML + calls {len(child_procs)} child procs').strip(' |')
            expanded.append(main_row)
        else:
            # Pure Orchestrator: main delegates entirely to children — mark ORCHESTRATOR_ONLY
            orch_row = dict(row)
            orch_row['LINEAGE_STATUS'] = 'ORCHESTRATOR_ONLY'
            orch_row['NOTES'] = ((orch_row.get('NOTES') or '')
                                 + f' | Orchestrator: calls {len(child_procs)} child procs').strip(' |')
            expanded.append(orch_row)

        # Add one child-proc row per lineage-relevant procedure (both patterns)
        for i, child_proc in enumerate(child_procs, 1):
            child_row = dict(row)
            child_row['PROC_NAME']        = child_proc
            child_row['SQL_OBJECT_TYPE']      = 'PKG_PROCEDURE'
            # Use dot format (PKG.PROC) for consistency — _find_llm_match converts to
            # underscore (PKG_PROC) at lookup time to match Gudu/LLM key format.
            child_row['FULL_OBJECT']  = f"{pkg_name}.{child_proc}"
            child_row['PARENT_CALLER']        = f"{pkg_name}.MAIN"
            child_row['LINEAGE_SOURCE']       = 'PKG_PROC_EXPANSION'
            child_row['PATTERN_TYPE']         = 'PKG_MAIN_CHILD_PROC'
            child_row['LINEAGE_STATUS']       = 'SQL_OBJECT_IDENTIFIED'
            child_row['SRC_TABLE']            = ''
            child_row['TGT_TABLE']            = ''
            child_row['SOURCE_COL']           = ''
            child_row['TARGET_COL']           = ''
            child_row['NOTES']                = (f"Expanded from {pkg_name}.MAIN "
                                                 f"(call order {i}/{len(child_procs)})")
            expanded.append(child_row)

    new_count = sum(1 for r in expanded if r.get('LINEAGE_SOURCE') == 'PKG_PROC_EXPANSION')
    orch_count = sum(1 for r in expanded if r.get('LINEAGE_STATUS') == 'ORCHESTRATOR_ONLY')
    print(f"  -> PKG expansion: {orch_count} orchestrator rows marked, {new_count} child proc rows added")
    return expanded


def tag_difw_framework_procs(combined: list[dict]) -> int:
    """Tag PKG_GRP_LOAD_DIFW child proc rows as DIFW_FRAMEWORK.

    These are generic infrastructure procs (GET_DELETE_LOAD, GET_MULTISRCSYS_INSERT,
    GET_SINGLESRCSYS_MERGE1, PRC_RECON, TBL_REC_COUNT_JOBWISE) that are called by
    every DIFW-based load job. They are not RPT-table-specific and have no
    meaningful column-level lineage, so they are tagged separately to distinguish
    them from actionable SQL_OBJECT_IDENTIFIED gaps.

    Returns count of tagged rows.
    """
    DIFW_PKG_PREFIXES = ('PKG_GRP_LOAD_DIFW',)   # covers DIFW and DIFW_PD variants
    tagged = 0
    for row in combined:
        pkg_upper = row.get('PACKAGE_NAME', '').upper()
        if any(pkg_upper.startswith(p) for p in DIFW_PKG_PREFIXES):
            if row.get('LINEAGE_STATUS') == 'SQL_OBJECT_IDENTIFIED':
                row['LINEAGE_STATUS'] = 'DIFW_FRAMEWORK'
                row['NOTES'] = (
                    (row.get('NOTES') or '')
                    + ' | Generic DIFW framework proc — no RPT-table-specific lineage'
                ).strip(' |')
                tagged += 1
    return tagged


# ── Delete-preprocess proc metadata ──────────────────────────────────────────
# Maps (PACKAGE_NAME_UPPER, PROC_NAME_UPPER) → (tgt_table, src_tables, description)
# These procs contain only DELETE statements so Gudu/LLM cannot produce SELECT→INSERT
# lineage for them.  We enrich them statically so they appear with meaningful context.
_DELETE_PROC_METADATA: dict[tuple[str, str], tuple[str, str, str]] = {
    (
        'PKG_GRP_LOAD_RPT_CLAIMANT_DTL_R_INC',
        'PRC_DEL_EXT_DATA',
    ): (
        'RPT_CLAIMANT_DTL_R',
        'DIM_GRP_CLAIM_DETAIL_R_MV_SSL_INC; DIM_GRP_PARTY_R',
        'Delete-before-insert: removes stale/duplicate records from RPT_CLAIMANT_DTL_R '
        'for insureds modified since last run (DIM_GRP_CLAIM_DETAIL_R_MV_SSL_INC) '
        'and physically deleted parties (DIM_GRP_PARTY_R).',
    ),
    (
        'PKG_GRP_LOAD_RPT_CLAIM_NOTE_R_INC',
        'PRC_DEL_EXT_DATA',
    ): (
        'RPT_CLAIM_NOTE_R',
        'FCT_GRP_CLAIM_NOTE_R_DRQ_MV_SSL_INC; ATOMIC.DIM_GRP_CLAIM_DIR_R',
        'Delete-before-insert: (1) removes existing RPT_CLAIM_NOTE_R records matching '
        'claims modified since last run (FCT_GRP_CLAIM_NOTE_R_DRQ_MV_SSL_INC); '
        '(2) removes physically deleted claims (ATOMIC.DIM_GRP_CLAIM_DIR_R).',
    ),
    (
        'PKG_GRP_LOAD_RPT_CLAIM_TASK_R_INC',
        'PRC_DEL_EXT_DATA',
    ): (
        'RPT_CLAIM_TASK_R',
        'RPT_CLAIM_TASK_R_DRQ_MV_SSL_INC; ATOMIC.FCT_GRP_PROCESS_CUSTOM_R; '
        'ATOMIC.DIM_GRP_CLAIM_DIR_R; ATOMIC.DIM_GRP_CLAIM_COVERAGE_R',
        'Delete-before-insert: (1) removes RPT_CLAIM_TASK_R records matching incremental '
        'changes in RPT_CLAIM_TASK_R_DRQ_MV_SSL_INC; '
        '(2) removes physically deleted claim tasks via FCT_GRP_PROCESS_CUSTOM_R / '
        'DIM_GRP_CLAIM_COVERAGE_R (V_CHANGE_REASON_R = Physically Deleted).',
    ),
    (
        'PKG_GRP_LOAD_RPT_DICS_CLAIMS_DETAIL_OAC_R_INC',
        'PRC_UPD_DEL_DATA',
    ): (
        'RPT_DICS_CLAIMS_DETAIL_OAC_R',
        'ATOMIC.DIM_TIME_R; SSL_PACKAGE_MILESTONE_TABLE',
        'Partition management: determines current fiscal month from DIM_TIME_R; '
        'on Fisc-Month-End +2 day marks prior-month records inactive; '
        'otherwise truncates the current-month partition via prc_trunc_partition.',
    ),
}

_DELETE_PROC_NAMES: frozenset[str] = frozenset(
    proc for (_, proc) in _DELETE_PROC_METADATA
)


def enrich_delete_preprocess_procs(combined: list[dict]) -> int:
    """Tag delete-before-insert helper procs as DELETE_PREPROCESS with SRC/TGT context.

    PRC_DEL_EXT_DATA and PRC_UPD_DEL_DATA contain only DELETE (or TRUNCATE)
    statements, so Gudu/LLM cannot produce SELECT→INSERT lineage for them and
    they remain SQL_OBJECT_IDENTIFIED.  This function enriches them with
    statically-known source/target table info and a meaningful description so
    they are visible in the lineage output with full context.

    Returns the number of rows enriched.
    """
    tagged = 0
    for row in combined:
        if row.get('LINEAGE_STATUS') != 'SQL_OBJECT_IDENTIFIED':
            continue
        proc_upper = row.get('PROC_NAME', '').upper()
        if proc_upper not in _DELETE_PROC_NAMES:
            continue

        pkg_upper = row.get('PACKAGE_NAME', '').upper()
        meta = _DELETE_PROC_METADATA.get((pkg_upper, proc_upper))

        row['LINEAGE_STATUS'] = 'DELETE_PREPROCESS'
        if meta:
            tgt, src, desc = meta
            row['TGT_TABLE'] = tgt
            row['SRC_TABLE'] = src
            row['NOTES'] = (
                (row.get('NOTES') or '').rstrip('; ')
                + ('; ' if row.get('NOTES') else '')
                + desc
            )
        else:
            # Unknown delete proc — tag status and add generic note
            row['NOTES'] = (
                (row.get('NOTES') or '').rstrip('; ')
                + ('; ' if row.get('NOTES') else '')
                + f'Delete/partition-management proc — no SELECT→INSERT lineage available. '
                  f'Review {pkg_upper}.{proc_upper} SQL definition for source/target tables.'
            )
        tagged += 1
    return tagged


def _parse_in_type_from_refs(table_refs_raw: str) -> tuple[str, str]:
    """Parse IN_TYPE:VALUE from a semicolon-delimited table_references string.

    shell_parsed_objects.csv may include an IN_TYPE annotation alongside the
    table name to indicate which branch/variant a parameterized proc call
    targets.  Example:

        'FCT_CLAIM_PAYMENT_DETAIL_ADJUSTMENT_R; IN_TYPE:ADJUSTMENT'

    Returns (in_type_value, clean_refs_without_in_type):
        ('ADJUSTMENT', 'FCT_CLAIM_PAYMENT_DETAIL_ADJUSTMENT_R')

    If no IN_TYPE token is found, returns ('', original_string).
    """
    if not table_refs_raw:
        return '', ''
    in_type = ''
    clean_parts = []
    for part in table_refs_raw.split(';'):
        part = part.strip()
        if part.upper().startswith('IN_TYPE:'):
            in_type = part.split(':', 1)[1].strip().upper()
        elif part:
            clean_parts.append(part)
    clean_refs = '; '.join(clean_parts)
    return in_type, clean_refs


def _find_llm_match(row: dict, llm_lookup: dict) -> dict | None:
    """Try to match a combined_lineage row to the Gudu/LLM merged lookup.

    Match strategy (case-insensitive, in priority order):
      0. FULL_OBJECT__IN_TYPE_PARAM (highest priority — when IN_TYPE is present in
         table_references, e.g. 'FCT_...ADJUSTMENT_R; IN_TYPE:ADJUSTMENT' produces
         PKG_..._PRC_...__ADJUSTMENT matching the Gudu branch key exactly).
         Also tries PRIMARY__IN_TYPE as fallback compound key.
      1. PROC_NAME__SHELL_TABLE_REFS or PACKAGE_NAME__SHELL_TABLE_REFS (compound key)
         MUST come first so that table-parameterized procs like PROC_REFRESH_GRP_M_VIEW_TBLS
         resolve to their table-specific Gudu entry (e.g. PROC_REFRESH_GRP_M_VIEW_TBLS__CLAIM_TIER_WFAM_MV)
         rather than the generic aggregate entry (PROC_REFRESH_GRP_M_VIEW_TBLS).
         Also tries stripping _TBL / _TABLE suffix from the ref name.
      2. FULL_OBJECT      (e.g. 'PKG_GRP_LOAD_RPT_CLAIM_DTL_R_INC_PRC_GET_CUR_DATA')
      3. FULL_OBJECT with dot->underscore  (e.g. 'PKG.PRC' -> 'PKG_PRC')
      4. Dotted parts of FULL_OBJECT (e.g. each segment of 'PKG.PRC')
      5. PACKAGE_NAME    (e.g. 'PKG_GRP_FULLLOAD_OFFSETS'; skipped for DBMS_MVIEW)
      6. PROC_NAME       (e.g. 'PRC_FULLLOAD_FCT_CLAIMPMNT_DET_OFFSET1_R'; skipped for MAIN/REFRESH)
      7. SHELL_TABLE_REFS alone  (hardcoded table/MV — UNKNOWN-pattern rows)
    """
    candidates = []
    res = (row.get('FULL_OBJECT') or '').strip()
    pkg = (row.get('PACKAGE_NAME') or '').strip()
    proc = (row.get('PROC_NAME') or '').strip()
    shell_refs = (row.get('SHELL_TABLE_REFS') or '').strip()
    in_type_param = (row.get('IN_TYPE_PARAM') or '').strip().upper()

    # ── Priority 0: FULL_OBJECT__IN_TYPE (when IN_TYPE present in table_references) ──
    # e.g. FULL_OBJECT='PKG_GRP_FULLLOAD...R.PRC_FULLLOAD...R', IN_TYPE='ADJUSTMENT'
    # -> try 'PKG_GRP_FULLLOAD...R_PRC_FULLLOAD...R__ADJUSTMENT' (exact Gudu branch key)
    # Shell scripts may use full descriptive names (e.g. DISBURSEMENT) while Gudu
    # uses abbreviated equivalents (e.g. DISBURS). _IN_TYPE_GUDU_ALIASES handles this.
    if in_type_param:
        gudu_in_type = _IN_TYPE_GUDU_ALIASES.get(in_type_param, in_type_param)
        if res:
            full_obj_under = res.upper().replace('.', '_')
            candidates.append(f"{full_obj_under}__{gudu_in_type}")
            if gudu_in_type != in_type_param:   # also try original spelling as fallback
                candidates.append(f"{full_obj_under}__{in_type_param}")
        # Fallback: PRIMARY (pkg or proc)__IN_TYPE
        primary_for_in_type = pkg if (pkg and pkg.upper() not in ('DBMS_MVIEW',)) else proc
        if primary_for_in_type and primary_for_in_type.upper() not in ('MAIN', 'REFRESH', 'DBMS_MVIEW'):
            candidates.append(f"{primary_for_in_type.upper()}__{gudu_in_type}")
            if gudu_in_type != in_type_param:
                candidates.append(f"{primary_for_in_type.upper()}__{in_type_param}")

    # ── Priority 1: Compound key PROC/PKG__TABLE (must come before generic proc/pkg) ──
    # When SHELL_TABLE_REFS is present, the row is for a specific table invocation of
    # a parameterized proc.  The compound Gudu key (e.g. PROC_REFRESH_GRP_M_VIEW_TBLS__
    # CLAIM_TIER_WFAM_MV) is table-specific and must be preferred over the generic
    # aggregate entry (PROC_REFRESH_GRP_M_VIEW_TBLS which merges all tables).
    primary = proc if (not pkg or pkg.upper() == 'DBMS_MVIEW') else pkg
    if primary and primary.upper() not in ('MAIN', 'REFRESH', 'DBMS_MVIEW') and shell_refs:
        for ref_name in (r.strip() for r in shell_refs.split(',') if r.strip()):
            ref_upper = ref_name.upper()
            candidates.append(f"{primary.upper()}__{ref_upper}")
            # Strip trailing _TBL or _TABLE suffix and try again
            for suffix in ('_TBL', '_TABLE'):
                if ref_upper.endswith(suffix):
                    candidates.append(f"{primary.upper()}__{ref_upper[:-len(suffix)]}")

    # ── Priority 2-4: FULL_OBJECT ──
    if res:
        candidates.append(res.upper())
        if '.' in res:
            # dot->underscore: PKG.PRC -> PKG_PRC (matches Gudu/LLM key format)
            candidates.append(res.upper().replace('.', '_'))
            for part in res.split('.'):
                p = part.strip().upper()
                if p and p not in ('MAIN', 'DBMS_MVIEW'):
                    candidates.append(p)

    # ── Priority 5: PACKAGE_NAME (skip system package DBMS_MVIEW) ──
    if pkg and pkg.upper() != 'DBMS_MVIEW':
        candidates.append(pkg.upper())

    # ── Priority 6: PROC_NAME (skip generic entry-point tokens) ──
    if proc and proc.upper() not in ('MAIN', 'REFRESH'):
        candidates.append(proc.upper())

    # ── Priority 7: SHELL_TABLE_REFS alone (UNKNOWN-pattern scripts) ──
    if shell_refs:
        for ref_name in (r.strip() for r in shell_refs.split(',') if r.strip()):
            candidates.append(ref_name.upper())

    for c in candidates:
        if c in llm_lookup:
            return llm_lookup[c]
    return None


def _parse_difw_col_list(col_str: str) -> list[str]:
    """Parse a DIFW parenthesized comma-separated column string.

    Input : '(COL_A,COL_B,COL_C)'  →  ['COL_A', 'COL_B', 'COL_C']
    Handles missing/empty strings gracefully.
    """
    s = (col_str or '').strip()
    if s.startswith('(') and s.endswith(')'):
        s = s[1:-1]
    return [c.strip() for c in s.split(',') if c.strip()]


def _build_col_mappings(col_map_raw: list) -> list:
    """Build ordered, deduplicated column-mapping groups from raw (tgt_t, tgt_c, src_t, src_c, transf) tuples.

    Returns a list of dicts in first-seen order:
        [{'tgt_table': str, 'tgt_col': str,
          'sources': [{'src_table': str, 'src_col': str, 'transformation': str}, ...]}, ...]

    A target column with no source rows (constant/expression) has sources=[].
    """
    from collections import OrderedDict
    groups: OrderedDict = OrderedDict()  # (tgt_table, tgt_col) -> group dict
    for tgt_t, tgt_c, src_t, src_c, transf in col_map_raw:
        gkey = (tgt_t, tgt_c)
        if gkey not in groups:
            groups[gkey] = {'tgt_table': tgt_t, 'tgt_col': tgt_c, 'sources': [], '_src_seen': set()}
        grp = groups[gkey]
        src_key = (src_t, src_c)
        if src_c and src_key not in grp['_src_seen']:
            grp['sources'].append({'src_table': src_t, 'src_col': src_c, 'transformation': transf})
            grp['_src_seen'].add(src_key)
    # Drop internal dedup set before returning
    result = []
    for grp in groups.values():
        grp.pop('_src_seen', None)
        result.append(grp)
    return result


def _col_mappings_to_aligned_strings(col_mappings: list) -> tuple[str, str]:
    """Convert col_mappings list to aligned pipe-separated TARGET_COL / SOURCE_COL strings.

    Each segment (separated by ' | ') corresponds to one target column.
    Source columns within one target use '; ' as separator.
    Target columns with no source mapping get '(CONST)' in the source position.

    Returns (target_col_str, source_col_str).
    """
    tgt_parts = []
    src_parts = []
    for grp in col_mappings:
        tgt_parts.append(grp['tgt_col'])
        srcs = [s['src_col'] for s in grp['sources'] if s.get('src_col')]
        src_parts.append('; '.join(srcs) if srcs else '(CONST)')
    return ' | '.join(tgt_parts), ' | '.join(src_parts)


def enrich_from_llm(combined: list[dict], llm_lookup: dict) -> int:
    """Enrich non-DIFW combined rows with SRC/TGT from LLM data.

    Only fills SRC_TABLE, TGT_TABLE, SOURCE_COL, TARGET_COL if they are
    currently empty. Updates LINEAGE_STATUS for newly enriched rows.

    TARGET_COL and SOURCE_COL are written as pipe-separated aligned strings:
      - Each '|'-delimited segment is one target column
      - The matching source segment holds all contributing source columns (; -separated)
      - Target columns derived from constants/expressions get '(CONST)' in the source segment

    Returns count of enriched rows.
    """
    enriched = 0
    for row in combined:
        # Skip DIFW rows — they already have lineage
        if row.get('LINEAGE_SOURCE') == 'DIFW_QUERY':
            continue
        # Skip rows that already have SRC/TGT data
        if row.get('SRC_TABLE', '').strip() and row.get('TGT_TABLE', '').strip():
            continue

        match = _find_llm_match(row, llm_lookup)
        if not match:
            continue

        row['SRC_TABLE'] = match['src_tables']
        # Preserve an existing TGT_TABLE (e.g. authoritative MV name from shell
        # script refs) — only fill in if currently empty.  This prevents procedure
        # side-effect tables (status/execution logs, audit tables) from overwriting
        # the real target table determined by the shell parameter.
        if not row.get('TGT_TABLE', '').strip():
            row['TGT_TABLE'] = match['tgt_tables']
        # Use aligned pipe-separated format when col_mappings are available;
        # fall back to flat comma-separated sets for legacy XLSX-sourced entries.
        col_mappings = match.get('col_mappings')
        if col_mappings:
            tgt_str, src_str = _col_mappings_to_aligned_strings(col_mappings)
            row['TARGET_COL'] = tgt_str
            row['SOURCE_COL'] = src_str
        else:
            row['SOURCE_COL'] = match['src_cols']
            row['TARGET_COL'] = match['tgt_cols']

        # Upgrade status
        old_status = row.get('LINEAGE_STATUS', '')
        if old_status in ('NEEDS_INVESTIGATION', 'RECURSIVE_ONLY'):
            row['LINEAGE_STATUS'] = 'SQL_OBJECT_IDENTIFIED'
        if old_status == 'SQL_OBJECT_IDENTIFIED':
            row['LINEAGE_STATUS'] = 'COMPLETE'
        if old_status == 'MV_REFRESH_ONLY':
            row['LINEAGE_STATUS'] = 'COMPLETE'
        # Hybrid Orchestrator MAIN: ORCHESTRATOR_ONLY was set because PKG_PROC_Analysis
        # previously missed DML in MAIN (forward-declaration / hint issue).  If LLM
        # now has a match for this row, promote it to COMPLETE.
        if old_status == 'ORCHESTRATOR_ONLY':
            row['LINEAGE_STATUS'] = 'COMPLETE'

        if not row.get('NOTES'):
            row['NOTES'] = 'Enriched from LLM/STM parsed lineage'
        else:
            row['NOTES'] = row['NOTES'] + '; Enriched from LLM/STM'

        enriched += 1
    return enriched


# ─────────────────────────────────────────────────────────────────────────────
# COMBINING LOGIC
# ─────────────────────────────────────────────────────────────────────────────

def extract_shell_filename(cmd: str) -> str:
    """Extract shell script filename from full CMD path.
    /home/jobs/group/EDW/edw_grp_load_rpt_claim_payment_r.sh -> edw_grp_load_rpt_claim_payment_r.sh
    """
    if not cmd:
        return ""
    # Handle double-slash typos
    cmd = cmd.replace("//", "/")
    return Path(cmd.strip()).name


def is_difw_cmd(cmd: str) -> bool:
    """Check if the CMD is a DIFW shell script."""
    if not cmd:
        return False
    return 'difw' in cmd.lower()


def resolve_tidal_params_for_mv(params: str) -> str:
    """Extract MV name from TIDAL params for MV refresh jobs.
    E.g., 'RPT_CLAIM_TASK_R_DRQ_MV_SSL_INC ' -> RPT_CLAIM_TASK_R_DRQ_MV_SSL_INC
    """
    if not params:
        return ""
    parts = params.strip().split()
    return parts[0].strip() if parts else ""


def resolve_tidal_params_for_difw(params: str) -> dict:
    """Parse TIDAL params for DIFW scripts.
    E.g., 'PKG_GRP_LOAD_DIFW SHINKA_LOAD_DIM_GRP_CLAIM_COVERAGE_GROUP_R CV DIM_MULTI_SRC'
    -> {pkg_name: PKG_GRP_LOAD_DIFW, job_name: ..., source_system: ..., sourcetype: ...}
    """
    if not params:
        return {}
    parts = params.strip().split()
    result = {}
    if len(parts) >= 1:
        result['pkg_name'] = parts[0]
    if len(parts) >= 2:
        result['job_name'] = parts[1]
    if len(parts) >= 3:
        result['source_system'] = parts[2]
    if len(parts) >= 4:
        result['sourcetype'] = parts[3]
    return result


def _is_nan_or_empty(val: object) -> bool:
    """Return True if val is None, NaN, or an empty/whitespace string."""
    if val is None:
        return True
    if isinstance(val, float):
        import math
        return math.isnan(val)
    return not str(val).strip()


def combine_lineage(rpt_data: list[dict], tidal: dict, shell_parsed: dict) -> tuple[list[dict], frozenset[str]]:
    """
    Main combining logic.
    For each row in RPT_ALL_Schema:
      - If DIFW and already has SRC/TGT: keep as-is (lineage_source = 'DIFW_QUERY')
      - If non-DIFW: look up shell parser results + TIDAL params to enrich

    Returns (combined_rows, orphaned_jobs) where orphaned_jobs is the frozenset of
    job names that were skipped because they have no CMD, no shell script, and no
    downstream connections — i.e., pure source-system placeholder jobs.
    """
    # Pre-scan: identify orphaned no-cmd no-downstream jobs using TIDAL reverse-dependency data.
    # A job is "orphaned" if it has no CMD/shell AND no EDP_* (EDW chain) job depends on it
    # as a prerequisite. Jobs only depended on by DataLake movers, PACS/Shinka source loaders,
    # or VUE copy jobs are source-system placeholders with no SQL lineage for the EDW chain.
    #
    # We check: does any job whose name starts with 'EDP_' (case-sensitive) list this job
    # as a DEPENDENT_JOB (prerequisite)? If yes → keep; if no → orphaned.
    _reverse_dep: dict[str, set[str]] = {}
    for _job_name, _entries in tidal.items():
        for _entry in _entries:
            _dj = _entry.get('DEPENDENT_JOB')
            if _dj:
                _reverse_dep.setdefault(str(_dj).strip(), set()).add(str(_job_name).strip())

    _edw_depended: set[str] = {
        _dep_j for _dep_j, _dependers in _reverse_dep.items()
        if any(_d.startswith('EDP_') for _d in _dependers)
    }

    orphaned_jobs: set[str] = set()
    for _row in rpt_data:
        _dep = str(_row.get('DEPENDENT_JOB', '') or '').strip()
        _cmd_raw = _row.get('CMD')
        _cmd = '' if _is_nan_or_empty(_cmd_raw) else str(_cmd_raw).strip()
        _sh = extract_shell_filename(_cmd)
        # Orphan: no CMD, no shell, and not a prerequisite of any EDP_* EDW chain job
        if not _cmd and not _sh and _dep not in _edw_depended:
            orphaned_jobs.add(_dep)

    combined = []

    for row in rpt_data:
        root_job = row.get('ROOT_JOB', '') or ''
        dep_job = row.get('DEPENDENT_JOB', '') or ''
        cmd = row.get('CMD', '') or ''
        # If rpt_data has no CMD for this job (e.g. it only appears in a
        # supplementary TIDAL export added after the schema was built), fall
        # back to the merged tidal dict so newly-discovered CMDs are used.
        if not cmd:
            for _e in tidal.get(dep_job, []):
                _c = str(_e.get('CMD', '') or '').strip()
                if _c:
                    cmd = _c
                    break
        src_table = row.get('SRC_TABLE', '') or ''
        tgt_table = row.get('TGT_TABLE', '') or ''
        source_col = row.get('SOURCE_COL', '') or ''
        target_col = row.get('TARGET_COL', '') or ''

        sh_filename = extract_shell_filename(cmd)

        # ── Exclusions: source extraction jobs, pure trigger jobs, and orphaned placeholders ──
        if sh_filename in _EXCLUDED_SHELL_SCRIPTS or dep_job in _EXCLUDED_JOB_NAMES or dep_job in orphaned_jobs:
            continue

        # ── Disabled TIDAL jobs — confirmed disabled in scheduler, flag explicitly ──
        if dep_job in _DISABLED_TIDAL_JOBS:
            combined.append({
                'ROOT_JOB': root_job,
                'RPT_TABLE': _root_to_rpt(root_job),
                'DEPENDENT_JOB': dep_job,
                'CMD': '',
                'SHELL_SCRIPT': '',
                'PARENT_CALLER': '',
                'LINEAGE_SOURCE': 'TIDAL_DISABLED',
                'SQL_OBJECT_TYPE': '',
                'SQL_OBJECT_SCHEMA': '',
                'PACKAGE_NAME': '',
                'PROC_NAME': '',
                'FULL_OBJECT': '',
                'PATTERN_TYPE': 'DISABLED',
                'SHELL_TABLE_REFS': '',
                'IN_TYPE_PARAM': '',
                'SRC_TABLE': '',
                'TGT_TABLE': '',
                'SOURCE_COL': '',
                'TARGET_COL': '',
                'TIDAL_PARAMS': '',
                'IS_PARAMETERIZED': False,
                'LINEAGE_STATUS': 'TIDAL_DISABLED',
                'NOTES': (
                    'Job confirmed DISABLED in TIDAL scheduler (verified 2026-07-14). '
                    'No CMD captured in any export — lineage cannot be traced until re-enabled.'
                ),
            })
            continue

        # ── CASE 1: DIFW with existing lineage ──
        if is_difw_cmd(cmd) and src_table:
            combined.append({
                'ROOT_JOB': root_job,
                'RPT_TABLE': _root_to_rpt(root_job),
                'DEPENDENT_JOB': dep_job,
                'CMD': cmd,
                'SHELL_SCRIPT': sh_filename,
                'PARENT_CALLER': sh_filename,
                'LINEAGE_SOURCE': 'DIFW_QUERY',
                'SQL_OBJECT_TYPE': 'PACKAGE',
                'SQL_OBJECT_SCHEMA': 'ATOMIC',
                'PACKAGE_NAME': _get_difw_pkg_from_tidal(dep_job, tidal),
                'PROC_NAME': 'main',
                'PATTERN_TYPE': 'EXECUTE_IMMEDIATE_DYNAMIC',
                'SHELL_TABLE_REFS': '',
                    'IN_TYPE_PARAM': '',
                    'SRC_TABLE': src_table,
                    'TGT_TABLE': tgt_table,
                    'SOURCE_COL': source_col,
                    'TARGET_COL': target_col,
                    'TIDAL_PARAMS': _get_tidal_params(dep_job, tidal),
                    'IS_PARAMETERIZED': True,
                    'LINEAGE_STATUS': 'COMPLETE',
                'NOTES': '',
            })
            continue

        # ── CASE 2: DIFW without lineage (rare edge case) ──
        if is_difw_cmd(cmd) and not src_table:
            combined.append({
                'ROOT_JOB': root_job,
                'RPT_TABLE': _root_to_rpt(root_job),
                'DEPENDENT_JOB': dep_job,
                'CMD': cmd,
                'SHELL_SCRIPT': sh_filename,
                'PARENT_CALLER': sh_filename,
                'LINEAGE_SOURCE': 'DIFW_QUERY',
                'SQL_OBJECT_TYPE': 'PACKAGE',
                'SQL_OBJECT_SCHEMA': 'ATOMIC',
                'PACKAGE_NAME': _get_difw_pkg_from_tidal(dep_job, tidal),
                'PROC_NAME': 'main',
                'PATTERN_TYPE': 'EXECUTE_IMMEDIATE_DYNAMIC',
                'SHELL_TABLE_REFS': '',
                    'IN_TYPE_PARAM': '',
                    'SRC_TABLE': '',
                    'TGT_TABLE': '',
                    'SOURCE_COL': '',
                    'TARGET_COL': '',
                    'TIDAL_PARAMS': _get_tidal_params(dep_job, tidal),
                    'IS_PARAMETERIZED': True,
                    'LINEAGE_STATUS': 'DIFW_MISSING_LINEAGE',
                'NOTES': 'DIFW job but no SRC/TGT extracted from query',
            })
            continue

        # ── CASE 3: Non-DIFW — enrich from shell parser ──
        shell_refs = shell_parsed.get(sh_filename, [])
        tidal_params = _get_tidal_params(dep_job, tidal)

        if shell_refs:
            for ref in shell_refs:
                obj_name = ref.get('object_name', '')
                obj_type = ref.get('object_type', '')
                pattern = ref.get('pattern_type', '')
                is_param = ref.get('is_parameterized', 'False') == 'True'

                # Resolve parameterized names from TIDAL params.
                # Handles is_parameterized=True flag AND shell parser placeholder '(from TIDAL param $1)'
                resolved_name = obj_name
                if (is_param or obj_name.startswith('(')) and tidal_params:
                    resolved_name = _resolve_param_from_tidal(obj_name, pattern, tidal_params)

                # Determine lineage status
                if obj_type in ('UNKNOWN', 'ERROR'):
                    status = 'NEEDS_INVESTIGATION'
                elif obj_type in ('MATERIALIZED_VIEW',):
                    status = 'MV_REFRESH_ONLY'
                elif obj_type in ('INSERT_SELECT',):
                    status = 'STATIC_SQL_PARSED'
                else:
                    status = 'SQL_OBJECT_IDENTIFIED'

                # Map resolved name to PACKAGE_NAME / PROC_NAME by object type.
                # PACKAGE/PKG_PROCEDURE: resolved_name is the package, sub_object is the proc.
                # MATERIALIZED_VIEW: DBMS_MVIEW is the Oracle system package, REFRESH the proc;
                #   the actual MV name is stored directly in FULL_OBJECT.
                # All other types (PROCEDURE, TABLE, etc.): no package wrapper.
                sub_raw = ref.get('sub_object', '')
                if obj_type == 'MATERIALIZED_VIEW':
                    pkg_name_val, proc_name_val = 'DBMS_MVIEW', 'REFRESH'
                elif obj_type in ('PACKAGE', 'PKG_PROCEDURE'):
                    pkg_name_val, proc_name_val = resolved_name, sub_raw
                else:
                    pkg_name_val, proc_name_val = '', resolved_name

                # Parse IN_TYPE from table_references (e.g. 'TABLE_R; IN_TYPE:ADJUSTMENT')
                # IN_TYPE takes precedence over the raw table name for Gudu/LLM matching.
                _in_type_val, _clean_refs = _parse_in_type_from_refs(ref.get('table_references', ''))

                combined.append({
                    'ROOT_JOB': root_job,
                    'RPT_TABLE': _root_to_rpt(root_job),
                    'DEPENDENT_JOB': dep_job,
                    'CMD': cmd,
                    'SHELL_SCRIPT': sh_filename,
                    'PARENT_CALLER': sh_filename,
                    'LINEAGE_SOURCE': 'SHELL_PARSER',
                    'SQL_OBJECT_TYPE': obj_type,
                    'SQL_OBJECT_SCHEMA': ref.get('schema', ''),
                    'PACKAGE_NAME': pkg_name_val,
                    'PROC_NAME': proc_name_val,
                    'FULL_OBJECT': resolved_name if obj_type == 'MATERIALIZED_VIEW' else '',
                    'PATTERN_TYPE': pattern,
                    'SHELL_TABLE_REFS': _clean_refs,
                    'IN_TYPE_PARAM': _in_type_val,
                    'SRC_TABLE': '',
                    # Only pre-populate TGT_TABLE from shell refs for MV/UNKNOWN/INSERT_SELECT
                    # types. For PROCEDURE/PACKAGE types, leave empty so gudu enrichment sets
                    # the correct targets. Prevents shell display variables from locking in
                    # wrong TGT_TABLE (e.g. TABLE_NAME=FCT_CLAIM_PAYMENT_DETAIL_R in a script
                    # that never passes it to the proc).
                    'TGT_TABLE': _clean_refs if obj_type in ('MATERIALIZED_VIEW', 'UNKNOWN', 'ERROR', 'INSERT_SELECT') else '',
                    'SOURCE_COL': '',
                    'TARGET_COL': '',
                    'TIDAL_PARAMS': tidal_params,
                    'IS_PARAMETERIZED': is_param,
                    'LINEAGE_STATUS': status,
                    'NOTES': ref.get('notes', ''),
                })
        else:
            # No shell parser result — might be a non-shell CMD (Windows .exe, extractor, etc.)
            # Note: orphaned no-cmd jobs are already excluded above via the orphaned_jobs pre-scan.
            combined.append({
                'ROOT_JOB': root_job,
                'RPT_TABLE': _root_to_rpt(root_job),
                'DEPENDENT_JOB': dep_job,
                'CMD': cmd,
                'SHELL_SCRIPT': sh_filename,
                'PARENT_CALLER': sh_filename,
                'LINEAGE_SOURCE': 'UNRESOLVED',
                'SQL_OBJECT_TYPE': '',
                'SQL_OBJECT_SCHEMA': '',
                'PACKAGE_NAME': '',
                'PROC_NAME': '',
                'PATTERN_TYPE': _classify_unresolved_cmd(cmd),
                'SHELL_TABLE_REFS': '',
                'IN_TYPE_PARAM': '',
                'SRC_TABLE': '',
                'TGT_TABLE': '',
                'SOURCE_COL': '',
                'TARGET_COL': '',
                'TIDAL_PARAMS': tidal_params,
                'IS_PARAMETERIZED': False,
                'LINEAGE_STATUS': 'NEEDS_INVESTIGATION',
                'NOTES': f'Shell script not found in parser output: {sh_filename}',
            })

    # Add FULL_OBJECT column to every row (MV rows already have it set directly)
    for row in combined:
        row['FULL_OBJECT'] = row.get('FULL_OBJECT') or _resolve_sql_object(row)

    return combined, frozenset(orphaned_jobs)


# ─────────────────────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_sql_object(row: dict) -> str:
    """Build the FULL_OBJECT key from PACKAGE_NAME + PROC_NAME.

    Rules:
      - MV rows have FULL_OBJECT pre-set (PACKAGE_NAME='DBMS_MVIEW') so the caller
        guards with 'row.get(FULL_OBJECT) or' — this function is not called for them.
      - Placeholder PROC_NAME (V_..., $, '(from TIDAL param ...)') is resolved from
        TIDAL_PARAMS (first token) as a safety net.
      - If both PACKAGE_NAME and PROC_NAME are set, combines as PACKAGE_NAME.PROC_NAME.
      - Standalone PROCEDURE / TABLE rows: PACKAGE_NAME is empty, PROC_NAME is the identifier.
    """
    pkg = (row.get('PACKAGE_NAME') or '').strip()
    proc = (row.get('PROC_NAME') or '').strip()
    params = (row.get('TIDAL_PARAMS') or '').strip()

    # Safety net: resolve placeholder PROC_NAME from TIDAL params
    if proc and (proc.upper().startswith('V_') or proc.upper().startswith('$')
                 or proc.startswith('(')):
        if params:
            return params.split()[0]
        return proc  # can't resolve without params, keep as-is

    # PKG.PROC combination (PACKAGE or PKG_PROCEDURE rows)
    if pkg and pkg.upper() != 'DBMS_MVIEW' and proc:
        return f"{pkg}.{proc}"

    result = (pkg if pkg.upper() != 'DBMS_MVIEW' else '') or proc or ''
    # Fallback: for UNKNOWN-pattern scripts where both are empty,
    # use the hardcoded table/MV name extracted from the shell script.
    if not result:
        shell_refs = (row.get('SHELL_TABLE_REFS') or '').strip()
        if shell_refs:
            return shell_refs.split(',')[0].strip()
    return result


def _root_to_rpt(root_job: str) -> str:
    """Map root job name back to RPT table name."""
    for table, job in RPT_ROOT_JOBS.items():
        if job == root_job:
            return table
    return root_job


def _get_tidal_params(job_name: str, tidal: dict) -> str:
    """Get PARAMS from TIDAL dump for a given job."""
    entries = tidal.get(job_name, [])
    for e in entries:
        params = e.get('PARAMS')
        if params:
            return str(params).strip()
    return ''


def _get_difw_pkg_from_tidal(job_name: str, tidal: dict) -> str:
    """For DIFW jobs, extract package name from TIDAL PARAMS ($1)."""
    params = _get_tidal_params(job_name, tidal)
    if params:
        parts = params.strip().split()
        if parts:
            return parts[0]  # First param is the package name
    return 'PKG_GRP_LOAD_DIFW'


def _resolve_param_from_tidal(obj_name: str, pattern: str, tidal_params: str) -> str:
    """Resolve parameterized object names using TIDAL params."""
    if not tidal_params:
        return obj_name

    parts = tidal_params.strip().split()

    if pattern == 'MV_REFRESH' or pattern == 'MV_REFRESH_WITH_TRUNCATE':
        # $1 is the MV name
        return parts[0] if parts else obj_name

    if pattern == 'EXECUTE_IMMEDIATE_DYNAMIC':
        # $1 is the package name
        return parts[0] if parts else obj_name

    # Generic: try first param
    if '$' in obj_name and parts:
        return parts[0]

    return obj_name


def _classify_unresolved_cmd(cmd: str) -> str:
    """Classify unresolved CMDs into categories."""
    if not cmd:
        return 'NO_CMD'
    cmd_lower = cmd.lower()
    if '.exe' in cmd_lower:
        return 'WINDOWS_EXE'
    if 'extractor' in cmd_lower:
        return 'DATA_EXTRACTOR'
    if 'datalake' in cmd_lower or '_dl_' in cmd_lower:
        return 'DATALAKE_SCRIPT'
    if '.sh' in cmd_lower:
        return 'SHELL_SCRIPT_NOT_AVAILABLE'
    return 'OTHER'


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

COMBINED_FIELDS = [
    'ROOT_JOB', 'RPT_TABLE', 'DEPENDENT_JOB', 'DEPTH', 'JOB_CATEGORY',
    'UPSTREAM_JOBS', 'DOWNSTREAM_JOBS',
    'Avg_runtime_Duration (Sec)',
    'CMD', 'SHELL_SCRIPT', 'PARENT_CALLER',
    'LINEAGE_SOURCE', 'SQL_OBJECT_TYPE', 'SQL_OBJECT_SCHEMA',
    'PACKAGE_NAME', 'PROC_NAME', 'FULL_OBJECT', 'PATTERN_TYPE',
    'SHELL_TABLE_REFS', 'IN_TYPE_PARAM',
    'SRC_TABLE', 'TGT_TABLE', 'SOURCE_COL', 'TARGET_COL',
    'TIDAL_PARAMS', 'IS_PARAMETERIZED', 'LINEAGE_STATUS', 'NOTES'
]

# Jobs whose shell script identifies them as source-system extraction jobs.
# These are upstream of the EDW layer and carry no EDW lineage — excluded from output.
_EXCLUDED_SHELL_SCRIPTS: frozenset[str] = frozenset({
    'extractor_client.sh',          # PACS source extraction jobs
    'CSSI.VUE.RLSI.ConsoleApp.exe', # VUE/EDP console service jobs (source-side Windows app)
    'edp_grp_datalake.sh',          # DataLake landing/move jobs (DL layer, not EDW)
    'EDPControlSummaryCount.bat',   # EDP control summary count — source-side Windows batch job
})

# Specific job names that are pure orchestration triggers with no SQL lineage.
_EXCLUDED_JOB_NAMES: frozenset[str] = frozenset({
    'Grp_Start',
    # PACS source-system extraction jobs (feed the PACS source layer, not EDW)
    'PACS_PROD_EDPinc_Entity',
    # VUE / EDP console service copy jobs (source-side Windows app, no EDW SQL)
    'COPY_VUE_EDPConsoleService_CUSTB_PROD',
    # EDP control summary count jobs (source-side Windows batch, multiple variants)
    'Copy_VUE_EDPControlSummary_Count CUST_PROD',
    'Copy_VUE_EDPControlSummary_Count POLB_PROD',
    'Copy_VUE_EDPControlSummary_Count APPL_PROD',
    'Copy_VUE_EDPControlSummary_Count INSRD_PROD',
    'Copy_VUE_EDPControlSummary_Count PREM_PROD',
})

# TIDAL jobs confirmed as DISABLED in the scheduler.
# They appear as DEPENDENT_JOB references (other jobs wait for them) but never
# run, so no CMD is captured in any export.  Flagged here so the lineage output
# shows a clear TIDAL_DISABLED status rather than a false NEEDS_INVESTIGATION gap.
# Source: manual TIDAL UI inspection 2026-07-14.
# TODO: ask data team to include a DISABLED flag in future TIDAL exports.
_DISABLED_TIDAL_JOBS: frozenset[str] = frozenset({
    'EDP_ODI_EDW_CV_SHINKA_LOAD_DIM_GRP_POLICY_DIR_R',
    'EDP_EDW_SHINKA_UPD_CV_PARTYSK_FCT_CLAIM_PAYMENT_DETAIL_BENEFIT_PAYMENT_R',
    'EDP_EDW_SHINKA_UPD_CV_PARTYSK_FCT_CLAIM_PAYMENT_DETAIL_GROSS_BENEFIT_R',
    'EDP_GRP_EDW_MV_REFRESH_RPT_CLAIM_PAYMENT_R_CHECKDT_MV_SSL-20.1',
    'EDP_GRP_EDW_MV_REFRESH_RPT_CLAIM_PAYMENT_R_PAIDDT_MV_SSL-20.2',
    'EDP_EDW_GRP_PACS_LOAD_FCT_CLAIM_PAYMENT_DETAIL_R_ADJUSTMENT',
    'EDP_GRP_EDW_MV_REFRESH_FCT_RPT_CROSS_SELL_ANNPREM_CUSTLINK_POL_EFF_DT_DRQ_MV-2',
    'EDP_GRP_EDW_MV_REFRESH_FCT_RPT_CROSS_SELL_ANNPREM_CUSTLINK_POL_EFF_DT_DRQ_MV2-3',
    'EDP_GRP_EDW_MV_REFRESH_FCT_RPT_CROSS_SELL_POLICY_EFF_DT_DRQ_MV-1',
    'EDP_GRP_EDW_MV_REFRESH_FCT_RPT_CROSS_SELL_REWRITEIND_DRQ_MV-4',
    'EDP_GRP_EDW_MV_REFRESH_FCT_RPT_CROSS_SELL_SUMMARY_DRQ_MV-5',
    'EDP_ODI_EDW_CV_SHINKA_LOAD_DIM_GRP_POLICY_DIR_R',
    'EDP_GRP_EDW_MV_REFRESH_FCT_RPT_CROSS_SELL_ANNPREM_CUSTLINK_POL_EFF_DT_DRQ_MV-2',
    'EDP_GRP_EDW_MV_REFRESH_FCT_RPT_CROSS_SELL_ANNPREM_CUSTLINK_POL_EFF_DT_DRQ_MV2-3',
    'EDP_GRP_EDW_MV_REFRESH_FCT_RPT_CROSS_SELL_POLICY_EFF_DT_DRQ_MV-1',
    'EDP_GRP_EDW_MV_REFRESH_FCT_RPT_CROSS_SELL_REWRITEIND_DRQ_MV-4',
    'EDP_GRP_EDW_MV_REFRESH_FCT_RPT_CROSS_SELL_SUMMARY_DRQ_MV-5',
})

# Audit / system-generated columns injected by the DIFW framework into every load.
# These carry no business lineage (values are SYSDATE, package name literals, or
# sequence constants) and are excluded from the normalized column-level output.
_DIFW_AUDIT_COLUMNS: frozenset[str] = frozenset({
    'T_CREATION_DATE_R',
    'V_CREATED_BY_R',
    'T_LAST_MODIFIED_DATE_R',
    'V_LAST_MODIFIED_BY_R',
    'N_LOAD_RUN_ID_R',
})


def load_runtime_lookup(path: Path) -> dict:
    """Load TidalJob_avg_runtime_duration.csv.

    Returns dict keyed by UPPER(Job Name) -> avg duration string (seconds).
    """
    if not path.exists():
        print(f"  WARNING: Runtime file not found: {path}")
        return {}
    lookup = {}
    with open(path, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            name = str(row.get('Job Name', '') or '').strip()
            dur  = str(row.get('Avg Duration (Second)', '') or '').strip()
            if name:
                lookup[name.upper()] = dur
    return lookup


def load_view_definitions(gudu_path: Path) -> dict:
    """Load VIEW object definitions from gudu_lineage_output.csv.

    Reads all rows where Object Type == 'View' and returns:
        dict: UPPER(view_name) -> set of UPPER(immediate_source_tables)

    Immediate sources are kept un-expanded so that cascading views appear as
    their own nodes (VIEW_A whose SRC = VIEW_B; VIEW_B whose SRC = base tables).
    """
    raw_views: dict[str, set] = {}

    if not gudu_path.exists():
        print(f"  WARNING: Gudu lineage file not found for view extraction: {gudu_path}")
        return raw_views

    with open(gudu_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for rec in reader:
            if str(rec.get('Object Type', '')).strip().lower() != 'view':
                continue
            view_name = str(rec.get('SQL Object Name', '') or '').strip().upper()
            src_table = str(rec.get('Source Table', '') or '').strip().upper()
            if not view_name or not src_table or src_table in ('', 'NONE', 'N/A', '-'):
                continue
            raw_views.setdefault(view_name, set()).add(src_table)

    if raw_views:
        print(f"  -> View definitions loaded: {len(raw_views)} views from gudu_lineage_output.csv")
        for vn in sorted(raw_views):
            print(f"       {vn}: {sorted(raw_views[vn])}")
    return raw_views


def inject_view_expansion_rows(combined: list[dict], view_defs: dict) -> tuple[list, int]:
    """Inject synthetic VIEW rows for non-Tidal views that appear as SRC_TABLE.

    Algorithm (BFS over cascading views):
      1. Collect every view name that appears in any SRC_TABLE of the existing rows.
      2. For each unique (RPT_TABLE, view_name) pair create one synthetic VIEW row:
           DEPENDENT_JOB = VIEW_{view_name}
           SRC_TABLE     = comma-joined immediate source tables of that view
           TGT_TABLE     = view_name
           JOB_CATEGORY  = VIEW
           LINEAGE_SOURCE = VIEW_EXPANSION
      3. If any of the newly injected VIEW rows' SRC_TABLE is itself a known view,
         add that view to the work queue and repeat (handles cascading views).

    The result is appended to `combined` and the count of injected rows returned.
    """
    if not view_defs:
        return combined, 0

    # Normalise view_defs keys to UPPER (already done in load_view_definitions)
    view_defs_upper = {k.upper(): v for k, v in view_defs.items()}

    injected: list[dict] = []
    seen_pairs: set[tuple] = set()   # (rpt_table, view_name) already injected

    # Seed: find views referenced in original combined rows
    # Build (RPT_TABLE, view_name) -> set of consuming job names
    to_process: list[tuple[str, str, str, str, str | int]] = []  # (rpt, root, view, consumer_job, depth)

    for row in combined:
        src = str(row.get('SRC_TABLE') or '').strip().upper()
        if not src:
            continue
        rpt = row.get('RPT_TABLE', '')
        root = row.get('ROOT_JOB', '')
        dep_job = row.get('DEPENDENT_JOB', '')
        depth_val = row.get('DEPTH', '')
        for part in (p.strip() for p in src.split(',') if p.strip()):
            if part in view_defs_upper:
                to_process.append((rpt, root, part, dep_job, depth_val))

    # BFS: process queue and collect new views triggered by injected VIEW rows
    queue = list(to_process)
    processed: set[tuple] = set()  # (rpt, view_name) guard against re-processing

    while queue:
        rpt, root, view_name, consumer_job, depth_val = queue.pop(0)
        pair_key = (rpt, view_name)
        if pair_key in seen_pairs:
            continue
        seen_pairs.add(pair_key)
        processed.add(pair_key)

        src_tables_set = view_defs_upper.get(view_name, set())
        if not src_tables_set:
            continue

        src_tables_str = ', '.join(sorted(src_tables_set))

        # Compute VIEW depth: one level deeper (upstream) than the consuming job
        try:
            view_depth: int | str = int(depth_val) + 1 if depth_val != '' else ''
        except (ValueError, TypeError):
            view_depth = ''

        view_row = {
            'ROOT_JOB':                   root,
            'RPT_TABLE':                  rpt,
            'DEPENDENT_JOB':              f"VIEW_{view_name}",
            'DEPTH':                      view_depth,
            'JOB_CATEGORY':               'VIEW',
            'UPSTREAM_JOBS':              '',
            'DOWNSTREAM_JOBS':            consumer_job,
            'Avg_runtime_Duration (Sec)': '',
            'CMD':                        '',
            'SHELL_SCRIPT':               '',
            'PARENT_CALLER':              '',
            'LINEAGE_SOURCE':             'VIEW_EXPANSION',
            'SQL_OBJECT_TYPE':            'VIEW',
            'SQL_OBJECT_SCHEMA':          'ATOMIC',
            'PACKAGE_NAME':               '',
            'PROC_NAME':                  '',
            'FULL_OBJECT':                view_name,
            'PATTERN_TYPE':               'VIEW_DEFINITION',
            'SHELL_TABLE_REFS':           '',
            'IN_TYPE_PARAM':              '',
            'SRC_TABLE':                  src_tables_str,
            'TGT_TABLE':                  view_name,
            'SOURCE_COL':                 '',
            'TARGET_COL':                 '',
            'TIDAL_PARAMS':               '',
            'IS_PARAMETERIZED':           False,
            'LINEAGE_STATUS':             'COMPLETE',
            'NOTES':                      (
                f"Virtual VIEW node — non-Tidal view '{view_name}' expanded from "
                f"gudu_lineage_output.csv. Consumes: {src_tables_str}"
            ),
        }
        injected.append(view_row)

        # Cascade: if any of the view's src_tables is itself a known view,
        # enqueue it so it gets its own VIEW row too.
        for src_tbl in src_tables_set:
            if src_tbl in view_defs_upper and (rpt, src_tbl) not in processed:
                queue.append((rpt, root, src_tbl, f"VIEW_{view_name}", view_depth))

    if injected:
        unique_views = {r['FULL_OBJECT'] for r in injected}
        unique_rpts  = {r['RPT_TABLE']   for r in injected}
        print(
            f"  -> VIEW expansion: injected {len(injected)} virtual VIEW node row(s) "
            f"({len(unique_views)} unique view(s): {sorted(unique_views)}) "
            f"across {len(unique_rpts)} RPT table(s)"
        )
    else:
        print("  -> VIEW expansion: no known views found in SRC_TABLE of combined lineage — nothing injected")

    return combined + injected, len(injected)


def apply_runtime_duration(combined: list[dict], runtime_lookup: dict) -> int:
    """Populate 'Avg_runtime_Duration (Sec)' for each row from runtime_lookup.

    Matches on DEPENDENT_JOB (case-insensitive). Returns count of rows filled.
    """
    filled = 0
    for row in combined:
        dep_job = (row.get('DEPENDENT_JOB') or '').strip().upper()
        val = runtime_lookup.get(dep_job, '')
        row['Avg_runtime_Duration (Sec)'] = val
        if val:
            filled += 1
    return filled


# All COMBINED_FIELDS except SOURCE_COL/TARGET_COL (aggregated aligned strings) and
# SRC_TABLE/TGT_TABLE (aggregated table sets), all replaced by the five granular
# per-mapping columns appended at the end.
_NORM_BASE = [f for f in COMBINED_FIELDS if f not in ('SOURCE_COL', 'TARGET_COL', 'SRC_TABLE', 'TGT_TABLE')]
NORMALIZED_FIELDS = _NORM_BASE + [
    'TARGET_TABLE', 'TARGET_COLUMN', 'SOURCE_TABLE', 'SOURCE_COLUMN', 'TRANSFORMATION',
]


def write_column_normalized(combined: list[dict], merged_lookup: dict, output_path: Path):
    """Write normalized column-level lineage: one row per (job, target_col, source_col) mapping.

    Every column from combined_lineage_latest is preserved (SOURCE_COL/TARGET_COL keep
    the aggregated aligned view). The five appended columns TARGET_TABLE, TARGET_COLUMN,
    SOURCE_TABLE, SOURCE_COLUMN, TRANSFORMATION hold the granular per-mapping values.
    Target columns derived from constants/expressions have SOURCE_COLUMN empty and
    TRANSFORMATION = 'Constant / Expression'.
    """
    norm_count = 0
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=NORMALIZED_FIELDS, extrasaction='ignore')
        writer.writeheader()
        for row in combined:
            # ── DIFW_QUERY rows: SOURCE_COL/TARGET_COL are parenthesized comma-separated
            # lists with 1:1 positional mapping. Normalize each pair into its own row.
            if row.get('LINEAGE_SOURCE') == 'DIFW_QUERY' and row.get('TARGET_COL'):
                tgt_cols = _parse_difw_col_list(row.get('TARGET_COL', ''))
                src_cols = _parse_difw_col_list(row.get('SOURCE_COL', ''))
                src_table = row.get('SRC_TABLE', '')
                tgt_table = row.get('TGT_TABLE', '')
                for i, tgt_col in enumerate(tgt_cols):
                    src_col = src_cols[i] if i < len(src_cols) else ''
                    # Skip audit/system columns — they have no business lineage
                    # (e.g. T_CREATION_DATE_R mapped to SYSDATE by the DIFW framework).
                    src_col_bare = src_col.split('.')[-1].strip().upper()
                    if src_col_bare in _DIFW_AUDIT_COLUMNS:
                        continue
                    writer.writerow({**row,
                        'TARGET_TABLE':  tgt_table,
                        'TARGET_COLUMN': tgt_col,
                        'SOURCE_TABLE':  src_table,
                        'SOURCE_COLUMN': src_col,
                        'TRANSFORMATION': 'Direct',
                    })
                    norm_count += 1
                continue

            # ── Gudu / LLM rows: use col_mappings from merged_lookup ──
            match = _find_llm_match(row, merged_lookup)
            if not match:
                continue
            col_mappings = match.get('col_mappings')
            if not col_mappings:
                continue
            for grp in col_mappings:
                sources = grp.get('sources', [])
                if not sources:
                    # Constant / Expression — no source column
                    writer.writerow({**row,
                        'TARGET_TABLE':  grp['tgt_table'],
                        'TARGET_COLUMN': grp['tgt_col'],
                        'SOURCE_TABLE':  '',
                        'SOURCE_COLUMN': '',
                        'TRANSFORMATION': 'Constant / Expression',
                    })
                    norm_count += 1
                else:
                    for src in sources:
                        writer.writerow({**row,
                            'TARGET_TABLE':  grp['tgt_table'],
                            'TARGET_COLUMN': grp['tgt_col'],
                            'SOURCE_TABLE':  src.get('src_table', ''),
                            'SOURCE_COLUMN': src.get('src_col', ''),
                            'TRANSFORMATION': src.get('transformation', ''),
                        })
                        norm_count += 1
    return norm_count


def write_combined_csv(combined: list[dict], output_path: Path):
    """Write combined lineage to CSV."""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=COMBINED_FIELDS)
        writer.writeheader()
        for row in combined:
            writer.writerow(row)


def write_combined_excel(combined: list[dict], output_path: Path):
    """Write combined lineage to Excel with formatting."""
    if not HAS_OPENPYXL:
        print("openpyxl not available; skipping Excel output")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined Lineage latest"

    # Header styling
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    # Write headers
    for col_idx, field in enumerate(COMBINED_FIELDS, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Status color mapping
    status_colors = {
        'COMPLETE': PatternFill(start_color="C6EFCE", fill_type="solid"),                # Green
        'SQL_OBJECT_IDENTIFIED': PatternFill(start_color="FFEB9C", fill_type="solid"),   # Yellow
        'MV_REFRESH_ONLY': PatternFill(start_color="BDD7EE", fill_type="solid"),         # Light blue
        'STATIC_SQL_PARSED': PatternFill(start_color="D9E2F3", fill_type="solid"),       # Pale blue
        'DIFW_MISSING_LINEAGE': PatternFill(start_color="FFC7CE", fill_type="solid"),    # Light red
        'NEEDS_INVESTIGATION': PatternFill(start_color="FFC7CE", fill_type="solid"),     # Light red
        'RECURSIVE_ONLY': PatternFill(start_color="E2EFDA", fill_type="solid"),          # Pale green
        'ORCHESTRATOR_ONLY': PatternFill(start_color="D9D9D9", fill_type="solid"),       # Grey
        'DIFW_FRAMEWORK': PatternFill(start_color="E2EFDA", fill_type="solid"),           # Pale green
        'DELETE_PREPROCESS': PatternFill(start_color="F4B942", fill_type="solid"),        # Amber
        'TIDAL_DISABLED': PatternFill(start_color="D9B3FF", fill_type="solid"),           # Light purple
    }
    # Category color mapping
    category_colors = {
        'RPT': PatternFill(start_color="4472C4", fill_type="solid"),     # Blue
        'MV': PatternFill(start_color="70AD47", fill_type="solid"),      # Green
        'FCT': PatternFill(start_color="ED7D31", fill_type="solid"),     # Orange
        'DIM': PatternFill(start_color="FFC000", fill_type="solid"),     # Gold
        'STG': PatternFill(start_color="A5A5A5", fill_type="solid"),     # Grey
        'REF': PatternFill(start_color="5B9BD5", fill_type="solid"),     # Light blue
        'MONTH_END': PatternFill(start_color="BF8F00", fill_type="solid"),# Dark gold
    }
    category_font = Font(color="FFFFFF")  # White text for category cells

    depth_col_idx = COMBINED_FIELDS.index('DEPTH') + 1
    category_col_idx = COMBINED_FIELDS.index('JOB_CATEGORY') + 1

    # Write data
    for row_idx, row in enumerate(combined, 2):
        for col_idx, field in enumerate(COMBINED_FIELDS, 1):
            val = row.get(field, '')
            # Use `val != ''` (not `if val`) so integer 0 (e.g. DEPTH=0) is written
            # correctly rather than treated as falsy and rendered as blank.
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val != '' else '')

        # Color status column
        status = row.get('LINEAGE_STATUS', '')
        status_col = COMBINED_FIELDS.index('LINEAGE_STATUS') + 1
        if status in status_colors:
            ws.cell(row=row_idx, column=status_col).fill = status_colors[status]

        # Color category column
        cat = row.get('JOB_CATEGORY', '')
        if cat in category_colors:
            c = ws.cell(row=row_idx, column=category_col_idx)
            c.fill = category_colors[cat]
            c.font = category_font

    # Auto-width (approximate)
    for col_idx, field in enumerate(COMBINED_FIELDS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(40, max(12, len(field) + 4))

    # Freeze top row
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(str(output_path))


def write_gaps(combined: list[dict], output_path: Path):
    """Write rows that still need lineage resolution (gaps)."""
    # DIFW_FRAMEWORK, DELETE_PREPROCESS, and TIDAL_DISABLED rows are intentionally
    # excluded — they are known situations with clear explanations, not open gaps.
    gaps = [r for r in combined if r.get('LINEAGE_STATUS') in
            ('NEEDS_INVESTIGATION', 'DIFW_MISSING_LINEAGE', 'SQL_OBJECT_IDENTIFIED')]

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=COMBINED_FIELDS)
        writer.writeheader()
        for row in gaps:
            writer.writerow(row)


def write_summary(combined: list[dict], output_path: Path):
    """Write per-RPT-table summary statistics."""
    summary = defaultdict(lambda: {
        'total_jobs': 0,
        'complete': 0,
        'sql_identified': 0,
        'mv_refresh': 0,
        'static_sql': 0,
        'difw_framework': 0,
        'needs_investigation': 0,
        'unique_sql_objects': set(),
    })

    for row in combined:
        rpt = row.get('RPT_TABLE', 'UNKNOWN')
        status = row.get('LINEAGE_STATUS', '')
        s = summary[rpt]
        s['total_jobs'] += 1
        if status in ('COMPLETE', 'DIFW_QUERY'):
            # DIFW_QUERY rows carry full column lineage from RPT_ALL_Schema
            s['complete'] += 1
        elif status == 'SQL_OBJECT_IDENTIFIED':
            s['sql_identified'] += 1
        elif status == 'DELETE_PREPROCESS':
            s['sql_identified'] += 1  # counted alongside SQL_OBJECT_IDENTIFIED in completion %
        elif status == 'TIDAL_DISABLED':
            s['needs_investigation'] += 1  # disabled = known gap, counts toward investigation
        elif status in ('MV_REFRESH_ONLY',):
            s['mv_refresh'] += 1
        elif status == 'STATIC_SQL_PARSED':
            s['static_sql'] += 1
        elif status == 'DIFW_FRAMEWORK':
            s['difw_framework'] += 1
        elif status == 'ORCHESTRATOR_ONLY':
            pass  # meta-row; child procs already counted as COMPLETE
        else:
            s['needs_investigation'] += 1

        obj = row.get('FULL_OBJECT', '') or row.get('PROC_NAME', '') or row.get('PACKAGE_NAME', '')
        if obj:
            s['unique_sql_objects'].add(obj)

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'RPT_TABLE', 'TOTAL_JOBS', 'COMPLETE_LINEAGE', 'SQL_OBJECT_IDENTIFIED',
            'MV_REFRESH', 'STATIC_SQL', 'DIFW_FRAMEWORK', 'NEEDS_INVESTIGATION',
            'UNIQUE_SQL_OBJECTS', 'COMPLETION_PCT'
        ])
        for rpt in sorted(summary.keys()):
            s = summary[rpt]
            total = s['total_jobs']
            resolved = s['complete'] + s['sql_identified'] + s['mv_refresh'] + s['static_sql'] + s['difw_framework']
            pct = f"{(resolved / total * 100):.1f}%" if total > 0 else "0%"
            writer.writerow([
                rpt, total, s['complete'], s['sql_identified'],
                s['mv_refresh'], s['static_sql'], s['difw_framework'], s['needs_investigation'],
                len(s['unique_sql_objects']), pct
            ])


def print_summary(combined: list[dict]):
    """Print summary to console."""
    print("\n" + "=" * 90)
    print("TIDAL + SHELL PARSER COMBINED LINEAGE — SUMMARY")
    print("=" * 90)

    total = len(combined)
    by_status = defaultdict(int)
    by_source = defaultdict(int)
    by_rpt = defaultdict(lambda: defaultdict(int))

    for row in combined:
        by_status[row.get('LINEAGE_STATUS', '')] += 1
        by_source[row.get('LINEAGE_SOURCE', '')] += 1
        by_rpt[row.get('RPT_TABLE', '')][row.get('LINEAGE_STATUS', '')] += 1

    print(f"\nTotal combined rows: {total}")

    print("\n--- By Lineage Source ---")
    for src, cnt in sorted(by_source.items(), key=lambda x: -x[1]):
        print(f"  {src:30s} : {cnt:4d}  ({cnt/total*100:.1f}%)")

    print("\n--- By Lineage Status ---")
    for s, cnt in sorted(by_status.items(), key=lambda x: -x[1]):
        print(f"  {s:30s} : {cnt:4d}  ({cnt/total*100:.1f}%)")

    print("\n--- Per RPT Table ---")
    print(f"  {'RPT TABLE':<40s} {'TOTAL':>6s} {'DONE':>6s} {'IDENT':>6s} {'DIFW_FW':>8s} {'DISABL':>7s} {'GAPS':>6s} {'%':>7s}")
    print("  " + "-" * 90)
    for rpt in sorted(by_rpt.keys()):
        statuses = by_rpt[rpt]
        t = sum(statuses.values())
        # DIFW_QUERY rows have full column lineage — count as done
        done = statuses.get('COMPLETE', 0) + statuses.get('DIFW_QUERY', 0)
        ident = statuses.get('SQL_OBJECT_IDENTIFIED', 0) + statuses.get('MV_REFRESH_ONLY', 0) + statuses.get('STATIC_SQL_PARSED', 0) + statuses.get('DELETE_PREPROCESS', 0)
        difw_fw = statuses.get('DIFW_FRAMEWORK', 0)
        disabled = statuses.get('TIDAL_DISABLED', 0)
        gaps = statuses.get('NEEDS_INVESTIGATION', 0) + statuses.get('DIFW_MISSING_LINEAGE', 0)
        # Exclude ORCHESTRATOR_ONLY from % denominator — their child procs are
        # the real lineage rows and are already counted as COMPLETE above.
        orch = statuses.get('ORCHESTRATOR_ONLY', 0)
        effective_total = t - orch
        pct = f"{((done + ident + difw_fw) / effective_total * 100):.1f}%" if effective_total > 0 else "0%"
        print(f"  {rpt:<40s} {t:6d} {done:6d} {ident:6d} {difw_fw:8d} {disabled:7d} {gaps:6d} {pct:>7s}")

    print("\n" + "=" * 90)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Combine TIDAL deps + Shell parser + DIFW lineage")
    parser.add_argument('--rpt-schema', type=str,
                        default=str(BASE_DIR.parent / "DIFW_Query_Results.xlsx"),
                        help="Path to DIFW_Query_Results.xlsx")
    parser.add_argument('--tidal', type=str,
                        default=str(BASE_DIR.parent / "TIDAL_15_April_updated.xlsx"),
                        help="Path to primary TIDAL file (.csv or .xlsx)")
    parser.add_argument('--tidal-supplement', type=str,
                        default='',
                        help="Optional supplementary TIDAL file to fill CMD/PARAMS for jobs missing from primary")
    parser.add_argument('--shell-parsed', type=str,
                        default=str(OUTPUT_DIR / "shell_parsed_objects.csv"),
                        help="Path to shell_parsed_objects.csv")
    parser.add_argument('--llm-lineage', type=str,
                        default=str(OUTPUT_DIR / "llm_lineage_output.csv"),
                        help="Path to llm_lineage_output.csv/.xlsx (LLM-parsed column lineage, fallback after Gudu)")
    parser.add_argument('--gudu-lineage', type=str,
                        default=str(OUTPUT_DIR / "gudu_lineage_output.csv"),
                        help="Path to gudu_lineage_output.csv/.xlsx (Gudu-parsed column lineage, takes priority)")
    parser.add_argument('--pkg-proc-analysis', type=str,
                        default=str(BASE_DIR / "PKG_PROC_Analysis_All_Metadata.xlsx"),
                        help="Path to PKG_PROC_Analysis_All_Metadata.xlsx (PKG->child proc call structure)")
    parser.add_argument('--runtime', type=str,
                        default=str(BASE_DIR.parent / "TidalJob_avg_runtime_duration.csv"),
                        help="Path to TidalJob_avg_runtime_duration.csv (job avg runtime in seconds)")
    parser.add_argument('--output-dir', type=str,
                        default=str(OUTPUT_DIR),
                        help="Output directory")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print("Loading data sources...")
    print(f"  RPT Schema:    {args.rpt_schema}")
    print(f"  TIDAL Deps:    {args.tidal}")
    if args.tidal_supplement:
        print(f"  TIDAL Suppl:   {args.tidal_supplement}")
    print(f"  Shell Parsed:  {args.shell_parsed}")
    print(f"  LLM Lineage:   {args.llm_lineage}")
    print(f"  Gudu Lineage:  {args.gudu_lineage}")
    print(f"  PKG Analysis:  {args.pkg_proc_analysis}")
    print(f"  Runtime CSV:   {args.runtime}")

    rpt_data = load_rpt_schema(Path(args.rpt_schema))
    print(f"  -> RPT_ALL_Schema: {len(rpt_data)} rows")

    tidal = load_tidal_deps(Path(args.tidal))
    print(f"  -> TIDAL primary: {len(tidal)} unique jobs")

    # Merge supplement file (covers jobs truncated from the primary CSV export)
    tidal_supp_path = Path(args.tidal_supplement) if args.tidal_supplement else None
    if tidal_supp_path and tidal_supp_path.exists():
        tidal_supp = load_tidal_deps(tidal_supp_path)
        tidal = merge_tidal_dicts(tidal, tidal_supp)
    print(f"  -> TIDAL merged:  {len(tidal)} unique jobs")

    shell_parsed = load_shell_parsed(Path(args.shell_parsed))
    print(f"  -> Shell Parsed: {len(shell_parsed)} shell scripts")

    # Load Gudu lineage (priority source)
    gudu_lineage_path = Path(args.gudu_lineage)
    gudu_lookup = {}
    if gudu_lineage_path.exists():
        gudu_lookup = load_gudu_lineage(gudu_lineage_path)
        print(f"  -> Gudu Lineage: {len(gudu_lookup)} unique SQL objects (priority)")
    else:
        print(f"  -> Gudu Lineage: file not found, skipping")

    # Load LLM lineage (fallback source)
    llm_lineage_path = Path(args.llm_lineage)
    llm_lookup = {}
    if llm_lineage_path.exists():
        llm_lookup = load_llm_lineage(llm_lineage_path)
        print(f"  -> LLM Lineage: {len(llm_lookup)} unique SQL objects (fallback)")
    else:
        print(f"  -> LLM Lineage: file not found, skipping")

    # Load PKG/PROC call structure analysis
    pkg_proc_path = Path(args.pkg_proc_analysis)
    pkg_analysis = {}
    if pkg_proc_path.exists():
        pkg_analysis = load_pkg_proc_analysis(pkg_proc_path)
        print(f"  -> PKG Analysis: {len(pkg_analysis)} packages loaded")
    else:
        print(f"  -> PKG Analysis: file not found, skipping expansion")

    # Merge: Gudu takes priority, LLM fills gaps
    merged_lookup = dict(llm_lookup)  # start with LLM as base
    gudu_override_count = 0
    for key, val in gudu_lookup.items():
        if key in merged_lookup:
            gudu_override_count += 1
        merged_lookup[key] = val  # Gudu overwrites LLM
    if gudu_lookup:
        print(f"  -> Merged lookup: {len(merged_lookup)} objects ({len(gudu_lookup)} from Gudu, {gudu_override_count} overridden from LLM)")

    print("\nCombining data...")
    combined, orphaned_jobs = combine_lineage(rpt_data, tidal, shell_parsed)
    print(f"  -> Combined: {len(combined)} rows")
    if orphaned_jobs:
        print(f"  -> Orphaned placeholder jobs excluded: {len(orphaned_jobs)} ({', '.join(sorted(orphaned_jobs))})")

    # ── Compute recursive depths from TIDAL BFS ──────────────────────────
    print("\nComputing recursive job depths (BFS from 8 root jobs, excluding _DL_)...")
    depth_map = build_recursive_depths(tidal)
    print(f"  -> {len(depth_map)} (root_job, dep_job) pairs with depth")

    # Enrich combined rows with DEPTH and JOB_CATEGORY
    for row in combined:
        key = (row.get('ROOT_JOB', ''), row.get('DEPENDENT_JOB', ''))
        info = depth_map.get(key)
        if info:
            row['DEPTH'] = info['depth']
            row['JOB_CATEGORY'] = info['category']
        else:
            row['DEPTH'] = ''
            row['JOB_CATEGORY'] = ''

    # Find jobs in the recursive chain that aren't in RPT_ALL_Schema.
    # NOTE: dep_job == root_job is the depth-0 seed row (the root job itself).
    # This is intentionally included so the CMD/shell script that directly loads
    # the RPT table is captured in the output (previously it was silently dropped).
    existing_keys = {(r.get('ROOT_JOB', ''), r.get('DEPENDENT_JOB', '')) for r in combined}
    new_recursive_rows = []
    for (root_job, dep_job), info in depth_map.items():
        if (root_job, dep_job) in existing_keys:
            continue
        # Skip source extraction jobs, pure trigger jobs, and orphaned placeholders
        if dep_job in _EXCLUDED_JOB_NAMES or dep_job in orphaned_jobs:
            continue
        # Get CMD/PARAMS from TIDAL
        tidal_entries = tidal.get(dep_job, [])
        cmd = ''
        tidal_params = ''
        for e in tidal_entries:
            if e.get('CMD'):
                cmd = str(e['CMD']).strip()
            if e.get('PARAMS'):
                tidal_params = str(e['PARAMS']).strip()
            if cmd:
                break
        sh_filename = extract_shell_filename(cmd)

        # Skip source extraction scripts (e.g. extractor_client.sh)
        if sh_filename in _EXCLUDED_SHELL_SCRIPTS:
            continue

        # Look up shell parser for this script
        shell_refs = shell_parsed.get(sh_filename, [])
        if shell_refs:
            for ref in shell_refs:
                obj_name = ref.get('object_name', '')
                is_param = ref.get('is_parameterized', 'False') == 'True'
                resolved_name = obj_name
                if is_param and tidal_params:
                    resolved_name = _resolve_param_from_tidal(
                        obj_name, ref.get('pattern_type', ''), tidal_params)
                obj_type = ref.get('object_type', '')
                if obj_type in ('UNKNOWN', 'ERROR'):
                    status = 'NEEDS_INVESTIGATION'
                elif obj_type == 'MATERIALIZED_VIEW':
                    status = 'MV_REFRESH_ONLY'
                else:
                    status = 'SQL_OBJECT_IDENTIFIED'
                sub_raw = ref.get('sub_object', '')
                if obj_type == 'MATERIALIZED_VIEW':
                    pkg_name_val, proc_name_val = 'DBMS_MVIEW', 'REFRESH'
                elif obj_type in ('PACKAGE', 'PKG_PROCEDURE'):
                    pkg_name_val, proc_name_val = resolved_name, sub_raw
                else:
                    pkg_name_val, proc_name_val = '', resolved_name
                is_root_entry = (dep_job == root_job)
                new_row = {
                    'ROOT_JOB': root_job,
                    'RPT_TABLE': _root_to_rpt(root_job),
                    'DEPENDENT_JOB': dep_job,
                    'DEPTH': info['depth'],
                    'JOB_CATEGORY': info['category'],
                    'CMD': cmd,
                    'SHELL_SCRIPT': sh_filename,
                    'PARENT_CALLER': sh_filename,
                    'LINEAGE_SOURCE': 'ROOT_JOB_ENTRY' if is_root_entry else 'RECURSIVE_BFS',
                    'SQL_OBJECT_TYPE': ref.get('object_type', ''),
                    'SQL_OBJECT_SCHEMA': ref.get('schema', ''),
                    'PACKAGE_NAME': pkg_name_val,
                    'PROC_NAME': proc_name_val,
                    'FULL_OBJECT': resolved_name if obj_type == 'MATERIALIZED_VIEW' else '',
                    'PATTERN_TYPE': ref.get('pattern_type', ''),
                    'SHELL_TABLE_REFS': ref.get('table_references', ''),
                    'SRC_TABLE': '',
                    # Only pre-populate TGT_TABLE from shell table_references for MV/UNKNOWN
                    # types — where the ref IS the actual target. For PROCEDURE/PACKAGE types,
                    # leave empty so gudu enrichment fills in the correct targets. This prevents
                    # shell display variables (e.g. TABLE_NAME=FCT_..._R used only in echo)
                    # from locking in an incorrect TGT_TABLE and blocking gudu correction.
                    'TGT_TABLE': ref.get('table_references', '') if obj_type in ('MATERIALIZED_VIEW', 'UNKNOWN', 'ERROR') else '',
                    'SOURCE_COL': '',
                    'TARGET_COL': '',
                    'TIDAL_PARAMS': tidal_params,
                    'IS_PARAMETERIZED': is_param,
                    'LINEAGE_STATUS': status,
                    'NOTES': ('Root job (depth=0) — direct RPT loader script' if is_root_entry
                              else 'Added from recursive BFS (not in RPT_ALL_Schema)'),
                }
                new_row['FULL_OBJECT'] = new_row.get('FULL_OBJECT') or _resolve_sql_object(new_row)
                new_recursive_rows.append(new_row)
        else:
            # No shell parser match
            is_root_entry = (dep_job == root_job)
            new_row = {
                'ROOT_JOB': root_job,
                'RPT_TABLE': _root_to_rpt(root_job),
                'DEPENDENT_JOB': dep_job,
                'DEPTH': info['depth'],
                'JOB_CATEGORY': info['category'],
                'CMD': cmd,
                'SHELL_SCRIPT': sh_filename,
                'PARENT_CALLER': sh_filename,
                'LINEAGE_SOURCE': 'ROOT_JOB_ENTRY' if is_root_entry else 'RECURSIVE_BFS',
                'SQL_OBJECT_TYPE': '',
                'SQL_OBJECT_SCHEMA': '',
                'PACKAGE_NAME': '',
                'PROC_NAME': '',
                'FULL_OBJECT': '',
                'PATTERN_TYPE': _classify_unresolved_cmd(cmd),
                'SHELL_TABLE_REFS': '',
                'SRC_TABLE': '',
                'TGT_TABLE': '',
                'SOURCE_COL': '',
                'TARGET_COL': '',
                'TIDAL_PARAMS': tidal_params,
                'IS_PARAMETERIZED': False,
                'LINEAGE_STATUS': 'NEEDS_INVESTIGATION' if is_root_entry else 'RECURSIVE_ONLY',
                'NOTES': ('Root job (depth=0) — direct RPT loader script; shell not in parser output'
                          if is_root_entry else 'Found in recursive BFS but no shell/SQL details'),
            }
            new_recursive_rows.append(new_row)

    combined.extend(new_recursive_rows)
    root_entry_count = sum(1 for r in new_recursive_rows if r.get('LINEAGE_SOURCE') == 'ROOT_JOB_ENTRY')
    print(f"  -> Added {len(new_recursive_rows)} new rows from recursive BFS "
          f"({root_entry_count} root-job depth-0 entries)")
    print(f"  -> Total combined: {len(combined)} rows")

    # ── Expand PKG.main orchestrator rows into child proc rows ───────────
    if pkg_analysis:
        print("\nExpanding PKG.main orchestrator rows into child proc rows...")
        combined = expand_pkg_main_rows(combined, pkg_analysis)
        print(f"  -> Total after expansion: {len(combined)} rows")

    # Sort by RPT_TABLE, then DEPTH
    def sort_key(r):
        d = r.get('DEPTH', '')
        return (r.get('RPT_TABLE', ''), int(d) if d != '' else 999, r.get('DEPENDENT_JOB', ''))
    combined.sort(key=sort_key)

    # ── Enrich with UPSTREAM_JOBS / DOWNSTREAM_JOBS from TIDAL edges ─────
    # Build dep_map from raw TIDAL: job -> set of upstream prerequisites
    dep_map_raw = defaultdict(set)
    for job_name, entries in tidal.items():
        for entry in entries:
            dj = entry.get('DEPENDENT_JOB')
            if dj:
                dep_map_raw[job_name].add(dj)

    # For each RPT table, get the set of jobs and compute edges
    jobs_per_rpt = defaultdict(set)
    for row in combined:
        rpt = row.get('RPT_TABLE', '')
        dj = row.get('DEPENDENT_JOB', '')
        if rpt and dj:
            jobs_per_rpt[rpt].add(dj)

    # Build upstream/downstream maps scoped to each RPT table's jobs
    upstream_map = defaultdict(set)   # (rpt, job) -> set of upstream jobs
    downstream_map = defaultdict(set) # (rpt, job) -> set of downstream jobs
    for rpt, jobs in jobs_per_rpt.items():
        for jn in jobs:
            for dep in dep_map_raw.get(jn, []):
                if dep in jobs:
                    # dep must run before jn -> dep is upstream of jn
                    upstream_map[(rpt, jn)].add(dep)
                    downstream_map[(rpt, dep)].add(jn)

    for row in combined:
        rpt = row.get('RPT_TABLE', '')
        dj = row.get('DEPENDENT_JOB', '')
        ups = sorted(upstream_map.get((rpt, dj), set()))
        downs = sorted(downstream_map.get((rpt, dj), set()))
        row['UPSTREAM_JOBS'] = ' | '.join(ups) if ups else ''
        row['DOWNSTREAM_JOBS'] = ' | '.join(downs) if downs else ''

    # ── Enrich from Gudu + LLM/STM parsed lineage ────────────────────────
    if merged_lookup:
        print("\nEnriching non-DIFW rows from Gudu + LLM/STM parsed lineage...")
        enriched_count = enrich_from_llm(combined, merged_lookup)
        print(f"  -> Enriched {enriched_count} rows with SRC/TGT data")

    # Tag DIFW framework procs (generic infrastructure — not RPT-table-specific)
    difw_tagged = tag_difw_framework_procs(combined)
    if difw_tagged:
        print(f"  -> Tagged {difw_tagged} DIFW_FRAMEWORK rows")

    # Tag delete-preprocess procs (PRC_DEL_EXT_DATA / PRC_UPD_DEL_DATA)
    del_tagged = enrich_delete_preprocess_procs(combined)
    if del_tagged:
        print(f"  -> Tagged {del_tagged} DELETE_PREPROCESS rows (delete-before-insert helpers)")

    # ── Inject virtual VIEW expansion rows for non-Tidal views ──────────
    # Views (e.g. FCT_CLAIM_PAYMENT_DETAIL_R) are not orchestrated by Tidal,
    # so their source-table dependencies are invisible in the standard lineage.
    # We inject a synthetic VIEW node per (RPT_TABLE, view_name) pair so that:
    #   1. The graph correctly shows the view layer and its underlying tables.
    #   2. Hop-reduction detector H (stale intermediate) correctly sees the
    #      underlying tables as consumed (fixes false "dead write" findings).
    print("\nInjecting virtual VIEW expansion rows for non-Tidal views...")
    view_defs = load_view_definitions(Path(args.gudu_lineage))
    if view_defs:
        combined, view_injected = inject_view_expansion_rows(combined, view_defs)
        if view_injected:
            print(f"  -> Total combined after VIEW expansion: {len(combined)} rows")
    else:
        print("  -> No view definitions loaded; skipping VIEW expansion")

    # Apply avg runtime duration
    runtime_lookup = load_runtime_lookup(Path(args.runtime))
    if runtime_lookup:
        filled = apply_runtime_duration(combined, runtime_lookup)
        total_rows = len(combined)
        print(f"  -> Runtime duration: filled {filled}/{total_rows} rows "
              f"({filled/total_rows*100:.1f}% match rate)")

    # Write outputs
    csv_path = output_dir / "combined_lineage_latest.csv"
    xlsx_path = output_dir / "combined_lineage_latest.xlsx"
    gaps_path = output_dir / "lineage_gaps_latest.csv"
    summary_path = output_dir / "lineage_summary_latest.csv"

    write_combined_csv(combined, csv_path)
    print(f"\n  CSV:     {csv_path}")

    write_combined_excel(combined, xlsx_path)
    print(f"  Excel:   {xlsx_path}")

    write_gaps(combined, gaps_path)
    gaps_count = sum(1 for r in combined if r.get('LINEAGE_STATUS') in
                     ('NEEDS_INVESTIGATION', 'DIFW_MISSING_LINEAGE', 'SQL_OBJECT_IDENTIFIED'))
    print(f"  Gaps:    {gaps_path} ({gaps_count} rows)")

    write_summary(combined, summary_path)
    print(f"  Summary: {summary_path}")

    if merged_lookup:
        norm_path = output_dir / "column_level_normalized.csv"
        norm_count = write_column_normalized(combined, merged_lookup, norm_path)
        print(f"  Normalized: {norm_path} ({norm_count} column-mapping rows)")

    print_summary(combined)


if __name__ == "__main__":
    main()
