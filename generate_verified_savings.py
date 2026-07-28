#!/usr/bin/env python3
"""
Verified Savings Simulation Generator
======================================

Reads validated recommendation XLSX files from:
  output/hop_reduction_recommendations_validated/

For each file:
  1. Reads the "Hop Recommendations" sheet (columns located by name, not position).
  2. Filters rows where "Recommendation Status" == "Good Recommendation - Consider".
  3. Runs the savings simulation using those validated recommendations.
  4. Adds (or replaces) a "Verified Savings Simulation" tab in the validated XLSX.

Also writes:
  output/verified_savings_simulation_aggregate.xlsx
    - Sheet "Summary"                  — one row per RPT (original vs verified metrics)
    - Sheet "Consider Recommendations" — all Consider recs across RPTs, sorted by priority
    - Sheet "Status Breakdown"         — status distribution per RPT

Run:
  python generate_verified_savings.py

Re-run safely: the tab in each validated XLSX is replaced on every run;
the aggregate file is fully regenerated.
"""

import json
from collections import defaultdict, deque
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR   = Path(__file__).resolve().parent
OUTPUT_DIR   = SCRIPT_DIR / "output"
VALIDATED_DIR = OUTPUT_DIR / "hop_reduction_recommendations_validated"
GRAPH_JSON   = OUTPUT_DIR / "tidal_dependency_graph.json"
AGGREGATE_XLSX = OUTPUT_DIR / "verified_savings_simulation_aggregate.xlsx"

STATUS_CONSIDER = "Good Recommendation - Consider"

# All four recognised status values (for ordered display)
STATUS_ORDER = [
    "Good Recommendation - Consider",
    "Good Recommendation - Not to consider",
    "Good Recommendation - Not applicable",
    "Not a Good Recommendation",
]

# ── Colour palette (shared) ───────────────────────────────────────────────────
_DARK_BLUE  = PatternFill(start_color="1F4E79", fill_type="solid")
_HL_FILL    = PatternFill(start_color="D6E4F0", fill_type="solid")   # highlight
_GN_FILL    = PatternFill(start_color="C6EFCE", fill_type="solid")   # green – Consider
_SEC_FILL   = PatternFill(start_color="EBF3FB", fill_type="solid")   # light section header
_HDR_FONT   = Font(color="FFFFFF", bold=True, size=11)
_RISK_FILLS = {
    "LOW":    PatternFill(start_color="C6EFCE", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFEB9C", fill_type="solid"),
    "HIGH":   PatternFill(start_color="FFC7CE", fill_type="solid"),
}
_WAVE_FILLS = {
    "WAVE_1":             PatternFill(start_color="C6EFCE", fill_type="solid"),
    "WAVE_2":             PatternFill(start_color="FFEB9C", fill_type="solid"),
    "WAVE_3":             PatternFill(start_color="FFC7CE", fill_type="solid"),
    "GLOBAL_WAVE_1_PLUS": PatternFill(start_color="CFE2F3", fill_type="solid"),
}


# ── Simulation ────────────────────────────────────────────────────────────────

def simulate_verified(recommendations: list, graph: dict) -> dict:
    """
    Run the savings simulation using ALL passed recommendations (no top-N cap).

    Identical removal/reachability logic as analyze_hop_reduction_merged.simulate_after_top_n,
    but without a top-N limit — callers control which recs are passed in.

    Only recs with verified_hop_savings > 0 contribute to job removals.
    D-category (serial chains) keeps the shallowest job; all others keep the first job.
    """
    node_map = {n["id"]: n for n in graph["nodes"]}
    root_set  = {nid for nid, n in node_map.items() if n.get("depth", -1) == 0}

    removed_jobs: set = set()
    for rec in recommendations:
        if rec.get("verified_hop_savings", 0) > 0:
            jobs = rec.get("affected_jobs", [])
            if len(jobs) > 1:
                if rec.get("category", "").startswith("D."):
                    # Serial chain: keep shallowest (last) job
                    candidates = jobs[:-1]
                else:
                    # Keep canonical first job, remove the rest
                    candidates = jobs[1:]
                removed_jobs.update(j for j in candidates if j not in root_set)

    remaining_nodes = {n["id"]: n for n in graph["nodes"] if n["id"] not in removed_jobs}
    remaining_links = [
        lnk for lnk in graph["links"]
        if lnk["source"] in remaining_nodes and lnk["target"] in remaining_nodes
    ]

    # BFS backward from roots to find all nodes still reachable
    reverse_map: dict = defaultdict(set)
    for lnk in remaining_links:
        reverse_map[lnk["target"]].add(lnk["source"])

    reachable: set = set(root_set & remaining_nodes.keys())
    queue = deque(reachable)
    while queue:
        nid = queue.popleft()
        for parent in reverse_map.get(nid, set()):
            if parent not in reachable and parent in remaining_nodes:
                reachable.add(parent)
                queue.append(parent)

    connected_depths = [
        remaining_nodes[nid].get("depth", 0)
        for nid in reachable
        if remaining_nodes[nid].get("depth", -1) >= 0
    ]

    return {
        "original_node_count": len(graph["nodes"]),
        "original_max_depth":  max((n["depth"] for n in graph["nodes"]), default=0),
        "removed_job_count":   len(removed_jobs),
        "new_node_count":      len(remaining_nodes),
        "new_max_depth":       max(connected_depths, default=0),
        "removed_jobs": sorted(removed_jobs),   # full list — no cap
    }


# ── Read validated XLSX ───────────────────────────────────────────────────────

def read_validated_xlsx(path: Path) -> tuple:
    """
    Read the "Hop Recommendations" sheet from a validated XLSX.

    Columns are located by header name (case-sensitive) so the function is
    robust to the data team inserting "Recommendation Status" at any position.

    Returns:
        (rpt_table, all_recs, consider_recs, status_counts)
        Each rec is a minimal dict: keys used by simulate_verified + display fields.
    """
    stem = path.stem   # e.g. hop_reduction_recommendations_RPT_CLAIM_PAYMENT_R
    prefix = "hop_reduction_recommendations_"
    rpt_table = stem[len(prefix):] if stem.startswith(prefix) else stem

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    sheet_name = "Hop Recommendations"
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: '{sheet_name}' not found in {path.name} — skipping")
        wb.close()
        return rpt_table, [], [], {}

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return rpt_table, [], [], {}

    # Build column index from header row
    raw_headers = rows[0]
    col = {str(h).strip(): i for i, h in enumerate(raw_headers) if h is not None}

    required = [
        "Category", "Affected Jobs", "Verified Hop Savings",
        "Estimated Runtime Saved (min)", "Priority Score",
    ]
    missing = [c for c in required if c not in col]
    if missing:
        print(f"  WARNING: Missing required columns in {path.name}: {missing}")

    if "Recommendation Status" not in col:
        print(f"  WARNING: 'Recommendation Status' column not found in {path.name} — "
              f"all rows will be treated as unvalidated")

    def _get(row, col_name, default=None):
        idx = col.get(col_name)
        if idx is None or idx >= len(row):
            return default
        return row[idx]

    all_recs = []
    status_counts: dict = defaultdict(int)

    for row in rows[1:]:
        if all(v is None for v in row):
            continue  # blank row

        status = str(_get(row, "Recommendation Status") or "").strip()
        status_counts[status if status else "(blank)"] += 1

        category = str(_get(row, "Category") or "")
        raw_jobs = str(_get(row, "Affected Jobs") or "")
        affected_jobs = [j.strip() for j in raw_jobs.split("\n") if j.strip()]

        try:
            hop_savings = int(_get(row, "Verified Hop Savings") or 0)
        except (ValueError, TypeError):
            hop_savings = 0

        try:
            rt_saved = float(_get(row, "Estimated Runtime Saved (min)") or 0.0)
        except (ValueError, TypeError):
            rt_saved = 0.0

        try:
            priority = int(_get(row, "Priority Score") or 0)
        except (ValueError, TypeError):
            priority = 0

        all_recs.append({
            "id":                    str(_get(row, "ID") or ""),
            "category":              category,
            "target_table":          str(_get(row, "Target Table") or ""),
            "affected_jobs":         affected_jobs,
            "verified_hop_savings":  hop_savings,
            "est_runtime_saved_minutes": rt_saved,
            "priority_score":        priority,
            "risk":                  str(_get(row, "Risk") or ""),
            "recommendation_type":   str(_get(row, "Recommendation Type") or ""),
            "implementation_decision": str(_get(row, "Implementation Decision") or ""),
            "execution_wave":        str(_get(row, "Execution Wave") or ""),
            "recommendation_status": status,
        })

    consider_recs = [r for r in all_recs if r["recommendation_status"] == STATUS_CONSIDER]
    return rpt_table, all_recs, consider_recs, dict(status_counts)


# ── Write "Verified Savings Simulation" tab into the validated XLSX ───────────

def write_verified_tab(
    path: Path,
    rpt_table: str,
    sim_baseline: dict,
    sim_verified: dict,
    all_recs: list,
    consider_recs: list,
    status_counts: dict,
) -> None:
    """
    Add (or replace) a 'Verified Savings Simulation' sheet in the validated XLSX.

    sim_baseline  — simulation run with the top-10 highest-priority HOP_REDUCTION recs
                    (mirrors what the original script put in "Savings Simulation")
    sim_verified  — simulation run with all "Consider" HOP_REDUCTION recs
    """
    TAB_NAME = "Verified Savings Simulation"

    wb = openpyxl.load_workbook(str(path))
    if TAB_NAME in wb.sheetnames:
        del wb[TAB_NAME]
    ws = wb.create_sheet(TAB_NAME)

    row = 1

    def _section_header(text: str) -> None:
        nonlocal row
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=12, color="1F4E79")
        c.fill = _SEC_FILL
        for ci in range(2, 6):
            ws.cell(row=row, column=ci).fill = _SEC_FILL
        row += 1

    def _kv(key, v_baseline, v_verified, note="", bold=False, highlight=False) -> None:
        nonlocal row
        fill = _HL_FILL if highlight else None
        font = Font(bold=True) if bold else None

        c1 = ws.cell(row=row, column=1, value=key)
        c1.font = Font(bold=True)
        if fill:
            c1.fill = fill

        c2 = ws.cell(row=row, column=2, value=v_baseline)
        if font:
            c2.font = font
        if fill:
            c2.fill = fill

        c3 = ws.cell(row=row, column=3, value=v_verified)
        if font:
            c3.font = font
        if fill:
            c3.fill = fill

        if note:
            ws.cell(row=row, column=4, value=note).font = Font(italic=True, color="595959")
        row += 1

    def _hdr_row(*labels) -> None:
        nonlocal row
        for ci, lbl in enumerate(labels, 1):
            c = ws.cell(row=row, column=ci, value=lbl)
            c.font = _HDR_FONT
            c.fill = _DARK_BLUE
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        row += 1

    # ── 1. File header ────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="Verified Savings Simulation").font = Font(bold=True, size=14, color="1F4E79")
    row += 1
    ws.cell(row=row, column=1, value=f"RPT Table: {rpt_table}").font = Font(bold=True, size=11)
    row += 1
    total = len(all_recs)
    consider_n = len(consider_recs)
    ws.cell(row=row, column=1,
            value=f"{consider_n} of {total} recommendations marked 'Good Recommendation - Consider'").font = Font(italic=True, color="595959")
    row += 2

    # ── 2. Comparison table ───────────────────────────────────────────────────
    _section_header("Savings Comparison: Baseline vs Verified")
    _hdr_row("Metric",
             f"Baseline Simulation\n(Top-10 scored recs)",
             f"Verified Simulation\n(Consider recs only — {consider_n} recs)",
             "Notes")
    ws.row_dimensions[row - 1].height = 32

    hop_recs_all     = sorted(
        [r for r in all_recs if r.get("recommendation_type") == "HOP_REDUCTION"],
        key=lambda r: -r.get("priority_score", 0),
    )
    hop_recs_consider = [r for r in consider_recs if r.get("recommendation_type") == "HOP_REDUCTION"]

    rt_baseline  = round(sum(r["est_runtime_saved_minutes"] for r in hop_recs_all[:10]), 2)
    rt_verified  = round(sum(r["est_runtime_saved_minutes"] for r in hop_recs_consider), 2)
    hop_baseline = sum(r["verified_hop_savings"] for r in hop_recs_all[:10])
    hop_verified = sum(r["verified_hop_savings"] for r in hop_recs_consider)

    depth_saved_b = sim_baseline["original_max_depth"] - sim_baseline["new_max_depth"]
    depth_saved_v = sim_verified["original_max_depth"]  - sim_verified["new_max_depth"]
    node_saved_b  = sim_baseline["original_node_count"] - sim_baseline["new_node_count"]
    node_saved_v  = sim_verified["original_node_count"]  - sim_verified["new_node_count"]

    metrics = [
        ("Total Nodes (original)",         sim_baseline["original_node_count"],  sim_verified["original_node_count"],
         "Total Tidal jobs in the dependency graph"),
        ("Max Depth (original)",            sim_baseline["original_max_depth"],   sim_verified["original_max_depth"],
         "Longest chain / critical path length"),
        ("Recommendations used",            f"Top 10 of {len(hop_recs_all)}", f"All {len(hop_recs_consider)} Consider recs",
         "Recs fed into the simulation"),
        ("Jobs removed by simulation",      sim_baseline["removed_job_count"],    sim_verified["removed_job_count"],
         "Jobs conceptually eliminated"),
        ("New node count after removal",    sim_baseline["new_node_count"],       sim_verified["new_node_count"],
         "Remaining jobs after applying recommendations"),
        ("New max depth after removal",     sim_baseline["new_max_depth"],        sim_verified["new_max_depth"],
         "New critical-path length"),
        ("Depth reduction",                 depth_saved_b,                        depth_saved_v,
         "Sequential hops removed from the critical path (pipeline latency improvement)"),
        ("Node reduction",                  node_saved_b,                         node_saved_v,
         "Jobs eliminated entirely (maintenance and scheduling overhead reduction)"),
        ("Est. runtime saved (min)",        rt_baseline,                          rt_verified,
         "Sum of est_runtime_saved_minutes for HOP_REDUCTION recs in this simulation"),
        ("Total verified hop savings",      hop_baseline,                         hop_verified,
         "Sum of verified_hop_savings (graph/lineage evidence only, excludes LLM)"),
    ]
    highlight_set = {"Depth reduction", "Node reduction", "Est. runtime saved (min)", "Total verified hop savings"}
    for label, v_b, v_v, note in metrics:
        _kv(label, v_b, v_v, note=note, bold=(label in highlight_set), highlight=(label in highlight_set))

    row += 1

    # ── 3. Status breakdown ───────────────────────────────────────────────────
    _section_header("Recommendation Status Breakdown")
    _hdr_row("Status", "Count", "% of Total")

    total_with_status = sum(status_counts.values())
    for s in STATUS_ORDER + [k for k in sorted(status_counts) if k not in STATUS_ORDER]:
        cnt = status_counts.get(s, 0)
        if cnt == 0:
            continue
        pct = f"{cnt / total_with_status * 100:.1f}%" if total_with_status else "—"
        ws.cell(row=row, column=1, value=s)
        ws.cell(row=row, column=2, value=cnt)
        ws.cell(row=row, column=3, value=pct)
        if s == STATUS_CONSIDER:
            for ci in range(1, 4):
                ws.cell(row=row, column=ci).fill = _GN_FILL
        row += 1

    row += 1

    # ── 4. Consider recs detail ───────────────────────────────────────────────
    _section_header(f"Recommendations Included in Verified Simulation ({consider_n})")
    _hdr_row("ID", "Category", "Target Table", "Risk", "Execution Wave",
             "Verified Hop Savings", "Est. Runtime Saved (min)", "Affected Jobs")
    ws.row_dimensions[row - 1].height = 28

    for rec in sorted(consider_recs, key=lambda r: -r.get("priority_score", 0)):
        ws.cell(row=row, column=1, value=rec["id"])
        ws.cell(row=row, column=2, value=rec["category"])
        ws.cell(row=row, column=3, value=rec["target_table"])
        rc = ws.cell(row=row, column=4, value=rec["risk"])
        rc.fill = _RISK_FILLS.get(rec["risk"], PatternFill())
        wave = rec["execution_wave"]
        wc = ws.cell(row=row, column=5, value=wave)
        wc.fill = _WAVE_FILLS.get(wave, PatternFill())
        ws.cell(row=row, column=6, value=rec["verified_hop_savings"])
        ws.cell(row=row, column=7, value=rec["est_runtime_saved_minutes"])
        ws.cell(row=row, column=8, value="\n".join(rec["affected_jobs"]))
        for ci in range(1, 9):
            ws.cell(row=row, column=ci).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # ── 5. All removed jobs (full list for before/after graph building) ────────
    removed_jobs = sim_verified["removed_jobs"]
    if removed_jobs:
        row += 1
        ws.cell(row=row, column=1,
                value=f"Jobs removed in verified simulation ({len(removed_jobs)} total):"
                ).font = Font(bold=True)
        row += 1
        for j in removed_jobs:
            ws.cell(row=row, column=2, value=j)
            row += 1

    # Column widths
    col_widths = {1: 38, 2: 30, 3: 30, 4: 85, 5: 22, 6: 22, 7: 26, 8: 45}
    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(str(path))
    print(f"  Tab '{TAB_NAME}' written → {path.name}")


# ── Write aggregate XLSX ──────────────────────────────────────────────────────

def write_aggregate_xlsx(results: list) -> None:
    """
    Write output/verified_savings_simulation_aggregate.xlsx with three sheets:
      Summary                  — one row per RPT, key metrics side-by-side
      Consider Recommendations — all Consider recs across RPTs, sorted by priority
      Status Breakdown         — per-RPT status distribution
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    sum_headers = [
        "RPT Table",                         # 1
        "Total Recs",                         # 2
        "Consider Recs",                      # 3
        "% Considered",                       # 4
        "Not Good Recs",                      # 5
        "Original Max Depth",                 # 6
        "Baseline New Depth\n(Top-10 all)",   # 7
        "Verified New Depth\n(Consider only)",# 8
        "Verified Depth Saved",               # 9
        "Original Nodes",                     # 10
        "Verified Nodes Remaining",           # 11
        "Verified Nodes Saved",               # 12
        "Baseline RT Saved (min)",            # 13
        "Verified RT Saved (min)",            # 14
        "Verified Hop Savings",               # 15
    ]
    for ci, h in enumerate(sum_headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = _HDR_FONT
        c.fill = _DARK_BLUE
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 42

    for ri, r in enumerate(sorted(results, key=lambda x: x["rpt_table"]), 2):
        sb = r["sim_baseline"]
        sv = r["sim_verified"]
        total   = r["total_recs"]
        consider = r["consider_count"]
        not_good = r["status_counts"].get("Not a Good Recommendation", 0)
        pct = f"{consider / total * 100:.1f}%" if total else "—"

        depth_saved_v = sb["original_max_depth"] - sv["new_max_depth"]
        node_saved_v  = sb["original_node_count"] - sv["new_node_count"]

        ws.cell(row=ri, column=1, value=r["rpt_table"])
        ws.cell(row=ri, column=2, value=total)
        wc = ws.cell(row=ri, column=3, value=consider)
        wc.fill = _GN_FILL
        ws.cell(row=ri, column=4, value=pct)
        ws.cell(row=ri, column=5, value=not_good)
        ws.cell(row=ri, column=6, value=sb["original_max_depth"])
        ws.cell(row=ri, column=7, value=sb["new_max_depth"])
        ws.cell(row=ri, column=8, value=sv["new_max_depth"])
        dc = ws.cell(row=ri, column=9, value=depth_saved_v)
        dc.fill = _HL_FILL
        dc.font = Font(bold=True)
        ws.cell(row=ri, column=10, value=sb["original_node_count"])
        ws.cell(row=ri, column=11, value=sv["new_node_count"])
        nc = ws.cell(row=ri, column=12, value=node_saved_v)
        nc.fill = _HL_FILL
        nc.font = Font(bold=True)
        ws.cell(row=ri, column=13, value=r["rt_saved_baseline"])
        vrc = ws.cell(row=ri, column=14, value=r["rt_saved_verified"])
        vrc.fill = _HL_FILL
        vrc.font = Font(bold=True)
        ws.cell(row=ri, column=15, value=r["hop_savings_verified"])

    # Aggregate totals row
    if results:
        tr = len(results) + 2
        agg_vals = {
            1:  "TOTAL / AGGREGATE",
            2:  sum(r["total_recs"]      for r in results),
            3:  sum(r["consider_count"]  for r in results),
            5:  sum(r["status_counts"].get("Not a Good Recommendation", 0) for r in results),
            9:  sum(r["sim_baseline"]["original_max_depth"] - r["sim_verified"]["new_max_depth"] for r in results),
            12: sum(r["sim_baseline"]["original_node_count"] - r["sim_verified"]["new_node_count"] for r in results),
            13: round(sum(r["rt_saved_baseline"] for r in results), 2),
            14: round(sum(r["rt_saved_verified"] for r in results), 2),
            15: sum(r["hop_savings_verified"] for r in results),
        }
        for ci, val in agg_vals.items():
            c = ws.cell(row=tr, column=ci, value=val)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = _DARK_BLUE

    col_widths = [32, 12, 14, 12, 14, 18, 20, 22, 18, 16, 22, 18, 22, 24, 20]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(sum_headers))}{len(results) + 1}"

    # ── Sheet 2: All Consider Recommendations ─────────────────────────────────
    ws2 = wb.create_sheet("Consider Recommendations")
    cr_headers = [
        "RPT Table", "ID", "Category", "Target Table", "Risk",
        "Execution Wave", "Verified Hop Savings", "Est. Runtime Saved (min)", "Priority Score",
    ]
    for ci, h in enumerate(cr_headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = _HDR_FONT
        c.fill = _DARK_BLUE
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    all_consider = [
        (r["rpt_table"], rec)
        for r in results
        for rec in r["consider_recs"]
    ]
    all_consider.sort(key=lambda x: -x[1].get("priority_score", 0))

    for ri, (rpt, rec) in enumerate(all_consider, 2):
        ws2.cell(row=ri, column=1, value=rpt)
        ws2.cell(row=ri, column=2, value=rec["id"])
        ws2.cell(row=ri, column=3, value=rec["category"])
        ws2.cell(row=ri, column=4, value=rec["target_table"])
        rc = ws2.cell(row=ri, column=5, value=rec["risk"])
        rc.fill = _RISK_FILLS.get(rec["risk"], PatternFill())
        wave = rec["execution_wave"]
        wc = ws2.cell(row=ri, column=6, value=wave)
        wc.fill = _WAVE_FILLS.get(wave, PatternFill())
        ws2.cell(row=ri, column=7, value=rec["verified_hop_savings"])
        ws2.cell(row=ri, column=8, value=rec["est_runtime_saved_minutes"])
        ws2.cell(row=ri, column=9, value=rec["priority_score"])
        for ci in range(1, 10):
            ws2.cell(row=ri, column=ci).alignment = Alignment(wrap_text=True, vertical="top")

    for ci, w in enumerate([32, 8, 42, 35, 10, 22, 20, 24, 14], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cr_headers))}{len(all_consider) + 1}"

    # ── Sheet 3: Status Breakdown ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Status Breakdown")
    sb_headers = ["RPT Table"] + STATUS_ORDER + ["(blank / unset)", "Total Recs"]
    for ci, h in enumerate(sb_headers, 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.font = _HDR_FONT
        c.fill = _DARK_BLUE
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws3.row_dimensions[1].height = 48

    for ri, r in enumerate(sorted(results, key=lambda x: x["rpt_table"]), 2):
        sc = r["status_counts"]
        ws3.cell(row=ri, column=1, value=r["rpt_table"])
        vals = [sc.get(s, 0) for s in STATUS_ORDER] + [sc.get("(blank)", 0), r["total_recs"]]
        for ci, v in enumerate(vals, 2):
            c = ws3.cell(row=ri, column=ci, value=v)
            if ci == 2:   # Consider column
                c.fill = _GN_FILL

    for ci, w in enumerate([32, 32, 34, 32, 28, 16, 10], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    wb.save(str(AGGREGATE_XLSX))
    print(f"Aggregate report written: {AGGREGATE_XLSX}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    print("=" * 70)
    print("Verified Savings Simulation Generator")
    print("=" * 70)
    print(f"Validated dir : {VALIDATED_DIR}")
    print(f"Graph JSON    : {GRAPH_JSON}")
    print(f"Aggregate out : {AGGREGATE_XLSX}\n")

    if not VALIDATED_DIR.exists():
        print(f"ERROR: Validated directory not found:\n  {VALIDATED_DIR}")
        print("Create it and place validated XLSX files there, then re-run.")
        return

    xlsx_files = sorted(VALIDATED_DIR.glob("hop_reduction_recommendations_*.xlsx"))
    if not xlsx_files:
        print("No validated XLSX files found in that directory.")
        return

    print(f"Found {len(xlsx_files)} validated file(s):")
    for f in xlsx_files:
        print(f"  {f.name}")
    print()

    if not GRAPH_JSON.exists():
        print(f"ERROR: Graph JSON not found: {GRAPH_JSON}")
        return

    print("Loading dependency graph...")
    with open(GRAPH_JSON, "r", encoding="utf-8") as fh:
        all_graphs = json.load(fh)
    print(f"  {len(all_graphs)} RPT graph(s) loaded\n")

    results = []

    for xlsx_path in xlsx_files:
        print(f"{'─' * 60}")
        print(f"Processing: {xlsx_path.name}")

        rpt_table, all_recs, consider_recs, status_counts = read_validated_xlsx(xlsx_path)

        if rpt_table not in all_graphs:
            print(f"  WARNING: '{rpt_table}' not found in graph JSON — skipping")
            continue

        graph = all_graphs[rpt_table]

        # Filter to HOP_REDUCTION recs (PLSQL_OPTIMISATION_REVIEW recs have
        # verified_hop_savings=0 and don't affect the graph simulation)
        hop_recs_all = sorted(
            [r for r in all_recs if r.get("recommendation_type") == "HOP_REDUCTION"],
            key=lambda r: -r.get("priority_score", 0),
        )
        hop_recs_consider = [
            r for r in consider_recs
            if r.get("recommendation_type") == "HOP_REDUCTION"
        ]

        # Baseline: top-10 scored recs (mirrors original script behaviour)
        sim_baseline = simulate_verified(hop_recs_all[:10], graph)
        # Verified:  all validated "Consider" recs (no cap)
        sim_verified = simulate_verified(hop_recs_consider, graph)

        rt_saved_baseline = round(sum(r["est_runtime_saved_minutes"] for r in hop_recs_all[:10]), 2)
        rt_saved_verified  = round(sum(r["est_runtime_saved_minutes"] for r in hop_recs_consider), 2)
        hop_savings_verified = sum(r["verified_hop_savings"] for r in hop_recs_consider)

        print(f"  RPT              : {rpt_table}")
        print(f"  Total recs       : {len(all_recs)}  |  HOP_REDUCTION: {len(hop_recs_all)}")
        print(f"  Consider recs    : {len(consider_recs)}  |  HOP_REDUCTION Consider: {len(hop_recs_consider)}")
        print(f"  Status breakdown : {dict(status_counts)}")
        print(f"  Depth (baseline) : {sim_baseline['original_max_depth']} → {sim_baseline['new_max_depth']}  "
              f"(saved {sim_baseline['original_max_depth'] - sim_baseline['new_max_depth']})")
        print(f"  Depth (verified) : {sim_verified['original_max_depth']} → {sim_verified['new_max_depth']}  "
              f"(saved {sim_verified['original_max_depth'] - sim_verified['new_max_depth']})")
        print(f"  RT saved (min)   : baseline={rt_saved_baseline}  verified={rt_saved_verified}")

        write_verified_tab(
            xlsx_path, rpt_table,
            sim_baseline, sim_verified,
            all_recs, consider_recs, status_counts,
        )

        results.append({
            "rpt_table":          rpt_table,
            "total_recs":         len(all_recs),
            "consider_count":     len(consider_recs),
            "consider_recs":      consider_recs,
            "status_counts":      status_counts,
            "sim_baseline":       sim_baseline,
            "sim_verified":       sim_verified,
            "rt_saved_baseline":  rt_saved_baseline,
            "rt_saved_verified":  rt_saved_verified,
            "hop_savings_verified": hop_savings_verified,
        })
        print()

    if not results:
        print("No RPTs successfully processed — aggregate file not written.")
        return

    print("Writing aggregate report...")
    write_aggregate_xlsx(results)

    # ── Console summary ──────────────────────────────────────────────────────
    print(f"\n{'=' * 70}")
    print(f"SUMMARY  ({len(results)} RPT(s) validated)")
    print(f"{'=' * 70}")
    hdr = f"{'RPT Table':<35} {'Total':>6} {'Consid':>7} {'DepthSaved':>11} {'NodeSaved':>10} {'RT Saved (min)':>15}"
    print(hdr)
    print("-" * len(hdr))
    for r in sorted(results, key=lambda x: x["rpt_table"]):
        sb = r["sim_baseline"]
        sv = r["sim_verified"]
        print(
            f"{r['rpt_table']:<35} {r['total_recs']:>6} {r['consider_count']:>7} "
            f"{sb['original_max_depth'] - sv['new_max_depth']:>11} "
            f"{sb['original_node_count'] - sv['new_node_count']:>10} "
            f"{r['rt_saved_verified']:>15.2f}"
        )
    print()
    print(f"Aggregate file: {AGGREGATE_XLSX}")


if __name__ == "__main__":
    main()
