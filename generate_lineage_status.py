"""
generate_lineage_status.py
--------------------------
Produces a post-run status report for the Gudu lineage pipeline.

Scans the three pipeline directories and classifies every original procedure
into one of four categories:

  COMPLETE        – single cleaned file ≤10K, Gudu produced JSON
  PARTIAL_SPLIT   – procedure was split into parts; at least one part has
                    a JSON but at least one part is missing (>10K or no JSON)
  SPLIT_ALL_OK    – all parts produced JSON (rare — split but all ≤10K)
  SKIPPED_NO_JSON – cleaned file(s) exist but NO JSON was produced at all
  NO_CLEANED      – procedure file in Stage-1 dir but produced no cleaned output

Usage
-----
  python generate_lineage_status.py [options]

Options
  --split-procs  PATH   Stage-1 dir  (default: Rest_Metadata_7RPT_SplitProcs)
  --cleaned      PATH   Stage-2 dir  (default: Rest_Metadata_7RPT_SplitProcs_cleaned)
  --output       PATH   Excel output (default: lineage_pipeline_status.xlsx)
  --csv                 Also write a CSV alongside the Excel
"""

import argparse
import glob
import json
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Category definitions ─────────────────────────────────────────────────────
CAT_COMPLETE        = "COMPLETE"
CAT_SPLIT_ALL_OK    = "SPLIT – All Parts OK"
CAT_PARTIAL_SPLIT   = "PARTIAL – Some Parts Skipped"
CAT_SKIPPED_NO_JSON = "SKIPPED – >10K, No JSON"
CAT_NO_CLEANED      = "NO CLEANED OUTPUT"

CAT_COLORS = {
    CAT_COMPLETE:        "E2EFDA",  # green
    CAT_SPLIT_ALL_OK:    "D9EAD3",  # light green
    CAT_PARTIAL_SPLIT:   "FFF2CC",  # yellow
    CAT_SKIPPED_NO_JSON: "FCE4D6",  # orange/red
    CAT_NO_CLEANED:      "F4CCCC",  # red
}

RECOMMENDED_ACTION = {
    CAT_COMPLETE:        "None – lineage complete",
    CAT_SPLIT_ALL_OK:    "Review – note cross-part CTE loss",
    CAT_PARTIAL_SPLIT:   "Manual review or further split at UNION ALL / CTE boundaries",
    CAT_SKIPPED_NO_JSON: "Manual review – needs different parsing strategy",
    CAT_NO_CLEANED:      "Check if procedure has any DML; may be utility/logging only",
}

# ── Helpers ──────────────────────────────────────────────────────────────────

def get_char_count(path: str) -> int:
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            return len(f.read())
    except Exception:
        return -1


def count_json_rows(json_path: str) -> int:
    """Count relationship rows in a Gudu lineage JSON (best-effort)."""
    try:
        with open(json_path, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if line.startswith("{"):
                    data = json.loads(line)
                    break
            else:
                return 0
        # Support both flat and wrapped Gudu formats
        if "data" in data and "sqlflow" in data.get("data", {}):
            rels = data["data"]["sqlflow"].get("relationships", [])
        else:
            rels = data.get("relationships", [])
        return len(rels)
    except Exception:
        return -1


def extract_proc_name_from_filename(filename: str) -> str:
    """
    Strip _partN_cleaned, _cleaned, _partN suffixes from a filename base
    to get the canonical procedure name.
    e.g. PKG_GRP_LOAD_X_PRC_GET_CUR_DATA_part3_cleaned -> PKG_GRP_LOAD_X_PRC_GET_CUR_DATA
    """
    name = os.path.splitext(filename)[0]
    name = re.sub(r'_cleaned$', '', name)
    name = re.sub(r'_part\d+$', '', name)
    name = re.sub(r'_lineage$', '', name)
    return name


def extract_package_proc(proc_name: str):
    """Split PKG_X_PRC_Y into (PKG_X, PRC_Y). Falls back gracefully."""
    m = re.search(r'_(PRC_[A-Z0-9_]+)$', proc_name, re.IGNORECASE)
    if m:
        proc = m.group(1)
        pkg  = proc_name[:-(len(proc) + 1)]
        return pkg, proc
    # Some procedures use MAIN
    m = re.search(r'_(MAIN[A-Z0-9_]*)$', proc_name, re.IGNORECASE)
    if m:
        proc = m.group(1)
        pkg  = proc_name[:-(len(proc) + 1)]
        return pkg, proc
    return proc_name, ""


# ── Core analysis ─────────────────────────────────────────────────────────────

def analyse(split_procs_dir: str, cleaned_dir: str):
    """
    Returns a list of dicts, one per original procedure, with full status info.
    """
    # ── Stage 1: all original procedure files ────────────────────────────────
    stage1_files = sorted(glob.glob(os.path.join(split_procs_dir, "*.sql")))
    stage1_procs = {}  # proc_name -> stage1_path
    for fp in stage1_files:
        base = os.path.splitext(os.path.basename(fp))[0]
        stage1_procs[base] = fp

    # ── Stage 2 / 3: all cleaned SQL + JSON files ────────────────────────────
    cleaned_sql   = sorted(glob.glob(os.path.join(cleaned_dir, "*_cleaned.sql")))
    json_files    = sorted(glob.glob(os.path.join(cleaned_dir, "*_lineage.json")))

    # Build proc_name → list of (part_label, cleaned_path, json_path_or_None, chars)
    proc_parts = {}   # canonical_proc_name -> list of part_info dicts

    for sql_path in cleaned_sql:
        fname   = os.path.basename(sql_path)
        proc_nm = extract_proc_name_from_filename(fname)
        chars   = get_char_count(sql_path)

        # Determine part label
        m = re.search(r'_part(\d+)_cleaned\.sql$', fname, re.IGNORECASE)
        part_label = f"part{m.group(1)}" if m else "single"

        # Look for corresponding JSON
        base_no_ext = os.path.splitext(sql_path)[0]
        json_path   = base_no_ext + "_lineage.json"
        json_exists = os.path.exists(json_path)
        json_rows   = count_json_rows(json_path) if json_exists else None

        entry = {
            "part_label":  part_label,
            "sql_path":    sql_path,
            "chars":       chars,
            "gudu_sent":   chars <= 10000,
            "json_exists": json_exists,
            "json_path":   json_path if json_exists else None,
            "json_rows":   json_rows,
        }

        proc_parts.setdefault(proc_nm, []).append(entry)

    # ── Merge stage1 procs that produced nothing (no cleaned output) ──────────
    all_proc_names = set(stage1_procs.keys()) | set(proc_parts.keys())

    rows = []
    for proc_name in sorted(all_proc_names):
        pkg, proc = extract_package_proc(proc_name)
        s1_path   = stage1_procs.get(proc_name)
        parts     = proc_parts.get(proc_name, [])

        if not parts:
            # Stage-1 file exists but no cleaned output
            rows.append({
                "package":          pkg,
                "procedure":        proc,
                "proc_name":        proc_name,
                "stage1_file":      os.path.basename(s1_path) if s1_path else "–",
                "is_split":         False,
                "total_parts":      0,
                "parts_sent_gudu":  0,
                "parts_with_json":  0,
                "parts_skipped":    0,
                "total_sql_chars":  get_char_count(s1_path) if s1_path else 0,
                "max_part_chars":   0,
                "total_json_rows":  0,
                "category":         CAT_NO_CLEANED,
                "parts_detail":     [],
            })
            continue

        is_split        = any(p["part_label"] != "single" for p in parts)
        parts_sent      = sum(1 for p in parts if p["gudu_sent"])
        parts_with_json = sum(1 for p in parts if p["json_exists"])
        parts_skipped   = len(parts) - parts_sent
        total_rows      = sum(p["json_rows"] or 0 for p in parts if p["json_rows"] is not None)
        max_chars       = max(p["chars"] for p in parts)
        total_chars_sum = sum(p["chars"] for p in parts)

        # Categorise
        if not is_split and parts_with_json == 1:
            cat = CAT_COMPLETE
        elif is_split and parts_skipped == 0 and parts_with_json == len(parts):
            cat = CAT_SPLIT_ALL_OK
        elif is_split and parts_with_json > 0 and parts_skipped > 0:
            cat = CAT_PARTIAL_SPLIT
        elif parts_with_json == 0 and parts_sent > 0:
            # Sent to Gudu but no JSON produced (parse failure)
            cat = CAT_SKIPPED_NO_JSON
        elif parts_sent == 0:
            cat = CAT_SKIPPED_NO_JSON
        else:
            cat = CAT_PARTIAL_SPLIT

        rows.append({
            "package":          pkg,
            "procedure":        proc,
            "proc_name":        proc_name,
            "stage1_file":      os.path.basename(s1_path) if s1_path else "–",
            "is_split":         is_split,
            "total_parts":      len(parts),
            "parts_sent_gudu":  parts_sent,
            "parts_with_json":  parts_with_json,
            "parts_skipped":    parts_skipped,
            "total_sql_chars":  total_chars_sum,
            "max_part_chars":   max_chars,
            "total_json_rows":  total_rows,
            "category":         cat,
            "parts_detail":     parts,
        })

    return rows


# ── Non-PKG analysis (standalone PRC + MView files) ─────────────────────────

def detect_object_type(fname: str) -> str:
    """Guess the object type from the filename prefix."""
    upper = fname.upper()
    if upper.startswith("PRC_"):
        return "Standalone Procedure"
    if any(upper.startswith(p) for p in ("DIM_", "FCT_", "MVW_", "RPT_", "VW_")):
        return "Materialized View / View"
    return "SQL Object"


def analyse_nonpkg(src_dir: str, cleaned_dir: str) -> list:
    """
    Analyse non-PKG files (standalone PRC + MViews).
    'src_dir'     = Rest_Metadata_7RPT  (original files, PKG excluded)
    'cleaned_dir' = Rest_Metadata_7RPT_NonPKG_cleaned
    """
    # Stage-1: all non-PKG source files
    src_files = sorted(
        f for f in glob.glob(os.path.join(src_dir, "*.sql"))
        if not os.path.basename(f).upper().startswith("PKG_")
    )
    src_map = {os.path.splitext(os.path.basename(f))[0]: f for f in src_files}

    # Stage-2/3: cleaned SQL files
    cleaned_sql = sorted(glob.glob(os.path.join(cleaned_dir, "*_cleaned.sql")))

    # Build obj_name → list of part_info dicts  (same logic as analyse())
    proc_parts: dict = {}
    for sql_path in cleaned_sql:
        fname   = os.path.basename(sql_path)
        obj_nm  = extract_proc_name_from_filename(fname)
        chars   = get_char_count(sql_path)
        m       = re.search(r'_part(\d+)_cleaned\.sql$', fname, re.IGNORECASE)
        part_label = f"part{m.group(1)}" if m else "single"

        base_no_ext = os.path.splitext(sql_path)[0]
        json_path   = base_no_ext + "_lineage.json"
        json_exists = os.path.exists(json_path)
        json_rows   = count_json_rows(json_path) if json_exists else None

        entry = {
            "part_label":  part_label,
            "sql_path":    sql_path,
            "chars":       chars,
            "gudu_sent":   chars <= 10000,
            "json_exists": json_exists,
            "json_path":   json_path if json_exists else None,
            "json_rows":   json_rows,
        }
        proc_parts.setdefault(obj_nm, []).append(entry)

    all_obj_names = set(src_map.keys()) | set(proc_parts.keys())
    rows = []

    for obj_name in sorted(all_obj_names):
        s1_path    = src_map.get(obj_name)
        parts      = proc_parts.get(obj_name, [])
        obj_type   = detect_object_type(obj_name)

        if not parts:
            rows.append({
                "package":          "",
                "procedure":        obj_name,
                "proc_name":        obj_name,
                "stage1_file":      os.path.basename(s1_path) if s1_path else "–",
                "object_type":      obj_type,
                "is_split":         False,
                "total_parts":      0,
                "parts_sent_gudu":  0,
                "parts_with_json":  0,
                "parts_skipped":    0,
                "total_sql_chars":  get_char_count(s1_path) if s1_path else 0,
                "max_part_chars":   0,
                "total_json_rows":  0,
                "category":         CAT_NO_CLEANED,
                "parts_detail":     [],
            })
            continue

        is_split        = any(p["part_label"] != "single" for p in parts)
        parts_sent      = sum(1 for p in parts if p["gudu_sent"])
        parts_with_json = sum(1 for p in parts if p["json_exists"])
        parts_skipped   = len(parts) - parts_sent
        total_rows      = sum(p["json_rows"] or 0 for p in parts if p["json_rows"] is not None)
        max_chars       = max(p["chars"] for p in parts)
        total_chars_sum = sum(p["chars"] for p in parts)

        if not is_split and parts_with_json == 1:
            cat = CAT_COMPLETE
        elif is_split and parts_skipped == 0 and parts_with_json == len(parts):
            cat = CAT_SPLIT_ALL_OK
        elif is_split and parts_with_json > 0 and parts_skipped > 0:
            cat = CAT_PARTIAL_SPLIT
        elif parts_with_json == 0 and parts_sent > 0:
            cat = CAT_SKIPPED_NO_JSON
        elif parts_sent == 0:
            cat = CAT_SKIPPED_NO_JSON
        else:
            cat = CAT_PARTIAL_SPLIT

        rows.append({
            "package":          "",
            "procedure":        obj_name,
            "proc_name":        obj_name,
            "stage1_file":      os.path.basename(s1_path) if s1_path else "–",
            "object_type":      obj_type,
            "is_split":         is_split,
            "total_parts":      len(parts),
            "parts_sent_gudu":  parts_sent,
            "parts_with_json":  parts_with_json,
            "parts_skipped":    parts_skipped,
            "total_sql_chars":  total_chars_sum,
            "max_part_chars":   max_chars,
            "total_json_rows":  total_rows,
            "category":         cat,
            "parts_detail":     parts,
        })

    return rows


# ── Summary counters ──────────────────────────────────────────────────────────

def build_summary(rows: list) -> dict:
    from collections import Counter
    cat_counts = Counter(r["category"] for r in rows)
    total      = len(rows)
    sent       = sum(r["parts_sent_gudu"] for r in rows)
    total_parts= sum(r["total_parts"] for r in rows)
    with_json  = sum(r["parts_with_json"] for r in rows)
    skipped    = sum(r["parts_skipped"] for r in rows)
    total_rows = sum(r["total_json_rows"] for r in rows)

    return {
        "total_procedures":   total,
        "total_cleaned_parts":total_parts,
        "parts_sent_gudu":    sent,
        "parts_with_json":    with_json,
        "parts_skipped":      skipped,
        "total_lineage_rows": total_rows,
        "by_category":        dict(cat_counts),
    }


# ── Excel writer ──────────────────────────────────────────────────────────────

def write_excel(rows: list, summary: dict, output_path: str):
    if not HAS_OPENPYXL:
        print("openpyxl not available — skipping Excel output")
        return

    wb   = openpyxl.Workbook()
    thin = Side(style="thin", color="B8CCE4")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    ctr  = Alignment(horizontal="center", vertical="top", wrap_text=True)
    hdr_fill = PatternFill("solid", fgColor="1A3C5A")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)

    def hcell(ws, row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = ctr; c.border = bdr
        return c

    def dcell(ws, row, col, val, fill=None, bold=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(size=9, bold=bold)
        c.alignment = wrap; c.border = bdr
        if fill:
            c.fill = fill
        return c

    # ── Sheet 1: Per-Procedure Status ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Procedure Status"

    h1 = [
        "Pipeline", "Object Type", "Package", "Procedure / Object Name", "Stage-1 File",
        "Split?", "Total Parts", "Parts Sent\nto Gudu",
        "Parts with\nJSON", "Parts\nSkipped (>10K)",
        "Max Part\nSize (chars)", "Total Lineage\nRows in JSON",
        "Category", "Recommended Action"
    ]
    widths1 = [14, 26, 38, 48, 52, 9, 11, 12, 12, 16, 14, 14, 28, 50]

    for col, (h, w) in enumerate(zip(h1, widths1), 1):
        hcell(ws1, 1, col, h)
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[1].height = 36
    ws1.freeze_panes = "A2"

    for ri, r in enumerate(rows, 2):
        cat       = r["category"]
        cat_color = CAT_COLORS.get(cat, "FFFFFF")
        cat_fill  = PatternFill("solid", fgColor=cat_color)
        row_fill  = PatternFill("solid", fgColor="EBF3FB") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        pipeline  = r.get("pipeline", "PKG")
        obj_type  = r.get("object_type", "Procedure (from PKG)")

        dcell(ws1, ri,  1, pipeline,     row_fill)
        dcell(ws1, ri,  2, obj_type,     row_fill)
        dcell(ws1, ri,  3, r["package"],        row_fill)
        dcell(ws1, ri,  4, r["procedure"],       row_fill)
        dcell(ws1, ri,  5, r["stage1_file"],     row_fill)
        dcell(ws1, ri,  6, "YES" if r["is_split"] else "NO",
              PatternFill("solid", fgColor="FFF2CC") if r["is_split"] else
              PatternFill("solid", fgColor="E2EFDA"))
        dcell(ws1, ri,  7, r["total_parts"] or "–", row_fill)
        dcell(ws1, ri,  8, r["parts_sent_gudu"],  row_fill)
        dcell(ws1, ri,  9, r["parts_with_json"],  row_fill)
        dcell(ws1, ri, 10, r["parts_skipped"] if r["parts_skipped"] else "–",
              PatternFill("solid", fgColor="FCE4D6") if r["parts_skipped"] else row_fill)
        dcell(ws1, ri, 11, r["max_part_chars"] if r["max_part_chars"] else "–", row_fill)
        dcell(ws1, ri, 12, r["total_json_rows"] if r["total_json_rows"] else "–", row_fill)
        dcell(ws1, ri, 13, cat, cat_fill, bold=True)
        dcell(ws1, ri, 14, RECOMMENDED_ACTION.get(cat, "–"), row_fill)
        ws1.row_dimensions[ri].height = 28

    # ── Sheet 2: Part-Level Detail ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Part-Level Detail")

    h2 = [
        "Pipeline", "Object Type", "Package", "Procedure / Object Name", "Part Label",
        "Cleaned SQL File", "Size (chars)", "≤10K?",
        "Sent to Gudu?", "JSON Exists?", "JSON Rows",
        "Status"
    ]
    widths2 = [14, 26, 38, 48, 10, 60, 14, 9, 13, 12, 10, 28]
    for col, (h, w) in enumerate(zip(h2, widths2), 1):
        hcell(ws2, 1, col, h)
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 30
    ws2.freeze_panes = "A2"

    ri2 = 2
    for r in rows:
        for part in r["parts_detail"]:
            ok_fill  = PatternFill("solid", fgColor="E2EFDA")
            bad_fill = PatternFill("solid", fgColor="FCE4D6")
            row_fill = PatternFill("solid", fgColor="EBF3FB") if ri2 % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

            chars    = part["chars"]
            sent     = part["gudu_sent"]
            has_json = part["json_exists"]

            if has_json:
                status = "✓ JSON generated"
                s_fill = ok_fill
            elif not sent:
                status = "✗ Skipped — >10K chars"
                s_fill = bad_fill
            else:
                status = "✗ Gudu failed — no JSON"
                s_fill = bad_fill

            dcell(ws2, ri2,  1, r.get("pipeline", "PKG"),          row_fill)
            dcell(ws2, ri2,  2, r.get("object_type", "Procedure (from PKG)"), row_fill)
            dcell(ws2, ri2,  3, r["package"],                       row_fill)
            dcell(ws2, ri2,  4, r["procedure"],                     row_fill)
            dcell(ws2, ri2,  5, part["part_label"],                 row_fill)
            dcell(ws2, ri2,  6, os.path.basename(part["sql_path"]), row_fill)
            dcell(ws2, ri2,  7, chars,                              row_fill)
            dcell(ws2, ri2,  8, "YES" if chars <= 10000 else "NO",
                  ok_fill if chars <= 10000 else bad_fill)
            dcell(ws2, ri2,  9, "YES" if sent     else "NO",      ok_fill if sent     else bad_fill)
            dcell(ws2, ri2, 10, "YES" if has_json else "NO",      ok_fill if has_json else bad_fill)
            dcell(ws2, ri2, 11, part["json_rows"] if part["json_rows"] is not None else "–", row_fill)
            dcell(ws2, ri2, 12, status, s_fill, bold=True)
            ws2.row_dimensions[ri2].height = 20
            ri2 += 1

    # ── Sheet 3: Summary ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 18

    title_fill  = PatternFill("solid", fgColor="1A3C5A")
    sec_fill    = PatternFill("solid", fgColor="2E75B6")
    val_font    = Font(size=11, bold=True)

    def sec(row, label):
        c = ws3.cell(row=row, column=1, value=label)
        c.fill = sec_fill; c.font = Font(bold=True, color="FFFFFF", size=10)
        c.border = bdr; c.alignment = wrap
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws3.row_dimensions[row].height = 22

    def kv(ws, row, label, value, fill=None):
        c1 = ws.cell(row=row, column=1, value=label)
        c1.border = bdr; c1.alignment = wrap; c1.font = Font(size=10)
        c2 = ws.cell(row=row, column=2, value=value)
        c2.border = bdr; c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.font = val_font
        if fill:
            c1.fill = fill; c2.fill = fill
        ws.row_dimensions[row].height = 22

    r = 1
    t = ws3.cell(row=r, column=1, value="Gudu Lineage Pipeline — Run Status Report")
    t.fill = title_fill; t.font = Font(bold=True, color="FFFFFF", size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws3.row_dimensions[r].height = 32

    r += 1; sec(r, "Overall Procedure Counts")
    r += 1; kv(ws3, r, "Total original procedures (Stage 1)",          summary["total_procedures"])
    r += 1; kv(ws3, r, "Procedures with cleaned SQL output (Stage 2)", summary["total_procedures"] - sum(1 for row in rows if row["category"] == CAT_NO_CLEANED))
    r += 1; kv(ws3, r, "Procedures with NO cleaned output",            sum(1 for row in rows if row["category"] == CAT_NO_CLEANED),
               PatternFill("solid", fgColor="F4CCCC"))

    r += 1; sec(r, "Part-File Counts (Stage 2 → Stage 3)")
    r += 1; kv(ws3, r, "Total cleaned SQL part-files generated",       summary["total_cleaned_parts"])
    r += 1; kv(ws3, r, "Part-files sent to Gudu (≤10K chars)",         summary["parts_sent_gudu"],
               PatternFill("solid", fgColor="E2EFDA"))
    r += 1; kv(ws3, r, "Part-files skipped — >10K (Gudu limit)",       summary["parts_skipped"],
               PatternFill("solid", fgColor="FCE4D6"))
    r += 1; kv(ws3, r, "Part-files with JSON output generated",        summary["parts_with_json"],
               PatternFill("solid", fgColor="E2EFDA"))

    r += 1; sec(r, "Procedure-Level Categories")
    for cat, count in sorted(summary["by_category"].items(), key=lambda x: -x[1]):
        r += 1
        kv(ws3, r, cat, count, PatternFill("solid", fgColor=CAT_COLORS.get(cat, "FFFFFF")))

    r += 1; sec(r, "Output Quality")
    r += 1; kv(ws3, r, "Total lineage relationship rows in all JSONs",  summary["total_lineage_rows"])
    complete_count = summary["by_category"].get(CAT_COMPLETE, 0)
    total_proc     = summary["total_procedures"]
    pct = f"{complete_count / total_proc * 100:.1f}%" if total_proc else "–"
    r += 1; kv(ws3, r, "% procedures with COMPLETE lineage",            pct,
               PatternFill("solid", fgColor="E2EFDA"))

    r += 1; sec(r, "Notes")
    notes = [
        ("COMPLETE",              "Single cleaned file ≤10K chars; Gudu produced JSON. Full lineage available."),
        ("SPLIT – All Parts OK",  "Procedure split into parts; all parts produced JSON. Lineage is mostly complete but cross-part CTE/subquery references may be missing."),
        ("PARTIAL – Some Parts Skipped", "Some parts exceeded 10K chars and were skipped. Lineage is INCOMPLETE. Recommend manual review or sub-splitting at UNION ALL / CTE boundaries."),
        ("SKIPPED – >10K, No JSON",      "All part-files exceeded 10K or Gudu failed. NO lineage captured. Needs different parsing strategy."),
        ("NO CLEANED OUTPUT",     "No DML was found after cleaning. Procedure may be utility-only (logging, index rebuild etc.)."),
    ]
    for label, note in notes:
        r += 1
        c1 = ws3.cell(row=r, column=1, value=label)
        c1.font = Font(bold=True, size=9)
        c1.fill = PatternFill("solid", fgColor=CAT_COLORS.get(label, "FFFFFF"))
        c1.border = bdr; c1.alignment = wrap
        c2 = ws3.cell(row=r, column=2, value=note)
        c2.font = Font(size=9)
        c2.fill = PatternFill("solid", fgColor=CAT_COLORS.get(label, "FFFFFF"))
        c2.border = bdr; c2.alignment = wrap
        ws3.row_dimensions[r].height = 36

    wb.save(output_path)
    print(f"Saved: {output_path}")


# ── CSV writer ────────────────────────────────────────────────────────────────

def write_csv(rows: list, output_path: str):
    import csv
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "package", "procedure", "stage1_file", "is_split",
            "total_parts", "parts_sent_gudu", "parts_with_json",
            "parts_skipped", "max_part_chars", "total_json_rows",
            "category", "recommended_action"
        ])
        writer.writeheader()
        for r in rows:
            writer.writerow({
                "package":            r["package"],
                "procedure":          r["procedure"],
                "stage1_file":        r["stage1_file"],
                "is_split":           r["is_split"],
                "total_parts":        r["total_parts"],
                "parts_sent_gudu":    r["parts_sent_gudu"],
                "parts_with_json":    r["parts_with_json"],
                "parts_skipped":      r["parts_skipped"],
                "max_part_chars":     r["max_part_chars"],
                "total_json_rows":    r["total_json_rows"],
                "category":           r["category"],
                "recommended_action": RECOMMENDED_ACTION.get(r["category"], ""),
            })
    print(f"Saved CSV: {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate Gudu pipeline status report")
    base = os.path.dirname(os.path.abspath(__file__))

    parser.add_argument(
        "--split-procs",
        default=os.path.join(base, "Rest_Metadata_7RPT_SplitProcs"),
        help="Stage-1 directory for PKG-split procedure files"
    )
    parser.add_argument(
        "--cleaned",
        default=os.path.join(base, "Rest_Metadata_7RPT_SplitProcs_cleaned"),
        help="Stage-2/3 directory for PKG-split cleaned SQL + lineage JSON files"
    )
    parser.add_argument(
        "--nonpkg-src",
        default=os.path.join(base, "Rest_Metadata_7RPT"),
        help="Source directory for non-PKG files (PRC + MView). PKG files are auto-excluded."
    )
    parser.add_argument(
        "--nonpkg-cleaned",
        default=os.path.join(base, "Rest_Metadata_7RPT_NonPKG_cleaned"),
        help="Stage-2/3 directory for non-PKG cleaned SQL + lineage JSON files"
    )
    parser.add_argument(
        "--output",
        default=os.path.join(os.path.dirname(base), "lineage_pipeline_status.xlsx"),
        help="Output Excel file path"
    )
    parser.add_argument("--csv", action="store_true", help="Also write a CSV file")
    args = parser.parse_args()

    # ── PKG pipeline rows ────────────────────────────────────────────────────
    print(f"\n[PKG Pipeline]")
    print(f"  Stage-1 dir  : {args.split_procs}")
    print(f"  Cleaned dir  : {args.cleaned}")
    rows_pkg = analyse(args.split_procs, args.cleaned)
    for r in rows_pkg:
        r["pipeline"] = "PKG"
        r["object_type"] = "Procedure (from PKG)"

    # ── Non-PKG pipeline rows ────────────────────────────────────────────────
    rows_nonpkg = []
    if os.path.isdir(args.nonpkg_src) and os.path.isdir(args.nonpkg_cleaned):
        print(f"\n[Non-PKG Pipeline]")
        print(f"  Source dir   : {args.nonpkg_src}")
        print(f"  Cleaned dir  : {args.nonpkg_cleaned}")
        # analyse() takes a stage-1 dir; for non-PKG the "stage-1" is the raw source dir
        # We need to exclude PKG files from the source dir
        rows_nonpkg = analyse_nonpkg(args.nonpkg_src, args.nonpkg_cleaned)
        for r in rows_nonpkg:
            r["pipeline"] = "Non-PKG"
    else:
        print(f"\n[Non-PKG Pipeline] skipped — directories not found")

    all_rows = rows_pkg + rows_nonpkg
    summary = build_summary(all_rows)

    # Print console summary
    print(f"\n{'='*60}")
    print(f"  PIPELINE STATUS SUMMARY")
    print(f"{'='*60}")
    print(f"  Total objects              : {summary['total_procedures']}")
    print(f"  Total cleaned part-files   : {summary['total_cleaned_parts']}")
    print(f"  Part-files sent to Gudu    : {summary['parts_sent_gudu']}")
    print(f"  Part-files with JSON output: {summary['parts_with_json']}")
    print(f"  Part-files skipped (>10K)  : {summary['parts_skipped']}")
    print(f"  Total lineage rows          : {summary['total_lineage_rows']}")
    print(f"\n  By Category:")
    for cat, cnt in sorted(summary["by_category"].items(), key=lambda x: -x[1]):
        print(f"    {cat:<38} {cnt}")
    print(f"{'='*60}\n")

    write_excel(all_rows, summary, args.output)
    if args.csv:
        csv_path = os.path.splitext(args.output)[0] + ".csv"
        write_csv(all_rows, csv_path)


if __name__ == "__main__":
    main()
