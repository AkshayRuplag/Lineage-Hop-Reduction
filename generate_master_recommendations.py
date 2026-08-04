#!/usr/bin/env python3
"""
generate_master_recommendations.py

Reads all 14 RPT-level hop-reduction workbooks and produces a single
MASTER workbook:

  Sheet 1 — Executive Summary    : headline stats, phase breakdown, top-10
  Sheet 2 — Master Recommendations: all unique recs (deduplicated), with
                                    "Appears in RPTs" column
  Sheet 3 — Implementation Roadmap: phased, ordered plan for dev team
  Sheet 4 — Cross-RPT Impact Map : which RPTs each rec touches (grid)
  Sheet 5 — Dev Work Orders      : per-Tidal-job action cards
  Sheet 6 — Global Savings Summary: cumulative savings by phase & category

Deduplication rule
------------------
  F (LLM) recs  : same script  → same rec  (appears in all 14, count once)
  Q recs         : same job     → same rec
  All others     : same (category, target_table, sorted affected jobs) → same rec
"""

import os, re, sys
from collections import defaultdict
import os, re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = Path(__file__).parent
WB_DIR      = SCRIPT_DIR / "output" / "hop_reduction_recommendations"
OUTPUT_FILE   = SCRIPT_DIR / "output" / "MASTER_Hop_Reduction_Recommendations.xlsx"
VALIDATED_DIR = SCRIPT_DIR / "output" / "hop_reduction_recommendations_validated"


# ── Phase assignment ─────────────────────────────────────────────────────────
#  Phase 0 — Decisions & Constraints (review only, no code)
#  Phase 1 — Foundation  (naming, dedup paths — prerequisite for everything)
#  Phase 2 — Quick Wins  (isolated, low-risk, each independently deployable)
#  Phase 3 — Consolidation (medium risk, refactoring, needs regression testing)
#  Phase 4 — MV Architecture (high risk, DBA + arch sign-off required)
#  Phase 5 — PL/SQL Optimization (continuous improvement, parallel track)

PHASE_MAP: dict[str, tuple[int, str]] = {
    # ── Phase 0 ─────────────────────────────────────────────────────────────
    "C. Source System Architectural Split (By Design)":              (0, "Decisions & Constraints"),
    "A. Source System Architectural Split (By Design)":              (0, "Decisions & Constraints"),
    "E. MV Shared Across RPTs (No Blind Elimination)":               (0, "Decisions & Constraints"),
    # "E. MV Zero Lineage Consumers — Verify Parser Gap" removed:
    # MVs in summaries have evidence of lineage → not a data-team recommendation.
    # ── Phase 1 ─────────────────────────────────────────────────────────────
    "O. Hybrid Naming Convention Violation":                         (1, "Foundation"),
    "O. View Naming Convention Violation":                           (1, "Foundation"),
    "G. Duplicate Column-Lineage Paths":                             (1, "Foundation"),
    # ── Phase 2 ─────────────────────────────────────────────────────────────
    "Q. Hardcoded Value Mapping — Externalize to Lookup Table":      (2, "Quick Wins"),
    "Q. Hardcoded Value Mapping - Externalize to Lookup Table":      (2, "Quick Wins"),
    "B. Pass-Through / Update-in-Place":                             (2, "Quick Wins"),
    "L. Overlapping Preprocessing":                                  (2, "Quick Wins"),
    "O. GTT Intra-Procedure Staging — Evaluate CTE Replacement":     (2, "Quick Wins"),
    "O. GTT Intra-Procedure Staging - Evaluate CTE Replacement":     (2, "Quick Wins"),
    "O. Intermediate/Working Table — Naming & Hop Reduction":        (2, "Quick Wins"),
    "O. Intermediate/Working Table - Naming & Hop Reduction":        (2, "Quick Wins"),
    # ── Phase 3 ─────────────────────────────────────────────────────────────
    "A. Redundant Multi-Source Load":                                (3, "Consolidation"),
    "C. Near-Duplicate Procedures":                                  (3, "Consolidation"),
    "C. OFFSET / Near-Duplicate Procedures":                         (3, "Consolidation"),
    "P. Post-Load Cursor UPDATE — Fold into INSERT":                 (3, "Consolidation"),
    "P. Post-Load Cursor UPDATE - Fold into INSERT":                 (3, "Consolidation"),
    "N. Circular Update Chain (Post-Load Self-Reference)":           (3, "Consolidation"),
    "N. Post-Consumption Update Pattern":                            (3, "Consolidation"),
    "D. Serial Chain — No Data Dependency":                          (3, "Consolidation"),
    # ── Phase 4 ─────────────────────────────────────────────────────────────
    "E. Dead MV — Zero Consumers (Decommission Candidate)":          (4, "MV Architecture"),
    "M. Orchestration Edge Without Data Handoff (MV->MV)":           (4, "MV Architecture"),
    "J. Over-Fanout MV":                                             (4, "MV Architecture"),
    "E. MV Elimination Candidate":                                   (4, "MV Architecture"),
    "E. MV Indicator-Chain Elimination Candidate":                   (4, "MV Architecture"),
    # ── Phase 5 ─────────────────────────────────────────────────────────────
    "F. LLM-Identified Issues":                                      (5, "PL/SQL Optimization"),
}

def get_phase(category: str) -> tuple[int, str]:
    cat = (category or "").strip()
    if cat in PHASE_MAP:
        return PHASE_MAP[cat]
    for k, v in PHASE_MAP.items():
        if cat.startswith(k[:18]):
            return v
    return (5, "PL/SQL Optimization")


# ── Colours ───────────────────────────────────────────────────────────────────
PHASE_ROW_FILL   = {0:"FFE5CCF5", 1:"FFFCE4EC", 2:"FFE8F5E9",
                    3:"FFFFF3E0", 4:"FFFBE9E7", 5:"FFE3F2FD"}
PHASE_HDR_FILL   = {0:"FF6A1B9A", 1:"FFC62828", 2:"FF2E7D32",
                    3:"FFE65100", 4:"FFB71C1C", 5:"FF0D47A1"}
PHASE_LABELS = {
    0: "Phase 0  —  DECISIONS & CONSTRAINTS",
    1: "Phase 1  —  FOUNDATION",
    2: "Phase 2  —  QUICK WINS",
    3: "Phase 3  —  CONSOLIDATION",
    4: "Phase 4  —  MV ARCHITECTURE",
    5: "Phase 5  —  PL/SQL OPTIMISATION",
}
NAVY     = "FF1A237E"
DARK     = "FF37474F"
WHITE    = "FFFFFFFF"
ALT_GRAY = "FFF5F5F5"

def hfill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)

def hdr_cell(ws, row: int, col: int, value, fill_hex: str = NAVY,
             font_col: str = WHITE, bold: bool = True, size: int = 10,
             wrap: bool = True, halign: str = "center") -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill = hfill(fill_hex)
    c.font = Font(bold=bold, color=font_col, size=size)
    c.alignment = Alignment(wrap_text=wrap, vertical="center", horizontal=halign)

def set_col_widths(ws, widths: dict) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Helpers ───────────────────────────────────────────────────────────────────
def _toint(v) -> int:
    try:   return int(v or 0)
    except: return 0

def _tofloat(v) -> float:
    try:   return float(v or 0)
    except: return 0.0

def _split_jobs(raw: str) -> list[str]:
    return [j.strip() for j in re.split(r'[\n,;]', raw or "") if j.strip()]


# ── SQL object type classifier ─────────────────────────────────────────────────
# PKG/PROC check runs FIRST so e.g. PRC_..._INTERMEDIATE_MV_TBL is CODE, not MV.
_CODE_PREFIXES = ("PKG_", "PRC_", "PROC_", "SP_", "USP_", "FN_", "FUNC_")

def _classify_sql_object(name: str) -> str:
    """
    'CODE' — PKG or standalone PROC/FUNC: open script, edit SQL, raise PR
    'MV'   — Materialized view: DDL DROP + remove Tidal refresh job (no code edit)
    'VIEW' — Logical view: DDL RENAME/ALTER (naming convention fix)
    Conservative default is 'CODE' so nothing is silently dropped.
    """
    u = name.upper().strip()
    left = u.split(".")[0] if "." in u else u
    if any(left.startswith(p) for p in _CODE_PREFIXES):
        return "CODE"
    if "_MV" in u or u.startswith(("MV_", "MVW_")):
        return "MV"
    if "_VW" in u or "_VIEW" in u:
        return "VIEW"
    return "CODE"


# ── Validation status helpers ─────────────────────────────────────────────────
# Canonical status labels (with emoji prefix for quick visual scanning in Excel)
VS_CONSIDER  = "✅ Consider"
VS_NOT_CONS  = "⏸ Not to Consider"
VS_NOT_GOOD  = "❌ Not a Good Recommendation"
VS_PENDING   = "⏳ Validation Pending"

VS_FILL = {
    VS_CONSIDER: "FFC6EFCE",   # green
    VS_NOT_CONS: "FFFFEB9C",   # amber
    VS_NOT_GOOD: "FFFFC7CE",   # red
    VS_PENDING:  "FFF2F2F2",   # light grey
}

def _normalize_status(raw: str) -> str:
    """Map free-text validation status to a canonical label."""
    s = (raw or "").strip().lower()
    if not s:
        return VS_PENDING
    if "not to consider" in s:
        return VS_NOT_CONS
    if "not a good" in s:
        return VS_NOT_GOOD
    if "consider" in s:   # catches "can be consider", "- Consider", "Consider" variants
        return VS_CONSIDER
    return raw.strip()    # unknown — keep as-is


def load_validation_statuses() -> dict:
    """
    Read all validated workbooks from VALIDATED_DIR.
    Each file has a "Hop Recommendations" sheet with a "Recommendation Status" column.

    Returns: dedup_key → canonical status string
    """
    if not VALIDATED_DIR.exists():
        print(f"  -> Validated dir not found: {VALIDATED_DIR}")
        return {}

    status_map: dict = {}
    files = sorted(VALIDATED_DIR.glob("*.xlsx"))
    if not files:
        print(f"  -> No validated workbooks found in {VALIDATED_DIR}")
        return {}

    matched = 0
    for fpath in files:
        wb = openpyxl.load_workbook(fpath, read_only=True)
        if "Hop Recommendations" not in wb.sheetnames:
            wb.close()
            continue
        ws = wb["Hop Recommendations"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            continue

        headers = [str(h) if h is not None else "" for h in rows[0]]
        h_idx   = {h: i for i, h in enumerate(headers) if h}

        # Find the four validation columns (case-insensitive, whitespace-tolerant)
        def _find_col(pattern: str) -> str | None:
            return next((h for h in h_idx if pattern in h.lower()), None)

        status_key    = _find_col("recommendation status")
        comments_key  = _find_col("comments")      # "Comments / Proposed Hop Reduction Approach"
        agreement_key = _find_col("agreement")     # "Agreement  Status/ Review Comments"
        reviewby_key  = _find_col("review by")

        if status_key is None:
            wb.close()
            continue

        for row in rows[1:]:
            if not row[0]:
                continue
            d = {h: (row[i] if i < len(row) else None) for h, i in h_idx.items()}
            raw_status = str(d.get(status_key, "") or "").strip()
            if not raw_status or raw_status.lower() == "recommendation status":
                continue
            cat  = d.get("Category", "") or ""
            tgt  = d.get("Target Table", "") or ""
            jobs = d.get("Affected Jobs", "") or ""
            sql  = d.get("SQL Objects Called", "") or ""
            key  = make_dedup_key(cat, tgt, jobs, sql)
            if key not in status_map:
                status_map[key] = {
                    "validation_status":   _normalize_status(raw_status),
                    "rec_status_raw":      raw_status,
                    "comments":            str(d.get(comments_key, "") or "").strip() if comments_key else "",
                    "agreement_status":    str(d.get(agreement_key, "") or "").strip() if agreement_key else "",
                    "review_by":           str(d.get(reviewby_key, "") or "").strip() if reviewby_key else "",
                }
                matched += 1

        wb.close()

    print(f"  -> Validation statuses loaded: {matched} entries from {len(files)} file(s)")
    return status_map


# ── Dedup key ─────────────────────────────────────────────────────────────────
def make_dedup_key(cat: str, tgt: str, jobs_raw: str, sql_obj: str):
    cat_clean = (cat or "").strip()
    tgt_clean = (tgt or "").strip().upper()
    jobs      = tuple(sorted(_split_jobs(jobs_raw)))
    sql_norm  = re.sub(r'\s+', ' ', (sql_obj or "").strip())[:120]
    prefix    = cat_clean.split('.')[0].strip()
    if prefix == 'F':
        # LLM findings: same script → same rec (it is attached to 95 scripts
        # and each script appears in all RPTs; dedup by script = sql_obj col)
        return ('F', sql_norm)
    elif prefix == 'Q':
        # Q recs: same job → same script → same rec
        return ('Q', jobs[:1], tgt_clean)
    else:
        return (prefix, tgt_clean, jobs)


# ── Read all workbooks ────────────────────────────────────────────────────────
def load_all_recommendations() -> list[dict]:
    files = sorted(WB_DIR.glob("hop_reduction_recommendations_*.xlsx"))
    all_rows: list[dict] = []
    for fpath in files:
        rpt_name = fpath.stem.replace("hop_reduction_recommendations_", "")
        wb = openpyxl.load_workbook(fpath, read_only=True)
        ws = wb["Hop Recommendations"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close(); continue
        headers = rows[0]
        h_idx   = {h: i for i, h in enumerate(headers) if h}
        for row in rows[1:]:
            if not row[0]:
                continue
            d = {h: row[i] for h, i in h_idx.items()}
            d["source_rpt"] = rpt_name
            all_rows.append(d)
        wb.close()
    return all_rows


# ── Deduplicate ───────────────────────────────────────────────────────────────
def deduplicate(all_rows: list[dict]) -> list[dict]:
    """
    Collapses cross-RPT duplicates. Returns list of master records, each with:
      appears_in_rpts   : sorted list of RPT names
      rpt_count         : int
      global_hop_savings: verified hop savings (counted once)
      global_est_min    : estimated minutes saved (counted once)
      phase             : int
      phase_name        : str
    """
    seen: dict = {}

    for row in all_rows:
        cat  = row.get("Category", "") or ""
        tgt  = row.get("Target Table", "") or ""
        jobs = row.get("Affected Jobs", "") or ""
        sql  = row.get("SQL Objects Called", "") or ""
        rpt  = row["source_rpt"]
        key  = make_dedup_key(cat, tgt, jobs, sql)

        if key not in seen:
            rec = dict(row)
            rec["appears_in_rpts"]    = [rpt]
            rec["rpt_count"]          = 1
            ph, ph_name               = get_phase(cat)
            rec["phase"]              = ph
            rec["phase_name"]         = ph_name
            rec["global_hop_savings"] = _toint(row.get("Verified Hop Savings", 0))
            rec["global_est_min"]     = _tofloat(row.get("Estimated Runtime Saved (min)", 0))
            seen[key] = rec
        else:
            if rpt not in seen[key]["appears_in_rpts"]:
                seen[key]["appears_in_rpts"].append(rpt)
                seen[key]["rpt_count"] += 1

    records = list(seen.values())
    # Sort: phase asc, priority score desc
    records.sort(key=lambda r: (r["phase"], -(r.get("Priority Score") or 0)))

    # ── Overlap resolution ────────────────────────────────────────────────────
    # When two recs share the same (target_table, affected_jobs) but come from
    # different detectors (e.g. A. Redundant Load + C. Near-Duplicate for the
    # same 3 OFFSET jobs), both may claim hop savings for the same physical
    # change. That double-counts the saving. Keep savings only for the
    # highest-priority rec (already sorted by score desc); zero the rest and
    # add a cross-reference note.
    jobs_tgt_seen: dict = {}  # (tgt_upper, jobs_tuple) → master_id of primary rec
    for i, rec in enumerate(records):
        tgt   = (rec.get("Target Table", "") or "").upper()
        jobs  = tuple(sorted(
            j.strip() for j in re.split(r"[\n,;]", rec.get("Affected Jobs", "") or "")
            if j.strip()
        ))
        hops  = rec.get("global_hop_savings", 0) or 0
        if not tgt or not jobs or hops == 0:
            continue
        overlap_key = (tgt, jobs)
        if overlap_key not in jobs_tgt_seen:
            jobs_tgt_seen[overlap_key] = i  # this rec is the primary
        else:
            # Secondary rec — same physical change already credited to primary
            primary_idx = jobs_tgt_seen[overlap_key]
            primary_cat = records[primary_idx].get("Category", "")
            # Zero out savings on secondary; add overlap note
            rec["global_hop_savings"]  = 0
            rec["global_est_min"]      = 0.0
            rec["_overlap_note"] = (
                f"Hop savings credited to primary rec M-{primary_idx+1:04d} "
                f"({primary_cat}) — same affected jobs and target table. "
                f"Implement together; do not count savings twice."
            )

    return records


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET WRITERS
# ══════════════════════════════════════════════════════════════════════════════

# ── Sheet 1: Executive Summary ────────────────────────────────────────────────
def write_executive_summary(wb: openpyxl.Workbook, records: list[dict],
                             rpt_names: list[str],
                             files: list | None = None) -> None:
    files = files or []
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "MASTER HOP REDUCTION RECOMMENDATIONS — Executive Summary"
    c.font  = Font(bold=True, size=16, color=WHITE)
    c.fill  = hfill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    ws["A2"].value = (f"Consolidated from {len(files)} RPT workbooks  |  "
                      f"{len(rpt_names)} RPTs have structural findings  |  "
                      f"Shared jobs counted once across all pipelines")
    ws["A2"].font  = Font(italic=True, size=10, color="FF616161")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # ── Helper: write a label/value row ──────────────────────────────────────
    def kv(label: str, value, note: str = "", bold_val: bool = False, row_ht: int = 18):
        r = ws.max_row + 1
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True, color="FF37474F", size=10)
        vc = ws.cell(row=r, column=2, value=value)
        if bold_val:
            vc.font = Font(bold=True, size=11, color=NAVY)
        if note:
            nc = ws.cell(row=r, column=3, value=note)
            nc.font = Font(italic=True, color="FF757575", size=9)
        ws.row_dimensions[r].height = row_ht

    def section(text: str):
        r = ws.max_row + 2
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(row=r, column=1, value=text)
        c.font  = Font(bold=True, size=12, color=NAVY)
        c.fill  = hfill("FFE8EAF6")
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r].height = 22

    # ── Overall stats ─────────────────────────────────────────────────────────
    section("OVERALL STATISTICS")
    actionable = [r for r in records if r["phase"] not in (0,)]
    non_f      = [r for r in actionable if not (r.get("Category","") or "").startswith("F")]
    all_jobs   = set()
    for rec in records:
        for j in _split_jobs(rec.get("Affected Jobs","") or ""):
            all_jobs.add(j)

    total_hops_all = sum(r["global_hop_savings"] for r in actionable)
    total_min_all  = sum(r["global_est_min"]     for r in actionable)
    hop_no_f       = sum(r["global_hop_savings"] for r in non_f)
    min_no_f       = sum(r["global_est_min"]     for r in non_f)

    kv("Total unique recommendations (deduplicated)", len(records),
       "Cross-RPT duplicates collapsed to one entry", bold_val=True)
    kv("  of which actionable (Phases 1-5)", len(actionable))
    kv("  of which graph/lineage-verified (excl. LLM)", len(non_f))
    kv("Total unique Tidal jobs touched", len(all_jobs),
       "If a job runs in multiple RPTs, it is counted once")
    kv("RPT pipelines in scope", len(rpt_names),
       "  |  ".join(sorted(rpt_names)))

    section("VALIDATION STATUS — Data Team Review Results")
    vs_counts: dict[str, int] = defaultdict(int)
    for rec in records:
        vs_counts[rec.get("validation_status", VS_PENDING)] += 1
    actionable_validated = vs_counts.get(VS_CONSIDER, 0)
    not_to_consider      = vs_counts.get(VS_NOT_CONS, 0)
    not_good             = vs_counts.get(VS_NOT_GOOD, 0)
    pending              = vs_counts.get(VS_PENDING, 0)
    validated_total      = actionable_validated + not_to_consider + not_good
    kv(f"{VS_CONSIDER}",  actionable_validated,
       "Validated by data team as a genuine improvement — include in work orders", bold_val=True)
    kv(f"{VS_NOT_CONS}", not_to_consider,
       "Validated as technically correct but not a priority for this programme cycle")
    kv(f"{VS_NOT_GOOD}",  not_good,
       "Validated as incorrect or not applicable — excluded from work orders")
    kv(f"{VS_PENDING}",   pending,
       "No matching entry found in validated workbooks — treat as actionable until reviewed")
    kv("Total reviewed by data team", validated_total,
       f"out of {len(records)} unique recs  ({round(validated_total/max(len(records),1)*100)}% coverage)")
    kv("Actionable (Consider + Pending)", actionable_validated + pending,
       "These are the recs in the Actionable Work Orders sheet (Sheet 10)", bold_val=True)

    section("VERIFIED SAVINGS (if all recommendations implemented)")
    kv("Global verified hop savings — all phases",
       f"{total_hops_all} pipeline hops eliminated",
       "Each hop = one Tidal job execution that can be removed", bold_val=True, row_ht=20)
    kv("  — graph/lineage-verified savings only (excl. LLM)",
       f"{hop_no_f} hops", bold_val=False)
    kv("Global estimated runtime saved — all phases",
       f"{total_min_all:.0f} min  ({total_min_all/60:.1f} hrs)",
       "Based on average Tidal job runtimes from dependency graph", bold_val=True, row_ht=20)
    kv("  — graph/lineage-verified savings only (excl. LLM)",
       f"{min_no_f:.0f} min  ({min_no_f/60:.1f} hrs)")

    # ── Savings by validation status ─────────────────────────────────────────
    # Break savings into validated-Consider, Pending (treated as actionable),
    # and excluded (Not to Consider + Not a Good Recommendation).
    def _savings(recs_subset) -> tuple[int, float]:
        h = sum(r["global_hop_savings"] for r in recs_subset
                if not (r.get("Category","") or "").startswith("F"))
        m = sum(r["global_est_min"] for r in recs_subset
                if not (r.get("Category","") or "").startswith("F"))
        return h, m

    consider_recs  = [r for r in records if r.get("validation_status") == VS_CONSIDER]
    pending_recs   = [r for r in records if r.get("validation_status") == VS_PENDING]
    excluded_recs  = [r for r in records if r.get("validation_status") in (VS_NOT_CONS, VS_NOT_GOOD)]

    h_con, m_con = _savings(consider_recs)
    h_pen, m_pen = _savings(pending_recs)
    h_exc, m_exc = _savings(excluded_recs)
    h_act = h_con + h_pen
    m_act = m_con + m_pen

    section("ACTUAL SAVINGS — By Validation Status")
    kv(f"  {VS_CONSIDER}  ({len(consider_recs)} recs)",
       f"{h_con} hops  |  {m_con:.0f} min  ({m_con/60:.1f} hrs)",
       "Data team confirmed these are genuine improvements", bold_val=True, row_ht=20)
    kv(f"  {VS_PENDING}  ({len(pending_recs)} recs)",
       f"{h_pen} hops  |  {m_pen:.0f} min  ({m_pen/60:.1f} hrs)",
       "Not yet reviewed — included in actionable total as conservative estimate", row_ht=20)
    kv("  ACTIONABLE TOTAL  (Consider + Pending)",
       f"{h_act} hops  |  {m_act:.0f} min  ({m_act/60:.1f} hrs)",
       "Realistic savings if all Consider + Pending recs are implemented", bold_val=True, row_ht=20)
    kv(f"  Excluded ({VS_NOT_CONS} + {VS_NOT_GOOD})  ({len(excluded_recs)} recs)",
       f"{h_exc} hops  |  {m_exc:.0f} min excluded",
       "Data team decided not to implement these — savings will NOT be realised", row_ht=20)
    # Show what % of savings the validated-Consider recs represent
    pct_realised = round(h_con / max(hop_no_f, 1) * 100)
    kv("  Validated savings as % of total programme",
       f"{pct_realised}% of verified hops are in 'Consider' recs",
       f"({h_con} of {hop_no_f} verified hops confirmed actionable by data team)")

    # ── Phase breakdown table ─────────────────────────────────────────────────
    section("IMPLEMENTATION PHASES — SNAPSHOT")
    ws.append([])

    phase_descs = {
        0: "Review constraints; document decisions. No code change.",
        1: "Fix naming violations; remove duplicate data paths. Prerequisite for all later phases.",
        2: "Externalize lookup tables; collapse pass-throughs; GTT→CTE. Each item deploys independently.",
        3: "Parameterise duplicate procedures; remove circular UPDATE jobs; fold cursor UPDs into INSERTs.",
        4: "Eliminate dead/chain MVs; remove scheduling edges. Needs DBA + architecture sign-off.",
        5: "PL/SQL script-level optimisations. Convert to dev tickets; run in parallel.",
    }
    phase_prereqs = {
        0: "None — start here",
        1: "Phase 0 decisions documented",
        2: "Phase 1 deployed to all environments",
        3: "Phase 2 validated; regression suite in place",
        4: "Phase 3 complete; DBA sign-off; rollback plan ready",
        5: "Can start any time (parallel track)",
    }

    hdr_row = ws.max_row + 1
    for col, h in enumerate(
        ["Phase", "Name", "# Unique Recs", "Hop Savings", "Est. Min Saved",
         "Est. Hrs Saved", "Risk", "Prerequisites", "Description"], 1):
        hdr_cell(ws, hdr_row, col, h, fill_hex=DARK, size=10)
    ws.row_dimensions[hdr_row].height = 22

    phase_groups: dict[int, list] = defaultdict(list)
    for rec in records:
        phase_groups[rec["phase"]].append(rec)

    for ph in sorted(phase_groups.keys()):
        ph_recs = phase_groups[ph]
        ph_hops = sum(r["global_hop_savings"] for r in ph_recs)
        ph_min  = sum(r["global_est_min"]     for r in ph_recs)
        risks   = [r.get("Risk","") for r in ph_recs]
        risk    = "HIGH" if "HIGH" in risks else ("MEDIUM" if "MEDIUM" in risks else "LOW")
        r_num   = ws.max_row + 1
        fill    = hfill(PHASE_ROW_FILL.get(ph, "FFFFFFFF"))
        vals    = [f"Phase {ph}", PHASE_LABELS[ph].split("—")[1].strip(),
                   len(ph_recs), ph_hops, round(ph_min, 1), round(ph_min/60, 2),
                   risk, phase_prereqs[ph], phase_descs[ph]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r_num, column=col, value=v)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if col == 1:
                c.font = Font(bold=True)
        ws.row_dimensions[r_num].height = 28

    # ── Top 10 ────────────────────────────────────────────────────────────────
    section("TOP 10 HIGHEST-IMPACT UNIQUE RECOMMENDATIONS")
    ws.append([])

    top10 = sorted(
        [r for r in records if r["phase"] < 5 and r["global_hop_savings"] > 0],
        key=lambda x: (-(x.get("Priority Score") or 0), x["phase"])
    )[:10]

    hdr_row = ws.max_row + 1
    for col, h in enumerate(
        ["Rank","Category","Target Table","Phase","Hop Savings",
         "Est. Min Saved","Risk","Appears in # RPTs","RPT List"], 1):
        hdr_cell(ws, hdr_row, col, h, fill_hex=DARK, size=10)
    ws.row_dimensions[hdr_row].height = 22

    for i, rec in enumerate(top10, 1):
        r_num = ws.max_row + 1
        fill  = hfill(PHASE_ROW_FILL.get(rec["phase"], "FFFFFFFF"))
        vals  = [
            i,
            rec.get("Category",""),
            rec.get("Target Table",""),
            f"Phase {rec['phase']} — {rec['phase_name']}",
            rec["global_hop_savings"],
            round(rec["global_est_min"], 1),
            rec.get("Risk",""),
            rec["rpt_count"],
            ", ".join(sorted(rec["appears_in_rpts"])),
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r_num, column=col, value=v)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 1:
                c.font = Font(bold=True)
        ws.row_dimensions[r_num].height = 36

    set_col_widths(ws, {"A":40,"B":22,"C":14,"D":8,"E":8,"F":8,"G":10,"H":35,"I":45})


# ── Sheet 2: Master Recommendations ──────────────────────────────────────────
def write_master_recommendations(wb: openpyxl.Workbook,
                                  records: list[dict]) -> None:
    ws = wb.create_sheet("Master Recommendations")

    # Banner
    ws.merge_cells("A1:U1")
    c = ws["A1"]
    c.value = ("MASTER RECOMMENDATIONS — Deduplicated across all RPTs  |  "
               "Same job/script appearing in multiple RPTs is listed once")
    c.font  = Font(bold=True, size=13, color=WHITE)
    c.fill  = hfill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    COLS = [
        "Master ID", "Phase", "Phase Name", "Orig. ID", "Category",
        "Target Table", "Affected Jobs", "SQL Objects Called",
        "Description",
        "What To Do (Recommendation)",
        "Appears in RPTs", "# RPTs",
        "Verified Hop Savings", "Est. Time Saved (min)",
        "Risk", "Priority Score",
        "Validation Status",                         # Q  normalised status
        "Comments / Proposed Hop Reduction Approach", # R
        "Recommendation Status (as entered)",         # S  raw text from data team
        "Agreement Status / Review Comments",         # T
        "Review By",                                  # U
    ]
    HUMAN_VALIDATION_HDR = "FF4A148C"   # deep purple — human-in-the-loop columns (Q-U)

    hdr_row = 2
    for col, h in enumerate(COLS, 1):
        # Columns Q-U (17-21) are human-validated fields — use purple header
        fill = HUMAN_VALIDATION_HDR if col >= 17 else NAVY
        hdr_cell(ws, hdr_row, col, h, fill_hex=fill, size=10)
    ws.row_dimensions[hdr_row].height = 30
    ws.freeze_panes = "A3"

    for i, rec in enumerate(records, 1):
        r_num    = ws.max_row + 1
        fill     = hfill(PHASE_ROW_FILL.get(rec["phase"], "FFFFFFFF"))
        rpts_str = "\n".join(sorted(rec["appears_in_rpts"]))
        overlap  = rec.get("_overlap_note", "")
        hops_val = rec["global_hop_savings"]

        # Append overlap note to recommendation text so reviewers see it
        rec_text = rec.get("Recommendation","") or rec.get("What To Do","")
        if overlap:
            rec_text = (rec_text or "") + f"\n\n[OVERLAP NOTE] {overlap}"

        vals    = [
            f"M-{i:04d}",
            rec["phase"],
            rec["phase_name"],
            rec.get("ID",""),
            rec.get("Category",""),
            rec.get("Target Table",""),
            rec.get("Affected Jobs",""),
            rec.get("SQL Objects Called",""),
            rec.get("Description",""),
            rec_text,
            rpts_str,
            rec["rpt_count"],
            hops_val if not overlap else "0 (see overlap note)",
            round(rec["global_est_min"], 1) if not overlap else "0.0",
            rec.get("Risk",""),
            rec.get("Priority Score",""),
            rec.get("validation_status", VS_PENDING),    # Q
            rec.get("val_comments", ""),                  # R
            rec.get("rec_status_raw", ""),               # S
            rec.get("val_agreement", ""),                 # T
            rec.get("val_review_by", ""),                 # U
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r_num, column=col, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            # Shade the phase/id columns for quick scanning
            if col in (1, 2, 3, 12, 13, 14, 15, 16):
                c.fill = fill
            # Highlight overlapping savings cells in orange to draw attention
            if col == 13 and overlap:
                c.fill = hfill("FFFFF3CD")  # light amber
            # Colour validation status cell
            if col == 17:
                c.fill = hfill(VS_FILL.get(v, "FFF2F2F2"))
        ws.row_dimensions[r_num].height = 100  # tall enough for full description

    set_col_widths(ws, {
        "A":12,"B":7,"C":22,"D":10,"E":45,"F":28,"G":40,"H":40,
        "I":70,"J":80,"K":35,"L":8,"M":12,"N":14,"O":10,"P":12,
        "Q":28,"R":55,"S":28,"T":55,"U":22,
    })
    ws.auto_filter.ref = f"A2:U{ws.max_row}"


# ── Sheet 3: Implementation Roadmap ──────────────────────────────────────────
def write_implementation_roadmap(wb: openpyxl.Workbook,
                                  records: list[dict]) -> None:
    ws = wb.create_sheet("Implementation Roadmap")
    ws.sheet_view.showGridLines = False

    # Banner
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "IMPLEMENTATION ROADMAP — Ordered, phased plan for the development team"
    c.font  = Font(bold=True, size=14, color=WHITE)
    c.fill  = hfill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    phase_details = {
        0: {
            "desc": (
                "These items represent architectural constraints or split-by-design patterns. "
                "No code change is required at this stage. The team should:\n"
                "  (a) Read each finding and confirm the 'By Design' classification is correct.\n"
                "  (b) Document the decision formally (ADR / confluence page).\n"
                "  (c) Use the outcome to scope later phases — e.g., if an A/C split is confirmed "
                "     By Design, those jobs are removed from Phases 3/4 scope."
            ),
            "prereq": "None — start here before any code changes",
            "exit_criteria": "All Phase 0 items reviewed and decisions documented",
        },
        1: {
            "desc": (
                "Rename tables/views that violate naming conventions and remove duplicate "
                "column-lineage paths. These changes are prerequisites for later phases: "
                "if a table is renamed after Phase 3 refactoring, every script touched in "
                "Phase 3 must be patched again. Do naming first.\n"
                "  (a) Co-ordinate rename with DBA team (DDL + synonyms).\n"
                "  (b) Search all PL/SQL packages for old name → update references.\n"
                "  (c) Deploy and validate lineage tooling still resolves correctly."
            ),
            "prereq": "Phase 0 decisions documented",
            "exit_criteria": "All renamed objects validated in DEV + UAT; no broken references",
        },
        2: {
            "desc": (
                "Independent, low-risk changes. Each item can be developed, tested, and deployed "
                "on its own without dependency on other items in this phase.\n"
                "  Q  — Create lookup/reference tables; replace CASE blocks with LEFT JOINs.\n"
                "  B  — Collapse pass-through jobs (read+write same table) into the upstream load.\n"
                "  L  — Merge overlapping pre-processing jobs into a single canonical prep step.\n"
                "  O  — Evaluate GTT→CTE replacement (only if GTT is read exactly once per call).\n"
                "  O  — Rename/reclassify intermediate working tables; evaluate hop removal."
            ),
            "prereq": "Phase 1 deployed and validated in all environments",
            "exit_criteria": "Each item regression-tested; row counts match ±0 across 3 load cycles",
        },
        3: {
            "desc": (
                "Medium-risk structural refactoring. Changes here affect multiple scripts or "
                "Tidal job flows; full end-to-end regression testing is required after each.\n"
                "  A  — Consolidate redundant multi-source loads into a single parameterized job.\n"
                "  C  — Merge near-duplicate procedures into one parameterized package procedure.\n"
                "  P  — Fold post-load cursor-based UPDATEs into the main INSERT SELECT.\n"
                "  N  — Remove circular post-load UPDATE jobs (self-referential on same table).\n"
                "Recommended approach: one recommendation per sprint, with full regression cycle."
            ),
            "prereq": "Phase 2 deployed and validated; regression test suite in place",
            "exit_criteria": "Each consolidated job validated end-to-end; Tidal schedule updated",
        },
        4: {
            "desc": (
                "HIGH blast radius — affects multiple Tidal job chains and may impact scheduled "
                "reporting. Requires DBA team involvement and architecture sign-off before any change.\n"
                "  E (Dead MV)         — Decommission MVs with zero downstream consumers.\n"
                "  M (Orchestration)   — Remove phantom Tidal scheduling edges (MV→MV with no data handoff).\n"
                "  J (Over-Fanout)     — Rationalize MVs that feed too many downstream consumers.\n"
                "  E (MV Elimination)  — Drop MVs that duplicate data already in the base table.\n"
                "  E (MV Indicator-Chain) — Eliminate UPD_IND_COLS → MV_REFRESH chains "
                "              by folding indicator logic into the base INSERT.\n"
                "Recommended order: Dead MVs first (lowest risk), then Orchestration, "
                "then MV Elimination, then MV Indicator-Chain."
            ),
            "prereq": "Phase 3 complete; DBA sign-off obtained; rollback plan documented; "
                       "change approved by architecture board",
            "exit_criteria": "All decommissioned jobs removed from Tidal; "
                              "no orphan schedule entries; reporting validated",
        },
        5: {
            "desc": (
                "LLM-identified PL/SQL script-level optimisations. These are independent of the "
                "structural changes above and can be converted into individual developer tickets "
                "that run on a parallel track with any phase.\n"
                "  Prioritise by: LLM finding count → redundancy signals → performance concerns.\n"
                "  Each ticket: (1) review LLM finding in context, (2) confirm with code review, "
                "  (3) implement + unit test, (4) regression test the affected RPT."
            ),
            "prereq": "Can start any time (parallel track)",
            "exit_criteria": "All high-priority tickets resolved; performance benchmarks met",
        },
    }

    # Precompute index for Master ID
    id_map = {id(rec): i for i, rec in enumerate(records, 1)}

    phase_groups: dict[int, list] = defaultdict(list)
    for rec in records:
        phase_groups[rec["phase"]].append(rec)

    global_step = 0

    for ph in sorted(phase_groups.keys()):
        ph_recs  = phase_groups[ph]
        ph_hops  = sum(r["global_hop_savings"] for r in ph_recs)
        ph_min   = sum(r["global_est_min"]     for r in ph_recs)
        det      = phase_details.get(ph, {})
        ph_color = PHASE_HDR_FILL.get(ph, DARK)
        row_col  = PHASE_ROW_FILL.get(ph, "FFFFFFFF")

        # ── Phase banner ──────────────────────────────────────────────────────
        ws.append([])
        ws.append([])
        ban_row = ws.max_row + 1
        ws.merge_cells(f"A{ban_row}:J{ban_row}")
        c = ws.cell(row=ban_row, column=1, value=PHASE_LABELS.get(ph, f"Phase {ph}"))
        c.fill  = hfill(ph_color)
        c.font  = Font(bold=True, size=14, color=WHITE)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[ban_row].height = 32

        # Phase meta rows
        meta = [
            ("Purpose",        det.get("desc",""),             True),
            ("Prerequisites",  det.get("prereq",""),           False),
            ("Exit Criteria",  det.get("exit_criteria",""),    False),
            ("Unique recs",    f"{len(ph_recs)} recommendations in this phase", False),
            ("Phase savings",  f"{ph_hops} verified hops  |  {ph_min:.0f} min  ({ph_min/60:.1f} hrs)", False),
        ]
        for label, val, big in meta:
            mr = ws.max_row + 1
            lc = ws.cell(row=mr, column=1, value=label)
            lc.font = Font(bold=True, size=10, color="FF37474F")
            lc.fill = hfill("FFE8EAF6")
            vc = ws.cell(row=mr, column=2, value=val)
            vc.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(f"B{mr}:J{mr}")
            ws.row_dimensions[mr].height = 52 if big else 20

        # Column headers for items
        ws.append([])
        hr = ws.max_row + 1
        hdrs = ["Global Step","Master ID","Category","Target Table","Affected Jobs",
                "Hop Savings","Est. Min","Risk","# RPTs","Description / Action"]
        for col, h in enumerate(hdrs, 1):
            hdr_cell(ws, hr, col, h, fill_hex="FF455A64", size=9)
        ws.row_dimensions[hr].height = 22

        for rec in ph_recs:
            global_step += 1
            nr   = ws.max_row + 1
            fill = hfill(row_col)
            # Use full description + recommendation text (no truncation)
            desc_full = rec.get("Description","") or ""
            rec_full  = rec.get("Recommendation","") or rec.get("What To Do","") or ""
            action_text = (desc_full + ("\n\n" + rec_full if rec_full.strip() else "")).strip()
            overlap = rec.get("_overlap_note","")
            if overlap:
                action_text = action_text + f"\n\n[OVERLAP NOTE] {overlap}"
            mid   = id_map.get(id(rec), 0)
            hops_disp = rec["global_hop_savings"]
            mins_disp = round(rec["global_est_min"], 1)
            vals  = [
                global_step,
                f"M-{mid:04d}",
                rec.get("Category",""),
                rec.get("Target Table",""),
                rec.get("Affected Jobs","") or "",
                hops_disp,
                mins_disp,
                rec.get("Risk",""),
                rec["rpt_count"],
                action_text,
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=nr, column=col, value=v)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 1:
                    cell.font = Font(bold=True)
            ws.row_dimensions[nr].height = 80

    set_col_widths(ws, {
        "A":9,"B":12,"C":45,"D":30,"E":45,"F":10,"G":10,"H":9,"I":8,"J":65
    })


# ── Sheet 4: Cross-RPT Impact Map ─────────────────────────────────────────────
def write_cross_rpt_impact(wb: openpyxl.Workbook, records: list[dict]) -> None:
    ws = wb.create_sheet("Cross-RPT Impact Map")

    # Collect all RPT names
    all_rpts = sorted({rpt for r in records for rpt in r["appears_in_rpts"]})
    n_fixed  = 7  # fixed columns before the RPT columns
    total_cols = n_fixed + len(all_rpts) + 1

    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value = ("CROSS-RPT IMPACT MAP — ✓ means this recommendation also appears in that RPT pipeline  "
               "|  Implement once to fix for ALL ticked RPTs")
    c.font  = Font(bold=True, size=12, color=WHITE)
    c.fill  = hfill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    HDRS = (["Master ID","Phase","Category","Target Table",
              "Hop Savings","Est. Min","Risk"]
             + [r.replace("RPT_","") for r in all_rpts]   # shorter column labels
             + ["# RPTs"])

    hdr_row = 2
    for col, h in enumerate(HDRS, 1):
        hdr_cell(ws, hdr_row, col, h, size=9)
        if col > n_fixed and col <= n_fixed + len(all_rpts):
            ws.cell(row=hdr_row, column=col).alignment = Alignment(
                wrap_text=True, text_rotation=60, horizontal="center", vertical="bottom")
    ws.row_dimensions[hdr_row].height = 75
    ws.freeze_panes = "A3"

    for i, rec in enumerate(records, 1):
        nr   = ws.max_row + 1
        fill = hfill(PHASE_ROW_FILL.get(rec["phase"], "FFFFFFFF"))
        base = [
            f"M-{i:04d}",
            rec["phase"],
            rec.get("Category",""),
            rec.get("Target Table",""),
            rec["global_hop_savings"],
            round(rec["global_est_min"], 1),
            rec.get("Risk",""),
        ]
        rpt_ticks = ["✓" if rpt in rec["appears_in_rpts"] else ""
                     for rpt in all_rpts]
        vals = base + rpt_ticks + [rec["rpt_count"]]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=nr, column=col, value=v)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col <= n_fixed:
                cell.fill = fill
            elif v == "✓":
                cell.fill  = hfill("FF81C784")
                cell.font  = Font(bold=True, color="FF1B5E20")

    # Column widths
    set_col_widths(ws, {"A":12,"B":7,"C":42,"D":28,"E":10,"F":10,"G":8})
    for col in range(n_fixed + 1, n_fixed + 1 + len(all_rpts)):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions[get_column_letter(n_fixed + len(all_rpts) + 1)].width = 8
    ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{ws.max_row}"


# ── Sheet 5: Dev Work Orders ──────────────────────────────────────────────────
def write_dev_work_orders(wb: openpyxl.Workbook, records: list[dict]) -> None:
    """
    Groups all recommendations by Tidal job so each developer can pull a
    single 'work card' showing everything that needs to change for their job.
    """
    ws = wb.create_sheet("Dev Work Orders")

    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = ("DEV WORK ORDERS — All actions grouped by Tidal job  |  "
               "Each job shows every recommendation that touches it, ordered by phase")
    c.font  = Font(bold=True, size=13, color=WHITE)
    c.fill  = hfill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Build job → [(master_idx, rec)] mapping
    id_map: dict = {id(rec): i for i, rec in enumerate(records, 1)}
    job_recs: dict[str, list] = defaultdict(list)
    for rec in records:
        for job in _split_jobs(rec.get("Affected Jobs","") or ""):
            job_recs[job].append(rec)

    # Sort jobs: lowest phase first, then highest score
    def job_sort_key(job: str):
        recs = job_recs[job]
        return (min(r["phase"] for r in recs),
                -max((r.get("Priority Score") or 0) for r in recs))

    sorted_jobs = sorted(job_recs.keys(), key=job_sort_key)

    HDRS = ["Tidal Job","# Actions","Lowest Phase","Step",
            "Master ID","Phase","Category","Target Table",
            "Risk","Hop Savings","Description / Action"]
    hdr_row = 2
    for col, h in enumerate(HDRS, 1):
        hdr_cell(ws, hdr_row, col, h, size=10)
    ws.row_dimensions[hdr_row].height = 28
    ws.freeze_panes = "A3"

    step = 0
    for job in sorted_jobs:
        recs_for_job = sorted(
            job_recs[job],
            key=lambda r: (r["phase"], -(r.get("Priority Score") or 0))
        )
        min_ph  = min(r["phase"] for r in recs_for_job)
        n_items = len(recs_for_job)

        for s_idx, rec in enumerate(recs_for_job, 1):
            step += 1
            nr   = ws.max_row + 1
            fill = hfill(PHASE_ROW_FILL.get(rec["phase"], "FFFFFFFF"))
            mid  = id_map.get(id(rec), 0)
            desc_full = rec.get("Description","") or ""
            rec_full  = rec.get("Recommendation","") or rec.get("What To Do","") or ""
            action_text = (desc_full + ("\n\n" + rec_full if rec_full.strip() else "")).strip()
            vals = [
                job       if s_idx == 1 else "",
                n_items   if s_idx == 1 else "",
                min_ph    if s_idx == 1 else "",
                step,
                f"M-{mid:04d}",
                rec["phase"],
                rec.get("Category",""),
                rec.get("Target Table",""),
                rec.get("Risk",""),
                rec["global_hop_savings"],
                action_text,
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=nr, column=col, value=v)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 1 and s_idx == 1:
                    cell.font = Font(bold=True)
            ws.row_dimensions[nr].height = 80

    set_col_widths(ws, {
        "A":50,"B":9,"C":10,"D":8,"E":12,"F":8,"G":44,"H":28,"I":8,"J":10,"K":75
    })
    ws.auto_filter.ref = f"A2:K{ws.max_row}"


# ── Sheet 6: Global Savings Summary ──────────────────────────────────────────
def write_global_savings(wb: openpyxl.Workbook, records: list[dict]) -> None:
    ws = wb.create_sheet("Global Savings Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "GLOBAL SAVINGS SUMMARY — Cumulative impact of each implementation phase"
    c.font  = Font(bold=True, size=13, color=WHITE)
    c.fill  = hfill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Cumulative scenario table ─────────────────────────────────────────────
    def section(text: str):
        r = ws.max_row + 2
        ws.merge_cells(f"A{r}:G{r}")
        c = ws.cell(row=r, column=1, value=text)
        c.font  = Font(bold=True, size=12, color=NAVY)
        c.fill  = hfill("FFE8EAF6")
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r].height = 22

    section("CUMULATIVE SAVINGS — What implementing each set of phases delivers")
    ws.append([])

    scenarios = [
        ("Phase 0 only (Decisions)",          [0]),
        ("+ Phase 1 Foundation",               [0,1]),
        ("+ Phase 2 Quick Wins",               [0,1,2]),
        ("+ Phase 3 Consolidation",            [0,1,2,3]),
        ("+ Phase 4 MV Architecture",          [0,1,2,3,4]),
        ("+ Phase 5 PL/SQL (all phases)",      [0,1,2,3,4,5]),
    ]
    hdr_row = ws.max_row + 1
    for col, h in enumerate(
        ["Scenario","Phases Included","Unique Recs","Verified Hops Saved",
         "Est. Min Saved","Est. Hrs Saved","Key Actions in Added Phase"], 1):
        hdr_cell(ws, hdr_row, col, h, size=10)
    ws.row_dimensions[hdr_row].height = 22

    phase_key_actions = {
        0: "Document architectural constraints; confirm By-Design splits",
        1: "Rename naming-convention violations; deduplicate lineage paths",
        2: "Create lookup tables; collapse pass-throughs; GTT evaluations; intermediate-table hops",
        3: "Parameterise duplicate procedures; remove circular UPD jobs; fold cursor UPDs into INSERTs",
        4: "Decommission dead MVs; remove scheduling edges; eliminate MV indicator-chains",
        5: "PL/SQL script-level optimisations from LLM analysis",
    }

    prev_hops, prev_min = 0, 0.0
    for label, phases in scenarios:
        recs_in = [r for r in records if r["phase"] in phases]
        hops    = sum(r["global_hop_savings"] for r in recs_in)
        mins    = sum(r["global_est_min"]     for r in recs_in)
        new_ph  = max(phases)
        ph_str  = ", ".join(f"Phase {p}" for p in sorted(phases))
        nr      = ws.max_row + 1
        fill    = hfill(PHASE_ROW_FILL.get(new_ph, "FFFFFFFF"))
        vals    = [label, ph_str, len(recs_in), hops, round(mins,1),
                   round(mins/60, 2), phase_key_actions.get(new_ph,"")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=nr, column=col, value=v)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if col in (4,5,6) and hops > prev_hops:
                c.font = Font(bold=True)
        ws.row_dimensions[nr].height = 28
        prev_hops, prev_min = hops, mins

    # ── Per-category breakdown ────────────────────────────────────────────────
    section("SAVINGS BY CATEGORY (unique recommendations only)")
    ws.append([])

    hdr_row = ws.max_row + 1
    for col, h in enumerate(
        ["Category","Phase","# Unique Recs","Verified Hops",
         "Est. Min","Multi-RPT Shared Recs","Note"], 1):
        hdr_cell(ws, hdr_row, col, h, size=10)
    ws.row_dimensions[hdr_row].height = 22

    cat_groups: dict[str, list] = defaultdict(list)
    for rec in records:
        cat_groups[rec.get("Category","")].append(rec)

    for cat, cat_recs in sorted(
            cat_groups.items(),
            key=lambda x: (x[1][0]["phase"],
                           -sum(r["global_hop_savings"] for r in x[1]))):
        ph      = cat_recs[0]["phase"]
        hops    = sum(r["global_hop_savings"] for r in cat_recs)
        mins    = sum(r["global_est_min"]     for r in cat_recs)
        risks   = [r.get("Risk","") for r in cat_recs]
        risk    = "HIGH" if "HIGH" in risks else ("MEDIUM" if "MEDIUM" in risks else "LOW")
        shared  = sum(1 for r in cat_recs if r["rpt_count"] > 1)
        note    = (f"{shared} recommendations appear in 2+ RPTs — implement once, "
                   f"fixes for all affected RPTs") if shared else ""
        nr   = ws.max_row + 1
        fill = hfill(PHASE_ROW_FILL.get(ph, "FFFFFFFF"))
        vals = [cat, f"Phase {ph}", len(cat_recs), hops, round(mins,1), shared, note]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=nr, column=col, value=v)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Cross-RPT multiplier note ─────────────────────────────────────────────
    section("HOW SAVINGS ARE COUNTED")
    ws.append([])
    notes = [
        ("Deduplication",
         "When the same Tidal job appears in multiple RPT analyses, its recommendation "
         "is counted ONCE. The hop saving is NOT multiplied across RPTs."),
        ("Verified vs LLM",
         "Verified savings (GRAPH_OR_LINEAGE evidence) are based on actual job runtime "
         "from the Tidal dependency graph. LLM-only (F) findings show 0 verified hops — "
         "their benefit is qualitative (maintainability, code clarity)."),
        ("Runtime estimates",
         "Est. minutes saved = average job runtime from the Tidal graph × hops removed. "
         "Actual savings depend on job criticality, parallelism, and Tidal scheduling."),
        ("Phase 0 savings",
         "Phase 0 (Decisions) has 0 hop savings by design — it is an analysis/decision "
         "phase. The value is in scoping the work correctly for later phases."),
    ]
    for label, text in notes:
        r_num = ws.max_row + 1
        lc    = ws.cell(row=r_num, column=1, value=label)
        lc.font = Font(bold=True, color="FF37474F")
        vc    = ws.cell(row=r_num, column=2, value=text)
        vc.alignment = Alignment(wrap_text=True)
        ws.merge_cells(f"B{r_num}:G{r_num}")
        ws.row_dimensions[r_num].height = 30

    set_col_widths(ws, {"A":50,"B":20,"C":12,"D":14,"E":12,"F":16,"G":65})


# ── Sheet 7: Script-Centric Work Orders ──────────────────────────────────────
def write_script_centric_work_orders(wb: openpyxl.Workbook,
                                      records: list[dict]) -> None:
    """
    Groups ALL recommendations by the SQL package/procedure they require code
    changes to.  One row = one script = one developer work package (one PR).

    For MV Elimination recs, splits the affected jobs into two buckets:
      • Consumer scripts (need SQL changes — shown as script rows)
      • MV refresh jobs  (need Tidal schedule removal only — listed in
        'Also Remove from Tidal' column, no SQL change required)

    Wave assignment — the "touch-once" rule:
      Wave = highest non-Phase-4, non-Phase-0 phase across all recs for the script.
      Phase 4 items are flagged in 'Wave 4 Awareness' so the developer knows a
      second (DBA-gated) PR is coming.

    Scripts where ALL SQL objects are DDL targets (MVs, views) or which have no
    SQL objects at all — i.e. no PL/SQL package/procedure to open — are grouped
    under a synthetic '[NO-CODE-CHANGE]' entry.  Actions for this row are:
      Phase 0 items  : review the constraint decision and document it.
      Phase 1 items  : DDL RENAME VIEW (listed in column O).
      Phase 4 items  : DBA — DROP MV objects (column O) + remove Tidal
                       MV_REFRESH jobs after parity validation.
    """
    ws = wb.create_sheet("Script-Centric Work Orders")
    ws.sheet_view.showGridLines = False

    WAVE_FILLS_SWO = {
        0: "FFE5CCF5", 1: "FFFCE4EC", 2: "FFE8F5E9",
        3: "FFFFF3E0", 4: "FFFBE9E7", 5: "FFE3F2FD",
    }
    WAVE_LABELS_SWO = {
        0: "Wave 0 — Decisions Only",
        1: "Wave 1 — DDL Foundation",
        2: "Wave 2 — Quick Wins",
        3: "Wave 3 — Consolidation",
        4: "Wave 4 — MV Architecture (DBA)",
        5: "Wave 5 — LLM Optimization (Parallel)",
    }

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:S1")
    c = ws["A1"]
    c.value = (
        "SCRIPT-CENTRIC WORK ORDERS  —  One row = one SQL Package/Procedure = ONE developer PR  "
        "|  Open each script exactly once; implement ALL bundled changes in that single PR  "
        "|  'Also Remove from Tidal' = separate Tidal-schedule action AFTER parity is confirmed"
    )
    c.font  = Font(bold=True, size=13, color=WHITE)
    c.fill  = hfill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Internal helpers ─────────────────────────────────────────────────────
    def _parse_scripts(rec: dict) -> list[str]:
        raw = rec.get("SQL Objects Called", "") or ""
        return [s.strip() for s in re.split(r'[\n,;]', raw) if s.strip()]

    def _mv_only_removals(rec: dict) -> list[str]:
        """
        Returns items for the 'DDL & Tidal Actions Required' column, using the
        module-level _classify_sql_object classifier uniformly for ALL recs:
          MV  SQL objects → 'DROP MV: <name>'
          VIEW SQL objects → 'RENAME VIEW: <name>'
          Tidal MV_REFRESH jobs → 'Remove from Tidal: <job>'
        """
        removals = []
        for obj in re.split(r'[\n,;]', rec.get("SQL Objects Called", "") or ""):
            obj = obj.strip()
            if not obj:
                continue
            cls = _classify_sql_object(obj)
            if cls == "MV":
                removals.append(f"DROP MV: {obj}")
            elif cls == "VIEW":
                removals.append(f"RENAME VIEW: {obj}")
        for j in _split_jobs(rec.get("Affected Jobs", "") or ""):
            if "MV_REFRESH" in j.upper():
                removals.append(f"Remove from Tidal: {j}")
        return removals

    def _assign_wave(phases: set) -> int:
        """Return implementation wave (1–5) for a script given its set of phases."""
        non_special = phases - {0, 4}   # exclude decisions-only and MV-arch
        if not non_special:
            return 4 if 4 in phases else 0
        if 3 in non_special: return 3
        if 2 in non_special: return 2
        if 1 in non_special: return 1
        return 5  # phase-5 only → parallel LLM track

    TIDAL_ONLY = "[NO-CODE-CHANGE]"   # sentinel: no PL/SQL package to open; DDL/Tidal/decisions only

    def _pkg_key(script: str) -> str:
        """
        Normalize to PACKAGE level so PKG.PROC1 and PKG.PROC2 collapse to PKG.
        Rule: if the name contains a dot AND the left part looks like a package
        name (starts with PKG_ / PRC_ / PROC_ or is uppercase-alpha with underscores),
        strip everything after the first dot.
        Plain table/MV/view names that happen to contain a dot (schema.TABLE) are
        left as-is so they don't get merged with unrelated packages.
        """
        if script == TIDAL_ONLY or '.' not in script:
            return script
        pkg_part, proc_part = script.split('.', 1)
        pkg_up = pkg_part.upper()
        # Heuristic: the left side is a PL/SQL package if it starts with a
        # well-known Oracle package prefix or if the right side looks like a
        # procedure name (starts with PRC_ / PROC_ / main / get_ etc.)
        pkg_prefixes = ("PKG_", "PRC_", "PROC_", "SP_", "USP_", "FN_", "FUNC_")
        proc_prefixes = ("PRC_", "PROC_", "PRC ", "MAIN", "GET_", "INSERT_", "LOAD_",
                         "UPDATE_", "DELETE_", "SEL_", "EXEC_", "RUN_")
        is_pkg = any(pkg_up.startswith(p) for p in pkg_prefixes)
        is_proc_rhs = any(proc_part.upper().startswith(p) for p in proc_prefixes)
        if is_pkg or is_proc_rhs:
            return pkg_part.strip()
        return script  # not a PKG.PROC pattern — keep as-is

    # ── Build package → rec mapping (grouped at package level) ───────────────
    id_map = {id(rec): i for i, rec in enumerate(records, 1)}
    script_map: dict[str, dict] = {}

    for rec in records:
        scripts_for_rec = _parse_scripts(rec)
        # Recs with no CODE-type SQL objects: no package to open — goes into NO-CODE-CHANGE row
        if not scripts_for_rec:
            scripts_for_rec = [TIDAL_ONLY]

        # Classify every SQL object — CODE only becomes a script row.
        # MV and VIEW objects are DDL targets captured by _mv_only_removals().
        # Applied uniformly to ALL recs regardless of category.
        code_scripts = [s for s in scripts_for_rec
                        if s == TIDAL_ONLY or _classify_sql_object(s) == "CODE"]
        scripts_for_rec = code_scripts if code_scripts else [TIDAL_ONLY]

        mv_removals = set(_mv_only_removals(rec))
        mid  = f"M-{id_map[id(rec)]:04d}"
        ph   = rec["phase"]
        rpts = set(rec["appears_in_rpts"])
        jobs = set(_split_jobs(rec.get("Affected Jobs", "") or ""))

        for script in scripts_for_rec:
            pkg = _pkg_key(script)           # group at package level
            if pkg not in script_map:
                script_map[pkg] = {
                    "procs": set(),          # individual procedures within this package
                    "recs": [],
                    "phases": set(),
                    "rpts": set(),
                    "tidal_jobs": set(),
                    "mv_removals": set(),
                    "hop_savings": 0,
                    "est_min": 0.0,
                }
            entry = script_map[pkg]
            # Track which procedure(s) inside the package are being changed
            if script != pkg:                # only add if it's genuinely PKG.PROC form
                entry["procs"].add(script)
            entry["recs"].append((mid, rec))
            entry["phases"].add(ph)
            entry["rpts"]      |= rpts
            entry["tidal_jobs"] |= jobs
            entry["mv_removals"]|= mv_removals
            entry["hop_savings"] += rec["global_hop_savings"]
            entry["est_min"]     += rec["global_est_min"]

    # Deduplicate recs within each package entry (a rec may have contributed
    # via multiple PROC names that all normalized to the same PKG key).
    # Also recompute hop_savings/est_min from the deduped list so that
    # PKG.PROC1 + PKG.PROC2 → same PKG doesn't double-count savings.
    for entry in script_map.values():
        seen_mids: set = set()
        deduped = []
        for item in entry["recs"]:
            mid_val = item[0]
            if mid_val not in seen_mids:
                seen_mids.add(mid_val)
                deduped.append(item)
        entry["recs"]        = deduped
        entry["hop_savings"] = sum(r["global_hop_savings"] for _, r in deduped)
        entry["est_min"]     = sum(r["global_est_min"]     for _, r in deduped)

    # ── Sort: wave asc, then GLOBAL before LOCAL, then script name ────────────
    def _sort_key(item):
        name, data = item
        w = _assign_wave(data["phases"])
        global_flag = 0 if len(data["rpts"]) > 1 else 1  # GLOBAL first
        return (w, global_flag, name)

    sorted_scripts = sorted(script_map.items(), key=_sort_key)

    # ── Compute bottom-up RPT ranking (mirrors Sheet 9 logic) ────────────────
    # Each script is assigned to its "primary RPT" — the RPT that appears
    # earliest in the bottom-up order — and scripts are then re-sorted and
    # grouped under RPT section banners.
    _all_rpts_seen = sorted({rpt for _, data in script_map.items() for rpt in data["rpts"]})
    _rpt_score: dict[str, float] = {}
    for _rpt in _all_rpts_seen:
        _lsc: set = set(); _gsc: set = set()
        _phs: dict = defaultdict(int); _rks: list = []
        # Gather unique recs for this RPT from the script_map entries
        _seen_mids: set = set()
        for _, _data in script_map.items():
            if _rpt not in _data["rpts"]:
                continue
            for _mid, _rec in _data["recs"]:
                if _mid in _seen_mids or _rpt not in _rec["appears_in_rpts"]:
                    continue
                _seen_mids.add(_mid)
                _phs[_rec["phase"]] += 1
                _rks.append(_rec.get("Risk", "LOW") or "LOW")
                _sql = _rec.get("SQL Objects Called", "") or ""
                for _s in re.split(r'[\n,;]', _sql):
                    _s = _s.strip()
                    if not _s: continue
                    if _rec["rpt_count"] > 1: _gsc.add(_s)
                    else:                     _lsc.add(_s)
        _lsc -= _gsc; _n_sc = len(_lsc | _gsc)
        _pct = round(len(_lsc) / max(_n_sc, 1) * 100)
        _mr = "HIGH" if "HIGH" in _rks else "MEDIUM" if "MEDIUM" in _rks else "LOW"
        _rw = {"LOW": 1, "MEDIUM": 2, "HIGH": 3}
        _rpt_score[_rpt] = (
            len(_seen_mids) * 1.5 + _n_sc * 2 + (100 - _pct) * 0.1
            + _phs.get(4, 0) * 3 + _phs.get(3, 0) * 2 + _rw.get(_mr, 2) * 5
        )
    _ranked_rpts = sorted(_all_rpts_seen, key=lambda r: _rpt_score.get(r, 999))
    _rpt_rank:   dict[str, int] = {rpt: i for i, rpt in enumerate(_ranked_rpts)}

    def _primary_rpt(data: dict) -> str:
        """
        Return the most relevant RPT for section grouping.
        Prefers the RPT whose name matches the target table(s) of bundled recs
        (e.g. PKG loading RPT_CLAIM_DTL_R should group under RPT_CLAIM_DTL_R, not
        RPT_CLIENT_DTL_R just because that happens to have a lower complexity rank).
        Falls back to lowest-ranked RPT when no target-table match is found.
        """
        rpts = data["rpts"]
        if len(rpts) == 1:
            return next(iter(rpts))
        # Score each RPT by how many bundled rec target tables mention its name
        rpt_scores: dict[str, int] = {rpt: 0 for rpt in rpts}
        for _mid, rec in data["recs"]:
            tgt = (rec.get("Target Table", "") or "").upper().strip()
            if not tgt:
                continue
            for rpt in rpts:
                rpt_up = rpt.upper()
                if rpt_up == tgt or rpt_up in tgt or tgt in rpt_up:
                    rpt_scores[rpt] += 1
        best = max(rpt_scores.values(), default=0)
        if best > 0:
            # Among those with the highest match score, use lowest rank as tiebreaker
            candidates = [r for r, s in rpt_scores.items() if s == best]
            return min(candidates, key=lambda r: _rpt_rank.get(r, 999))
        # No target-table match — fall back to lowest-ranked RPT
        return min(rpts, key=lambda r: _rpt_rank.get(r, 999))

    # Re-sort: primary_rpt_rank → wave → global_before_local → name
    sorted_scripts = sorted(
        script_map.items(),
        key=lambda item: (
            _rpt_rank.get(_primary_rpt(item[1]), 999),
            _assign_wave(item[1]["phases"]),
            0 if len(item[1]["rpts"]) > 1 else 1,
            item[0],
        ),
    )

    # Tier colours for RPT section banners (mirrors Sheet 9)
    _n_rpts = len(_ranked_rpts)
    def _rpt_tier(rank: int) -> int:
        if rank < max(3, _n_rpts // 4):     return 1
        if rank < max(7, _n_rpts * 2 // 3): return 2
        return 3
    _TIER_BANNER_FILL = {1: "FF2E7D32", 2: "FFE65100", 3: "FFC62828"}

    COLS_SWO = [
        "Package Name",                      # A  normalized to PKG level
        "Procedures Touched in Package",     # B  NEW — individual PROCs within the PKG
        "Wave",                              # C
        "Scope",                             # D  GLOBAL / LOCAL
        "Affected RPTs",                     # E
        "Tidal Jobs Calling This Package",   # F
        "# Changes Bundled",                 # G
        "Bundled Master IDs",                # H
        "Phase Range",                       # I
        "Categories of Changes",             # J
        "All Changes — Implement Together",  # K  KEY COLUMN
        "Combined Hop Savings",              # L
        "Est. Min Saved",                    # M
        "Max Risk",                          # N
        "Also Remove from Tidal / DDL DROP (after parity)", # O  DROP MV + RENAME VIEW + Remove Tidal jobs
        "Wave 4 Awareness",                  # P  Phase 4 pending for this package
        "Prerequisites",                     # Q
        "Validation Status",                 # R  ← from validated workbooks
        "Dev Note",                          # S
    ]
    hdr_row = 2
    for col, h in enumerate(COLS_SWO, 1):
        hdr_cell(ws, hdr_row, col, h, size=10)
    ws.row_dimensions[hdr_row].height = 30
    ws.freeze_panes = "A3"

    risk_fill_hex = {"HIGH": "FFFFC7CE", "MEDIUM": "FFFFEB9C", "LOW": "FFC6EFCE"}
    scope_fill_hex = {"GLOBAL": "FFCFE2F3", "LOCAL": "FFF3F3F3"}

    # ── Pre-compute per-RPT script/rec counts for the banner summaries ────────
    _rpt_script_count: dict[str, int] = defaultdict(int)
    _rpt_rec_count:    dict[str, set] = defaultdict(set)
    _rpt_hop_count:    dict[str, int] = defaultdict(int)
    for script, data in sorted_scripts:
        prpt = _primary_rpt(data)
        _rpt_script_count[prpt] += 1
        for mid, rec in data["recs"]:
            _rpt_rec_count[prpt].add(mid)
            _rpt_hop_count[prpt] += rec["global_hop_savings"]

    # ── Write rows with RPT section banners ───────────────────────────────────
    current_rpt: str = ""
    n_cols = len(COLS_SWO)

    for script, data in sorted_scripts:
        prpt  = _primary_rpt(data)

        # ── Insert RPT section banner when RPT changes ────────────────────────
        if prpt != current_rpt:
            current_rpt = prpt
            rank  = _rpt_rank.get(prpt, 0)
            tier  = _rpt_tier(rank)
            n_sc  = _rpt_script_count[prpt]
            n_rc  = len(_rpt_rec_count[prpt])
            n_hp  = _rpt_hop_count[prpt]
            tier_label = {1: "Tier 1 — Start Here", 2: "Tier 2 — Build Confidence",
                          3: "Tier 3 — Do Last (mostly regression)"}

            if ws.max_row > 2:
                ws.append([])  # blank spacer before each new RPT section

            ban_row = ws.max_row + 1
            ws.merge_cells(f"A{ban_row}:{get_column_letter(n_cols)}{ban_row}")
            bc = ws.cell(row=ban_row, column=1,
                         value=(f"  {prpt}   ▸ "
                                f"Rank {rank+1}  |  {tier_label.get(tier,'')}  |  "
                                f"{n_sc} script(s) in this section  |  "
                                f"{n_rc} unique rec(s)  |  "
                                f"{n_hp} verified hops"))
            bc.fill  = hfill(_TIER_BANNER_FILL.get(tier, DARK))
            bc.font  = Font(bold=True, size=11, color=WHITE)
            bc.alignment = Alignment(vertical="center", indent=1)
            ws.row_dimensions[ban_row].height = 22

        recs_list   = data["recs"]
        phases      = data["phases"]
        rpts        = sorted(data["rpts"])
        tidal_jobs  = sorted(data["tidal_jobs"])
        mv_removals = sorted(data["mv_removals"])

        wave  = _assign_wave(phases)
        scope = "GLOBAL" if len(rpts) > 1 else "LOCAL"
        has_ph4_pending = 4 in phases and wave != 4  # non-Phase-4 wave but Phase 4 also pending

        # Max risk across bundled recs
        max_risk = "LOW"
        for lvl in ("HIGH", "MEDIUM", "LOW"):
            if any((r.get("Risk", "") or "") == lvl for _, r in recs_list):
                max_risk = lvl
                break

        # Build the "All Changes" summary — one block per bundled rec, sorted by phase
        change_blocks = []
        cats_set: set = set()
        master_ids: list[str] = []
        for mid, rec in sorted(recs_list, key=lambda x: (x[1]["phase"], x[0])):
            cat   = rec.get("Category", "") or ""
            tgt   = rec.get("Target Table", "") or ""
            recmd = rec.get("Recommendation", "") or rec.get("What To Do", "") or ""
            ph    = rec["phase"]
            hops  = rec["global_hop_savings"]
            cats_set.add(cat)
            master_ids.append(mid)
            # First line of recommendation text as action summary
            first_line = next(
                (ln.strip() for ln in recmd.split("\n") if ln.strip()), cat
            )[:220]
            change_blocks.append(
                f"── {mid} | Phase {ph} | {cat}\n"
                f"   Target : {tgt}\n"
                f"   Hops   : {hops}\n"
                f"   Action : {first_line}"
            )

        # Wave 4 awareness note
        ph4_text = "(none)"
        if has_ph4_pending:
            ph4_entries = [(mid, r) for mid, r in recs_list if r["phase"] == 4]
            ph4_ids  = [mid for mid, _ in ph4_entries]
            ph4_cats = sorted({r.get("Category", "") for _, r in ph4_entries})
            ph4_text = (
                f"⚠ PHASE 4 PENDING — after Wave {wave} changes are deployed and validated,\n"
                f"this SAME script needs a SECOND PR in Wave 4 (requires DBA sign-off):\n"
                + "\n".join(f"  • {mid}: {r.get('Category','')}" for mid, r in ph4_entries)
                + f"\n\nDesign Wave {wave} changes to leave the MV table reference intact;\n"
                f"the MV query will be inlined in Wave 4.\n"
                f"Wave 4 Master IDs: {', '.join(ph4_ids)}"
            )

        # Prerequisites
        prereq_parts = []
        if 4 in phases:
            prereq_parts.append("Phase 3 validated; DBA + architecture sign-off; rollback plan ready")
        elif 3 in phases:
            prereq_parts.append("Phase 2 deployed and validated (3+ cycles)")
        elif 2 in phases:
            prereq_parts.append("Phase 1 DDL renames deployed to all environments")
        if mv_removals:
            prereq_parts.append(
                "SQL changes deployed + 3 parity cycles completed BEFORE removing Tidal jobs"
            )

        nr   = ws.max_row + 1
        wfill = hfill(WAVE_FILLS_SWO.get(wave, "FFFFFFFF"))

        phase_range = (
            f"Phase {min(phases)}–{max(phases)}" if len(phases) > 1
            else f"Phase {min(phases)}"
        )

        # Procedures column: sorted list of individual PROC names, or note if
        # the entire package is the unit (no dot-notation in SQL Objects Called)
        procs_sorted = sorted(data["procs"])
        procs_display = (
            "\n".join(procs_sorted) if procs_sorted
            else "(entire package — no specific procedure identified)"
        )

        # ── Validation status summary for this script ────────────────────
        vs_counts: dict[str, int] = defaultdict(int)
        for mid, rec in recs_list:
            vs_counts[rec.get("validation_status", VS_PENDING)] += 1
        # Build concise summary, e.g. "✅ 2  ⏳ 1"
        vs_parts = []
        for label in (VS_CONSIDER, VS_NOT_CONS, VS_NOT_GOOD, VS_PENDING):
            if vs_counts.get(label, 0):
                vs_parts.append(f"{label}: {vs_counts[label]}")
        vs_summary = "  |  ".join(vs_parts) if vs_parts else VS_PENDING
        # Dominant status for cell colour
        vs_dominant = (VS_CONSIDER if vs_counts.get(VS_CONSIDER)
                       else VS_NOT_CONS if vs_counts.get(VS_NOT_CONS)
                       else VS_NOT_GOOD if vs_counts.get(VS_NOT_GOOD)
                       else VS_PENDING)

        vals = [
            script,                          # A  Package Name
            procs_display,                   # B  Procedures Touched
            WAVE_LABELS_SWO.get(wave, f"Wave {wave}"),  # C
            scope,                           # D
            "\n".join(rpts),                 # E
            ("\n".join(tidal_jobs[:10]) + ("\n  ...+" + str(len(tidal_jobs)-10) + " more"
             if len(tidal_jobs) > 10 else "")),  # F
            len(recs_list),                  # G
            "\n".join(master_ids),           # H
            phase_range,                     # I
            "\n".join(sorted(cats_set)),     # J
            "\n\n".join(change_blocks),      # K
            data["hop_savings"],             # L
            round(data["est_min"], 1),       # M
            max_risk,                        # N
            "\n".join(mv_removals) if mv_removals else "(none — no Tidal or DDL removals needed)",  # O
            ph4_text,                        # P
            "; ".join(prereq_parts) if prereq_parts else "See wave prerequisites in Implementation Roadmap",  # Q
            vs_summary,                      # R  Validation Status
            ("Implement ALL changes in this row in a SINGLE PR — do not split across sprints."
             if script != TIDAL_ONLY else
             ("No SQL package/procedure to open.  Actions required (see change blocks + column O):\n"
             "  Phase 0 items  : Review the constraint/decision and document it.\n"
             "  Phase 1 items  : DDL RENAME VIEW objects listed in column O.\n"
             "  Phase 2/3 items: Tidal schedule changes only (no code edit).\n"
             "  Phase 4 items  : DBA work — DROP MV objects in column O, THEN\n"
             "                   remove Tidal MV_REFRESH jobs after 3 parity cycles.\n"
             "  Sequence: Phase 0 first, then Phase 1 DDL renames, then Phase 4 (DBA gate).")),  # S
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=nr, column=col, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.fill = wfill  # wave colour by default
            if col == 1:
                c.font = Font(bold=True, size=10)
            if col == 2:   # Procedures — light tint to distinguish from package name
                c.fill = hfill("FFF9FBE7")
            if col == 4:   # Scope override
                c.fill = hfill(scope_fill_hex.get(scope, "FFF3F3F3"))
            if col == 14:  # Risk override
                c.fill = hfill(risk_fill_hex.get(max_risk, "FFFFFFFF"))
            if col == 18:  # Validation Status override
                c.fill = hfill(VS_FILL.get(vs_dominant, "FFF2F2F2"))
        ws.row_dimensions[nr].height = max(75, 30 * len(recs_list))

    # ── Legend row ────────────────────────────────────────────────────────────
    ws.append([])
    leg_row = ws.max_row + 1
    ws.merge_cells(f"A{leg_row}:S{leg_row}")
    lc = ws.cell(row=leg_row, column=1, value=(
        "LEGEND:  GLOBAL = script is called across >1 RPT (co-ordinate ownership before starting)  "
        "|  LOCAL = only 1 RPT owns this script  "
        "|  Wave 4 Awareness = Phase 4 MV elimination also touches this script (separate DBA PR)  "
        "|  'Also Remove from Tidal / DDL DROP' = DROP MV (after parity) + RENAME VIEW + Remove Tidal MV_REFRESH jobs  "
        "|  Validation Status: ✅ Consider = validated as actionable  "
        "⏸ Not to Consider = validated as low-priority  "
        "❌ Not a Good Rec = validated as incorrect  "
        "⏳ Pending = not yet reviewed (treat as actionable until reviewed)"
    ))
    lc.font = Font(italic=True, size=9, color="FF616161")
    lc.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[leg_row].height = 40

    set_col_widths(ws, {
        "A": 50, "B": 45, "C": 30, "D": 10, "E": 35, "F": 52, "G": 10,
        "H": 25, "I": 12, "J": 55, "K": 85, "L": 14, "M": 12,
        "N": 10, "O": 50, "P": 58, "Q": 55, "R": 30, "S": 55,
    })
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS_SWO))}{ws.max_row}"


# ── Sheet 8: RPT Pipeline Summary ─────────────────────────────────────────────
def write_rpt_pipeline_summary(wb: openpyxl.Workbook,
                                records: list[dict],
                                rpt_names: list[str]) -> None:
    """
    One row per RPT pipeline — the program-management / RPT-lead view.

    Answers: 'How much work is in MY RPT?'
    Complements Sheet 7 (developer's script-centric view) with a capacity-
    planning view per RPT team.

    Does NOT contradict the script-centric approach:
      • A GLOBAL script appears in multiple RPTs' rows.
      • Only ONE team owns the code change (co-ordinated via Sheet 7).
      • All other RPT teams are reviewers and run their own regression suites.
    """
    ws = wb.create_sheet("RPT Pipeline Summary")
    ws.sheet_view.showGridLines = False

    # Banner
    ws.merge_cells("A1:S1")
    c = ws["A1"]
    c.value = (
        "RPT PIPELINE SUMMARY  —  One row per RPT pipeline  "
        "|  Use for sprint capacity planning per RPT team  "
        "|  GLOBAL scripts appear in multiple rows — only ONE team owns the code change; "
        "all others run regression"
    )
    c.font  = Font(bold=True, size=13, color=WHITE)
    c.fill  = hfill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    COLS_RPT = [
        "RPT Pipeline",                      # A
        "Total Recs in this RPT",            # B  includes shared (GLOBAL) recs
        "Unique Scripts to Change",          # C
        "  of which LOCAL (this RPT only)",  # D
        "  of which GLOBAL (shared)",        # E
        "Total Tidal Jobs Affected",         # F
        "Phase 0 — Decisions",               # G
        "Phase 1 — DDL / Naming",            # H
        "Phase 2 — Quick Wins",              # I
        "Phase 3 — Consolidation",           # J
        "Phase 4 — MV Architecture",         # K
        "Phase 5 — LLM Optimization",        # L
        "Verified Hop Savings",              # M
        "Est. Min Saved",                    # N
        "Max Risk",                          # O
        "LOCAL Scripts (sprint this RPT's team)", # P
        "GLOBAL Scripts (co-ordinate ownership)", # Q
        "Phase 4 MV Targets",                # R
        "Co-ordinate With RPTs",             # S  other RPTs that share GLOBAL scripts
    ]
    hdr_row = 2
    for col, h in enumerate(COLS_RPT, 1):
        hdr_cell(ws, hdr_row, col, h, size=10)
    ws.row_dimensions[hdr_row].height = 32
    ws.freeze_panes = "A3"

    risk_fill_hex = {"HIGH": "FFFFC7CE", "MEDIUM": "FFFFEB9C", "LOW": "FFC6EFCE"}
    TIDAL_ONLY = "[NO-CODE-CHANGE]"

    # ── One row per RPT ───────────────────────────────────────────────────────
    for rpt in sorted(rpt_names):
        rpt_recs = [r for r in records if rpt in r["appears_in_rpts"]]
        if not rpt_recs:
            continue

        # Phase counts
        phase_counts: dict[int, int] = defaultdict(int)
        for rec in rpt_recs:
            phase_counts[rec["phase"]] += 1

        # Scripts — classify LOCAL vs GLOBAL
        local_scripts: set  = set()
        global_scripts: set = set()
        for rec in rpt_recs:
            raw = rec.get("SQL Objects Called", "") or ""
            scripts = [s.strip() for s in re.split(r'[\n,;]', raw) if s.strip()]
            if not scripts:
                scripts = [TIDAL_ONLY]
            for s in scripts:
                if rec["rpt_count"] > 1:
                    global_scripts.add(s)
                else:
                    local_scripts.add(s)
        # If a script has both LOCAL and GLOBAL recs, promote it to GLOBAL
        local_scripts -= global_scripts
        all_scripts = local_scripts | global_scripts

        # Tidal jobs
        all_jobs: set = set()
        for rec in rpt_recs:
            all_jobs |= set(_split_jobs(rec.get("Affected Jobs", "") or ""))

        # Savings
        hop_savings = sum(r["global_hop_savings"] for r in rpt_recs)
        est_min     = sum(r["global_est_min"]     for r in rpt_recs)

        # Max risk
        risks  = [r.get("Risk", "LOW") for r in rpt_recs]
        max_risk = ("HIGH" if "HIGH" in risks
                    else "MEDIUM" if "MEDIUM" in risks
                    else "LOW")

        # Phase 4 targets
        ph4_targets = sorted({
            r.get("Target Table", "") for r in rpt_recs
            if r["phase"] == 4 and r.get("Target Table")
        })

        # Other RPTs that share GLOBAL scripts
        other_rpts: set = set()
        for rec in rpt_recs:
            if rec["rpt_count"] > 1:
                for other in rec["appears_in_rpts"]:
                    if other != rpt:
                        other_rpts.add(other)

        nr = ws.max_row + 1
        vals = [
            rpt,
            len(rpt_recs),
            len(all_scripts),
            len(local_scripts),
            len(global_scripts),
            len(all_jobs),
            phase_counts.get(0, 0),
            phase_counts.get(1, 0),
            phase_counts.get(2, 0),
            phase_counts.get(3, 0),
            phase_counts.get(4, 0),
            phase_counts.get(5, 0),
            hop_savings,
            round(est_min, 1),
            max_risk,
            ("\n".join(sorted(s for s in local_scripts if s != TIDAL_ONLY)[:20])
             + ("\n  ...+" + str(len(local_scripts) - 20) + " more"
                if len(local_scripts) > 20 else "")),
            ("\n".join(sorted(s for s in global_scripts if s != TIDAL_ONLY)[:20])
             + ("\n  ...+" + str(len(global_scripts) - 20) + " more"
                if len(global_scripts) > 20 else "")),
            "\n".join(ph4_targets[:12]),
            "\n".join(sorted(other_rpts)) if other_rpts else "(none — all changes are LOCAL)",
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=nr, column=col, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 1:
                c.font = Font(bold=True)
            if col == 15:
                c.fill = hfill(risk_fill_hex.get(max_risk, "FFFFFFFF"))
        ws.row_dimensions[nr].height = 80

    # ── Footer note ───────────────────────────────────────────────────────────
    ws.append([])
    note_row = ws.max_row + 1
    ws.merge_cells(f"A{note_row}:S{note_row}")
    nc = ws.cell(row=note_row, column=1, value=(
        "HOW TO READ:  "
        "'Unique Scripts to Change' = count of SQL packages/procedures needing at least one code edit in this RPT.  "
        "'LOCAL' = this RPT team owns and deploys the change.  "
        "'GLOBAL' = script is shared across RPTs — only ONE team writes the PR; all others do regression.  "
        "The Script-Centric Work Orders sheet (Sheet 7) is the developer's primary ticket source.  "
        "This sheet gives the RPT lead a capacity/risk overview for sprint planning."
    ))
    nc.font = Font(italic=True, size=9, color="FF616161")
    nc.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[note_row].height = 40

    set_col_widths(ws, {
        "A": 35, "B": 14, "C": 14, "D": 16, "E": 16, "F": 14,
        "G": 12, "H": 12, "I": 12, "J": 12, "K": 12, "L": 12,
        "M": 14, "N": 12, "O": 12, "P": 55, "Q": 55, "R": 40, "S": 45,
    })


# ── Sheet 9: RPT Implementation Plan (Bottom-Up) ─────────────────────────────
def write_rpt_implementation_plan(wb: openpyxl.Workbook,
                                   records: list[dict]) -> None:
    """
    Bottom-up RPT implementation plan with credit rollover.

    Section A — Ranked order:  Each RPT scored by complexity (lower = easier start).
    Section B — Credit rollover: As you implement RPTs in order, GLOBAL scripts
                changed for RPT-N are automatically free for all later RPTs that
                share them. This table shows the NET NEW work each RPT actually
                adds once earlier RPTs have been completed.

    Key insight: the programme appears to span ~700 rec-instances across 12 RPTs,
    but the credit rollover reduces it to 122 unique recs and 61 unique scripts.
    By the end of Tier 2 (rank 7), 101/122 recs and 59/61 scripts are done.
    """
    ws = wb.create_sheet("RPT Implementation Plan")
    ws.sheet_view.showGridLines = False

    # ── Shared pkg-key normaliser (same rule as Sheet 7) ─────────────────────
    def _pk(s: str) -> str:
        if '.' not in s:
            return s
        left, right = s.split('.', 1)
        pkg_pfx  = ("PKG_", "PRC_", "PROC_", "SP_", "USP_", "FN_", "FUNC_")
        proc_pfx = ("PRC_", "PROC_", "MAIN", "GET_", "INSERT_", "LOAD_",
                    "UPDATE_", "DELETE_", "SEL_", "EXEC_", "RUN_")
        if any(left.upper().startswith(p) for p in pkg_pfx) or \
           any(right.upper().startswith(p) for p in proc_pfx):
            return left.strip()
        return s

    # ── Build per-rec index from master records ───────────────────────────────
    rec_data: dict[str, dict] = {}
    for i, rec in enumerate(records, 1):
        mid = f"M-{i:04d}"
        sql_raw  = rec.get("SQL Objects Called", "") or ""
        scripts  = {_pk(s.strip()) for s in re.split(r'[\n,;]', sql_raw) if s.strip()}
        try:    hops = int(rec.get("global_hop_savings", 0) or 0)
        except: hops = 0
        try:    est_min = float(rec.get("global_est_min", 0.0) or 0.0)
        except: est_min = 0.0
        rec_data[mid] = {
            "rpts":    set(rec["appears_in_rpts"]),
            "scripts": scripts,
            "phase":   rec["phase"],
            "hops":    hops,
            "est_min": est_min,
            "risk":    rec.get("Risk", "LOW") or "LOW",
            "rpt_cnt": rec["rpt_count"],
        }

    # Build RPT → list of MIDs
    rpt_mids: dict[str, list] = defaultdict(list)
    for mid, d in rec_data.items():
        for rpt in d["rpts"]:
            rpt_mids[rpt].append(mid)

    all_rpts = sorted(rpt_mids.keys())

    # ── Score each RPT (lower = simpler to start) ─────────────────────────────
    rpt_stats: dict[str, dict] = {}
    for rpt in all_rpts:
        mids = rpt_mids[rpt]
        local_sc: set = set()
        global_sc: set = set()
        ph_counts: dict[int, int] = defaultdict(int)
        risk_vals = []
        for m in mids:
            d = rec_data[m]
            ph_counts[d["phase"]] += 1
            risk_vals.append(d["risk"])
            for s in d["scripts"]:
                if d["rpt_cnt"] > 1: global_sc.add(s)
                else:                local_sc.add(s)
        local_sc -= global_sc
        all_sc = local_sc | global_sc
        n_sc = len(all_sc)
        pct_local = round(len(local_sc) / max(n_sc, 1) * 100)
        max_risk = ("HIGH" if "HIGH" in risk_vals
                    else "MEDIUM" if "MEDIUM" in risk_vals else "LOW")
        rw = {"LOW": 1, "MEDIUM": 2, "HIGH": 3}
        p2 = ph_counts.get(2, 0); p3 = ph_counts.get(3, 0); p4 = ph_counts.get(4, 0)
        score = (len(mids)*1.5) + (n_sc*2) + ((100-pct_local)*0.1) + (p4*3) + (p3*2) + (rw.get(max_risk,2)*5)
        rpt_stats[rpt] = {
            "mids": mids, "all_sc": all_sc, "local_sc": local_sc,
            "global_sc": global_sc, "n_sc": n_sc, "pct_local": pct_local,
            "ph_counts": ph_counts, "max_risk": max_risk, "score": score,
            "p2": p2, "p3": p3, "p4": p4,
        }

    # Sort by score → bottom-up order
    ranked = sorted(all_rpts, key=lambda r: rpt_stats[r]["score"])

    # ── Banner ────────────────────────────────────────────────────────────────
    TOTAL_COLS = 18
    ws.merge_cells(f"A1:{get_column_letter(TOTAL_COLS)}1")
    c = ws["A1"]
    c.value = (
        "RPT IMPLEMENTATION PLAN  —  Bottom-Up Sequence  "
        "|  Start with the simplest RPT, build confidence, let credit rollover reduce later work  "
        "|  A GLOBAL script fixed for RPT-N is automatically FREE for all later RPTs that share it"
    )
    c.font  = Font(bold=True, size=13, color=WHITE)
    c.fill  = hfill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Tier colour map ───────────────────────────────────────────────────────
    TIER_FILLS = {
        1: "FFE8F5E9",   # Tier 1 green  — easiest
        2: "FFFFF3E0",   # Tier 2 amber  — medium
        3: "FFFCE4EC",   # Tier 3 red    — complex
    }
    TIER_HDR_FILLS = {
        1: "FF2E7D32",
        2: "FFE65100",
        3: "FFC62828",
    }

    def _tier(rank: int, total: int) -> int:
        if rank <= max(3, total // 4):        return 1
        if rank <= max(7, total * 2 // 3):    return 2
        return 3

    # ─── SECTION A: Complexity ranking ───────────────────────────────────────
    sec_a = ws.max_row + 2
    ws.merge_cells(f"A{sec_a}:{get_column_letter(TOTAL_COLS)}{sec_a}")
    sa = ws.cell(row=sec_a, column=1,
                 value="SECTION A — COMPLEXITY RANKING  (lower score = safer/easier to start)")
    sa.font  = Font(bold=True, size=12, color=WHITE)
    sa.fill  = hfill("FF37474F")
    sa.alignment = Alignment(indent=1, vertical="center")
    ws.row_dimensions[sec_a].height = 24

    rank_hdrs = [
        "Rank", "Tier", "RPT Pipeline", "Complexity\nScore",
        "Total\nRecs", "Total\nScripts", "LOCAL\nScripts%",
        "Phase 2\nQuick Wins", "Phase 3\nConsolidat.", "Phase 4\nMV Arch",
        "Max Risk", "Notes",
    ]
    rh = ws.max_row + 1
    for col, h in enumerate(rank_hdrs, 1):
        hdr_cell(ws, rh, col, h, fill_hex=DARK, size=9)
    ws.row_dimensions[rh].height = 30

    tier_notes = {
        1: "START HERE — low recs, few scripts, mostly independent. Use as proof-of-concept and playbook.",
        2: "BUILD CONFIDENCE — moderate complexity; credit from Tier 1 RPTs reduces net work.",
        3: "DO LAST — high volume; most work is FREE by the time you arrive here.",
    }
    for rank, rpt in enumerate(ranked, 1):
        st   = rpt_stats[rpt]
        tier = _tier(rank, len(ranked))
        fill = hfill(TIER_FILLS[tier])
        nr   = ws.max_row + 1
        risk_fill_hex_local = {"HIGH":"FFFFC7CE","MEDIUM":"FFFFEB9C","LOW":"FFC6EFCE"}
        vals = [
            rank, f"Tier {tier}", rpt.replace("RPT_",""),
            round(st["score"], 0),
            len(st["mids"]), st["n_sc"], f"{st['pct_local']}%",
            st["p2"], st["p3"], st["p4"],
            st["max_risk"],
            tier_notes[tier],
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=nr, column=col, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.fill = fill
            if col == 1:
                c.font = Font(bold=True)
            if col == 11:
                c.fill = hfill(risk_fill_hex_local.get(st["max_risk"], "FFFFFFFF"))
        ws.row_dimensions[nr].height = 22

    # ─── SECTION B: Credit rollover ───────────────────────────────────────────
    ws.append([])
    sec_b = ws.max_row + 2
    ws.merge_cells(f"A{sec_b}:{get_column_letter(TOTAL_COLS)}{sec_b}")
    sb = ws.cell(row=sec_b, column=1,
                 value=(
                     "SECTION B — CREDIT ROLLOVER  "
                     "A GLOBAL script changed for an earlier RPT is FREE for every later RPT that uses it.  "
                     "'Net New' = genuine new work this RPT adds once all earlier RPTs are done."
                 ))
    sb.font  = Font(bold=True, size=12, color=WHITE)
    sb.fill  = hfill("FF1A237E")
    sb.alignment = Alignment(indent=1, vertical="center")
    ws.row_dimensions[sec_b].height = 24

    cr_hdrs = [
        "Rank", "Tier", "RPT Pipeline",
        "Total\nRecs", "Free Recs\n(already done)", "Net New\nRecs  ← key",
        "Total\nScripts", "Free Sc\n(already done)", "Net New\nScripts ← key",
        "Net New\nHops", "Net New\nMin Saved",
        "Cumul.\nRecs Done", "Cumul.\nScripts Done",
        "% Programme\nComplete (Recs)", "Recommended Action",
    ]
    ch = ws.max_row + 1
    for col, h in enumerate(cr_hdrs, 1):
        hdr_cell(ws, ch, col, h, fill_hex="FF0D47A1", size=9)
    ws.row_dimensions[ch].height = 32

    done_scripts: set = set()
    done_recs:    set = set()
    total_unique_recs    = len(rec_data)
    total_unique_scripts = len({s for d in rec_data.values() for s in d["scripts"]})

    tier_actions = {
        1: "Assign to one team. Build the deployment playbook here. Regression test all 3+ RPT pipelines it touches.",
        2: "Assign. Check Script-Centric Work Orders for FREE scripts — those are regression-only, no code needed.",
        3: "Mostly regression testing by the time you get here. Confirm FREE scripts via Sheet 7 before coding.",
    }

    for rank, rpt in enumerate(ranked, 1):
        st   = rpt_stats[rpt]
        tier = _tier(rank, len(ranked))
        mids = st["mids"]

        free_recs_list = [m for m in mids if m in done_recs]
        new_recs_list  = [m for m in mids if m not in done_recs]

        all_sc    = st["all_sc"]
        free_sc   = all_sc & done_scripts
        new_sc    = all_sc - done_scripts

        net_hops = sum(rec_data[m]["hops"]    for m in new_recs_list)
        net_min  = sum(rec_data[m]["est_min"] for m in new_recs_list)

        done_recs    |= set(new_recs_list)
        done_scripts |= new_sc

        pct_done = round(len(done_recs) / max(total_unique_recs, 1) * 100)

        fill = hfill(TIER_FILLS[tier])
        nr   = ws.max_row + 1

        # Highlight Net New Recs column green if ≤5, amber if ≤20, red if >20
        net_new_color = (
            "FFC6EFCE" if len(new_recs_list) <= 5
            else "FFFFEB9C" if len(new_recs_list) <= 20
            else "FFFFC7CE"
        )

        vals = [
            rank, f"Tier {tier}", rpt.replace("RPT_",""),
            len(mids), len(free_recs_list), len(new_recs_list),
            len(all_sc), len(free_sc), len(new_sc),
            net_hops, round(net_min, 1),
            len(done_recs), len(done_scripts),
            f"{pct_done}%",
            tier_actions[tier],
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=nr, column=col, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.fill = fill
            if col == 1:
                c.font = Font(bold=True)
            if col == 6:   # Net New Recs — key metric
                c.fill = hfill(net_new_color)
                c.font = Font(bold=True)
            if col == 9:   # Net New Scripts
                sc_color = "FFC6EFCE" if len(new_sc) <= 2 else "FFFFEB9C" if len(new_sc) <= 10 else "FFFFC7CE"
                c.fill = hfill(sc_color)
        ws.row_dimensions[nr].height = 28

    # Totals row
    tot_row = ws.max_row + 1
    tot_vals = [
        "TOTAL", "", "All 12 RPTs",
        sum(len(rpt_stats[r]["mids"]) for r in ranked),
        "", len(done_recs),
        "", "", len(done_scripts),
        sum(rec_data[m]["hops"]    for m in done_recs),
        round(sum(rec_data[m]["est_min"] for m in done_recs), 1),
        len(done_recs), len(done_scripts), "100%",
        "Full programme complete. FCT/HIST/DTL RPTs should be regression-only at this point.",
    ]
    for col, v in enumerate(tot_vals, 1):
        c = ws.cell(row=tot_row, column=col, value=v)
        c.fill  = hfill(NAVY)
        c.font  = Font(bold=True, color=WHITE)
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[tot_row].height = 28

    # ── Footnote ─────────────────────────────────────────────────────────────
    fn_row = ws.max_row + 2
    ws.merge_cells(f"A{fn_row}:{get_column_letter(TOTAL_COLS)}{fn_row}")
    fn = ws.cell(row=fn_row, column=1, value=(
        "HOW TO USE:  (1) Use Section A to prioritise your first sprint — pick a Tier 1 RPT.  "
        "(2) Use Section B to plan each sprint: 'Net New Recs' and 'Net New Scripts' are your actual ticket count.  "
        "(3) For FREE scripts (already done), open the Script-Centric Work Orders sheet (Sheet 7) and confirm scope=GLOBAL "
        "— those packages need regression-only, no code change.  "
        "(4) By end of Rank 7, 59/61 scripts are done. Ranks 8–12 are primarily regression testing."
    ))
    fn.font = Font(italic=True, size=9, color="FF616161")
    fn.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[fn_row].height = 36

    set_col_widths(ws, {
        "A": 7, "B": 8, "C": 35, "D": 9, "E": 12, "F": 12,
        "G": 9, "H": 12, "I": 12, "J": 9, "K": 10,
        "L": 11, "M": 11, "N": 12, "O": 55,
    })


# ── Sheet 10: Actionable Work Orders (validation-filtered) ───────────────────
def write_actionable_work_orders(wb: openpyxl.Workbook,
                                  records: list[dict]) -> None:
    """
    Filtered version of the Script-Centric Work Orders.
    Only includes recs with validation_status = ✅ Consider OR ⏳ Validation Pending.
    Recs marked ⏸ Not to Consider or ❌ Not a Good Recommendation are excluded.

    Scripts where ALL bundled recs are excluded are omitted entirely.
    Scripts where SOME recs are excluded show only the remaining recs, with a
    note on how many were filtered out.

    Same column structure as Sheet 7 (Script-Centric Work Orders) for consistency.
    """
    ws = wb.create_sheet("Actionable Work Orders")
    ws.sheet_view.showGridLines = False

    ACTIONABLE_STATUSES = {VS_CONSIDER, VS_PENDING}

    # ── Reuse the same PKG-level grouping logic ───────────────────────────────
    def _pk(s: str) -> str:
        if '.' not in s:
            return s
        left, right = s.split('.', 1)
        pkg_pfx  = ("PKG_", "PRC_", "PROC_", "SP_", "USP_", "FN_", "FUNC_")
        proc_pfx = ("PRC_", "PROC_", "MAIN", "GET_", "INSERT_", "LOAD_",
                    "UPDATE_", "DELETE_", "SEL_", "EXEC_", "RUN_")
        if any(left.upper().startswith(p) for p in pkg_pfx) or \
           any(right.upper().startswith(p) for p in proc_pfx):
            return left.strip()
        return s

    def _assign_wave(phases: set) -> int:
        non_special = phases - {0, 4}
        if not non_special:
            return 4 if 4 in phases else 0
        if 3 in non_special: return 3
        if 2 in non_special: return 2
        if 1 in non_special: return 1
        return 5

    TIDAL_ONLY = "[NO-CODE-CHANGE]"

    WAVE_FILLS_AWO = {
        0: "FFE5CCF5", 1: "FFFCE4EC", 2: "FFE8F5E9",
        3: "FFFFF3E0", 4: "FFFBE9E7", 5: "FFE3F2FD",
    }
    WAVE_LABELS_AWO = {
        0: "Wave 0 — Decisions Only",
        1: "Wave 1 — DDL Foundation",
        2: "Wave 2 — Quick Wins",
        3: "Wave 3 — Consolidation",
        4: "Wave 4 — MV Architecture (DBA)",
        5: "Wave 5 — LLM Optimization (Parallel)",
    }

    # Banner
    ws.merge_cells("A1:R1")
    c = ws["A1"]
    c.value = (
        "ACTIONABLE WORK ORDERS  —  Validation-filtered view  "
        f"|  Only '{VS_CONSIDER}' and '{VS_PENDING}' recs are shown  "
        f"|  '{VS_NOT_CONS}' and '{VS_NOT_GOOD}' recs are excluded  "
        "|  This is the developer's delivery list after data team review"
    )
    c.font  = Font(bold=True, size=13, color=WHITE)
    c.fill  = hfill("FF1B5E20")   # dark green — actionable
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Column headers (same as Sheet 7 minus one column for simplicity)
    COLS_AWO = [
        "Package Name",                    # A
        "Procedures Touched",              # B
        "Wave",                            # C
        "Scope",                           # D
        "Affected RPTs",                   # E
        "Tidal Jobs",                      # F
        "# Actionable Recs",               # G
        "# Excluded Recs",                 # H  ← new: how many were filtered
        "Bundled Master IDs",              # I
        "Phase Range",                     # J
        "Categories",                      # K
        "All Actionable Changes",          # L  KEY
        "Hop Savings",                     # M
        "Est. Min Saved",                  # N
        "Max Risk",                        # O
        "Also Remove from Tidal",          # P
        "Wave 4 Awareness",                # Q
        "Validation Status (per rec)",     # R
    ]
    hdr_row = 2
    for col, h in enumerate(COLS_AWO, 1):
        hdr_cell(ws, hdr_row, col, h, fill_hex="FF1B5E20", size=10)
    ws.row_dimensions[hdr_row].height = 28
    ws.freeze_panes = "A3"

    # Build script_map — identical to Sheet 7 logic
    id_map = {id(rec): i for i, rec in enumerate(records, 1)}
    script_map: dict[str, dict] = {}
    for rec in records:
        raw = rec.get("SQL Objects Called", "") or ""
        scripts_for_rec = [s.strip() for s in re.split(r'[\n,;]', raw) if s.strip()] or [TIDAL_ONLY]
        # Classify every SQL object uniformly — CODE only becomes a script row.
        code_scripts = [s for s in scripts_for_rec
                        if s == TIDAL_ONLY or _classify_sql_object(s) == "CODE"]
        scripts_for_rec = code_scripts if code_scripts else [TIDAL_ONLY]
        # DDL & Tidal removals using the same classifier
        mv_rems: set[str] = set()
        for obj in re.split(r'[\n,;]', raw):
            obj = obj.strip()
            if not obj: continue
            cls = _classify_sql_object(obj)
            if cls == "MV":   mv_rems.add(f"DROP MV: {obj}")
            elif cls == "VIEW": mv_rems.add(f"RENAME VIEW: {obj}")
        for j in _split_jobs(rec.get("Affected Jobs", "") or ""):
            if "MV_REFRESH" in j.upper():
                mv_rems.add(f"Remove from Tidal: {j}")
        mid  = f"M-{id_map[id(rec)]:04d}"
        ph   = rec["phase"]
        rpts = set(rec["appears_in_rpts"])
        jobs = set(_split_jobs(rec.get("Affected Jobs","") or ""))
        for script in scripts_for_rec:
            pkg = _pk(script)
            if pkg not in script_map:
                script_map[pkg] = {
                    "procs": set(), "recs": [],
                    "phases": set(), "rpts": set(),
                    "tidal_jobs": set(), "mv_removals": set(),
                    "hop_savings": 0, "est_min": 0.0,
                }
            e = script_map[pkg]
            if script != pkg:
                e["procs"].add(script)
            e["recs"].append((mid, rec))
            e["phases"].add(ph)
            e["rpts"]       |= rpts
            e["tidal_jobs"] |= jobs
            e["mv_removals"]|= mv_rems
            e["hop_savings"] += rec["global_hop_savings"]
            e["est_min"]     += rec["global_est_min"]

    # Dedup per package; recompute savings from the deduped list
    for e in script_map.values():
        seen: set = set()
        e["recs"]        = [(m, r) for m, r in e["recs"] if m not in seen and not seen.add(m)]
        e["hop_savings"] = sum(r["global_hop_savings"] for _, r in e["recs"])
        e["est_min"]     = sum(r["global_est_min"]     for _, r in e["recs"])

    # RPT ranking (same score formula as Sheet 9)
    _all_rpts = sorted({rpt for _, d in script_map.items() for rpt in d["rpts"]})
    _rpt_score: dict[str, float] = {}
    for _rpt in _all_rpts:
        _sm: set = set(); _gm: set = set(); _ph: dict = defaultdict(int); _rk = []
        _seen: set = set()
        for _, _d in script_map.items():
            if _rpt not in _d["rpts"]: continue
            for _m, _r in _d["recs"]:
                if _m in _seen or _rpt not in _r["appears_in_rpts"]: continue
                _seen.add(_m); _ph[_r["phase"]] += 1; _rk.append(_r.get("Risk","LOW") or "LOW")
                for _s in re.split(r'[\n,;]', _r.get("SQL Objects Called","") or ""):
                    _s = _s.strip()
                    if _s: (_gm if _r["rpt_count"] > 1 else _sm).add(_s)
        _sm -= _gm; _n = len(_sm | _gm); _pct = round(len(_sm)/max(_n,1)*100)
        _mr = "HIGH" if "HIGH" in _rk else "MEDIUM" if "MEDIUM" in _rk else "LOW"
        _rw = {"LOW":1,"MEDIUM":2,"HIGH":3}
        _rpt_score[_rpt] = (len(_seen)*1.5 + _n*2 + (100-_pct)*0.1
                            + _ph.get(4,0)*3 + _ph.get(3,0)*2 + _rw.get(_mr,2)*5)
    _ranked = sorted(_all_rpts, key=lambda r: _rpt_score.get(r, 999))
    _rpt_rank: dict[str, int] = {rpt: i for i, rpt in enumerate(_ranked)}

    def _prim(data: dict) -> str:
        """
        Return the most relevant RPT for section grouping.
        Prefers the RPT whose name matches the target table(s) of bundled recs;
        falls back to lowest-ranked RPT when no target-table match is found.
        """
        rpts = data["rpts"]
        if len(rpts) == 1:
            return next(iter(rpts))
        rpt_scores: dict[str, int] = {rpt: 0 for rpt in rpts}
        for _mid, rec in data["recs"]:
            tgt = (rec.get("Target Table", "") or "").upper().strip()
            if not tgt:
                continue
            for rpt in rpts:
                rpt_up = rpt.upper()
                if rpt_up == tgt or rpt_up in tgt or tgt in rpt_up:
                    rpt_scores[rpt] += 1
        best = max(rpt_scores.values(), default=0)
        if best > 0:
            candidates = [r for r, s in rpt_scores.items() if s == best]
            return min(candidates, key=lambda r: _rpt_rank.get(r, 999))
        return min(rpts, key=lambda r: _rpt_rank.get(r, 999))

    sorted_scripts = sorted(
        script_map.items(),
        key=lambda item: (
            _rpt_rank.get(_prim(item[1]), 999),
            _assign_wave(item[1]["phases"]),
            0 if len(item[1]["rpts"]) > 1 else 1,
            item[0],
        ),
    )

    _n_rpts = len(_ranked)
    def _tier(rank: int) -> int:
        if rank < max(3, _n_rpts // 4):     return 1
        if rank < max(7, _n_rpts * 2 // 3): return 2
        return 3
    _TIER_FILL = {1: "FF2E7D32", 2: "FFE65100", 3: "FFC62828"}

    risk_fill_hex = {"HIGH": "FFFFC7CE", "MEDIUM": "FFFFEB9C", "LOW": "FFC6EFCE"}
    scope_fill_hex = {"GLOBAL": "FFCFE2F3", "LOCAL": "FFF3F3F3"}

    current_rpt = ""
    n_cols = len(COLS_AWO)
    total_actionable = 0

    for script, data in sorted_scripts:
        prpt = _prim(data)

        # Split recs into actionable vs excluded
        all_recs        = data["recs"]
        actionable_recs = [(m, r) for m, r in all_recs
                           if r.get("validation_status", VS_PENDING) in ACTIONABLE_STATUSES]
        excluded_recs   = [(m, r) for m, r in all_recs
                           if r.get("validation_status", VS_PENDING) not in ACTIONABLE_STATUSES]

        # Skip scripts with no actionable recs
        if not actionable_recs:
            continue

        total_actionable += len(actionable_recs)

        # RPT section banner
        if prpt != current_rpt:
            current_rpt = prpt
            rank  = _rpt_rank.get(prpt, 0)
            tier  = _tier(rank)
            if ws.max_row > 2:
                ws.append([])
            ban_row = ws.max_row + 1
            ws.merge_cells(f"A{ban_row}:{get_column_letter(n_cols)}{ban_row}")
            bc = ws.cell(row=ban_row, column=1,
                         value=f"  {prpt}   ▸ Rank {rank+1}  |  Tier {tier}")
            bc.fill  = hfill(_TIER_FILL.get(tier, DARK))
            bc.font  = Font(bold=True, size=11, color=WHITE)
            bc.alignment = Alignment(vertical="center", indent=1)
            ws.row_dimensions[ban_row].height = 20

        phases      = {r["phase"] for _, r in actionable_recs}
        # Derive rpts, tidal_jobs AND mv_removals from ACTIONABLE recs only —
        # excluded recs must not leak RPTs, jobs, or DDL targets into this view.
        rpts        = sorted({rpt for _, r in actionable_recs for rpt in r["appears_in_rpts"]})
        tidal_jobs  = sorted({j for _, r in actionable_recs
                              for j in _split_jobs(r.get("Affected Jobs", "") or "")})
        _mv_rem_set: set[str] = set()
        for _, _r in actionable_recs:
            for _obj in re.split(r'[\n,;]', _r.get("SQL Objects Called", "") or ""):
                _obj = _obj.strip()
                if not _obj: continue
                _cls = _classify_sql_object(_obj)
                if _cls == "MV":   _mv_rem_set.add(f"DROP MV: {_obj}")
                elif _cls == "VIEW": _mv_rem_set.add(f"RENAME VIEW: {_obj}")
            for _j in _split_jobs(_r.get("Affected Jobs", "") or ""):
                if "MV_REFRESH" in _j.upper():
                    _mv_rem_set.add(f"Remove from Tidal: {_j}")
        mv_removals = sorted(_mv_rem_set)
        wave        = _assign_wave(phases)
        scope       = "GLOBAL" if len(rpts) > 1 else "LOCAL"
        has_ph4     = 4 in phases and wave != 4
        max_risk    = "LOW"
        for lvl in ("HIGH","MEDIUM","LOW"):
            if any((r.get("Risk","") or "") == lvl for _, r in actionable_recs):
                max_risk = lvl; break

        # Build change blocks (actionable only)
        change_blocks = []; cats_set: set = set(); master_ids: list = []
        for mid, rec in sorted(actionable_recs, key=lambda x: (x[1]["phase"], x[0])):
            cat   = rec.get("Category","") or ""
            tgt   = rec.get("Target Table","") or ""
            recmd = rec.get("Recommendation","") or rec.get("What To Do","") or ""
            ph    = rec["phase"]
            hops  = rec["global_hop_savings"]
            vs    = rec.get("validation_status", VS_PENDING)
            cats_set.add(cat); master_ids.append(mid)
            first_line = next((ln.strip() for ln in recmd.split("\n") if ln.strip()), cat)[:200]
            change_blocks.append(
                f"── {mid} | Phase {ph} | {cat}  [{vs}]\n"
                f"   Target : {tgt}  |  Hops: {hops}\n"
                f"   Action : {first_line}"
            )

        ph4_text = "(none)"
        if has_ph4:
            ph4e = [(m, r) for m, r in actionable_recs if r["phase"] == 4]
            ph4_ids = [m for m, _ in ph4e]
            ph4_text = (
                f"⚠ PHASE 4 PENDING — DBA sign-off required:\n"
                + "\n".join(f"  • {m}: {r.get('Category','')}" for m, r in ph4e)
                + f"\nWave 4 IDs: {', '.join(ph4_ids)}"
            )

        excluded_note = ""
        if excluded_recs:
            ex_ids = [m for m, _ in excluded_recs]
            ex_vs  = [r.get("validation_status","") for _, r in excluded_recs]
            excluded_note = (f"{len(excluded_recs)} rec(s) excluded from this script by data team: "
                             f"{', '.join(ex_ids)}  "
                             f"({'; '.join(ex_vs[:3])})")

        vs_per_rec = "\n".join(
            f"{mid}: {rec.get('validation_status', VS_PENDING)}"
            for mid, rec in sorted(actionable_recs, key=lambda x: x[0])
        )

        procs_disp = ("\n".join(sorted(data["procs"])) if data["procs"]
                      else "(entire package)")
        phase_range = (f"Phase {min(phases)}–{max(phases)}" if len(phases) > 1
                       else f"Phase {min(phases)}")

        nr    = ws.max_row + 1
        wfill = hfill(WAVE_FILLS_AWO.get(wave, "FFFFFFFF"))

        vals = [
            script,                             # A
            procs_disp,                         # B
            WAVE_LABELS_AWO.get(wave, f"Wave {wave}"),  # C
            scope,                              # D
            "\n".join(rpts),                    # E
            ("\n".join(tidal_jobs[:8]) + ("\n  ...+" + str(len(tidal_jobs)-8)
             if len(tidal_jobs) > 8 else "")),  # F
            len(actionable_recs),               # G
            f"{len(excluded_recs)} excluded — {excluded_note}" if excluded_recs else "0",  # H
            "\n".join(master_ids),              # I
            phase_range,                        # J
            "\n".join(sorted(cats_set)),        # K
            "\n\n".join(change_blocks),         # L
            sum(rec["global_hop_savings"] for _, rec in actionable_recs),  # M
            round(sum(rec["global_est_min"]    for _, rec in actionable_recs), 1),  # N
            max_risk,                           # O
            "\n".join(mv_removals) if mv_removals else "(none)",  # P
            ph4_text,                           # Q
            vs_per_rec,                         # R
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=nr, column=col, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.fill = wfill
            if col == 1:
                c.font = Font(bold=True, size=10)
            if col == 2:
                c.fill = hfill("FFF9FBE7")
            if col == 4:
                c.fill = hfill(scope_fill_hex.get(scope, "FFF3F3F3"))
            if col == 15:
                c.fill = hfill(risk_fill_hex.get(max_risk, "FFFFFFFF"))
        ws.row_dimensions[nr].height = max(70, 28 * len(actionable_recs))

    # Footer stats
    ws.append([])
    foot_row = ws.max_row + 1
    ws.merge_cells(f"A{foot_row}:R{foot_row}")
    fc = ws.cell(row=foot_row, column=1,
                 value=(f"SUMMARY:  {total_actionable} actionable rec(s) shown  |  "
                        f"Excluded: {VS_NOT_CONS} and {VS_NOT_GOOD} recs (see Sheet 7 for complete view)  |  "
                        f"{VS_PENDING} recs are treated as actionable until the data team reviews them."))
    fc.font = Font(italic=True, size=9, color="FF616161")
    fc.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[foot_row].height = 30

    set_col_widths(ws, {
        "A": 50, "B": 40, "C": 30, "D": 10, "E": 35, "F": 50, "G": 10,
        "H": 35, "I": 25, "J": 12, "K": 55, "L": 85, "M": 12, "N": 12,
        "O": 10, "P": 45, "Q": 50, "R": 35,
    })
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{ws.max_row}"


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main() -> None:
    # Ensure emoji in print output doesn't crash on Windows consoles
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    print("=" * 70)
    print("  MASTER RECOMMENDATION GENERATOR")
    print("=" * 70)

    print("\n[1/4] Reading all RPT workbooks...")
    all_rows = load_all_recommendations()
    rpt_names = sorted({r["source_rpt"] for r in all_rows})
    files = sorted(WB_DIR.glob("hop_reduction_recommendations_*.xlsx"))
    print(f"      Workbook files found       : {len(files)}")
    print(f"      RPTs with hop findings     : {len(rpt_names)}  "
          f"(RPTs with no pipeline structure — e.g. RPT_EMPLOYEE_R — contribute 0 rows and are excluded)")
    print(f"      Total rows (cross-RPT dups): {len(all_rows)}")

    print("\n[2/4] Deduplicating...")
    records = deduplicate(all_rows)
    print(f"      Unique recommendations: {len(records)}")

    print("\n      Loading validation statuses from data team review...")
    validation_statuses = load_validation_statuses()
    matched_count = 0
    for rec in records:
        cat  = rec.get("Category", "") or ""
        tgt  = rec.get("Target Table", "") or ""
        jobs = rec.get("Affected Jobs", "") or ""
        sql  = rec.get("SQL Objects Called", "") or ""
        key  = make_dedup_key(cat, tgt, jobs, sql)
        entry = validation_statuses.get(key)
        if entry:
            rec["validation_status"]   = entry["validation_status"]
            rec["rec_status_raw"]      = entry["rec_status_raw"]
            rec["val_comments"]        = entry["comments"]
            rec["val_agreement"]       = entry["agreement_status"]
            rec["val_review_by"]       = entry["review_by"]
            matched_count += 1
        else:
            rec["validation_status"]   = VS_PENDING
            rec["rec_status_raw"]      = ""
            rec["val_comments"]        = ""
            rec["val_agreement"]       = ""
            rec["val_review_by"]       = ""
    print(f"      Matched: {matched_count}/{len(records)} recs  "
          f"({len(records)-matched_count} marked {VS_PENDING})")

    # Print phase summary to console
    phase_groups: dict[int, list] = defaultdict(list)
    for rec in records:
        phase_groups[rec["phase"]].append(rec)

    print("\n      Phase breakdown:")
    for ph in sorted(phase_groups.keys()):
        ph_recs = phase_groups[ph]
        ph_hops = sum(r["global_hop_savings"] for r in ph_recs)
        ph_min  = sum(r["global_est_min"]     for r in ph_recs)
        print(f"        Phase {ph} ({PHASE_LABELS[ph].split('—')[1].strip():30s}) "
              f"{len(ph_recs):4d} recs  |  {ph_hops:4d} hops  |  {ph_min:7.1f} min")

    non_f_recs = [r for r in records if not (r.get("Category","") or "").startswith("F")]
    total_hops = sum(r["global_hop_savings"] for r in non_f_recs)
    total_min  = sum(r["global_est_min"]     for r in non_f_recs)
    print(f"\n      TOTAL (excl. LLM): {total_hops} verified hops  |  "
          f"{total_min:.0f} min  ({total_min/60:.1f} hrs) saved if all implemented")

    print("\n[3/4] Building Excel workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_executive_summary(wb, records, rpt_names, files)
    print("      OK Executive Summary")
    write_master_recommendations(wb, records)
    print("      OK Master Recommendations")
    write_implementation_roadmap(wb, records)
    print("      OK Implementation Roadmap")
    write_cross_rpt_impact(wb, records)
    print("      OK Cross-RPT Impact Map")
    write_dev_work_orders(wb, records)
    print("      OK Dev Work Orders")
    write_global_savings(wb, records)
    print("      OK Global Savings Summary")
    write_script_centric_work_orders(wb, records)
    print("      OK Script-Centric Work Orders")
    write_rpt_pipeline_summary(wb, records, rpt_names)
    print("      OK RPT Pipeline Summary")
    write_rpt_implementation_plan(wb, records)
    print("      OK RPT Implementation Plan (bottom-up)")
    write_actionable_work_orders(wb, records)
    print("      OK Actionable Work Orders (validation-filtered)")

    print(f"\n[4/4] Saving...")
    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"\n  Output: {OUTPUT_FILE}")
    print(f"  Size  : {OUTPUT_FILE.stat().st_size / 1024:.0f} KB")
    print("\nDone.\n")


if __name__ == "__main__":
    main()
