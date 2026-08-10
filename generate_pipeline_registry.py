#!/usr/bin/env python3
"""
generate_pipeline_registry.py
══════════════════════════════
Reads the validated MASTER_Hop_Reduction_Recommendations.xlsx and builds:

  1. pipeline_registry.json   — per-package state + recommendations (full text)
  2. pipeline/                — directory scaffold with source SQL copies

Output root: configurable via PIPELINE_ROOT (defaults to RSLI-DataLineage-VDI/output/pipeline)

Schema
──────
{
  "schema_version": "1.1",
  "generated_at": "YYYY-MM-DD",
  "implementation_order": ["PKG_A", "PKG_B", ...],   # bottom-up RPT order
  "packages": {
    "PKG_GRP_LOAD_RPT_CLAIM_TASK_R_INC": {
      "rpt":               "RPT_CLAIM_TASK_R",
      "impl_rank":         2,
      "impl_tier":         1,
      "scope":             "LOCAL",
      "source_sql":        "All_Metadata/PKG_GRP_LOAD_RPT_CLAIM_TASK_R_INC.sql",
      "pipeline_stage":    "raw",            # raw | merged | optimized | standardized
      "applicable_agents": ["merger"],       # merger | optimizer | standardizer
      "recommendations": [
        {
          "master_id":          "M-0061",
          "phase":              3,
          "phase_name":         "Consolidation",
          "category":           "C. OFFSET / Near-Duplicate Procedures",
          "agent":              "merger",       # which agent handles this rec
          "sub_agent":          "offset-consolidator",
          "target_table":       "FCT_CLAIM_PAYMENT_DETAIL_OFFSETS_R",
          "affected_jobs":      ["EDP_EDW_GRP_PACS_LOAD_..._OFFSET1", "...OFFSET2"],
          "sql_objects_called": ["PRC_FULLLOAD_FCT_CLAIMPMNT_DET_OFFSET1_R"],
          "hop_savings":        2,
          "est_min_saved":      17.2,
          "risk":               "MEDIUM",
          "validation_status":  "✅ Consider",
          "description":        "Full description text...",
          "recommendation":     "Full recommendation / what-to-do text...",
          "comments":           "Data team comments...",
          "review_by":          "...",
          "appears_in_rpts":    ["RPT_CLAIM_TASK_R"],
          "rpt_count":          1
        }
      ]
    }
  }
}

Agent routing
─────────────
  merger    ← C. OFFSET/Near-Duplicate, A. Redundant Multi-Source, N. Circular, D. Serial Chain
  optimizer ← B. Pass-Through, Q. Hardcoded Values, P. Cursor UPDATE, O. GTT, F. LLM Findings
  (standardizer always runs last — driven by agent 3, not by a specific rec category)

Sub-agent routing (within merger/optimizer)
───────────────────────────────────────────
  offset-consolidator    ← C. OFFSET / Near-Duplicate Procedures
  multisrc-consolidator  ← A. Redundant Multi-Source Load
  circular-breaker       ← N. Circular Update Chain
  q-lookup-externalizer  ← Q. Hardcoded Value Mapping
  p-cursor-folder        ← P. Post-Load Cursor UPDATE
  (b, o, f handled inline in optimizer — no dedicated sub-agent)

Usage:
    python generate_pipeline_registry.py
    python generate_pipeline_registry.py --pipeline-root "C:/path/to/output/pipeline"
    python generate_pipeline_registry.py --master "path/to/MASTER.xlsx"
    python generate_pipeline_registry.py --no-table-meta   # skip table metadata enrichment
"""

import argparse
import json
import re
import shutil
from datetime import date
from pathlib import Path

import openpyxl

from config import (
    ALL_METADATA_DIR,
    INPUT_DIR,
    OUTPUT_DIR,
    PROJECT_ROOT,
)

# ── Default paths ─────────────────────────────────────────────────────────────
SCRIPT_DIR        = PROJECT_ROOT
MASTER_XLSX       = OUTPUT_DIR / "MASTER_Hop_Reduction_Recommendations.xlsx"

# VDI output root — override with --pipeline-root
VDI_ROOT          = SCRIPT_DIR
DEFAULT_PIPELINE  = OUTPUT_DIR / "pipeline"

# Table-level metadata sources
TABLE_INFO_XLSX      = INPUT_DIR / "Table_Info.xlsx"
INDEX_INFO_XLSX      = INPUT_DIR / "Index_Allobject.xlsx"
TIDAL_GRAPH_JSON     = OUTPUT_DIR / "tidal_dependency_graph.json"
SCRIPT_SUMMARIES_JSON = OUTPUT_DIR / "script_summaries_all.json"

# ── Agent routing tables ─────────────────────────────────────────────────────
# Maps category prefix → (agent, sub_agent)
AGENT_ROUTING: dict[str, tuple[str, str]] = {
    "C. OFFSET":                         ("merger",    "offset-consolidator"),
    "C. Near-Duplicate":                 ("merger",    "offset-consolidator"),
    "A. Redundant Multi-Source":         ("merger",    "multisrc-consolidator"),
    "N. Circular Update Chain":          ("merger",    "circular-breaker"),
    "N. Post-Load Updater":              ("merger",    "circular-breaker"),
    "N. Post-Consumption Update":        ("merger",    "circular-breaker"),
    "D. Serial Chain":                   ("merger",    "inline"),            # Tidal-only
    "D. Serial Orchestration":           ("merger",    "inline"),
    "Q. Hardcoded Value Mapping":        ("optimizer", "q-lookup-externalizer"),
    "P. Post-Load Cursor UPDATE":        ("optimizer", "p-cursor-folder"),
    "P. Post-Load Cursor Update":        ("optimizer", "p-cursor-folder"),
    "B. Pass-Through":                   ("optimizer", "inline"),
    "O. GTT Intra-Procedure":            ("optimizer", "inline"),
    "O. Intermediate/Working Table":     ("optimizer", "inline"),
    "O. Hybrid Naming":                  ("optimizer", "inline"),           # rename only
    "O. View Naming":                    ("standardizer", "inline"),         # DDL rename
    "F. LLM-Identified":                 ("optimizer", "inline"),
    "G. Duplicate Column-Lineage":       ("optimizer", "inline"),
    "H. Stale Intermediate":             ("optimizer", "inline"),
    "L. Overlapping Preprocessing":      ("optimizer", "inline"),
    "K. Repeated Transformations":       ("optimizer", "inline"),
    "M. Orchestration Edge":             ("merger",    "inline"),            # Tidal edge
    "J. Over-Fanout MV":                 ("skip",      ""),                 # Phase 4 / DBA — no package change
    # ── E. Materialized View recs ────────────────────────────────────────────
    # All E. MV recs with ✅ Consider are included; agent handles PL/SQL side,
    # flags DBA DDL + Tidal job removal in decisions.md.
    # "Shared Across RPTs" recs are GLOBAL migrations — optimizer adds extra
    # cross-RPT coordination warning. No E. recs remain as skip.
    "E. MV Shared Across RPTs":          ("optimizer", "inline"),           # GLOBAL: needs perf analysis + coord
    "E. MV Indicator-Chain":             ("optimizer", "inline"),
    "E. MV Elimination":                 ("optimizer", "inline"),
    "E. Dead MV":                        ("optimizer", "inline"),           # decommission candidate
    "E.":                                ("optimizer", "inline"),           # catch-all for any other E. Phase 4
    # ── Phase 0 architectural decisions ──────────────────────────────────────
    "A. Source System Architectural":    ("skip",      ""),                 # Phase 0
    "C. Source System Architectural":    ("skip",      ""),                 # Phase 0
}

# Only process "✅ Consider" or "⏳ Validation Pending" recs
ACTIONABLE_STATUSES = {"✅ Consider", "⏳ Validation Pending"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _route(category: str) -> tuple[str, str]:
    """Return (agent, sub_agent) for a recommendation category."""
    cat = (category or "").strip()
    for prefix, routing in AGENT_ROUTING.items():
        if cat.startswith(prefix):
            return routing
    return ("optimizer", "inline")   # default: treat as optimizer inline


def _pkg_key(sql_obj: str) -> str:
    """Normalise a SQL object string to package-level key."""
    if not sql_obj or '.' not in sql_obj:
        return sql_obj or ""
    left, right = sql_obj.split('.', 1)
    pkg_pfx  = ("PKG_", "PRC_", "PROC_", "SP_", "USP_", "FN_", "FUNC_")
    proc_pfx = ("PRC_", "PROC_", "MAIN", "GET_", "INSERT_", "LOAD_",
                "UPDATE_", "DELETE_", "SEL_", "EXEC_", "RUN_")
    if any(left.upper().startswith(p) for p in pkg_pfx) or \
       any(right.upper().startswith(p) for p in proc_pfx):
        return left.strip()
    return sql_obj


def _split_newline(raw: str) -> list[str]:
    if not raw:
        return []
    return [s.strip() for s in re.split(r'[\n,;]', raw) if s.strip()]


def _find_source_sql(pkg_name: str, all_metadata: Path) -> str | None:
    """Return relative path from RSLI-DataLineage-VDI root to the .sql file."""
    if not pkg_name or not all_metadata.exists():
        return None
    candidates = [
        all_metadata / f"{pkg_name}.sql",
        all_metadata / f"{pkg_name.upper()}.sql",
        all_metadata / f"{pkg_name.lower()}.sql",
    ]
    for c in candidates:
        if c.exists():
            return str(c).replace("\\", "/")
    # Fuzzy: case-insensitive scan
    target = pkg_name.upper()
    for f in all_metadata.glob("*.sql"):
        if f.stem.upper() == target:
            return str(f).replace("\\", "/")
    return None


def _load_table_info(xlsx_path: Path) -> dict[str, dict]:
    """Load Table_Info.xlsx → {TABLE_NAME_UPPER: {owner, row_count, column_count, size_mb}}."""
    if not xlsx_path.exists():
        print(f"  [WARN] Table_Info.xlsx not found: {xlsx_path} — skipping table metadata")
        return {}
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    result: dict[str, dict] = {}
    for row in rows[1:]:
        if not row[1]:
            continue
        name = str(row[1]).strip().upper()
        result[name] = {
            "owner":        str(row[0]).strip() if row[0] else "ATOMIC",
            "row_count":    int(row[2]) if row[2] is not None else None,
            "column_count": int(row[3]) if row[3] is not None else None,
            "size_mb":      float(row[4]) if row[4] is not None else None,
        }
    print(f"  -> Loaded {len(result)} table-level metadata entries from Table_Info.xlsx")
    return result


def _load_index_info(xlsx_path: Path) -> dict[str, list[dict]]:
    """Load Index_Allobject.xlsx → {TABLE_NAME_UPPER: [{index_name, uniqueness, columns: [...]}]}."""
    if not xlsx_path.exists():
        print(f"  [WARN] Index_Allobject.xlsx not found: {xlsx_path} — skipping index metadata")
        return {}
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    result: dict[str, list[dict]] = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        tbl = str(row[0]).strip().upper()
        cols_raw = str(row[3]).strip() if row[3] else ""
        result.setdefault(tbl, []).append({
            "index_name": str(row[1]).strip() if row[1] else "",
            "uniqueness":  str(row[2]).strip() if row[2] else "NONUNIQUE",
            "columns":     [c.strip() for c in cols_raw.split(",") if c.strip()],
        })
    total = sum(len(v) for v in result.values())
    print(f"  -> Loaded {total} index records across {len(result)} tables from Index_Allobject.xlsx")
    return result


def _build_tidal_table_lookup(tidal_path: Path) -> dict[str, dict]:
    """Parse tidal_dependency_graph.json → {SQL_OBJ_UPPER: {src: [tables], tgt: [tables]}}."""
    if not tidal_path.exists():
        print(f"  [WARN] tidal_dependency_graph.json not found: {tidal_path} — skipping tidal table lookup")
        return {}
    with open(tidal_path, encoding="utf-8") as f:
        tidal = json.load(f)
    lookup: dict[str, dict] = {}
    for graph in tidal.values():
        for node in graph.get("nodes", []):
            src_tables = [t for t in node.get("src_tables", []) if t]
            tgt_tables = [t for t in node.get("tgt_tables", []) if t]
            if not src_tables and not tgt_tables:
                continue
            for sobj in node.get("sql_objects", []):
                for key_raw in [sobj.get("package"), sobj.get("name")]:
                    if not key_raw:
                        continue
                    key = key_raw.strip().upper()
                    if key not in lookup:
                        lookup[key] = {"src": set(), "tgt": set()}
                    lookup[key]["src"].update(src_tables)
                    lookup[key]["tgt"].update(tgt_tables)
    # Convert sets to sorted lists for JSON serialisation
    return {k: {"src": sorted(v["src"]), "tgt": sorted(v["tgt"])} for k, v in lookup.items()}


def _make_table_entry(name: str, table_info: dict[str, dict], index_info: dict[str, list]) -> dict:
    """Return a metadata dict for a single table (None-valued fields when not in Table_Info)."""
    meta = table_info.get(name.upper())
    indexes = index_info.get(name.upper(), [])
    base = {"table_name": name, **(meta or {"owner": None, "row_count": None, "column_count": None, "size_mb": None})}
    base["index_count"]      = len(indexes)
    base["has_unique_index"] = any(i["uniqueness"] == "UNIQUE" for i in indexes)
    base["indexes"]          = indexes
    return base


def _enrich_table_metadata(
    pkg_map: dict[str, dict],
    table_info: dict[str, dict],
    tidal_lookup: dict[str, dict],
    index_info: dict[str, list] | None = None,
) -> None:
    """
    Mutates pkg_map in-place:
      - adds ``table_metadata`` block to each package
      - adds ``target_table_meta`` to each recommendation
    """
    for pkg_name, entry in pkg_map.items():
        up = pkg_name.upper()

        # ── Gather source / target table names from tidal graph ──────────────
        tidal_entry = tidal_lookup.get(up, {})
        src_names: set[str] = set(tidal_entry.get("src", []))
        tgt_names: set[str] = set(tidal_entry.get("tgt", []))

        # Also add target_table from each recommendation
        for rec in entry.get("recommendations", []):
            for t in _split_newline(rec.get("target_table", "")):
                tgt_names.add(t)
            # Add sql_objects_called as potential source lookup keys
            for obj_raw in _split_newline(rec.get("sql_objects_called", "")):
                obj_up = obj_raw.strip().upper()
                if obj_up in tidal_lookup:
                    src_names.update(tidal_lookup[obj_up].get("src", []))
                    tgt_names.update(tidal_lookup[obj_up].get("tgt", []))

        # Remove obviously non-table entries (db links, semicolons, commas)
        def _is_clean_table(t: str) -> bool:
            return bool(t) and "@" not in t and ";" not in t and "," not in t and " " not in t.strip()

        src_clean = sorted(t for t in src_names if _is_clean_table(t))
        tgt_clean = sorted(t for t in tgt_names if _is_clean_table(t))

        idx = index_info or {}
        src_entries = [_make_table_entry(t, table_info, idx) for t in src_clean]
        tgt_entries = [_make_table_entry(t, table_info, idx) for t in tgt_clean]

        def _sum_field(entries: list, field: str) -> float | None:
            vals = [e[field] for e in entries if e.get(field) is not None]
            return round(sum(vals), 2) if vals else None

        entry["table_metadata"] = {
            "target_tables": tgt_entries,
            "source_tables":  src_entries,
            "summary": {
                "total_target_rows":         _sum_field(tgt_entries, "row_count"),
                "total_source_rows":         _sum_field(src_entries, "row_count"),
                "total_target_size_mb":      _sum_field(tgt_entries, "size_mb"),
                "total_source_size_mb":      _sum_field(src_entries, "size_mb"),
                "target_table_count":        len(tgt_entries),
                "source_table_count":        len(src_entries),
                "total_target_indexes":      sum(e.get("index_count", 0) for e in tgt_entries),
                "total_source_indexes":      sum(e.get("index_count", 0) for e in src_entries),
                "meta_coverage": (
                    f"{sum(1 for e in tgt_entries + src_entries if e.get('row_count') is not None)}"
                    f"/{len(tgt_entries) + len(src_entries)} tables in Table_Info"
                ),
            },
        }

        # Enrich each recommendation with target table metadata including indexes
        for rec in entry.get("recommendations", []):
            tbl = rec.get("target_table", "").strip()
            primary = next((t for t in _split_newline(tbl) if _is_clean_table(t)), None)
            if primary and primary.upper() in table_info:
                rec["target_table_meta"] = {
                    **table_info[primary.upper()],
                    "index_count":      len(idx.get(primary.upper(), [])),
                    "has_unique_index": any(i["uniqueness"] == "UNIQUE" for i in idx.get(primary.upper(), [])),
                    "indexes":          idx.get(primary.upper(), []),
                }
            else:
                rec["target_table_meta"] = {
                    "owner": None, "row_count": None, "column_count": None, "size_mb": None,
                    "index_count": 0, "has_unique_index": False, "indexes": [],
                }


def _build_script_summary_consumer_map(summaries_path: Path) -> dict[str, list[str]]:
    """
    Parse script_summaries_all.json INPUT_DEPENDENCIES → reverse map:
    {OBJECT_NAME_UPPER: [script_stems_that_list_it_as_input]}.
    Captures ALL references (FROM, JOIN, subquery) — not only FROM-clause lineage.
    """
    if not summaries_path.exists():
        print(f"  [WARN] script_summaries_all.json not found: {summaries_path} — skipping consumer gap analysis")
        return {}
    with open(summaries_path, encoding="utf-8", errors="replace") as f:
        summaries = json.load(f)
    rev: dict[str, list[str]] = {}
    for script_name, data in summaries.items():
        if not isinstance(data, dict):
            continue
        st = data.get("INPUT_DEPENDENCIES", {}).get("source_tables", {})
        if not isinstance(st, dict):
            continue
        stem = script_name.replace(".sql", "").replace(".SQL", "")
        for bucket in st.values():
            if not isinstance(bucket, list):
                continue
            for item in bucket:
                # Strip alias hints like "FCT_X (aliased y)" and schema prefixes like "ATOMIC."
                raw = str(item).split("(")[0].split("@")[0].strip()
                raw = re.sub(r"^ATOMIC\.", "", raw, flags=re.IGNORECASE).strip().upper()
                if raw and len(raw) > 2:
                    rev.setdefault(raw, []).append(stem)
    total = sum(len(v) for v in rev.values())
    print(f"  -> Built script-summaries consumer map: {len(rev)} objects, {total} total references")
    return rev


def _enrich_consumer_gaps(
    pkg_map: dict[str, dict],
    consumer_map: dict[str, list[str]],
) -> None:
    """
    For every global-scope E. (MV elimination) or O. (naming) recommendation,
    cross-reference consumer_map to find consumers missed by lineage-only analysis.

    Adds to each matching rec:
      gap_consumers       — scripts referencing the target object that are NOT in sql_objects_called
      all_consumers       — union of sql_objects_called + gap_consumers (deduplicated)
      cascade_chain       — gap/all consumers that are ALSO targets of other global E./O. recs
      has_consumer_gaps   — bool flag for quick agent detection
    """
    # Build a set of all global target objects in this registry
    global_targets: dict[str, str] = {}  # {TARGET_UPPER: master_id}
    for entry in pkg_map.values():
        for rec in entry.get("recommendations", []):
            if rec.get("rpt_count", 1) > 1:
                cat = rec.get("category", "")
                if cat.startswith("E.") or cat.startswith("O."):
                    tgt = rec.get("target_table", "").strip().upper()
                    if tgt:
                        global_targets[tgt] = rec["master_id"]

    processed: set[str] = set()  # dedup: (master_id, target)
    for entry in pkg_map.values():
        for rec in entry.get("recommendations", []):
            if rec.get("rpt_count", 1) <= 1:
                continue
            cat = rec.get("category", "")
            if not (cat.startswith("E.") or cat.startswith("O.")):
                continue
            mid = rec["master_id"]
            tgt = rec.get("target_table", "").strip()
            key = (mid, tgt.upper())
            if key in processed:
                # Already enriched a previous occurrence — copy from earlier
                # (same rec appears in multiple packages when global)
                rec.setdefault("gap_consumers", [])
                rec.setdefault("all_consumers", [])
                rec.setdefault("cascade_chain", [])
                rec.setdefault("has_consumer_gaps", False)
                continue
            processed.add(key)

            so_raw = rec.get("sql_objects_called", "") or ""
            listed_raw = [s.strip() for s in re.split(r'[\n,;]', so_raw) if s.strip()]
            listed_stems = {s.split(".")[0].upper() for s in listed_raw if s}  # strip proc suffix

            # All script_summaries consumers of this object
            ss_consumers = consumer_map.get(tgt.upper(), [])

            # Gaps: in script_summaries but NOT already in sql_objects_called
            gap_consumers = sorted(set(
                c for c in ss_consumers
                if c.upper() not in listed_stems and c.upper() != tgt.upper()
            ))

            # Union of both sources
            all_consumers = sorted(set(listed_stems) | set(c.upper() for c in ss_consumers))

            # Cascade: gap consumers that are ALSO global targets of E./O. recs
            cascade_chain = [
                {"consumer": c, "also_targets": global_targets[c.upper()],
                 "note": "This consumer is itself a global E./O. target — sequence elimination carefully"}
                for c in gap_consumers
                if c.upper() in global_targets
            ]

            rec["gap_consumers"]     = gap_consumers
            rec["all_consumers"]     = all_consumers
            rec["cascade_chain"]     = cascade_chain
            rec["has_consumer_gaps"] = bool(gap_consumers)

    n_with_gaps = sum(
        1 for e in pkg_map.values()
        for r in e.get("recommendations", [])
        if r.get("has_consumer_gaps")
    )
    print(f"  -> Consumer gap analysis: {n_with_gaps} global E./O. recs have gap consumers")

    cascades = sum(
        len(r.get("cascade_chain", []))
        for e in pkg_map.values()
        for r in e.get("recommendations", [])
    )
    print(f"  -> Cascade chains detected: {cascades} gap consumers are ALSO global E./O. targets")

def _compute_rpt_rank(records: list[dict]) -> dict[str, int]:
    """Return {rpt_name: rank_0indexed} sorted by implementation complexity (ascending)."""
    from collections import defaultdict
    rpt_data: dict = defaultdict(lambda: {"mids": set(), "sc_local": set(), "sc_global": set(),
                                          "ph": defaultdict(int), "risks": []})
    for rec in records:
        rpts = rec.get("appears_in_rpts", [])
        rpt_cnt = rec.get("rpt_count", 1)
        for rpt in rpts:
            d = rpt_data[rpt]
            d["mids"].add(rec["master_id"])
            d["ph"][rec["phase"]] += 1
            d["risks"].append(rec.get("risk", "LOW") or "LOW")
            sql_raw = rec.get("sql_objects_called", "") or ""
            for s in _split_newline(sql_raw):
                pkg = _pkg_key(s)
                if rpt_cnt > 1: d["sc_global"].add(pkg)
                else:           d["sc_local"].add(pkg)
    scores: dict[str, float] = {}
    for rpt, d in rpt_data.items():
        local  = d["sc_local"] - d["sc_global"]
        all_sc = local | d["sc_global"]
        n_sc   = max(len(all_sc), 1)
        pct    = round(len(local) / n_sc * 100)
        mr = "HIGH" if "HIGH" in d["risks"] else "MEDIUM" if "MEDIUM" in d["risks"] else "LOW"
        rw = {"LOW":1, "MEDIUM":2, "HIGH":3}
        p3 = d["ph"].get(3, 0); p4 = d["ph"].get(4, 0)
        scores[rpt] = (len(d["mids"])*1.5 + n_sc*2 + (100-pct)*0.1 +
                       p4*3 + p3*2 + rw.get(mr,2)*5)
    ranked = sorted(scores, key=lambda r: scores[r])
    return {rpt: i for i, rpt in enumerate(ranked)}


# ── Main reader ───────────────────────────────────────────────────────────────

def read_master_xlsx(master_path: Path) -> list[dict]:
    """Read 'Master Recommendations' sheet → list of rec dicts."""
    tmp = master_path.parent / "_reg_tmp.xlsx"
    shutil.copy2(master_path, tmp)
    wb  = openpyxl.load_workbook(str(tmp), read_only=True)
    ws  = wb["Master Recommendations"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close(); tmp.unlink()

    # Row 0 = banner, Row 1 = headers
    hdrs = rows[1]
    h    = {str(v).strip(): i for i, v in enumerate(hdrs) if v}

    def _g(row, col_name: str) -> str:
        idx = h.get(col_name)
        return str(row[idx] or "").strip() if idx is not None and idx < len(row) else ""

    records = []
    for row in rows[2:]:
        if not row[0]:
            continue
        mid = _g(row, "Master ID")
        if not mid:
            continue
        try:
            phase = int(row[h.get("Phase", 1)] or 0)
        except Exception:
            phase = 0

        vs = _g(row, "Validation Status")
        if vs not in ACTIONABLE_STATUSES:
            continue
        # Note: phase is NOT used to exclude — the routing table (AGENT_ROUTING) handles
        # architectural/DBA categories (E, J, A/C Source-System) via agent=="skip".
        # Data team ✅ Consider overrides phase concerns.

        cat = _g(row, "Category")
        agent, sub_agent = _route(cat)
        if agent == "skip":
            continue

        rpts_raw   = _g(row, "Appears in RPTs")
        rpts_list  = [r.strip() for r in rpts_raw.split("\n") if r.strip()]
        try:
            rpt_count = int(row[h.get("# RPTs", 11)] or 1)
        except Exception:
            rpt_count = len(rpts_list) or 1

        try:
            hops = int(row[h.get("Verified Hop Savings", 12)] or 0)
        except Exception:
            hops = 0
        try:
            est_min = float(row[h.get("Est. Time Saved (min)", 13)] or 0.0)
        except Exception:
            est_min = 0.0

        records.append({
            "master_id":          mid,
            "phase":              phase,
            "phase_name":         _g(row, "Phase Name"),
            "category":           cat,
            "agent":              agent,
            "sub_agent":          sub_agent,
            "target_table":       _g(row, "Target Table"),
            "affected_jobs":      _split_newline(_g(row, "Affected Jobs")),
            "sql_objects_called": _g(row, "SQL Objects Called"),
            "hop_savings":        hops,
            "est_min_saved":      round(est_min, 1),
            "risk":               _g(row, "Risk"),
            "validation_status":  vs,
            "description":        _g(row, "Description"),
            "recommendation":     _g(row, "What To Do (Recommendation)"),
            "comments":           _g(row, "Comments / Proposed Hop Reduction Approach"),
            "review_by":          _g(row, "Review By"),
            "appears_in_rpts":    rpts_list,
            "rpt_count":          rpt_count,
        })

    print(f"  -> Loaded {len(records)} actionable recs from master workbook")
    return records


# ── Build registry ─────────────────────────────────────────────────────────────

def build_registry(
    records: list[dict],
    all_metadata: Path,
    table_info: dict[str, dict] | None = None,
    tidal_lookup: dict[str, dict] | None = None,
    index_info: dict[str, list] | None = None,
    consumer_map: dict[str, list] | None = None,
) -> dict:
    """Group records by package and build the full registry dict."""
    from collections import defaultdict

    rpt_rank = _compute_rpt_rank(records)

    pkg_map: dict[str, dict] = {}

    for rec in records:
        sql_raw = rec["sql_objects_called"]
        pkgs    = [_pkg_key(s) for s in _split_newline(sql_raw)] if sql_raw else ["[TIDAL-ONLY]"]
        rpts    = rec["appears_in_rpts"]
        primary_rpt = min(rpts, key=lambda r: rpt_rank.get(r, 999)) if rpts else "UNKNOWN"

        for pkg in pkgs:
            if pkg not in pkg_map:
                src = _find_source_sql(pkg, all_metadata)
                pkg_map[pkg] = {
                    "rpt":               primary_rpt,
                    "impl_rank":         rpt_rank.get(primary_rpt, 999) + 1,   # 1-based
                    "impl_tier":         (1 if rpt_rank.get(primary_rpt, 999) < 3
                                          else 2 if rpt_rank.get(primary_rpt, 999) < 7
                                          else 3),
                    "scope":             "GLOBAL" if rec["rpt_count"] > 1 else "LOCAL",
                    "source_sql":        src or f"All_Metadata/{pkg}.sql (not found)",
                    "pipeline_stage":    "raw",
                    "source_type":       "recommended",
                    "applicable_agents": [],
                    "recommendations":   [],
                    "pending_cross_impacts": [],  # populated by PL/SQL Planner when global recs flow in
                }
            entry = pkg_map[pkg]
            if rec["agent"] not in entry["applicable_agents"]:
                entry["applicable_agents"].append(rec["agent"])
            # Avoid duplicate recs (same master_id may appear via multiple proc names)
            if not any(r["master_id"] == rec["master_id"] for r in entry["recommendations"]):
                entry["recommendations"].append(rec)

    # Sort recommendations within each package by phase then priority
    # Also set standardizer_mode and ensure standardizer is in applicable_agents
    for pkg_name, entry in pkg_map.items():
        entry["recommendations"].sort(key=lambda r: (r["phase"], r["master_id"]))
        if entry.get("source_type") == "recommended":
            has_std_rec = any(r["agent"] == "standardizer" for r in entry["recommendations"])
            entry["standardizer_mode"] = "rec-driven" if has_std_rec else "auto-final-pass"
            # Standardizer always runs last on every recommended package
            if "standardizer" not in entry["applicable_agents"]:
                entry["applicable_agents"].append("standardizer")
        else:
            entry["standardizer_mode"] = "free-scan"

    # ── Add standalone scripts (All_Metadata scripts with no explicit rec) ────
    standalone_added = 0
    if all_metadata.exists():
        existing_upper = {k.upper() for k in pkg_map}
        for sql_file in sorted(all_metadata.glob("*.sql")):
            pkg_name = sql_file.stem
            if pkg_name.upper() in existing_upper:
                continue   # already covered by a recommendation-driven entry
            pkg_map[pkg_name] = {
                "rpt":               "STANDALONE",
                "impl_rank":         999,
                "impl_tier":         4,
                "scope":             "STANDALONE",
                "source_type":       "standalone",
                "source_sql":        str(sql_file).replace("\\", "/"),
                "pipeline_stage":    "raw",
                "applicable_agents": ["optimizer", "standardizer"],
                "recommendations":   [],
                "pending_cross_impacts": [],
            }
            existing_upper.add(pkg_name.upper())
            standalone_added += 1
    print(f"  -> Standalone scripts added from All_Metadata: {standalone_added}")

    # ── Enrich with table-level metadata ──────────────────────────────────────
    has_table_meta = bool(table_info or tidal_lookup)
    if has_table_meta:
        _enrich_table_metadata(pkg_map, table_info or {}, tidal_lookup or {}, index_info=index_info)
        enriched = sum(
            1 for e in pkg_map.values()
            if e.get("table_metadata", {}).get("target_tables")
            or e.get("table_metadata", {}).get("source_tables")
        )
        print(f"  -> Table metadata enriched: {enriched} packages have table info")

    # ── Enrich global E./O. recs with consumer gap analysis ─────────────────────────
    has_consumer_gaps = bool(consumer_map)
    if has_consumer_gaps:
        _enrich_consumer_gaps(pkg_map, consumer_map)
    impl_order = (
        sorted(
            [p for p in pkg_map if pkg_map[p].get("source_type") == "recommended"],
            key=lambda p: (pkg_map[p]["impl_rank"], p)
        )
        + sorted(p for p in pkg_map if pkg_map[p].get("source_type") == "standalone")
    )

    n_recommended = sum(1 for e in pkg_map.values() if e.get("source_type") == "recommended")
    n_standalone  = standalone_added

    return {
        "schema_version":              "1.5" if has_consumer_gaps else ("1.4" if has_table_meta else "1.2"),
        "generated_at":                str(date.today()),
        "total_packages":              len(pkg_map),
        "total_recommended_packages":  n_recommended,
        "total_standalone_packages":   n_standalone,
        "total_actionable_recs":       len(records),
        "table_info_source":           str(TABLE_INFO_XLSX) if has_table_meta else None,
        "index_info_source":           str(INDEX_INFO_XLSX) if (has_table_meta and index_info) else None,
        "script_summaries_source":     str(SCRIPT_SUMMARIES_JSON) if has_consumer_gaps else None,
        "implementation_order":        impl_order,
        "packages":                    pkg_map,
    }


# ── Scaffold pipeline directories ─────────────────────────────────────────────

def scaffold_pipeline(registry: dict, pipeline_root: Path, all_metadata: Path) -> None:
    """Create directory structure and copy source SQL files."""
    pipeline_root.mkdir(parents=True, exist_ok=True)
    copied = 0; missing = 0

    for pkg_name, entry in registry["packages"].items():
        rpt = entry["rpt"]
        pkg_dir = pipeline_root / rpt / pkg_name
        pkg_dir.mkdir(parents=True, exist_ok=True)

        # Copy source SQL
        src_raw = entry.get("source_sql", "")
        if src_raw and "not found" not in src_raw:
            src_path = Path(src_raw)
            if not src_path.is_absolute():
                # Relative to VDI root or All_Metadata
                src_path = all_metadata.parent / src_raw
            dest = pkg_dir / "00_source.sql"
            if not dest.exists() and src_path.exists():
                shutil.copy2(src_path, dest)
                copied += 1
            elif not src_path.exists():
                missing += 1
        else:
            missing += 1

        # Create decisions log placeholder
        decisions = pkg_dir / "decisions.md"
        if not decisions.exists():
            if entry.get("source_type") == "standalone":
                decisions.write_text(
                    f"# Pipeline Decisions — {pkg_name}\n\n"
                    f"**Type:** STANDALONE — no specific hop-reduction recommendation.\n"
                    f"Candidate for:\n"
                    f"- **Optimization pass** (agent scans for PL/SQL signals: cursors, pass-through UPDATEs, hardcoded values, GTTs, etc.)\n"
                    f"- **Standardization pass** (PL/SQL coding standards review)\n\n"
                    f"## Stage Log\n\n| Stage | Date | Decision | Notes |\n|-------|------|----------|-------|\n",
                    encoding="utf-8"
                )
            else:
                decisions.write_text(
                    f"# Pipeline Decisions — {pkg_name}\n\n"
                    f"**RPT:** {rpt}  |  **Rank:** {entry['impl_rank']}  |  "
                    f"**Tier:** {entry['impl_tier']}  |  **Scope:** {entry['scope']}\n\n"
                    f"## Recommendations\n\n"
                    + "\n".join(
                        f"- {r['master_id']} | {r['category']} | {r['validation_status']}"
                        for r in entry["recommendations"]
                    )
                    + "\n\n## Stage Log\n\n| Stage | Date | Decision | Notes |\n|-------|------|----------|-------|\n",
                    encoding="utf-8"
                )

    print(f"  -> Scaffold: {copied} source SQL files copied, {missing} not found in All_Metadata")


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Build pipeline registry from master XLSX")
    parser.add_argument("--master",        default=str(MASTER_XLSX),       help="Path to MASTER_Hop_Reduction_Recommendations.xlsx")
    parser.add_argument("--pipeline-root", default=str(DEFAULT_PIPELINE),  help="Output pipeline root directory")
    parser.add_argument("--all-metadata",  default=str(ALL_METADATA_DIR),  help="Path to All_Metadata/ folder with .sql files")
    parser.add_argument("--table-info",    default=str(TABLE_INFO_XLSX),   help="Path to Table_Info.xlsx (row/column/size metadata)")
    parser.add_argument("--index-info",    default=str(INDEX_INFO_XLSX),   help="Path to Index_Allobject.xlsx (index metadata)")
    parser.add_argument("--tidal-graph",      default=str(TIDAL_GRAPH_JSON),      help="Path to tidal_dependency_graph.json")
    parser.add_argument("--script-summaries",  default=str(SCRIPT_SUMMARIES_JSON),  help="Path to script_summaries_all.json (consumer gap analysis)")
    parser.add_argument("--no-scaffold",       action="store_true",                help="Skip directory creation and SQL file copying")
    parser.add_argument("--no-table-meta",     action="store_true",                help="Skip table metadata enrichment")
    parser.add_argument("--no-consumer-gaps",  action="store_true",                help="Skip consumer gap and cascade analysis")
    args = parser.parse_args()

    master_path    = Path(args.master)
    pipeline_root  = Path(args.pipeline_root)
    all_metadata   = Path(args.all_metadata)
    table_info_path       = Path(args.table_info)
    index_info_path       = Path(args.index_info)
    tidal_graph_path      = Path(args.tidal_graph)
    script_summaries_path = Path(args.script_summaries)

    print("=" * 65)
    print("  PIPELINE REGISTRY GENERATOR")
    print("=" * 65)
    print(f"  Master XLSX    : {master_path}")
    print(f"  Pipeline root  : {pipeline_root}")
    print(f"  All_Metadata   : {all_metadata}")
    if not args.no_table_meta:
        print(f"  Table_Info     : {table_info_path}")
        print(f"  Index_Info     : {index_info_path}")
        print(f"  Tidal Graph    : {tidal_graph_path}")
    if not args.no_consumer_gaps:
        print(f"  Script Summaries: {script_summaries_path}")
    print()

    if not master_path.exists():
        print(f"ERROR: Master file not found: {master_path}")
        return

    # Load optional table metadata sources
    table_info: dict = {}
    tidal_lookup: dict = {}
    index_info: dict = {}
    consumer_map: dict = {}
    if not args.no_table_meta:
        print("[1/4] Loading table metadata sources...")
        table_info   = _load_table_info(table_info_path)
        index_info   = _load_index_info(index_info_path)
        tidal_lookup = _build_tidal_table_lookup(tidal_graph_path)
        print(f"  -> Tidal SQL-object \u2192 table mappings: {len(tidal_lookup)}")
        n_steps = 4
    else:
        n_steps = 3
    if not args.no_consumer_gaps:
        if args.no_table_meta:  # adjust step count
            print("[1/4] Loading consumer gap analysis sources...")
        consumer_map = _build_script_summary_consumer_map(script_summaries_path)
        n_steps = 4

    step = 2 if not args.no_table_meta else 1
    print(f"[{step}/{n_steps}] Reading master recommendations...")
    records = read_master_xlsx(master_path)

    step += 1
    print(f"[{step}/{n_steps}] Building registry...")
    registry = build_registry(records, all_metadata, table_info=table_info, tidal_lookup=tidal_lookup, index_info=index_info, consumer_map=consumer_map)
    print(f"  -> {registry['total_recommended_packages']} recommended packages  +  "
          f"{registry['total_standalone_packages']} standalone candidates  "
          f"= {registry['total_packages']} total")
    print(f"  -> Schema version: {registry['schema_version']}")
    print(f"  -> Recommended order (first 5): "
          + ", ".join(registry["implementation_order"][:5]))

    # Write JSON
    registry_path = pipeline_root / "pipeline_registry.json"
    pipeline_root.mkdir(parents=True, exist_ok=True)
    with open(registry_path, "w", encoding="utf-8") as f:
        json.dump(registry, f, indent=2, ensure_ascii=False)
    print(f"  -> Registry written: {registry_path}")

    if not args.no_scaffold:
        print(f"[{n_steps}/{n_steps}] Scaffolding pipeline directories...")
        scaffold_pipeline(registry, pipeline_root, all_metadata)
    else:
        print(f"[{n_steps}/{n_steps}] Skipped scaffold (--no-scaffold)")

    # Summary
    # Recommended: split rec-driven agents vs auto-final-pass standardizer
    xlsxbacked: dict = {}           # agents explicitly from a rec in the XLSX
    auto_std_count = 0              # packages where standardizer = auto-final-pass only
    std_standalone: dict = {}       # standalone agent counts
    for entry in registry["packages"].values():
        if entry.get("source_type") == "recommended":
            std_mode = entry.get("standardizer_mode", "auto-final-pass")
            for a in entry["applicable_agents"]:
                if a == "standardizer" and std_mode == "auto-final-pass":
                    auto_std_count += 1
                else:
                    xlsxbacked[a] = xlsxbacked.get(a, 0) + 1
        else:
            for a in entry["applicable_agents"]:
                std_standalone[a] = std_standalone.get(a, 0) + 1
    print()
    print("  Recommended packages — agents with XLSX-backed recs:")
    for agent, count in sorted(xlsxbacked.items()):
        print(f"    {agent:15s}: {count} package(s)")
    print(f"    {'standardizer':15s}: {auto_std_count} package(s)  [auto final-pass on all recommended, no specific XLSX rec]")
    print("  Standalone candidates (no XLSX rec, agent scans for signals):")
    for agent, count in sorted(std_standalone.items()):
        print(f"    {agent:15s}: {count} package(s)")
    print()
    print(f"Done. Pipeline registry → {registry_path}\n")


if __name__ == "__main__":
    main()
