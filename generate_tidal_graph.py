"""
Tidal Job Dependency Graph Builder & Visualizer
================================================
Reads combined_lineage_latest.csv (output of tidal_shell_combiner.py) and generates
an interactive HTML visualization with:
  - RPT table dropdown to filter the graph
  - Job nodes colored by category (RPT, FCT, DIM, MV, STG, etc.)
  - SQL object labels on each node
  - Depth-based layered layout
  - Click-to-highlight upstream/downstream chains
  - Detail panel with full job metadata

Usage:
  python generate_tidal_graph.py
"""
import csv
import json
import os
import shutil
from collections import defaultdict
from pathlib import Path

csv.field_size_limit(10 * 1024 * 1024)  # 10 MB to handle large column lists

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
COMBINED_CSV = os.path.join(OUTPUT_DIR, "combined_lineage_latest.csv")
TIDAL_FILE = os.path.join(SCRIPT_DIR, "..", "TidalDeps_April15.xlsx")
OUTPUT_HTML = os.path.join(OUTPUT_DIR, "tidal_dependency_graph.html")
COL_NORMALIZED_CSV = os.path.join(OUTPUT_DIR, "column_level_normalized.csv")


def _copy_if_locked(path):
    try:
        with open(path, 'rb') as f:
            f.read(1)
        return path
    except PermissionError:
        p = Path(path)
        copy_path = p.parent / (p.stem + "_copy" + p.suffix)
        shutil.copy2(str(p), str(copy_path))
        return str(copy_path)


def load_combined_lineage(path):
    csv.field_size_limit(10_000_000)   # SOURCE_COL can exceed the 131KB default
    rows_by_rpt = defaultdict(list)
    with open(path, 'r', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            rpt = row.get('RPT_TABLE', '')
            if rpt:
                rows_by_rpt[rpt].append(row)
    return dict(rows_by_rpt)


def load_tidal_edges():
    """Load TidalDeps to get actual job-to-job dependency edges."""
    import openpyxl
    tidal_path = _copy_if_locked(TIDAL_FILE)
    wb = openpyxl.load_workbook(str(tidal_path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = [str(h).strip() for h in rows[0]]
    dep_map = defaultdict(set)
    for row in rows[1:]:
        rec = {headers[i]: (row[i] if i < len(headers) else None) for i in range(len(headers))}
        jn = rec.get('JOB_NAME', '')
        dj = rec.get('DEPENDENT_JOB')
        if jn and dj:
            dep_map[jn].add(dj)
    return dep_map


def build_graph_per_rpt(rows_by_rpt, dep_map):
    all_graphs = {}

    for rpt_table, rows in rows_by_rpt.items():
        job_info = {}

        for row in rows:
            dep_job = row.get('DEPENDENT_JOB', '')
            root_job = row.get('ROOT_JOB', '')
            if not dep_job:
                continue

            depth_str = row.get('DEPTH', '')
            try:
                depth = int(float(depth_str)) if depth_str else None
            except (ValueError, TypeError):
                depth = None

            sql_obj = row.get('FULL_OBJECT', '') or ''
            status = row.get('LINEAGE_STATUS', '')
            category = row.get('JOB_CATEGORY', '') or 'OTHER'
            source = row.get('LINEAGE_SOURCE', '')
            cmd = row.get('CMD', '')
            shell_script = row.get('SHELL_SCRIPT', '')
            params = row.get('TIDAL_PARAMS', '')
            obj_type = row.get('SQL_OBJECT_TYPE', '')
            pattern = row.get('PATTERN_TYPE', '')
            src_table = row.get('SRC_TABLE', '') or ''
            tgt_table = row.get('TGT_TABLE', '') or ''
            pkg_name = row.get('PACKAGE_NAME', '') or ''
            proc_name = row.get('PROC_NAME', '') or ''
            parent_caller = row.get('PARENT_CALLER', '') or ''
            schema = row.get('SQL_OBJECT_SCHEMA', '') or ''
            notes = row.get('NOTES', '') or ''
            runtime_raw = row.get('Avg_runtime_Duration (Sec)', '') or ''
            try:
                avg_runtime = round(float(runtime_raw), 1) if runtime_raw.strip() else None
            except (ValueError, TypeError):
                avg_runtime = None

            if dep_job not in job_info:
                job_info[dep_job] = {
                    'category': category,
                    'depth': depth,
                    'sql_objects': [],
                    'status': status,
                    'source': source,
                    'cmd': cmd,
                    'shell_script': shell_script,
                    'params': params,
                    'parent_caller': parent_caller,
                    'schema': schema,
                    'notes': notes,
                    'avg_runtime': avg_runtime,
                    'src_tables': set(),
                    'tgt_tables': set(),
                }

            if sql_obj:
                entry = f"[{obj_type}] {sql_obj}" if obj_type else sql_obj
                if entry not in [o['label'] for o in job_info[dep_job]['sql_objects']]:
                    job_info[dep_job]['sql_objects'].append({
                        'label': entry, 'name': sql_obj,
                        'type': obj_type, 'pattern': pattern, 'status': status,
                        'package': pkg_name, 'proc': proc_name, 'schema': schema,
                    })

            # Collect SRC/TGT tables — normalise to uppercase so case-only
            # differences (e.g. gudu lowercase vs shell-script uppercase) never
            # create phantom "no consumer" findings in downstream detectors.
            if src_table:
                for t in src_table.split(','):
                    t = t.strip()
                    if t and t not in ('None', 'N/A', '-'):
                        job_info[dep_job]['src_tables'].add(t.upper())
            if tgt_table:
                for t in tgt_table.split(','):
                    t = t.strip()
                    if t and t not in ('None', 'N/A', '-'):
                        job_info[dep_job]['tgt_tables'].add(t.upper())

            if depth is not None:
                existing = job_info[dep_job]['depth']
                if existing is None or depth < existing:
                    job_info[dep_job]['depth'] = depth

            if root_job and root_job not in job_info:
                job_info[root_job] = {
                    'category': 'RPT', 'depth': 0, 'sql_objects': [],
                    'status': 'ROOT', 'source': '', 'cmd': '',
                    'shell_script': '', 'params': '',
                    'src_tables': set(), 'tgt_tables': {rpt_table},
                    'avg_runtime': None,
                }
            elif root_job and root_job in job_info:
                job_info[root_job]['tgt_tables'].add(rpt_table)

        # Build edges scoped to this RPT's jobs
        jobs_in_rpt = set(job_info.keys())
        edges = set()
        for jn in jobs_in_rpt:
            for dep in dep_map.get(jn, []):
                if dep in jobs_in_rpt:
                    edges.add((dep, jn))

        # ── Synthetic edges for virtual VIEW nodes ────────────────────────
        # VIEW nodes (LINEAGE_SOURCE=VIEW_EXPANSION) are not in Tidal, so they
        # have no entries in dep_map.  We infer their edges from table matching:
        #   - producer_job.tgt_tables ∩ view_node.src_tables → producer → VIEW
        #   - view_node.tgt_tables    ∩ consumer_job.src_tables → VIEW → consumer
        view_node_ids = {
            jn for jn, info in job_info.items()
            if info.get('source') == 'VIEW_EXPANSION'
        }
        if view_node_ids:
            # Build reverse table lookup over non-VIEW nodes only
            tbl_to_producers: dict = defaultdict(set)
            tbl_to_consumers: dict = defaultdict(set)
            for jn, info in job_info.items():
                if jn in view_node_ids:
                    continue
                for t in info.get('tgt_tables', set()):
                    tbl_to_producers[t.upper()].add(jn)
                for t in info.get('src_tables', set()):
                    tbl_to_consumers[t.upper()].add(jn)

            for view_id in view_node_ids:
                vinfo = job_info[view_id]
                # Edge: VIEW → consumer  (consumer reads the view by name)
                for tgt in vinfo.get('tgt_tables', set()):
                    for consumer in tbl_to_consumers.get(tgt.upper(), set()):
                        edges.add((view_id, consumer))
                # Edge: producer → VIEW  (producer writes one of the view's src tables)
                for src in vinfo.get('src_tables', set()):
                    for producer in tbl_to_producers.get(src.upper(), set()):
                        edges.add((producer, view_id))
                    # Also handle cascading: another VIEW node may produce this src
                    for other_view_id in view_node_ids:
                        if other_view_id == view_id:
                            continue
                        other_tgts = {t.upper() for t in job_info[other_view_id].get('tgt_tables', set())}
                        if src.upper() in other_tgts:
                            edges.add((other_view_id, view_id))

        # Populate root job's src_tables from upstream jobs' tgt_tables
        # so that table-level lineage shows edges into the RPT table
        for jn, info in job_info.items():
            if info['depth'] == 0 and not info['src_tables']:
                for dep, target in edges:
                    if target == jn and dep in job_info:
                        info['src_tables'].update(
                            job_info[dep].get('tgt_tables', set())
                        )

        nodes = []
        for jn, info in sorted(job_info.items()):
            sql_objs = info['sql_objects']
            # Only the real root (depth=0 from BFS) stays at 0.
            # Jobs with no BFS depth get -1 ("unlinked" column).
            depth_val = info['depth'] if info['depth'] is not None else -1
            nodes.append({
                'id': jn,
                'category': info['category'],
                'depth': depth_val,
                'sql_objects': sql_objs,
                'primary_sql': sql_objs[0]['name'] if sql_objs else '',
                'status': info['status'],
                'source': info['source'],
                'cmd': info['cmd'],
                'shell_script': info['shell_script'],
                'params': info['params'],
                'parent_caller': info.get('parent_caller', ''),
                'schema': info.get('schema', ''),
                'notes': info.get('notes', ''),
                'avg_runtime': info.get('avg_runtime'),
                'src_tables': sorted(info.get('src_tables', set())),
                'tgt_tables': sorted(info.get('tgt_tables', set())),
            })

        all_graphs[rpt_table] = {
            'nodes': nodes,
            'links': [{'source': s, 'target': t} for s, t in sorted(edges)],
        }

    return all_graphs


def load_col_data(path):
    """
    Read column_level_normalized.csv and compress into a per-RPT shared-string format.
    Each RPT entry: {jobs, tbls, cols, trs, rows}
    rows: list of [job_idx, src_tbl_idx, tgt_tbl_idx, src_col_idx, tgt_col_idx, tr_idx, depth]
    This compression reduces size ~6x vs raw JSON objects.
    """
    if not os.path.exists(path):
        print(f"  WARNING: {path} not found — column lineage will be disabled")
        return {}
    raw = defaultdict(list)
    with open(path, "r", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            rpt = row.get("RPT_TABLE", "")
            job = row.get("DEPENDENT_JOB", "")
            st  = row.get("SOURCE_TABLE", "")
            tt  = row.get("TARGET_TABLE", "")
            sc  = row.get("SOURCE_COLUMN", "")
            tc  = row.get("TARGET_COLUMN", "")
            tr  = row.get("TRANSFORMATION", "Direct") or "Direct"
            dep = row.get("DEPTH", "")
            if not (rpt and job and st and tt and sc and tc):
                continue
            try:
                d = int(float(dep)) if dep else -1
            except (ValueError, TypeError):
                d = -1
            raw[rpt].append((job, st, tt, sc, tc, tr, d))

    result = {}
    for rpt, records in raw.items():
        jobs_lst, tbls_lst, cols_lst, trs_lst = [], [], [], []
        jobs_idx, tbls_idx, cols_idx, trs_idx = {}, {}, {}, {}

        def _idx(val, lst, idx):
            if val not in idx:
                idx[val] = len(lst)
                lst.append(val)
            return idx[val]

        rows = []
        for (job, st, tt, sc, tc, tr, d) in records:
            rows.append([
                _idx(job, jobs_lst, jobs_idx),
                _idx(st,  tbls_lst, tbls_idx),
                _idx(tt,  tbls_lst, tbls_idx),
                _idx(sc,  cols_lst, cols_idx),
                _idx(tc,  cols_lst, cols_idx),
                _idx(tr,  trs_lst,  trs_idx),
                d,
            ])
        result[rpt] = {"jobs": jobs_lst, "tbls": tbls_lst,
                       "cols": cols_lst, "trs": trs_lst, "rows": rows}
    return result


def generate_html(all_graphs, col_data=None):
    """Build a self-contained HTML file using string concatenation (no f-strings)
    to avoid escaping issues with inlined D3 and JSON data."""

    data_json = json.dumps(all_graphs, indent=None)
    # Escape </script> sequences that might appear in inlined data
    data_json = data_json.replace("</", "<\\/")

    # Read D3.js for inline embedding
    d3_path = os.path.join(OUTPUT_DIR, "d3.v7.min.js")
    if os.path.exists(d3_path):
        with open(d3_path, "r", encoding="utf-8") as fh:
            d3_inline = fh.read()
        d3_inline = d3_inline.replace("</", "<\\/")
    else:
        d3_inline = ""
        print("  WARNING: d3.v7.min.js not found — graph will not render")

    # --- Build HTML in parts (no f-string) ---
    parts = []
    parts.append(r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Data Lineage — RPT Tables</title>
<style>
* { margin: 0; padding: 0; box-sizing: border-box; }
body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #0a0a1a; color: #e0e0e0; display: flex; height: 100vh; overflow: hidden; }

#sidebar { width: 320px; background: #12122a; border-right: 1px solid #2a2a4a; display: flex; flex-direction: column; flex-shrink: 0; }
#sidebar-header { padding: 14px 16px; border-bottom: 1px solid #2a2a4a; }
#sidebar-header h2 { font-size: 14px; color: #7b8cff; margin-bottom: 10px; }
#search-box { width: 100%; padding: 8px 10px; background: #1a1a3a; border: 1px solid #3a3a5a; color: #e0e0e0; border-radius: 6px; font-size: 13px; }
#search-box:focus { outline: none; border-color: #7b8cff; }
#rpt-select { width: 100%; padding: 8px 10px; background: #1a1a3a; border: 1px solid #3a3a5a; color: #7b8cff; border-radius: 6px; font-size: 12px; margin-bottom: 8px; cursor: pointer; font-weight: 600; }
#rpt-select:focus { outline: none; border-color: #7b8cff; }
#rpt-select option { background: #1a1a3a; color: #e0e0e0; }
#stats { padding: 8px 16px; font-size: 11px; color: #888; border-bottom: 1px solid #2a2a4a; }
#job-list { flex: 1; overflow-y: auto; padding: 4px 0; }
.job-item { padding: 6px 14px; font-size: 11px; cursor: pointer; border-left: 3px solid transparent; display: flex; align-items: flex-start; gap: 8px; }
.job-item:hover { background: #1a1a3a; }
.job-item.selected { background: #1e1e4a; border-left-color: #7b8cff; }
.job-item .dot { width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; margin-top: 3px; }
.job-item .info { overflow: hidden; }
.job-item .name { white-space: nowrap; overflow: hidden; text-overflow: ellipsis; font-weight: 500; }
.job-item .sql { font-size: 10px; color: #69db7c; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; margin-top: 1px; }
.cat-header { padding: 6px 16px; font-size: 10px; font-weight: 700; color: #7b8cff; text-transform: uppercase; letter-spacing: 1px; background: #0e0e22; position: sticky; top: 0; z-index: 1; }

#graph-container { flex: 1; position: relative; overflow: hidden; display: flex; flex-direction: column; }
svg { flex: 1; width: 100%; min-height: 0; }
.link-path { fill: none; stroke: #5a6aaa; stroke-width: 1.8; stroke-opacity: 0.45; }
.link-path.highlighted { stroke: #8b9fff; stroke-width: 3; stroke-opacity: 0.9; }
.link-path.dimmed { stroke-opacity: 0.06; }
.node-group { cursor: pointer; }
.node-group .node-box { rx: 6; ry: 6; }
.node-group .node-box.highlighted { stroke-width: 3px; stroke: #fff; }
.node-group .job-label { font-size: 9px; fill: #ddd; pointer-events: none; }
.node-group .sql-label { font-size: 8px; fill: #69db7c; pointer-events: none; font-style: italic; }
.node-group .tbl-label { font-size: 7px; fill: #74c0fc; pointer-events: none; }
.node-group .depth-badge { font-size: 7px; fill: #fff; pointer-events: none; font-weight: 700; }
.node-group .runtime-badge { font-size: 7px; fill: #ffa94d; pointer-events: none; font-weight: 600; }
.node-group .difw-badge { font-size: 7px; fill: #f59e0b; pointer-events: none; font-weight: 700; }
.node-group .disabled-badge { font-size: 7px; fill: #9ca3af; pointer-events: none; font-weight: 700; }
.depth-col-label { font-size: 11px; fill: #5a5a8a; font-weight: 600; text-anchor: middle; }

#detail-panel { position: absolute; top: 12px; right: 12px; width: 380px; background: #15152e; border: 1px solid #2a2a4a; border-radius: 8px; padding: 16px; display: none; font-size: 12px; max-height: 85vh; overflow-y: auto; box-shadow: 0 4px 20px rgba(0,0,0,0.5); z-index: 10; }
#detail-panel h3 { color: #7b8cff; font-size: 13px; margin-bottom: 10px; word-break: break-all; }
.field { margin-bottom: 8px; }
.label { color: #888; font-size: 10px; text-transform: uppercase; letter-spacing: 0.5px; }
.value { color: #e0e0e0; margin-top: 2px; word-break: break-all; font-size: 12px; }
.sql-obj-item { padding: 4px 8px; margin: 3px 0; background: #1a2a1a; border-left: 3px solid #69db7c; border-radius: 4px; font-size: 11px; }
.sql-obj-item .sql-type { color: #888; font-size: 10px; }
.dep-item { padding: 4px 8px; margin: 2px 0; background: #1a1a3a; border-radius: 4px; cursor: pointer; font-size: 11px; }
.dep-item:hover { background: #2a2a5a; }
.close-btn { position: absolute; top: 8px; right: 12px; cursor: pointer; color: #888; font-size: 18px; }
.close-btn:hover { color: #fff; }
.status-badge { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 10px; font-weight: 600; }
.status-COMPLETE { background: #1b4332; color: #69db7c; }
.status-SQL_OBJECT_IDENTIFIED { background: #3d2e00; color: #ffd43b; }
.status-MV_REFRESH_ONLY { background: #1a2e44; color: #74c0fc; }
.status-NEEDS_INVESTIGATION { background: #3d1a1a; color: #ff6b6b; }
.status-RECURSIVE_ONLY { background: #1a3d2e; color: #8ce99a; }
.status-ROOT { background: #2e1a44; color: #da77f2; }
.status-ORCHESTRATOR_ONLY { background: #2e2e1a; color: #ffe066; }
.status-DIFW_MISSING_LINEAGE { background: #3d2a1a; color: #ffa94d; }

#legend { position: absolute; bottom: 12px; left: 12px; background: #15152eee; border: 1px solid #2a2a4a; border-radius: 8px; padding: 12px 16px; font-size: 11px; z-index: 5; }
#legend .leg-item { display: flex; align-items: center; gap: 8px; margin: 3px 0; }
#legend .leg-dot { width: 10px; height: 10px; border-radius: 2px; }

.tooltip { position: absolute; background: #1e1e3e; border: 1px solid #3a3a5a; padding: 8px 12px; border-radius: 6px; font-size: 11px; pointer-events: none; display: none; z-index: 100; max-width: 400px; word-break: break-all; }

#toolbar { position: relative; padding: 6px 14px; background: #0e0e22; border-bottom: 1px solid #2a2a4a; display: flex; gap: 10px; z-index: 10; align-items: center; flex-shrink: 0; }
.toolbar-toggle { display: flex; align-items: center; gap: 6px; background: #15152eee; border: 1px solid #2a2a4a; border-radius: 6px; padding: 6px 12px; font-size: 11px; color: #ccc; cursor: pointer; user-select: none; }
.toolbar-toggle input { accent-color: #7b8cff; cursor: pointer; }
.toolbar-toggle:hover { background: #1e1e4a; }
.toolbar-btn { background: #15152eee; border: 1px solid #2a2a4a; border-radius: 6px; padding: 6px 14px; font-size: 11px; color: #ccc; cursor: pointer; }
.toolbar-btn:hover { background: #1e1e4a; color: #fff; }

#stats-modal { display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.7); z-index: 200; justify-content: center; align-items: center; }
#stats-modal.show { display: flex; }
#stats-content { background: #15152e; border: 1px solid #2a2a4a; border-radius: 12px; padding: 24px 32px; max-width: 700px; width: 90%; max-height: 85vh; overflow-y: auto; box-shadow: 0 8px 40px rgba(0,0,0,0.6); }
#stats-content h2 { color: #7b8cff; font-size: 16px; margin-bottom: 16px; }
.stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 12px; margin-bottom: 20px; }
.stat-card { background: #1a1a3a; border: 1px solid #2a2a4a; border-radius: 8px; padding: 14px 16px; text-align: center; }
.stat-card .stat-value { font-size: 28px; font-weight: 700; color: #7b8cff; }
.stat-card .stat-label { font-size: 11px; color: #888; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.5px; }
.stats-section { margin-top: 16px; }
.stats-section h3 { color: #aaa; font-size: 13px; margin-bottom: 8px; border-bottom: 1px solid #2a2a4a; padding-bottom: 4px; }
.stats-table { width: 100%; font-size: 12px; border-collapse: collapse; }
.stats-table th { text-align: left; color: #888; font-size: 10px; text-transform: uppercase; padding: 6px 8px; border-bottom: 1px solid #2a2a4a; }
.stats-table td { padding: 5px 8px; border-bottom: 1px solid #1a1a3a; }
.stats-table td:last-child { text-align: right; color: #7b8cff; font-weight: 600; }
.stats-close { position: absolute; top: 16px; right: 20px; cursor: pointer; color: #888; font-size: 22px; }
.stats-close:hover { color: #fff; }

/* ── Column Lineage Mode ── */
#col-bar { display:none; padding:8px 14px 0; background:#0a0a1a; border-bottom:1px solid #2a2a4a; position:relative; }
#col-bar.active { display:block; }
#col-input-wrap { display:flex; gap:8px; align-items:center; }
#col-input { flex:1; padding:7px 12px; background:#1a1a3a; border:1px solid #3a3a5a; color:#e0e0e0; border-radius:6px; font-size:13px; }
#col-input:focus { outline:none; border-color:#ffd43b; }
.col-clear-btn { padding:5px 10px; background:#2a2a3a; border:1px solid #3a3a5a; color:#aaa; border-radius:6px; cursor:pointer; font-size:11px; }
.col-clear-btn:hover { background:#3a3a5a; color:#fff; }
#col-ac { position:absolute; top:100%; left:14px; right:14px; background:#1a1a3a; border:1px solid #3a3a5a; border-radius:0 0 6px 6px; z-index:50; max-height:200px; overflow-y:auto; display:none; }
.col-ac-item { padding:6px 12px; font-size:12px; cursor:pointer; border-bottom:1px solid #1a1a2a; }
.col-ac-item:hover,.col-ac-item.ac-sel { background:#2a2a5a; }
.col-ac-more { padding:5px 12px; font-size:10px; color:#888; }
#col-flow { display:none; padding:10px 14px; background:#060616; border-bottom:2px solid #ffd43b44; overflow-x:auto; max-height:180px; }
#col-flow.active { display:block; }
.cf-title { font-size:10px; color:#ffd43b; font-weight:700; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:8px; }
.cf-chain { display:flex; align-items:center; gap:0; min-width:max-content; flex-wrap:nowrap; }
.cf-node { background:#1a2a4a; border:1px solid #3a5a8a; border-radius:6px; padding:5px 10px; font-size:10px; text-align:center; min-width:130px; cursor:pointer; flex-shrink:0; }
.cf-node:hover { border-color:#ffd43b; background:#1e2e5a; }
.cf-node.cf-src { border-color:#20c997; }
.cf-node.cf-rpt { border-color:#ff6b6b; background:#2a1a1a; }
.cf-node .cf-tbl { color:#74c0fc; font-size:9px; font-weight:600; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
.cf-node .cf-col { color:#e0e0e0; font-weight:700; margin-top:2px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
.cf-edge { display:flex; flex-direction:column; align-items:center; padding:0 6px; flex-shrink:0; }
.cf-ej { background:#2a1a3a; border:1px solid #5a3a7a; border-radius:4px; padding:2px 7px; font-size:8px; color:#da77f2; cursor:pointer; white-space:nowrap; max-width:110px; overflow:hidden; text-overflow:ellipsis; }
.cf-ej:hover { border-color:#ffd43b; }
.cf-tr { font-size:8px; color:#888; margin-top:2px; white-space:nowrap; }
.cf-arr { color:#3a4a8a; font-size:12px; margin:2px 0; }
.cf-nodata { color:#888; font-size:11px; padding:6px 0; }
/* Job node highlight in column mode */
.node-group.col-hi .node-box { stroke:#ffd43b !important; stroke-width:3px !important; }
.node-group.col-dim { opacity:0.18; pointer-events:none; }
.col-flow-edge { stroke:#ffd43b !important; stroke-width:3 !important; stroke-opacity:0.85 !important; }
.node-group.col-dim-tbl { opacity:0.18; }
</style>
</head>
<body>

<div id="sidebar">
  <div id="sidebar-header">
    <h2 id="sidebar-title">Job Lineage</h2>
    <select id="rpt-select"></select>
    <input type="text" id="search-box" placeholder="Search jobs or SQL objects...">
  </div>
  <div id="stats"></div>
  <div id="job-list"></div>
</div>

<div id="graph-container">
  <div id="toolbar">
    <label class="toolbar-toggle"><input type="checkbox" id="toggle-unlinked"> Show Unlinked Nodes</label>
    <label class="toolbar-toggle"><input type="checkbox" id="toggle-table-view"> Table Lineage</label>
    <button class="toolbar-btn" id="btn-stats">Stats</button>
    <button class="toolbar-btn" id="btn-col" onclick="toggleColMode()">&#9670; Column Lineage</button>
    <button class="toolbar-btn" id="btn-difw" onclick="cycleDifwMode()">&#11041; DIFW Loads</button>
    <button class="toolbar-btn" id="btn-disabled" onclick="cycleDisabledMode()">&#8856; Disabled Jobs</button>
  </div>
  <div id="col-bar">
    <div id="col-input-wrap">
      <input id="col-input" autocomplete="off" placeholder="Search column name (e.g. CLAIM_ID_R)..." oninput="onColInput(this.value)" onkeydown="onColKey(event)">
      <button class="col-clear-btn" onclick="clearColMode()">&#10005; Clear</button>
    </div>
    <div id="col-ac"></div>
  </div>
  <div id="col-flow"></div>
  <svg id="graph-svg"></svg>
  <div id="stats-modal" onclick="if(event.target===this)this.classList.remove('show')">
    <div id="stats-content" style="position:relative"></div>
  </div>
  <div id="detail-panel">
    <span class="close-btn" onclick="document.getElementById('detail-panel').style.display='none'">&times;</span>
    <div id="detail-content"></div>
  </div>
  <div id="legend"></div>
  <div class="tooltip" id="tooltip"></div>
</div>

<script>
""")
    # Inline D3
    parts.append(d3_inline)
    parts.append("\n</" + "script>\n<script>\n")

    # Data
    parts.append("var ALL_GRAPHS = ")
    parts.append(data_json)
    parts.append(";\n")

    # Column lineage data (compressed per-RPT shared-string format)
    if col_data:
        col_json = json.dumps(col_data, separators=(",", ":")).replace("</", "<\\/")
        parts.append("var COL_DATA = ")
        parts.append(col_json)
        parts.append(";\n")
    else:
        parts.append("var COL_DATA = {};\n")

    # Application JS (raw string, no f-string escaping needed)
    parts.append(r"""
var COLORS = {
  RPT: "#ff6b6b", FCT: "#ffa94d", DIM: "#69db7c", STG: "#74c0fc",
  MV: "#da77f2", REF: "#ffd43b", MONTH_END: "#868e96",
  SOURCE: "#20c997", OTHER: "#495057"
};
var CATEGORY_LABELS = {
  RPT: "Report Tables", FCT: "Fact Tables", DIM: "Dimension Tables",
  STG: "Staging Tables", MV: "Materialized Views", REF: "Reference Tables",
  MONTH_END: "Month-End Jobs", SOURCE: "Source/Extract", OTHER: "Other"
};

var rptKeys = Object.keys(ALL_GRAPHS).sort();
var rptTable = rptKeys[0];
var currentGraph = null, nodeMap = new Map();
var upstreamMap = new Map(), downstreamMap = new Map();

var svg = d3.select("#graph-svg");
var gRoot = svg.append("g");
var zoom = d3.zoom().scaleExtent([0.05, 5]).on("zoom", function(e){ gRoot.attr("transform", e.transform); });
svg.call(zoom);

var showUnlinked = false;
var tableViewMode = false;
var tableGraph = null;
var statsData = null;

/* ---- Fixed horizontal layout constants ---- */
var COL_WIDTH  = 320;   /* horizontal spacing between depth columns */
var ROW_HEIGHT = 72;    /* vertical spacing between nodes in a column */
var NODE_W     = 260;   /* node box width */
var NODE_H     = 58;    /* node box height */
var PAD_LEFT   = 60;
var PAD_TOP    = 50;

function loadGraph(rpt) {
  var data = ALL_GRAPHS[rpt];
  if (!data) { console.error("No data for", rpt); return; }
  currentGraph = JSON.parse(JSON.stringify(data));
  nodeMap = new Map();
  currentGraph.nodes.forEach(function(n){ nodeMap.set(n.id, n); });
  upstreamMap = new Map(); downstreamMap = new Map();
  currentGraph.links.forEach(function(l){
    var s = typeof l.source === 'object' ? l.source.id : l.source;
    var t = typeof l.target === 'object' ? l.target.id : l.target;
    if (!upstreamMap.has(t)) upstreamMap.set(t, []);
    upstreamMap.get(t).push(s);
    if (!downstreamMap.has(s)) downstreamMap.set(s, []);
    downstreamMap.get(s).push(t);
  });
  rptTable = rpt;
  buildTableGraph();
  computeStats(rpt);
  document.getElementById("sidebar-title").textContent = rpt + " \u2014 Job Lineage";
  document.title = "Data Lineage \u2014 " + rpt;
  document.getElementById("stats").textContent =
    currentGraph.nodes.length + " jobs \u00b7 " + currentGraph.links.length + " dependencies";
  renderSidebar();
  if (tableViewMode) { renderTableView(); } else { renderGraph(); }
}

function renderSidebar(filter) {
  var list = document.getElementById("job-list");
  list.innerHTML = "";
  if (!currentGraph) return;
  var groups = {};
  currentGraph.nodes.forEach(function(n){
    if (filter) {
      var f = filter.toLowerCase();
      if (n.id.toLowerCase().indexOf(f) === -1 &&
          !n.sql_objects.some(function(o){ return o.name.toLowerCase().indexOf(f) !== -1; })) return;
    }
    var cat = n.category || "OTHER";
    if (!groups[cat]) groups[cat] = [];
    groups[cat].push(n);
  });
  ["RPT","FCT","MV","DIM","STG","REF","SOURCE","MONTH_END","OTHER"].forEach(function(cat){
    if (!groups[cat] || groups[cat].length === 0) return;
    var header = document.createElement("div");
    header.className = "cat-header";
    header.textContent = (CATEGORY_LABELS[cat]||cat) + " (" + groups[cat].length + ")";
    list.appendChild(header);
    groups[cat].sort(function(a,b){ return (a.depth||0)-(b.depth||0) || a.id.localeCompare(b.id); }).forEach(function(n){
      var item = document.createElement("div");
      item.className = "job-item"; item.dataset.id = n.id;
      var sql = n.primary_sql || '';
      item.innerHTML = '<span class="dot" style="background:' + (COLORS[n.category]||COLORS.OTHER) + '"></span>' +
        '<div class="info"><div class="name" title="' + n.id + '">' + n.id + '</div>' +
        (sql ? '<div class="sql" title="' + sql + '">\u2192 ' + sql + '</div>' : '') + '</div>';
      item.onclick = function(){ selectJob(n.id); };
      list.appendChild(item);
    });
  });
}

document.getElementById("search-box").addEventListener("input", function(e){ renderSidebar(e.target.value); });

function renderGraph() {
  gRoot.selectAll("*").remove();
  var width  = document.getElementById("graph-container").clientWidth;
  var height = document.getElementById("graph-container").clientHeight;

  /* --- Filter nodes based on toggle --- */
  var visibleNodes = currentGraph.nodes;
  if (!showUnlinked) {
    visibleNodes = currentGraph.nodes.filter(function(n){ return n.depth !== -1; });
  }
  if (difwMode === 2) {
    visibleNodes = visibleNodes.filter(function(n){ return !isDifwNode(n); });
  }
  if (disabledMode === 2) {
    visibleNodes = visibleNodes.filter(function(n){ return !isDisabledNode(n); });
  }
  var visibleIds = new Set(visibleNodes.map(function(n){ return n.id; }));
  var visibleLinks = currentGraph.links.filter(function(l){
    var s = typeof l.source === 'object' ? l.source.id : l.source;
    var t = typeof l.target === 'object' ? l.target.id : l.target;
    return visibleIds.has(s) && visibleIds.has(t);
  });

  /* --- Assign fixed x,y positions by depth column --- */
  var minDepth = d3.min(visibleNodes, function(d){ return d.depth; }) || 0;
  var maxDepth = d3.max(visibleNodes, function(d){ return d.depth; }) || 1;

  /* Group nodes into depth buckets */
  var columns = {};
  visibleNodes.forEach(function(n){
    var d = n.depth;
    if (!columns[d]) columns[d] = [];
    columns[d].push(n);
  });

  /* Sort within each column by category then name for stable layout */
  var catOrder = {RPT:0, FCT:1, MV:2, DIM:3, STG:4, REF:5, SOURCE:6, MONTH_END:7, OTHER:8};
  Object.keys(columns).forEach(function(d){
    columns[d].sort(function(a,b){
      var ca = catOrder[a.category] !== undefined ? catOrder[a.category] : 9;
      var cb = catOrder[b.category] !== undefined ? catOrder[b.category] : 9;
      return ca - cb || a.id.localeCompare(b.id);
    });
  });

  /* Map each depth to a column index (0-based), handle -1 as rightmost column */
  var depths = Object.keys(columns).map(Number).sort(function(a,b){return a-b;});
  var depthToCol = {};
  var colIdx = 0;
  depths.forEach(function(d){
    if (d >= 0) { depthToCol[d] = colIdx; colIdx++; }
  });
  /* Put -1 (unlinked) as the last column on the right */
  if (columns[-1]) { depthToCol[-1] = colIdx; colIdx++; }
  var numCols = colIdx;

  /* Assign x,y to each node */
  Object.keys(columns).forEach(function(d){
    var col = columns[d];
    var ci = depthToCol[parseInt(d)];
    var x = PAD_LEFT + ci * COL_WIDTH;
    col.forEach(function(n, i){
      n._x = x + NODE_W / 2;
      n._y = PAD_TOP + 24 + i * ROW_HEIGHT + NODE_H / 2;
    });
  });

  var totalW = PAD_LEFT + numCols * COL_WIDTH + 60;
  var maxColLen = d3.max(Object.values(columns), function(c){ return c.length; }) || 1;
  var totalH = PAD_TOP + 24 + maxColLen * ROW_HEIGHT + 40;

  /* --- Depth column headers --- */
  gRoot.selectAll(".depth-col-label").data(depths).join("text")
    .attr("class","depth-col-label")
    .attr("x", function(d){ return PAD_LEFT + depthToCol[d] * COL_WIDTH + NODE_W/2; })
    .attr("y", PAD_TOP)
    .text(function(d){
      var label = d === -1 ? "Unlinked" : "Depth " + d;
      return label + " (" + columns[d].length + ")";
    });

  /* --- Draw links as horizontal bezier curves --- */
  gRoot.append("defs").append("marker")
    .attr("id","arrow").attr("viewBox","0 -4 8 8")
    .attr("refX",8).attr("refY",0).attr("markerWidth",6).attr("markerHeight",6)
    .attr("orient","auto").append("path").attr("d","M0,-3.5L8,0L0,3.5").attr("fill","#6a7acc");

  var linkData = visibleLinks.map(function(l){
    var sn = nodeMap.get(typeof l.source === 'object' ? l.source.id : l.source);
    var tn = nodeMap.get(typeof l.target === 'object' ? l.target.id : l.target);
    return { source: sn, target: tn };
  }).filter(function(l){ return l.source && l.target; });

  var linkPaths = gRoot.append("g").selectAll("path")
    .data(linkData).join("path")
    .attr("class","link-path")
    .attr("marker-end","url(#arrow)")
    .attr("d", function(l){
      var sx = l.source._x - NODE_W/2;       /* left edge of source  */
      var sy = l.source._y;
      var tx = l.target._x + NODE_W/2;       /* right edge of target */
      var ty = l.target._y;
      var mx = (sx + tx) / 2;
      /* P2 must be right of tx so the bezier always approaches from the right,
         ensuring the arrowhead points left (into the node) on every edge. */
      var P2x = Math.max(mx, tx + NODE_W * 0.35);
      return "M"+sx+","+sy+" C"+mx+","+sy+" "+P2x+","+ty+" "+tx+","+ty;
    });

  /* --- Draw node boxes --- */
  var nodeG = gRoot.append("g").selectAll("g")
    .data(visibleNodes).join("g").attr("class","node-group")
    .attr("transform", function(d){ return "translate("+d._x+","+d._y+")"; });

  nodeG.append("rect").attr("class","node-box")
    .attr("width", NODE_W).attr("height", NODE_H)
    .attr("x", -NODE_W/2).attr("y", -NODE_H/2)
    .attr("fill", function(d){ var c = d3.color(COLORS[d.category]||COLORS.OTHER); c.opacity=0.18; return c; })
    .attr("stroke", function(d){
      if (disabledMode===1&&isDisabledNode(d)) return "#6b7280";
      if (difwMode===1&&isDifwNode(d)) return "#f59e0b";
      return COLORS[d.category]||COLORS.OTHER;
    })
    .attr("stroke-width", function(d){
      return (difwMode===1&&isDifwNode(d)||(disabledMode===1&&isDisabledNode(d))) ? 2 : (d.depth===0 ? 2.5 : 1.2);
    })
    .attr("stroke-dasharray", function(d){ return (disabledMode===1&&isDisabledNode(d)) ? "5,3" : null; })
    .attr("opacity", function(d){ return (disabledMode===1&&isDisabledNode(d)) ? 0.45 : 1; });

  /* Job name label */
  nodeG.append("text").attr("class","job-label").attr("text-anchor","middle")
    .attr("dy", function(d){ return d.primary_sql ? -10 : (d.tgt_tables && d.tgt_tables.length ? -4 : 4); })
    .text(function(d){ return d.id.length > 40 ? d.id.substring(0,38)+"\u2026" : d.id; });

  /* SQL object label */
  nodeG.filter(function(d){ return d.primary_sql; }).append("text").attr("class","sql-label")
    .attr("text-anchor","middle").attr("dy", 2)
    .text(function(d){ return d.primary_sql.length > 36 ? d.primary_sql.substring(0,34)+"\u2026" : d.primary_sql; });

  /* SRC → TGT table label */
  nodeG.filter(function(d){ return (d.tgt_tables && d.tgt_tables.length > 0) || (d.src_tables && d.src_tables.length > 0); })
    .append("text").attr("class","tbl-label").attr("text-anchor","middle").attr("dy", 14)
    .text(function(d){
      var src = d.src_tables && d.src_tables.length > 0 ? d.src_tables[0].replace(/^ATOMIC\./i,'') : '';
      var tgt = d.tgt_tables && d.tgt_tables.length > 0 ? d.tgt_tables[0].replace(/^ATOMIC\./i,'') : '';
      var extra = '';
      var totalTbls = (d.src_tables ? d.src_tables.length : 0) + (d.tgt_tables ? d.tgt_tables.length : 0);
      if (totalTbls > 2) extra = ' +' + (totalTbls - 2) + ' more';
      if (src && tgt) { var t = src.substring(0,16) + ' \u2192 ' + tgt.substring(0,16) + extra; return t.length > 42 ? t.substring(0,40)+'\u2026' : t; }
      if (tgt) { var t = '\u2192 ' + tgt.substring(0,24) + extra; return t.length > 42 ? t.substring(0,40)+'\u2026' : t; }
      if (src) { var t = src.substring(0,24) + ' \u2192' + extra; return t.length > 42 ? t.substring(0,40)+'\u2026' : t; }
      return '';
    });

  /* Depth badge circle */
  nodeG.append("circle").attr("r",8)
    .attr("cx", NODE_W/2 - 6).attr("cy", -NODE_H/2 + 6)
    .attr("fill","#1a1a3a").attr("stroke","#666").attr("stroke-width",0.8);
  nodeG.append("text").attr("class","depth-badge").attr("text-anchor","middle")
    .attr("x", NODE_W/2 - 6).attr("y", -NODE_H/2 + 10)
    .text(function(d){ return d.depth != null ? d.depth : "?"; });

  /* Runtime badge (bottom-right corner) */
  nodeG.filter(function(d){ return d.avg_runtime != null; })
    .append("text").attr("class","runtime-badge").attr("text-anchor","end")
    .attr("x", NODE_W/2 - 4).attr("y", NODE_H/2 - 4)
    .text(function(d){
      var s = d.avg_runtime;
      if (s >= 3600) return (s/3600).toFixed(1) + 'h';
      if (s >= 60) return (s/60).toFixed(1) + 'm';
      return s.toFixed(0) + 's';
    });

  /* DIFW badge (bottom-left corner) — only in highlight mode */
  if (difwMode === 1) {
    nodeG.filter(function(d){ return isDifwNode(d); })
      .append("text").attr("class","difw-badge").attr("text-anchor","start")
      .attr("x", -NODE_W/2 + 4).attr("y", NODE_H/2 - 4)
      .text("\u29c6 DIFW");
  }

  /* Disabled badge (centre) — only in highlight mode */
  if (disabledMode === 1) {
    nodeG.filter(function(d){ return isDisabledNode(d); })
      .append("text").attr("class","disabled-badge").attr("text-anchor","middle")
      .attr("x", 0).attr("y", 4)
      .text("\u2296 DISABLED");
  }

  /* --- Tooltip on hover --- */
  var tooltip = document.getElementById("tooltip");
  nodeG.on("mouseover",function(e,d){
    tooltip.style.display = "block";
    var sl = d.sql_objects.map(function(o){ return o.label; }).join("<br>");
    var srcTip = d.src_tables && d.src_tables.length > 0 ? "<br><br><strong>Source Tables:</strong><br>" + d.src_tables.join("<br>") : "";
    var tgtTip = d.tgt_tables && d.tgt_tables.length > 0 ? "<br><br><strong>Target Tables:</strong><br>" + d.tgt_tables.join("<br>") : "";
    var parentTip = d.parent_caller ? "<br>Parent: "+d.parent_caller : "";
    var rtTip = d.avg_runtime != null ? "<br>Avg Runtime: "+fmtRuntime(d.avg_runtime) : "";
    var difwTip = isDifwNode(d) ? "<br><span style='color:#f59e0b;font-weight:bold;'>&#11041; DIFW Load (STG \u2192 DIM/FCT via PKG_GRP_LOAD_DIFW)</span>" : "";
    var disabledTip = isDisabledNode(d) ? "<br><span style='color:#9ca3af;font-weight:bold;'>&#8856; DISABLED in TIDAL — lineage cannot be traced</span>" : "";
    tooltip.innerHTML = "<strong>"+d.id+"</strong><br>Category: "+d.category+
      "<br>Depth: "+d.depth+"<br>Status: "+d.status+
      "<br>Source: "+(d.source||"\u2014")+parentTip+rtTip+difwTip+disabledTip+
      (sl ? "<br><br><strong>SQL Objects:</strong><br>"+sl : "") + srcTip + tgtTip;
  }).on("mousemove",function(e){
    tooltip.style.left = (e.pageX+12)+"px"; tooltip.style.top = (e.pageY-12)+"px";
  }).on("mouseout",function(){ tooltip.style.display = "none"; });

  nodeG.on("click",function(e,d){ e.stopPropagation(); selectJob(d.id); });
  svg.on("click",function(){ clearSelection(); });

  /* --- Fit view to content --- */
  setTimeout(function(){
    var sc = Math.min(0.95, Math.min(width / totalW, height / totalH));
    svg.transition().duration(600).call(zoom.transform,
      d3.zoomIdentity.translate(10, 10).scale(sc));
  }, 100);

  window._linkPaths = linkPaths; window._nodeG = nodeG;
}

function getConnected(jobId) {
  var up = new Set(), down = new Set();
  var q = [jobId];
  while(q.length){ var j=q.pop(); (upstreamMap.get(j)||[]).forEach(function(u){ if(!up.has(u)){up.add(u);q.push(u);} }); }
  q = [jobId];
  while(q.length){ var j=q.pop(); (downstreamMap.get(j)||[]).forEach(function(d){ if(!down.has(d)){down.add(d);q.push(d);} }); }
  return { upstream: up, downstream: down };
}

function selectJob(jobId) {
  var conn = getConnected(jobId);
  var all = new Set([jobId, ...conn.upstream, ...conn.downstream]);
  window._nodeG.select(".node-box").classed("highlighted",function(d){return d.id===jobId;})
    .attr("opacity",function(d){return all.has(d.id)?1:0.08;});
  window._nodeG.selectAll("text").attr("opacity",function(d){return all.has(d.id)?1:0.06;});
  window._linkPaths.each(function(l){
    var sid = l.source.id, tid = l.target.id;
    var inChain = all.has(sid) && all.has(tid);
    d3.select(this).classed("highlighted", inChain).classed("dimmed", !inChain);
  });
  document.querySelectorAll(".job-item").forEach(function(el){ el.classList.toggle("selected",el.dataset.id===jobId); });
  showDetail(jobId, conn);
}

function clearSelection() {
  if(!window._nodeG) return;
  window._nodeG.select(".node-box").classed("highlighted",false).attr("opacity",1);
  window._nodeG.selectAll("text").attr("opacity",1);
  window._linkPaths.classed("highlighted",false).classed("dimmed",false);
  document.getElementById("detail-panel").style.display="none";
  document.querySelectorAll(".job-item").forEach(function(el){el.classList.remove("selected");});
}

function showDetail(jobId, conn) {
  var n = nodeMap.get(jobId);
  if(!n) return;
  var h = '<h3>'+n.id+'</h3>';
  h += '<div class="field"><span class="status-badge status-'+n.status+'">'+n.status+'</span></div>';
  h += field("Category", n.category);
  h += field("Depth", n.depth);
  h += field("Schema", n.schema || "\u2014");
  h += field("Parent Caller", n.parent_caller || "\u2014");
  h += field("Shell Script", n.shell_script || "\u2014");
  h += field("Command", n.cmd || "\u2014");
  h += field("TIDAL Params", n.params || "\u2014");
  h += field("Lineage Source", n.source || "\u2014");
  h += field("Avg Runtime", n.avg_runtime != null ? fmtRuntime(n.avg_runtime) : "\u2014");
  if(n.notes) { h += field("Notes", n.notes); }

  if(n.src_tables && n.src_tables.length > 0) {
    h += '<div class="field"><div class="label">Source Tables ('+n.src_tables.length+')</div>';
    n.src_tables.forEach(function(t){
      h += '<div style="padding:2px 8px;margin:2px 0;background:#1a2a2a;border-left:3px solid #20c997;border-radius:4px;font-size:11px;">'+t+'</div>';
    });
    h += '</div>';
  }
  if(n.tgt_tables && n.tgt_tables.length > 0) {
    h += '<div class="field"><div class="label">Target Tables ('+n.tgt_tables.length+')</div>';
    n.tgt_tables.forEach(function(t){
      h += '<div style="padding:2px 8px;margin:2px 0;background:#2a1a2a;border-left:3px solid #ff6b6b;border-radius:4px;font-size:11px;">'+t+'</div>';
    });
    h += '</div>';
  }

  if(n.sql_objects.length > 0) {
    h += '<div class="field"><div class="label">SQL Objects ('+n.sql_objects.length+')</div>';
    n.sql_objects.forEach(function(o){
      h += '<div class="sql-obj-item"><strong>'+o.name+'</strong>' +
        (o.type ? '<br><span class="sql-type">'+o.type+' \u00b7 '+o.pattern+'</span>' : '') +
        (o.package ? '<br><span class="sql-type">Pkg: '+o.package+(o.proc?' \u2192 '+o.proc:'')+'</span>' : '') +
        (o.schema ? '<br><span class="sql-type">Schema: '+o.schema+'</span>' : '') + '</div>';
    });
    h += '</div>';
  }

  var ups = [...conn.upstream].sort();
  h += '<div class="field"><div class="label">Upstream ('+ups.length+')</div>';
  ups.slice(0,30).forEach(function(u){
    var un = nodeMap.get(u); var color = un ? COLORS[un.category] : "#888";
    var sql = un && un.primary_sql ? ' \u2192 '+un.primary_sql : '';
    h += '<div class="dep-item" onclick="selectJob(\''+u+'\')" style="border-left:3px solid '+color+'">'+u+sql+'</div>';
  });
  if(ups.length>30) h += '<div class="dep-item" style="color:#888">... +'+(ups.length-30)+' more</div>';
  h += '</div>';

  var downs = [...conn.downstream].sort();
  h += '<div class="field"><div class="label">Downstream ('+downs.length+')</div>';
  downs.slice(0,30).forEach(function(d){
    var dn = nodeMap.get(d); var color = dn ? COLORS[dn.category] : "#888";
    var sql = dn && dn.primary_sql ? ' \u2192 '+dn.primary_sql : '';
    h += '<div class="dep-item" onclick="selectJob(\''+d+'\')" style="border-left:3px solid '+color+'">'+d+sql+'</div>';
  });
  if(downs.length>30) h += '<div class="dep-item" style="color:#888">... +'+(downs.length-30)+' more</div>';
  h += '</div>';

  // Column mappings (from COL_DATA if available)
  if (typeof COL_DATA !== 'undefined' && COL_DATA[rptTable]) {
    var cc = _getColCache(rptTable);
    if (cc && cc.jobToRows && cc.jobToRows[jobId]) {
      var crows = cc.jobToRows[jobId];
      var tblGrps = {};
      crows.forEach(function(r){
        var gk = r.st + '\u2192' + r.tt;
        if (!tblGrps[gk]) tblGrps[gk] = {st:r.st, tt:r.tt, pairs:[]};
        tblGrps[gk].pairs.push(r);
      });
      h += '<div class="field"><div class="label" style="color:#ffd43b;">Column Mappings (' + crows.length + ' columns)</div>';
      Object.values(tblGrps).forEach(function(g){
        h += '<div style="background:#161626;border-left:3px solid #ffd43b44;border-radius:4px;padding:6px 8px;margin:3px 0;">';
        h += '<div style="color:#74c0fc;font-size:10px;margin-bottom:4px;">' + g.st + ' \u2192 ' + g.tt + '</div>';
        var show = Math.min(g.pairs.length, 12);
        g.pairs.slice(0, show).forEach(function(p){
          var isDirect = (p.tr === 'Direct');
          h += '<div style="display:flex;gap:5px;align-items:baseline;margin:1px 0;font-size:10px;">';
          h += '<span style="color:#69db7c;">'+p.sc+'</span>';
          h += '<span style="color:#444;">&rarr;</span>';
          h += '<span style="color:#e0e0e0;">'+p.tc+'</span>';
          if (!isDirect) h += '<span style="color:#ffd43b;font-size:9px;margin-left:4px;">['+p.tr+']</span>';
          h += '</div>';
        });
        if (g.pairs.length > show)
          h += '<div style="color:#888;font-size:10px;margin-top:2px;">\u2026 +' + (g.pairs.length-show) + ' more</div>';
        h += '</div>';
      });
      h += '</div>';
    }
  }

  document.getElementById("detail-content").innerHTML = h;
  document.getElementById("detail-panel").style.display = "block";
}

function field(label, value) {
  return '<div class="field"><div class="label">'+label+'</div><div class="value">'+value+'</div></div>';
}

function fmtRuntime(sec) {
  if (sec == null) return '\u2014';
  if (sec >= 3600) return (sec/3600).toFixed(1) + ' hrs (' + Math.round(sec) + 's)';
  if (sec >= 60) return (sec/60).toFixed(1) + ' min (' + Math.round(sec) + 's)';
  return sec.toFixed(1) + ' sec';
}

// Legend
var legend = document.getElementById("legend");
var legHtml = "<strong>Legend</strong><br>";
Object.entries(CATEGORY_LABELS).forEach(function(pair){
  legHtml += '<div class="leg-item"><span class="leg-dot" style="background:'+COLORS[pair[0]]+'"></span>'+pair[1]+'</div>';
});
legHtml += '<br><div style="font-size:10px;color:#888">Badge = BFS depth from root<br>Green italic = SQL object called<br>Scroll/zoom to navigate</div>';
legend.innerHTML = legHtml;

function buildTableGraph() {
  var tblMap = {};
  var tblEdges = new Set();

  currentGraph.nodes.forEach(function(n) {
    var srcTables = n.src_tables || [];
    var tgtTables = n.tgt_tables || [];
    srcTables.forEach(function(t) {
      if (!tblMap[t]) tblMap[t] = { id: t, category: guessTableCategory(t), depths: [], jobs: new Set() };
      tblMap[t].jobs.add(n.id);
      if (n.depth != null && n.depth >= 0) tblMap[t].depths.push(n.depth + 1);
    });
    tgtTables.forEach(function(t) {
      if (!tblMap[t]) tblMap[t] = { id: t, category: guessTableCategory(t), depths: [], jobs: new Set() };
      tblMap[t].jobs.add(n.id);
      if (n.depth != null && n.depth >= 0) tblMap[t].depths.push(n.depth);
    });
    srcTables.forEach(function(s) {
      tgtTables.forEach(function(t) {
        if (s !== t) tblEdges.add(s + '|||' + t);
      });
    });
  });

  var tblNodes = [];
  Object.keys(tblMap).forEach(function(t) {
    var info = tblMap[t];
    var minD = info.depths.length > 0 ? Math.min.apply(null, info.depths) : -1;
    tblNodes.push({
      id: info.id, category: info.category, depth: minD,
      jobs: Array.from(info.jobs).sort(),
      sql_objects: [], primary_sql: '', status: '', src_tables: [], tgt_tables: [],
    });
  });
  var tblLinks = [];
  tblEdges.forEach(function(e) {
    var p = e.split('|||');
    if (tblMap[p[0]] && tblMap[p[1]]) tblLinks.push({ source: p[0], target: p[1] });
  });
  tableGraph = { nodes: tblNodes, links: tblLinks };
  tblUpstreamMap = new Map(); tblDownstreamMap = new Map();
  tblLinks.forEach(function(l){
    var s = l.source, t = l.target;
    if (!tblUpstreamMap.has(t)) tblUpstreamMap.set(t, []);
    tblUpstreamMap.get(t).push(s);
    if (!tblDownstreamMap.has(s)) tblDownstreamMap.set(s, []);
    tblDownstreamMap.get(s).push(t);
  });
}

var tblUpstreamMap = new Map(), tblDownstreamMap = new Map();

function getTableConnected(tblId) {
  var up = new Set(), down = new Set();
  var q = [tblId];
  while(q.length){ var j=q.pop(); (tblUpstreamMap.get(j)||[]).forEach(function(u){ if(!up.has(u)){up.add(u);q.push(u);} }); }
  q = [tblId];
  while(q.length){ var j=q.pop(); (tblDownstreamMap.get(j)||[]).forEach(function(d){ if(!down.has(d)){down.add(d);q.push(d);} }); }
  return { upstream: up, downstream: down };
}

function selectTable(tblId) {
  var conn = getTableConnected(tblId);
  var all = new Set([tblId, ...conn.upstream, ...conn.downstream]);
  window._tblNodeG.select(".node-box").classed("highlighted",function(d){return d.id===tblId;})
    .attr("opacity",function(d){return all.has(d.id)?1:0.08;});
  window._tblNodeG.selectAll("text").attr("opacity",function(d){return all.has(d.id)?1:0.06;});
  window._tblLinkPaths.each(function(l){
    var sid = l.source.id, tid = l.target.id;
    var inChain = all.has(sid) && all.has(tid);
    d3.select(this).classed("highlighted", inChain).classed("dimmed", !inChain);
  });
  showTableDetail(tblId, conn);
}

function clearTableSelection() {
  if(!window._tblNodeG) return;
  window._tblNodeG.select(".node-box").classed("highlighted",false).attr("opacity",1);
  window._tblNodeG.selectAll("text").attr("opacity",1);
  window._tblLinkPaths.classed("highlighted",false).classed("dimmed",false);
  document.getElementById("detail-panel").style.display="none";
}

function showTableDetail(tblId, conn) {
  var tblNodeMap = new Map();
  if (tableGraph) tableGraph.nodes.forEach(function(n){ tblNodeMap.set(n.id, n); });
  var n = tblNodeMap.get(tblId);
  if(!n) return;
  var h = '<h3>'+n.id+'</h3>';
  h += field("Category", n.category);
  h += field("Depth", n.depth);
  if(n.jobs && n.jobs.length > 0) {
    h += '<div class="field"><div class="label">Used by Jobs ('+n.jobs.length+')</div>';
    n.jobs.forEach(function(j){
      h += '<div style="padding:2px 8px;margin:2px 0;background:#1a1a2e;border-left:3px solid #6a7acc;border-radius:4px;font-size:11px;">'+j+'</div>';
    });
    h += '</div>';
  }
  var ups = Array.from(conn.upstream).sort();
  h += '<div class="field"><div class="label">Upstream Tables ('+ups.length+')</div>';
  ups.slice(0,40).forEach(function(u){
    var cat = guessTableCategory(u); var color = COLORS[cat]||COLORS.OTHER;
    h += '<div class="dep-item" onclick="selectTable(\''+u.replace(/'/g,"\\'")+'\')" style="border-left:3px solid '+color+'">'+u+'</div>';
  });
  if(ups.length>40) h += '<div class="dep-item" style="color:#888">... +'+(ups.length-40)+' more</div>';
  h += '</div>';
  var downs = Array.from(conn.downstream).sort();
  h += '<div class="field"><div class="label">Downstream Tables ('+downs.length+')</div>';
  downs.slice(0,40).forEach(function(d){
    var cat = guessTableCategory(d); var color = COLORS[cat]||COLORS.OTHER;
    h += '<div class="dep-item" onclick="selectTable(\''+d.replace(/'/g,"\\'")+'\')" style="border-left:3px solid '+color+'">'+d+'</div>';
  });
  if(downs.length>40) h += '<div class="dep-item" style="color:#888">... +'+(downs.length-40)+' more</div>';
  h += '</div>';
  var dp = document.getElementById("detail-panel");
  document.getElementById("detail-content").innerHTML = h;
  dp.style.display = "block";
}

function guessTableCategory(name) {
  var t = name.toUpperCase();
  if (t.indexOf('RPT_') === 0 || t.indexOf('.RPT_') > -1) return 'RPT';
  if (t.indexOf('FCT_') === 0 || t.indexOf('.FCT_') > -1) return 'FCT';
  if (t.indexOf('DIM_') === 0 || t.indexOf('.DIM_') > -1) return 'DIM';
  if (t.indexOf('STG_') === 0 || t.indexOf('.STG_') > -1 || t.indexOf('_STG') > -1) return 'STG';
  if (t.indexOf('_MV') > -1 || t.indexOf('MV_') === 0 || t.indexOf('.MV_') > -1) return 'MV';
  if (t.indexOf('REF_') === 0 || t.indexOf('.REF_') > -1) return 'REF';
  return 'OTHER';
}

function computeStats(rptName) {
  /* Apply the same visibility filters as renderGraph() so stats reflect what's on screen */
  var nodes = currentGraph.nodes;
  if (!showUnlinked) nodes = nodes.filter(function(n){ return n.depth !== -1; });
  if (difwMode === 2)      nodes = nodes.filter(function(n){ return !isDifwNode(n); });
  if (disabledMode === 2) nodes = nodes.filter(function(n){ return !isDisabledNode(n); });
  var visNodeIds = new Set(nodes.map(function(n){ return n.id; }));
  /* Build active-filter label for the stats header */
  var filterLabels = [];
  if (difwMode === 2)      filterLabels.push('DIFW hidden');
  if (difwMode === 1)      filterLabels.push('DIFW highlighted');
  if (disabledMode === 2) filterLabels.push('Disabled hidden');
  if (disabledMode === 1) filterLabels.push('Disabled highlighted');
  if (!showUnlinked)      filterLabels.push('Unlinked hidden');
  var uniqueJobs = new Set();
  var uniqueSqlObjs = new Set();
  var uniquePackages = new Set();
  var uniqueProcs = new Set();
  var allTables = new Set();
  var tablesByType = { RPT: new Set(), FCT: new Set(), DIM: new Set(), MV: new Set(), STG: new Set(), REF: new Set(), OTHER: new Set() };
  var objsByType = { PACKAGE: 0, PKG_PROCEDURE: 0, PROCEDURE: 0, MATERIALIZED_VIEW: 0 };
  var sourceBreakdown = {};
  var statusBreakdown = {};
  var runtimes = [];

  nodes.forEach(function(n) {
    uniqueJobs.add(n.id);
    if (n.avg_runtime != null) runtimes.push(n.avg_runtime);
    var src = n.source || 'UNKNOWN';
    sourceBreakdown[src] = (sourceBreakdown[src] || 0) + 1;
    var st = n.status || 'UNKNOWN';
    statusBreakdown[st] = (statusBreakdown[st] || 0) + 1;
    (n.sql_objects || []).forEach(function(o) {
      uniqueSqlObjs.add(o.name);
      if (o.package) uniquePackages.add(o.package);
      if (o.proc) uniqueProcs.add(o.proc);
      if (o.type && objsByType[o.type] !== undefined) objsByType[o.type]++;
    });
    (n.src_tables || []).forEach(function(t) {
      allTables.add(t);
      var cat = guessTableCategory(t);
      if (tablesByType[cat]) tablesByType[cat].add(t); else tablesByType.OTHER.add(t);
    });
    (n.tgt_tables || []).forEach(function(t) {
      allTables.add(t);
      var cat = guessTableCategory(t);
      if (tablesByType[cat]) tablesByType[cat].add(t); else tablesByType.OTHER.add(t);
    });
  });

  statsData = {
    rptName: rptName,
    totalJobs: uniqueJobs.size,
    totalSqlObjects: uniqueSqlObjs.size,
    totalPackages: uniquePackages.size,
    totalProcs: uniqueProcs.size,
    totalTables: allTables.size,
    objsByType: objsByType,
    sourceBreakdown: sourceBreakdown,
    statusBreakdown: statusBreakdown,
    tablesByType: {},
    tableNames: {},
    sqlObjectNames: Array.from(uniqueSqlObjs).sort(),
    linkedJobs: nodes.filter(function(n){ return n.depth !== -1; }).length,
    unlinkedJobs: nodes.filter(function(n){ return n.depth === -1; }).length,
    totalEdges: currentGraph.links.length,
    totalRuntime: runtimes.reduce(function(a,b){return a+b;}, 0),
    avgRuntime: runtimes.length > 0 ? runtimes.reduce(function(a,b){return a+b;}, 0) / runtimes.length : 0,
    maxRuntime: runtimes.length > 0 ? Math.max.apply(null, runtimes) : 0,
    jobsWithRuntime: runtimes.length,
    activeFilters: filterLabels,
    totalEdges: currentGraph.links.filter(function(l){
      var s=typeof l.source==='object'?l.source.id:l.source;
      var t=typeof l.target==='object'?l.target.id:l.target;
      return visNodeIds.has(s)&&visNodeIds.has(t);
    }).length,
  };
  ['RPT','FCT','DIM','MV','STG','REF','OTHER'].forEach(function(k){
    statsData.tablesByType[k] = tablesByType[k].size;
    statsData.tableNames[k] = Array.from(tablesByType[k]).sort();
  });
}

function showStatsPanel() {
  if (!statsData) return;
  var s = statsData;
  var h = '<span class="stats-close" onclick="document.getElementById(\'stats-modal\').classList.remove(\'show\')">&times;</span>';
  h += '<h2>' + s.rptName + ' \u2014 Lineage Statistics</h2>';
  h += '<div class="stats-grid">';
  h += statCard(s.totalJobs, 'TIDAL Jobs');
  h += statCard(s.totalSqlObjects, 'SQL Objects');
  h += statCard(s.totalPackages, 'Packages');
  h += statCard(s.totalProcs, 'Procedures');
  h += statCard(s.totalTables, 'Tables/Views/MVs');
  h += statCard(s.totalEdges, 'Job Dependencies');
  h += statCard(s.linkedJobs, 'Linked Jobs');
  h += statCard(s.unlinkedJobs, 'Unlinked Jobs');
  h += '</div>';

  // Runtime summary
  if (s.jobsWithRuntime > 0) {
    h += '<div class="stats-section"><h3>Avg Runtime Summary (' + s.jobsWithRuntime + ' jobs with data)</h3>';
    h += '<div class="stats-grid">';
    h += statCard(fmtRuntime(s.totalRuntime), 'Total Runtime');
    h += statCard(fmtRuntime(s.avgRuntime), 'Avg per Job');
    h += statCard(fmtRuntime(s.maxRuntime), 'Max Single Job');
    h += '</div></div>';
  }

  // Lineage Status breakdown
  h += '<div class="stats-section"><h3>Lineage Status</h3>';
  h += '<table class="stats-table"><tr><th>Status</th><th>Jobs</th></tr>';
  Object.keys(s.statusBreakdown).sort().forEach(function(st) {
    h += '<tr><td><span class="status-badge status-'+st+'">'+st+'</span></td><td>'+s.statusBreakdown[st]+'</td></tr>';
  });
  h += '</table></div>';

  // Lineage Source breakdown
  h += '<div class="stats-section"><h3>Lineage Source</h3>';
  h += '<table class="stats-table"><tr><th>Source</th><th>Jobs</th></tr>';
  Object.keys(s.sourceBreakdown).sort().forEach(function(src) {
    h += '<tr><td>'+src+'</td><td>'+s.sourceBreakdown[src]+'</td></tr>';
  });
  h += '</table></div>';

  // SQL Object Type breakdown
  h += '<div class="stats-section"><h3>SQL Object Types</h3>';
  h += '<table class="stats-table"><tr><th>Type</th><th>Count</th></tr>';
  Object.keys(s.objsByType).forEach(function(t) {
    if (s.objsByType[t] > 0) h += '<tr><td>'+t+'</td><td>'+s.objsByType[t]+'</td></tr>';
  });
  h += '</table></div>';

  h += '<div class="stats-section"><h3>Tables / Views / MVs Breakdown</h3>';
  h += '<table class="stats-table"><tr><th>Type</th><th>Count</th></tr>';
  var types = [['RPT','Report Tables'],['FCT','Fact Tables'],['DIM','Dimension Tables'],['MV','Materialized Views'],['STG','Staging Tables'],['REF','Reference Tables'],['OTHER','Other']];
  types.forEach(function(tp) {
    var cnt = s.tablesByType[tp[0]] || 0;
    if (cnt > 0) h += '<tr><td><span style="color:' + (COLORS[tp[0]]||COLORS.OTHER) + '">\u25cf</span> ' + tp[1] + '</td><td>' + cnt + '</td></tr>';
  });
  h += '</table></div>';

  types.forEach(function(tp) {
    var names = s.tableNames[tp[0]];
    if (names && names.length > 0) {
      h += '<div class="stats-section"><h3>' + tp[1] + ' (' + names.length + ')</h3>';
      h += '<div style="max-height:150px;overflow-y:auto;font-size:11px;">';
      names.forEach(function(n) {
        h += '<div style="padding:2px 8px;margin:1px 0;background:#1a1a3a;border-left:3px solid ' + (COLORS[tp[0]]||COLORS.OTHER) + ';border-radius:3px;">' + n + '</div>';
      });
      h += '</div></div>';
    }
  });

  if (s.sqlObjectNames.length > 0) {
    h += '<div class="stats-section"><h3>SQL Objects (' + s.sqlObjectNames.length + ')</h3>';
    h += '<div style="max-height:150px;overflow-y:auto;font-size:11px;">';
    s.sqlObjectNames.forEach(function(n) {
      h += '<div style="padding:2px 8px;margin:1px 0;background:#1a2a1a;border-left:3px solid #69db7c;border-radius:3px;">' + n + '</div>';
    });
    h += '</div></div>';
  }
  document.getElementById('stats-content').innerHTML = h;
  document.getElementById('stats-modal').classList.add('show');
}

function statCard(value, label) {
  return '<div class="stat-card"><div class="stat-value">' + value + '</div><div class="stat-label">' + label + '</div></div>';
}

function renderTableView() {
  gRoot.selectAll("*").remove();
  if (!tableGraph || tableGraph.nodes.length === 0) {
    gRoot.append("text").attr("x",100).attr("y",100).attr("fill","#888").attr("font-size","14px").text("No table-level data available");
    return;
  }
  var width  = document.getElementById("graph-container").clientWidth;
  var height = document.getElementById("graph-container").clientHeight;

  var tNodes = tableGraph.nodes;
  if (!showUnlinked) {
    tNodes = tNodes.filter(function(n){ return n.depth !== -1; });
  }
  var tNodeIds = new Set(tNodes.map(function(n){ return n.id; }));
  var tLinks = tableGraph.links.filter(function(l){
    return tNodeIds.has(l.source) && tNodeIds.has(l.target);
  });
  var tNodeMap = new Map();
  tNodes.forEach(function(n){ tNodeMap.set(n.id, n); });

  var columns = {};
  tNodes.forEach(function(n){
    var d = n.depth;
    if (!columns[d]) columns[d] = [];
    columns[d].push(n);
  });

  var catOrder = {RPT:0, FCT:1, MV:2, DIM:3, STG:4, REF:5, OTHER:8};
  Object.keys(columns).forEach(function(d){
    columns[d].sort(function(a,b){
      var ca = catOrder[a.category] !== undefined ? catOrder[a.category] : 9;
      var cb = catOrder[b.category] !== undefined ? catOrder[b.category] : 9;
      return ca - cb || a.id.localeCompare(b.id);
    });
  });

  var depths = Object.keys(columns).map(Number).sort(function(a,b){return a-b;});
  var depthToCol = {};
  var colIdx = 0;
  depths.forEach(function(d){ if (d >= 0) { depthToCol[d] = colIdx; colIdx++; } });
  if (columns[-1]) { depthToCol[-1] = colIdx; colIdx++; }
  var numCols = colIdx;
  var TBL_W = 240, TBL_H = 44, TBL_ROW = 56;

  Object.keys(columns).forEach(function(d){
    var col = columns[d];
    var ci = depthToCol[parseInt(d)];
    var x = PAD_LEFT + ci * COL_WIDTH;
    col.forEach(function(n, i){
      n._x = x + TBL_W / 2;
      n._y = PAD_TOP + 24 + i * TBL_ROW + TBL_H / 2;
    });
  });

  var totalW = PAD_LEFT + numCols * COL_WIDTH + 60;
  var maxColLen = d3.max(Object.values(columns), function(c){ return c.length; }) || 1;
  var totalH = PAD_TOP + 24 + maxColLen * TBL_ROW + 40;

  gRoot.selectAll(".depth-col-label").data(depths).join("text")
    .attr("class","depth-col-label")
    .attr("x", function(d){ return PAD_LEFT + depthToCol[d] * COL_WIDTH + TBL_W/2; })
    .attr("y", PAD_TOP)
    .text(function(d){
      var label = d === -1 ? "Unlinked" : "Depth " + d;
      return label + " (" + columns[d].length + ")";
    });

  gRoot.append("defs").append("marker")
    .attr("id","arrow-tbl").attr("viewBox","0 -4 8 8")
    .attr("refX",8).attr("refY",0).attr("markerWidth",6).attr("markerHeight",6)
    .attr("orient","auto").append("path").attr("d","M0,-3.5L8,0L0,3.5").attr("fill","#6a7acc");

  var linkData = tLinks.map(function(l){
    var sn = tNodeMap.get(l.source);
    var tn = tNodeMap.get(l.target);
    return { source: sn, target: tn };
  }).filter(function(l){ return l.source && l.target; });

  gRoot.append("g").selectAll("path")
    .data(linkData).join("path")
    .attr("class","link-path")
    .attr("marker-end","url(#arrow-tbl)")
    .attr("d", function(l){
      var sx = l.source._x - TBL_W/2, sy = l.source._y;
      var tx = l.target._x + TBL_W/2, ty = l.target._y;
      var mx = (sx + tx) / 2;
      var P2x = Math.max(mx, tx + TBL_W * 0.35);
      return "M"+sx+","+sy+" C"+mx+","+sy+" "+P2x+","+ty+" "+tx+","+ty;
    });

  var nodeG = gRoot.append("g").selectAll("g")
    .data(tNodes).join("g").attr("class","node-group")
    .attr("transform", function(d){ return "translate("+d._x+","+d._y+")"; });

  nodeG.append("rect").attr("class","node-box")
    .attr("width", TBL_W).attr("height", TBL_H)
    .attr("x", -TBL_W/2).attr("y", -TBL_H/2)
    .attr("fill", function(d){ var c = d3.color(COLORS[d.category]||COLORS.OTHER); c.opacity=0.22; return c; })
    .attr("stroke", function(d){ return COLORS[d.category]||COLORS.OTHER; })
    .attr("stroke-width", 1.5);

  nodeG.append("text").attr("class","job-label").attr("text-anchor","middle").attr("dy", -2)
    .attr("fill", function(d){ return COLORS[d.category]||COLORS.OTHER; })
    .style("font-weight","600").style("font-size","10px")
    .text(function(d){
      var name = d.id.replace(/^ATOMIC\./i,'');
      return name.length > 34 ? name.substring(0,32)+'\u2026' : name;
    });

  nodeG.filter(function(d){ return d.jobs && d.jobs.length > 0; })
    .append("text").attr("class","sql-label").attr("text-anchor","middle").attr("dy", 12)
    .style("font-size","8px").attr("fill","#888")
    .text(function(d){ return d.jobs.length + ' job' + (d.jobs.length > 1 ? 's' : ''); });

  var tooltip = document.getElementById("tooltip");
  nodeG.on("mouseover",function(e,d){
    tooltip.style.display = "block";
    var jobList = d.jobs ? d.jobs.join("<br>") : '';
    tooltip.innerHTML = "<strong>"+d.id+"</strong><br>Category: "+d.category+
      "<br>Depth: "+d.depth+
      (jobList ? "<br><br><strong>Used by Jobs:</strong><br>"+jobList : "");
  }).on("mousemove",function(e){
    tooltip.style.left = (e.pageX+12)+"px"; tooltip.style.top = (e.pageY-12)+"px";
  }).on("mouseout",function(){ tooltip.style.display = "none"; });

  nodeG.on("click",function(e,d){ e.stopPropagation(); selectTable(d.id); });
  svg.on("click",function(){ clearTableSelection(); });

  window._tblLinkPaths = gRoot.selectAll(".link-path");
  window._tblNodeG = nodeG;

  setTimeout(function(){
    var sc = Math.min(0.95, Math.min(width / totalW, height / totalH));
    svg.transition().duration(600).call(zoom.transform,
      d3.zoomIdentity.translate(10, 10).scale(sc));
  }, 100);
}

// Toggle event handlers
document.getElementById("toggle-unlinked").addEventListener("change", function(e){
  showUnlinked = e.target.checked;
  if (tableViewMode) { renderTableView(); } else { renderGraph(); }
});
document.getElementById("toggle-table-view").addEventListener("change", function(e){
  tableViewMode = e.target.checked;
  if (tableViewMode) { renderTableView(); } else { renderGraph(); }
});
document.getElementById("btn-stats").addEventListener("click", function(){ showStatsPanel(); });

/* ── DIFW Loads filter/highlight ── */
var difwMode = 0; // 0 = off  |  1 = highlight  |  2 = hide
function isDifwNode(n) {
  var p = (n.params || '').toUpperCase();
  if (p.indexOf('PKG_GRP_LOAD_DIFW') !== -1) return true;
  return (n.sql_objects || []).some(function(o) {
    return o.package && o.package.toUpperCase().startsWith('PKG_GRP_LOAD_DIFW');
  });
}
function cycleDifwMode() {
  difwMode = (difwMode + 1) % 3;
  var btn = document.getElementById('btn-difw');
  var labels  = ['\u29c6 DIFW Loads', '\u29c6 DIFW: Highlighted', '\u29c6 DIFW: Hidden'];
  var bgs     = ['', '#f59e0b',       '#7f1d1d'];
  var fgColors= ['', '#000',          '#fca5a5'];
  btn.textContent      = labels[difwMode];
  btn.style.background = bgs[difwMode];
  btn.style.color      = fgColors[difwMode];
  btn.style.borderColor = difwMode ? bgs[difwMode] : '';
  computeStats(rptTable);
  if (tableViewMode) { renderTableView(); } else { renderGraph(); }
}

var disabledMode = 0; // 0 = off  |  1 = highlight  |  2 = hide
function isDisabledNode(n) {
  return !!(n.notes && n.notes.toUpperCase().includes('DISABLED'));
}
function cycleDisabledMode() {
  disabledMode = (disabledMode + 1) % 3;
  var btn = document.getElementById('btn-disabled');
  var labels   = ['\u2296 Disabled Jobs', '\u2296 Disabled: Highlighted', '\u2296 Disabled: Hidden'];
  var bgs      = ['', '#374151',           '#450a0a'];
  var fgColors = ['', '#d1d5db',           '#fca5a5'];
  btn.textContent      = labels[disabledMode];
  btn.style.background = bgs[disabledMode];
  btn.style.color      = fgColors[disabledMode];
  btn.style.borderColor = disabledMode ? bgs[disabledMode] : '';
  computeStats(rptTable);
  if (tableViewMode) { renderTableView(); } else { renderGraph(); }
}

// Populate RPT dropdown
var rptSelect = document.getElementById("rpt-select");
rptKeys.forEach(function(rpt){
  var opt = document.createElement("option");
  opt.value = rpt; opt.textContent = rpt;
  rptSelect.appendChild(opt);
});
rptSelect.value = rptTable;
rptSelect.addEventListener("change", function(e){ loadGraph(e.target.value); });

// Auto-load
loadGraph(rptTable);

/* ── Column Lineage Mode ─────────────────────────────────────────────── */
var _colMode = false;
var _colCache2 = {};
var _colTblCache = {};
var _colTimer = null;
var _colAcIdx = -1;
var _expandedTbls = new Set();
var _selectedCol = null;

function _getColCache(rpt) {
  if (_colCache2[rpt] !== undefined) return _colCache2[rpt];
  var d = COL_DATA[rpt];
  if (!d || !d.rows || !d.rows.length) { _colCache2[rpt] = null; return null; }
  var tcIdx = {}, scIdx = {}, jIdx = {};
  d.rows.forEach(function(r) {
    var obj = { job:d.jobs[r[0]], st:d.tbls[r[1]], tt:d.tbls[r[2]],
                sc:d.cols[r[3]], tc:d.cols[r[4]], tr:d.trs[r[5]], depth:r[6] };
    var tc = obj.tc, sc = obj.sc;
    if (!tcIdx[tc]) tcIdx[tc] = [];
    tcIdx[tc].push(obj);
    if (sc !== tc) { if (!scIdx[sc]) scIdx[sc] = []; scIdx[sc].push(obj); }
    var j = obj.job;
    if (!jIdx[j]) jIdx[j] = [];
    jIdx[j].push({ st:obj.st, tt:obj.tt, sc:obj.sc, tc:obj.tc, tr:obj.tr, depth:obj.depth });
  });
  var colList = d.cols.slice().sort();
  _colCache2[rpt] = { tcIdx:tcIdx, scIdx:scIdx, jobToRows:jIdx, colList:colList };
  return _colCache2[rpt];
}

function toggleColMode() {
  _colMode = !_colMode;
  var btn = document.getElementById('btn-col');
  var bar = document.getElementById('col-bar');
  btn.style.background = _colMode ? '#ffd43b' : '';
  btn.style.color      = _colMode ? '#000'    : '';
  bar.classList.toggle('active', _colMode);
  if (!_colMode) {
    document.getElementById('col-flow').classList.remove('active');
    document.getElementById('col-input').value = '';
    document.getElementById('col-ac').style.display = 'none';
    _expandedTbls.clear();
    _selectedCol = null;
    if (tableViewMode) { renderTableView(); } else { renderGraph(); }
  } else {
    buildTableGraph();
    renderColumnView();
    document.getElementById('col-input').focus();
  }
}

function clearColMode() {
  document.getElementById('col-input').value = '';
  document.getElementById('col-ac').style.display = 'none';
  document.getElementById('col-flow').classList.remove('active');
  document.getElementById('col-flow').innerHTML = '';
  _selectedCol = null;
  if (_colMode) renderColumnView();
}

function onColInput(val) {
  clearTimeout(_colTimer);
  _colTimer = setTimeout(function(){ _showAc(val); }, 100);
}

function _showAc(q) {
  var ac = document.getElementById('col-ac');
  if (!q || q.length < 2) { ac.style.display='none'; return; }
  var cache = _getColCache(rptTable);
  if (!cache) { ac.style.display='none'; return; }
  var qu = q.toUpperCase();
  var hits = cache.colList.filter(function(c){ return c.toUpperCase().indexOf(qu) !== -1; });
  if (!hits.length) { ac.style.display='none'; return; }
  _colAcIdx = -1;
  var re = new RegExp('(' + q.replace(/[.*+?^${}()|[\]\\]/g,'\\$&') + ')','gi');
  ac.innerHTML = hits.slice(0,30).map(function(c,i){
    return '<div class="col-ac-item" data-col="'+c+'" onclick="selectCol(\'' + c.replace(/'/g,"\\'") + '\')">'
           + c.replace(re,'<strong>$1</strong>') + '</div>';
  }).join('') + (hits.length>30 ? '<div class="col-ac-more">...+'+(hits.length-30)+' more</div>' : '');
  ac.style.display = 'block';
}

function onColKey(e) {
  var ac = document.getElementById('col-ac');
  if (ac.style.display === 'none') return;
  var items = ac.querySelectorAll('.col-ac-item');
  if (e.key === 'ArrowDown') { _colAcIdx = Math.min(_colAcIdx+1, items.length-1); }
  else if (e.key === 'ArrowUp')  { _colAcIdx = Math.max(_colAcIdx-1, -1); }
  else if (e.key === 'Enter' && _colAcIdx >= 0) { selectCol(items[_colAcIdx].dataset.col); return; }
  else if (e.key === 'Escape') { ac.style.display='none'; return; }
  items.forEach(function(el,i){ el.classList.toggle('ac-sel', i===_colAcIdx); });
  if (_colAcIdx >= 0) items[_colAcIdx].scrollIntoView({block:'nearest'});
}

function selectCol(colName) {
  document.getElementById('col-ac').style.display = 'none';
  selectColInTable(colName);
}

function _gatherHops(colName) {
  var cache = _getColCache(rptTable);
  if (!cache) return [];
  var hops = [], seen = new Set();
  var writes = (cache.tcIdx[colName] || []).concat(cache.scIdx[colName] || []);
  writes.sort(function(a,b){ return (b.depth||0)-(a.depth||0); });
  writes.forEach(function(w){
    var k = w.job+'|'+w.st+'|'+w.tt;
    if (!seen.has(k)) { seen.add(k); hops.push(w); }
  });
  return hops;
}

function _renderColFlow(colName) {
  var fp = document.getElementById('col-flow');
  var hops = _gatherHops(colName);
  if (!hops.length) {
    fp.innerHTML = '<div class="cf-title">Column: ' + colName + '</div><div class="cf-nodata">No column lineage found for this RPT.</div>';
    fp.classList.add('active'); return;
  }
  var h = '<div class="cf-title">Flow: <strong>' + colName + '</strong>&nbsp;&middot;&nbsp;' + hops.length + ' hop(s) &nbsp;&middot;&nbsp; depth ' + (hops[0].depth||'?') + ' &rarr; ' + (hops[hops.length-1].depth||'0') + '</div>';
  h += '<div class="cf-chain">';
  var shownNodes = new Set();
  hops.forEach(function(hop) {
    var snk = hop.st+'.'+hop.sc, tnk = hop.tt+'.'+hop.tc;
    var srcIsRpt = hop.st && hop.st.toUpperCase().startsWith('RPT_');
    if (!shownNodes.has(snk)) {
      shownNodes.add(snk);
      h += '<div class="cf-node '+(srcIsRpt?'cf-rpt':'cf-src')+'" onclick="selectJob(\'' + hop.job.replace(/'/g,"\\'") + '\')"><div class="cf-tbl">' + hop.st + '</div><div class="cf-col">.' + hop.sc + '</div></div>';
    }
    var jShort = hop.job.length>30 ? hop.job.substring(0,28)+'\u2026' : hop.job;
    h += '<div class="cf-edge"><div class="cf-arr">&#9658;</div><div class="cf-ej" onclick="selectJob(\'' + hop.job.replace(/'/g,"\\'") + '\')" title="'+hop.job+'">'+jShort+'</div><div class="cf-tr">'+(hop.tr||'Direct')+'</div></div>';
    if (!shownNodes.has(tnk)) {
      shownNodes.add(tnk);
      var tgtIsRpt = hop.tt && hop.tt.toUpperCase().startsWith('RPT_');
      h += '<div class="cf-node '+(tgtIsRpt?'cf-rpt':'')+'" onclick="selectJob(\'' + hop.job.replace(/'/g,"\\'") + '\')"><div class="cf-tbl">' + hop.tt + '</div><div class="cf-col">.' + hop.tc + '</div></div>';
    }
  });
  h += '</div>';
  fp.innerHTML = h;
  fp.classList.add('active');
}

function _highlightColJobs(colName) {
  if (_colMode) return;  // column view handles its own highlighting
  if (!window._nodeG) return;
  var cache = _getColCache(rptTable);
  if (!cache) return;
  var colJobs = new Set();
  (cache.tcIdx[colName]||[]).forEach(function(r){ colJobs.add(r.job); });
  (cache.scIdx[colName]||[]).forEach(function(r){ colJobs.add(r.job); });
  window._nodeG
    .classed('col-hi',  function(d){ return colJobs.has(d.id); })
    .classed('col-dim', function(d){ return !colJobs.has(d.id); });
}

/* ── Column Table View (expandable table nodes) ── */
var _MAX_COL_SHOW = 15;
var _COL_H = 16;
var _COMPACT_H = 54;
var _TBL_W_CV = 240;

function _buildColTblData(rpt) {
  if (_colTblCache[rpt] !== undefined) return _colTblCache[rpt];
  var d = COL_DATA[rpt];
  if (!d || !d.rows || !d.rows.length) { _colTblCache[rpt] = null; return null; }
  var tblCols = {};
  d.rows.forEach(function(r) {
    var tt=d.tbls[r[2]], tc=d.cols[r[4]], st=d.tbls[r[1]], sc=d.cols[r[3]];
    if (!tblCols[tt]) tblCols[tt]=new Set(); tblCols[tt].add(tc);
    if (!tblCols[st]) tblCols[st]=new Set(); tblCols[st].add(sc);
  });
  var sorted = {};
  Object.keys(tblCols).forEach(function(t){ sorted[t]=Array.from(tblCols[t]).sort(); });
  _colTblCache[rpt] = {tblCols:sorted};
  return _colTblCache[rpt];
}

function _nodeHCV(tblId) {
  if (!_expandedTbls.has(tblId)) return _COMPACT_H;
  var ctd=_buildColTblData(rptTable);
  var cols=(ctd&&ctd.tblCols[tblId])?ctd.tblCols[tblId]:[];
  var shown=Math.min(cols.length,_MAX_COL_SHOW);
  return _COMPACT_H+(shown+(cols.length>_MAX_COL_SHOW?1:0))*_COL_H+10;
}

function renderColumnView() {
  gRoot.selectAll("*").remove();
  if (!tableGraph||!tableGraph.nodes.length) {
    gRoot.append("text").attr("x",100).attr("y",80).attr("fill","#888").attr("font-size","13px").text("No table data available.");
    return;
  }
  var width=document.getElementById('graph-container').clientWidth;
  var height=document.getElementById('graph-container').clientHeight;
  var ctd=_buildColTblData(rptTable), tblCols=ctd?ctd.tblCols:{};
  var flowTbls=new Set();
  if (_selectedCol) _gatherHops(_selectedCol).forEach(function(h){flowTbls.add(h.st);flowTbls.add(h.tt);});
  var anyFlow=flowTbls.size>0;
  var tNodes=tableGraph.nodes.filter(function(n){return showUnlinked||n.depth!==-1;});
  var tNodeMap=new Map(); tNodes.forEach(function(n){tNodeMap.set(n.id,n);});
  var tLinks=tableGraph.links.filter(function(l){return tNodeMap.has(l.source)&&tNodeMap.has(l.target);});
  var columns={};
  tNodes.forEach(function(n){if(!columns[n.depth])columns[n.depth]=[];columns[n.depth].push(n);});
  var catOrd={RPT:0,FCT:1,MV:2,DIM:3,STG:4,REF:5,OTHER:8};
  Object.keys(columns).forEach(function(d){
    columns[d].sort(function(a,b){var ca=catOrd[a.category]!==undefined?catOrd[a.category]:9,cb=catOrd[b.category]!==undefined?catOrd[b.category]:9;return ca-cb||a.id.localeCompare(b.id);});
  });
  var depths=Object.keys(columns).map(Number).sort(function(a,b){return a-b;});
  var depthToCol={},ci=0;
  depths.forEach(function(d){if(d>=0){depthToCol[d]=ci++;}});
  if(columns[-1]){depthToCol[-1]=ci++;}
  var numCols=ci, PAD_L=60, PAD_T=55, ROW_GAP=10;
  Object.keys(columns).forEach(function(d){
    var col=columns[d],cci=depthToCol[parseInt(d)],x=PAD_L+cci*COL_WIDTH,cy=PAD_T+24;
    col.forEach(function(n){var h=_nodeHCV(n.id);n._x=x+_TBL_W_CV/2;n._y=cy+h/2;cy+=h+ROW_GAP;});
  });
  var maxH=0;
  Object.keys(columns).forEach(function(d){var cy=PAD_T+24;columns[d].forEach(function(n){cy+=_nodeHCV(n.id)+ROW_GAP;});if(cy>maxH)maxH=cy;});
  var totalW=PAD_L+numCols*COL_WIDTH+60, totalH=maxH+40;
  gRoot.selectAll('.depth-col-label').data(depths).join('text')
    .attr('class','depth-col-label')
    .attr('x',function(d){return PAD_L+depthToCol[d]*COL_WIDTH+_TBL_W_CV/2;})
    .attr('y',PAD_T)
    .text(function(d){return (d===-1?'Unlinked':'Depth '+d)+' ('+columns[d].length+')';});
  var defs=gRoot.append('defs');
  ['#6a7acc','#ffd43b'].forEach(function(clr,i){
    defs.append('marker').attr('id','acv-arr'+(i?'-hl':'')).attr('viewBox','0 -4 8 8')
      .attr('refX',8).attr('refY',0).attr('markerWidth',6).attr('markerHeight',6).attr('orient','auto')
      .append('path').attr('d','M0,-3.5L8,0L0,3.5').attr('fill',clr);
  });
  var flowEdges=new Set();
  if (_selectedCol) _gatherHops(_selectedCol).forEach(function(h){flowEdges.add(h.st+'|||'+h.tt);});
  var linkData=tLinks.map(function(l){return {src:tNodeMap.get(l.source),tgt:tNodeMap.get(l.target),ek:l.source+'|||'+l.target};}).filter(function(l){return l.src&&l.tgt;});
  gRoot.append('g').selectAll('path').data(linkData).join('path')
    .attr('class',function(l){return 'link-path'+(flowEdges.has(l.ek)?' col-flow-edge':'');})
    .attr('marker-end',function(l){return 'url(#acv-arr'+(flowEdges.has(l.ek)?'-hl':'')+')';  })
    .attr('d',function(l){
      var sx=l.src._x-_TBL_W_CV/2,sy=l.src._y,tx=l.tgt._x+_TBL_W_CV/2,ty=l.tgt._y,mx=(sx+tx)/2,P2x=Math.max(mx,tx+_TBL_W_CV*0.35);
      return 'M'+sx+','+sy+' C'+mx+','+sy+' '+P2x+','+ty+' '+tx+','+ty;
    });
  var nodeG=gRoot.append('g').selectAll('g').data(tNodes).join('g')
    .attr('class',function(n){return 'node-group'+(anyFlow&&!flowTbls.has(n.id)?' col-dim-tbl':'');})
    .attr('transform',function(n){return 'translate('+n._x+','+n._y+')';});
  nodeG.append('rect').attr('class','node-box')
    .attr('width',_TBL_W_CV).attr('height',function(n){return _nodeHCV(n.id);})
    .attr('x',-_TBL_W_CV/2).attr('y',function(n){return -_nodeHCV(n.id)/2;})
    .attr('fill',function(n){var c=d3.color(COLORS[n.category]||COLORS.OTHER);c.opacity=flowTbls.has(n.id)?0.28:0.18;return c;})
    .attr('stroke',function(n){return flowTbls.has(n.id)?'#ffd43b':(COLORS[n.category]||COLORS.OTHER);})
    .attr('stroke-width',function(n){return flowTbls.has(n.id)?2.5:1.5;});
  /* Clickable header overlay */
  nodeG.append('rect')
    .attr('x',-_TBL_W_CV/2).attr('y',function(n){return -_nodeHCV(n.id)/2;})
    .attr('width',_TBL_W_CV).attr('height',_COMPACT_H)
    .attr('fill','transparent').style('cursor','pointer')
    .on('click',function(e,n){e.stopPropagation();toggleTblExpand(n.id);});
  nodeG.append('text').attr('text-anchor','middle')
    .attr('y',function(n){return -_nodeHCV(n.id)/2+15;})
    .attr('fill',function(n){return COLORS[n.category]||COLORS.OTHER;})
    .style('font-weight','600').style('font-size','10px').style('cursor','pointer')
    .text(function(n){var nm=n.id.replace(/^ATOMIC\./i,'');return nm.length>32?nm.substring(0,30)+'\u2026':nm;})
    .on('click',function(e,n){e.stopPropagation();toggleTblExpand(n.id);});
  nodeG.append('text').attr('text-anchor','middle')
    .attr('y',function(n){return -_nodeHCV(n.id)/2+30;})
    .attr('fill','#777').style('font-size','9px').style('cursor','pointer')
    .text(function(n){
      var cnt=(tblCols[n.id]||[]).length;
      return cnt>0?cnt+' cols  '+(_expandedTbls.has(n.id)?'\u25bc collapse':'\u25b6 expand'):'';
    })
    .on('click',function(e,n){e.stopPropagation();toggleTblExpand(n.id);});
  /* Column rows inside expanded nodes */
  nodeG.filter(function(n){return _expandedTbls.has(n.id);}).each(function(n){
    var el=d3.select(this), cols=tblCols[n.id]||[], nh=_nodeHCV(n.id);
    var startY=-nh/2+_COMPACT_H+2;
    cols.slice(0,_MAX_COL_SHOW).forEach(function(col,i){
      var isHl=(_selectedCol&&col===_selectedCol);
      var g=el.append('g').attr('transform','translate(0,'+(startY+i*_COL_H)+')').style('cursor','pointer')
        .on('click',function(e){e.stopPropagation();selectColInTable(col);});
      if(isHl) g.append('rect').attr('x',-_TBL_W_CV/2+3).attr('y',-1).attr('width',_TBL_W_CV-6).attr('height',_COL_H-2).attr('fill','#ffd43b22').attr('rx',2);
      g.append('text').attr('x',-_TBL_W_CV/2+8).attr('y',_COL_H-5)
        .attr('fill',isHl?'#ffd43b':'#ccc').style('font-size','9px')
        .text(col.length>32?col.substring(0,30)+'\u2026':col);
    });
    if(cols.length>_MAX_COL_SHOW)
      el.append('text').attr('x',-_TBL_W_CV/2+8).attr('y',startY+_MAX_COL_SHOW*_COL_H+10)
        .attr('fill','#666').style('font-size','9px').text('\u2026 +'+(cols.length-_MAX_COL_SHOW)+' more (use search)');
  });
  window._tblNodeG=nodeG;
  setTimeout(function(){
    var sc=Math.min(0.95,Math.min(width/totalW,height/totalH));
    svg.transition().duration(500).call(zoom.transform,d3.zoomIdentity.translate(10,10).scale(sc));
  },80);
}

function toggleTblExpand(tblId) {
  if (_expandedTbls.has(tblId)) { _expandedTbls.delete(tblId); } else { _expandedTbls.add(tblId); }
  renderColumnView();
}

function selectColInTable(colName) {
  _selectedCol = colName;
  document.getElementById('col-input').value = colName;
  document.getElementById('col-ac').style.display = 'none';
  _renderColFlow(colName);
  renderColumnView();
}
""")

    parts.append("</" + "script>\n</body>\n</html>")
    return "".join(parts)


def main():
    import sys as _sys

    # --html-only: regenerate the HTML from the existing JSON without re-reading CSVs
    if "--html-only" in _sys.argv:
        json_path = os.path.join(OUTPUT_DIR, "tidal_dependency_graph.json")
        if not os.path.exists(json_path):
            print(f"ERROR: {json_path} not found. Run without --html-only first.")
            return
        print(f"Loading existing JSON: {json_path}")
        with open(json_path, "r", encoding="utf-8") as fh:
            all_graphs = json.load(fh)
        print(f"  {len(all_graphs)} RPT graphs loaded")
        print("Loading column lineage data ...")
        col_data = load_col_data(COL_NORMALIZED_CSV)
        print("Regenerating HTML visualization ...")
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        html = generate_html(all_graphs, col_data)
        with open(OUTPUT_HTML, "w", encoding="utf-8") as fh:
            fh.write(html)
        print(f"  Output: {OUTPUT_HTML}")
        print("Done!")
        return

    print("Loading combined_lineage.csv ...")
    if not os.path.exists(COMBINED_CSV):
        print(f"  ERROR: {COMBINED_CSV} not found. Run tidal_shell_combiner.py first.")
        return

    rows_by_rpt = load_combined_lineage(COMBINED_CSV)
    print(f"  {len(rows_by_rpt)} RPT tables loaded")
    for rpt in sorted(rows_by_rpt.keys()):
        print(f"    {rpt}: {len(rows_by_rpt[rpt])} rows")

    print("\nLoading TIDAL dependency edges ...")
    dep_map = load_tidal_edges()
    print(f"  {len(dep_map)} jobs with dependencies")

    print("\nBuilding graph data per RPT table ...")
    all_graphs = build_graph_per_rpt(rows_by_rpt, dep_map)
    for rpt in sorted(all_graphs.keys()):
        g = all_graphs[rpt]
        print(f"  {rpt}: {len(g['nodes'])} nodes, {len(g['links'])} edges")

    print("\nGenerating HTML visualization ...")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("Loading column lineage data ...")
    col_data = load_col_data(COL_NORMALIZED_CSV)
    html = generate_html(all_graphs, col_data)
    with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  Output: {OUTPUT_HTML}")

    json_path = os.path.join(OUTPUT_DIR, "tidal_dependency_graph.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(all_graphs, f, indent=2)
    print(f"  JSON: {json_path}")

    print("Done!")


if __name__ == "__main__":
    main()
